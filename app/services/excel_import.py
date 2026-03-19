from __future__ import annotations

"""
Excel import service for parsing SVJ owner data.

Supports dynamic column mapping via `mapping` dict parameter.
Falls back to DEFAULT_OWNER_MAPPING when no mapping is provided.
"""
import logging

from openpyxl import load_workbook
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

from app.models.owner import Owner, OwnerType, OwnerUnit, Unit
from app.utils import build_name_with_titles, strip_diacritics

# Default column indices (0-based) — legacy layout
DEFAULT_OWNER_MAPPING = {
    "fields": {
        "unit_kn": 0,
        "building_number": 1,
        "podil_scd": 2,
        "floor_area": 3,
        "room_count": 4,
        "space_type": 5,
        "section": 6,
        "orientation_number": 7,
        "address": 8,
        "lv_number": 9,
        "ownership_type": 10,
        "first_name": 11,
        "last_name": 12,
        "title": 13,
        "birth_or_ic": 14,
        "perm_street": 15,
        "perm_district": 16,
        "perm_city": 17,
        "perm_zip": 18,
        "perm_country": 19,
        "corr_street": 20,
        "corr_district": 21,
        "corr_city": 22,
        "corr_zip": 23,
        "corr_country": 24,
        "phone_gsm": 25,
        "phone_landline": 26,
        "email_evidence": 27,
        "email_contacts": 28,
        "owner_since": 29,
        "note": 30,
    },
    "sheet_name": "Vlastnici_SVJ",
    "start_row": 2,
}


def _build_field_map(mapping: dict | None) -> dict[str, int | None]:
    """Extract field→column_index dict from mapping, filtering out None values."""
    if not mapping or "fields" not in mapping:
        return DEFAULT_OWNER_MAPPING["fields"]
    return {k: v for k, v in mapping["fields"].items() if v is not None}


def _cell(row: tuple, idx: int | None) -> str | None:
    """Safely get cell value as stripped string, or None."""
    if idx is None or idx >= len(row) or row[idx] is None:
        return None
    val = str(row[idx]).strip()
    return val if val else None


def _cell_int(row: tuple, idx: int | None) -> int | None:
    """Safely get cell value as integer."""
    raw = _cell(row, idx)
    if raw is None:
        return None
    try:
        return int(float(raw))
    except (ValueError, TypeError):
        return None


def _cell_float(row: tuple, idx: int | None) -> float | None:
    """Safely get cell value as float."""
    raw = _cell(row, idx)
    if raw is None:
        return None
    try:
        return float(raw)
    except (ValueError, TypeError):
        return None


def _normalize_name(text: str) -> str:
    """Normalize name for matching: lowercase, no diacritics, single spaces."""
    result = strip_diacritics(text)
    return " ".join(result.split())


def _is_birth_number(value: str) -> bool:
    """Check if value looks like Czech birth number (XXXXXX/XXXX or 10 digits)."""
    clean = value.replace(" ", "")
    if "/" in clean:
        parts = clean.split("/")
        return len(parts) == 2 and parts[0].isdigit() and len(parts[0]) == 6 and parts[1].isdigit()
    return clean.isdigit() and len(clean) == 10


def _is_company_id(value: str) -> bool:
    """Check if value looks like Czech IČ (8-digit number, no slash)."""
    clean = value.replace(" ", "")
    return clean.isdigit() and len(clean) == 8 and "/" not in value


def _detect_owner_type(birth_or_ic: str | None) -> OwnerType:
    """Detect owner type: legal if IČ (8-digit), otherwise physical."""
    if birth_or_ic and _is_company_id(birth_or_ic):
        return OwnerType.LEGAL_ENTITY
    return OwnerType.PHYSICAL


def _normalize_ownership_type(raw: str | None) -> str | None:
    """Normalize ownership type: ANO → SJM, keep others as-is."""
    if not raw:
        return None
    val = raw.strip()
    if val.upper() == "ANO":
        return "SJM"
    return val


def _build_name_normalized(first_name: str, last_name: str | None) -> str:
    """Build normalized name for search: 'last_name first_name' lowercased, no diacritics."""
    parts = []
    if last_name:
        parts.append(last_name)
    if first_name:
        parts.append(first_name)
    return _normalize_name(" ".join(parts))


def _owner_group_key(first_name: str | None, last_name: str | None, birth_or_ic: str | None) -> str:
    """Generate a grouping key for identifying unique owners across rows."""
    if birth_or_ic:
        clean = birth_or_ic.replace(" ", "").strip()
        if clean:
            return f"id:{clean}"
    fn = _normalize_name(first_name or "")
    ln = _normalize_name(last_name or "")
    return f"name:{ln}|{fn}"


def _describe_skip_error(row: tuple, row_idx: int, fm: dict) -> str:
    """Build a detailed error message for a skipped row."""
    unit_kn = _cell(row, fm.get("unit_kn"))
    first_name = _cell(row, fm.get("first_name"))
    last_name = _cell(row, fm.get("last_name"))

    missing = []
    if not unit_kn:
        missing.append("číslo jednotky")
    if not first_name:
        missing.append("jméno")

    present = []
    if unit_kn:
        present.append(f"jednotka={unit_kn}")
    if first_name:
        present.append(f"jméno={first_name}")
    if last_name:
        present.append(f"příjmení={last_name}")
    title = _cell(row, fm.get("title"))
    if title:
        present.append(f"titul={title}")
    birth_ic = _cell(row, fm.get("birth_or_ic"))
    if birth_ic:
        present.append(f"RČ/IČ={birth_ic}")
    ownership = _cell(row, fm.get("ownership_type"))
    if ownership:
        present.append(f"vlastnictví={ownership}")
    space_type = _cell(row, fm.get("space_type"))
    if space_type:
        present.append(f"typ={space_type}")

    msg = f"Řádek {row_idx}: chybí {', '.join(missing)}"
    if present:
        msg += f" (nalezeno: {', '.join(present)})"
    return msg


def _get_worksheet(file_path: str, sheet_name: str | None = None):
    """Open workbook and return the correct worksheet."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif "Vlastnici_SVJ" in wb.sheetnames:
        ws = wb["Vlastnici_SVJ"]
    else:
        ws = wb.active
    return wb, ws


def _parse_row(row: tuple, row_idx: int, fm: dict) -> dict | None:
    """Parse a single row into a structured dict. Returns None if row should be skipped."""
    unit_kn = _cell(row, fm.get("unit_kn"))
    first_name = _cell(row, fm.get("first_name"))

    if not unit_kn or not first_name:
        return None

    # Strip building prefix: "1098/115" → "115"
    if "/" in unit_kn:
        unit_kn = unit_kn.split("/")[-1].strip()
    # Convert to integer
    try:
        unit_kn = int(unit_kn)
    except (ValueError, TypeError):
        logger.debug("Cannot convert unit_kn '%s' to int, keeping as string", unit_kn)

    return {
        "row_idx": row_idx,
        # Unit data
        "unit_kn": unit_kn,
        "building_number": _cell(row, fm.get("building_number")),
        "podil_scd": _cell_float(row, fm.get("podil_scd")),
        "floor_area": _cell_float(row, fm.get("floor_area")),
        "room_count": _cell(row, fm.get("room_count")),
        "space_type": _cell(row, fm.get("space_type")),
        "section": _cell(row, fm.get("section")),
        "orientation_number": _cell_int(row, fm.get("orientation_number")),
        "address": _cell(row, fm.get("address")),
        "lv_number": _cell_int(row, fm.get("lv_number")),
        # Owner data
        "ownership_type": _cell(row, fm.get("ownership_type")),
        "first_name": first_name,
        "last_name": _cell(row, fm.get("last_name")),
        "title": _cell(row, fm.get("title")),
        "birth_or_ic": _cell(row, fm.get("birth_or_ic")),
        # Addresses
        "perm_street": _cell(row, fm.get("perm_street")),
        "perm_district": _cell(row, fm.get("perm_district")),
        "perm_city": _cell(row, fm.get("perm_city")),
        "perm_zip": _cell(row, fm.get("perm_zip")),
        "perm_country": _cell(row, fm.get("perm_country")),
        "corr_street": _cell(row, fm.get("corr_street")),
        "corr_district": _cell(row, fm.get("corr_district")),
        "corr_city": _cell(row, fm.get("corr_city")),
        "corr_zip": _cell(row, fm.get("corr_zip")),
        "corr_country": _cell(row, fm.get("corr_country")),
        # Contacts
        "phone_gsm": _cell(row, fm.get("phone_gsm")),
        "phone_landline": _cell(row, fm.get("phone_landline")),
        "email_evidence": _cell(row, fm.get("email_evidence")),
        "email_contacts": _cell(row, fm.get("email_contacts")),
        # Other
        "owner_since": _cell(row, fm.get("owner_since")),
        "note": _cell(row, fm.get("note")),
    }


def preview_owners_from_excel(file_path: str, mapping: dict | None = None) -> dict:
    """Parse Excel and return preview data without saving to DB."""
    m = mapping or DEFAULT_OWNER_MAPPING
    fm = _build_field_map(m)
    sheet_name = m.get("sheet_name")
    start_row = m.get("start_row", 2)

    wb, ws = _get_worksheet(file_path, sheet_name)

    owner_keys = set()
    unit_numbers = set()
    rows_processed = 0
    errors = []
    preview_rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        parsed = _parse_row(row, row_idx, fm)
        if parsed is None:
            if row and any(c is not None for c in row[:15]):
                errors.append(_describe_skip_error(row, row_idx, fm))
            continue

        rows_processed += 1
        unit_numbers.add(parsed["unit_kn"])

        key = _owner_group_key(parsed["first_name"], parsed["last_name"], parsed["birth_or_ic"])
        owner_keys.add(key)

        owner_type = _detect_owner_type(parsed["birth_or_ic"])

        last = parsed["last_name"] or ""
        first = parsed["first_name"] or ""
        preview_rows.append({
            "row": row_idx,
            "name": build_name_with_titles(parsed["title"], first, last),
            "sort_name": f"{last} {first}".strip().lower(),
            "owner_type": owner_type.value,
            "unit_number": parsed["unit_kn"],
            "building_number": parsed["building_number"] or "",
            "ownership_type_raw": _normalize_ownership_type(parsed["ownership_type"]) or "",
            "podil_scd": parsed["podil_scd"] or 0,
            "section": parsed["section"] or "",
            "email": parsed["email_evidence"] or parsed["email_contacts"] or "",
            "phone": parsed["phone_gsm"] or "",
        })

    wb.close()

    return {
        "rows_processed": rows_processed,
        "owners_count": len(owner_keys),
        "units_count": len(unit_numbers),
        "preview_rows": preview_rows,
        "errors": errors,
    }


def import_owners_from_excel(db: Session, file_path: str, mapping: dict | None = None) -> dict:
    """Parse Excel and save owners, units, and relationships to DB."""
    m = mapping or DEFAULT_OWNER_MAPPING
    fm = _build_field_map(m)
    sheet_name = m.get("sheet_name")
    start_row = m.get("start_row", 2)

    wb, ws = _get_worksheet(file_path, sheet_name)

    # First pass: collect all rows grouped by owner key
    owner_groups: dict[str, list[dict]] = {}
    rows_processed = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        parsed = _parse_row(row, row_idx, fm)
        if parsed is None:
            if row and any(c is not None for c in row[:15]):
                errors.append(_describe_skip_error(row, row_idx, fm))
            continue

        rows_processed += 1
        key = _owner_group_key(parsed["first_name"], parsed["last_name"], parsed["birth_or_ic"])
        owner_groups.setdefault(key, []).append(parsed)

    wb.close()

    # Second pass: create DB records
    owners_created = 0
    units_created = 0
    unit_cache: dict[str, Unit] = {}

    for key, rows in owner_groups.items():
        first_row = rows[0]

        # Detect owner type
        owner_type = _detect_owner_type(first_row["birth_or_ic"])

        # Parse birth number vs company ID
        birth_number = None
        company_id_val = None
        birth_or_ic = first_row["birth_or_ic"]
        if birth_or_ic:
            if _is_company_id(birth_or_ic):
                company_id_val = birth_or_ic.strip()
            elif _is_birth_number(birth_or_ic):
                birth_number = birth_or_ic.strip()
            else:
                birth_number = birth_or_ic.strip()

        # Build names
        name_with_titles = build_name_with_titles(
            first_row["title"], first_row["first_name"], first_row["last_name"]
        )
        name_normalized = _build_name_normalized(first_row["first_name"], first_row["last_name"])

        # Pick best email/phone from all rows for this owner
        email = None
        email_secondary = None
        phone = None
        phone_landline = None
        for r in rows:
            if not email and r["email_evidence"]:
                email = r["email_evidence"]
            if not email_secondary and r["email_contacts"]:
                email_secondary = r["email_contacts"]
            if not phone and r["phone_gsm"]:
                phone = r["phone_gsm"]
            if not phone_landline and r["phone_landline"]:
                phone_landline = r["phone_landline"]

        owner = Owner(
            first_name=first_row["first_name"],
            last_name=first_row["last_name"],
            title=first_row["title"],
            name_with_titles=name_with_titles,
            name_normalized=name_normalized,
            owner_type=owner_type,
            birth_number=birth_number,
            company_id=company_id_val,
            perm_street=first_row["perm_street"],
            perm_district=first_row["perm_district"],
            perm_city=first_row["perm_city"],
            perm_zip=first_row["perm_zip"],
            perm_country=first_row["perm_country"],
            corr_street=first_row["corr_street"],
            corr_district=first_row["corr_district"],
            corr_city=first_row["corr_city"],
            corr_zip=first_row["corr_zip"],
            corr_country=first_row["corr_country"],
            phone=phone,
            phone_landline=phone_landline,
            email=email,
            email_secondary=email_secondary,
            owner_since=first_row["owner_since"],
            note=first_row["note"],
        )
        db.add(owner)
        db.flush()
        owners_created += 1

        # Create units and owner-unit links
        for row_data in rows:
            unit_kn = row_data["unit_kn"]

            if unit_kn not in unit_cache:
                existing_unit = db.query(Unit).filter_by(unit_number=unit_kn).first()
                if existing_unit:
                    unit_cache[unit_kn] = existing_unit
                else:
                    unit = Unit(
                        unit_number=unit_kn,
                        building_number=row_data["building_number"],
                        podil_scd=row_data["podil_scd"],
                        floor_area=row_data["floor_area"],
                        room_count=row_data["room_count"],
                        space_type=row_data["space_type"],
                        section=row_data["section"],
                        orientation_number=row_data["orientation_number"],
                        address=row_data["address"],
                        lv_number=row_data["lv_number"],
                    )
                    db.add(unit)
                    db.flush()
                    unit_cache[unit_kn] = unit
                    units_created += 1

            unit_obj = unit_cache[unit_kn]
            votes = unit_obj.podil_scd or 0

            owner_unit = OwnerUnit(
                owner_id=owner.id,
                unit_id=unit_obj.id,
                ownership_type=_normalize_ownership_type(row_data["ownership_type"]),
                share=1.0,
                votes=votes,
                excel_row_number=row_data["row_idx"],
            )
            db.add(owner_unit)

    db.commit()

    return {
        "owners_created": owners_created,
        "units_created": units_created,
        "rows_processed": rows_processed,
        "errors": errors,
    }
