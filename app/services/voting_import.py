"""
Service for importing voting results from Excel.

Expected Excel format:
- Row 1: headers
- Row 2+: data rows with owner name, unit number, and vote columns (FOR/AGAINST values)

The mapping dict describes which columns map to which roles:
{
    "owner_col": 0,
    "unit_col": 2,
    "items": [
        {"item_id": 5, "for_col": 3, "against_col": 4},
        {"item_id": 6, "for_col": 5, "against_col": 6}
    ]
}
"""
from __future__ import annotations

import json
import re
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.owner import Owner, OwnerUnit, Unit
from app.models.voting import (
    Ballot, BallotStatus, BallotVote, Voting, VoteValue,
)
from app.utils import utcnow
from app.utils import strip_diacritics as _strip_diacritics


def _cell(row: tuple, idx: int) -> str | None:
    """Safely get cell value as stripped string, or None."""
    if idx is None or idx >= len(row) or row[idx] is None:
        return None
    val = str(row[idx]).strip()
    return val if val else None


def _cell_numeric(row: tuple, idx: int) -> float | None:
    """Safely get cell value as a number."""
    raw = _cell(row, idx)
    if raw is None:
        return None
    try:
        return float(raw)
    except (ValueError, TypeError):
        return None


def _parse_unit_number(raw: str) -> int | None:
    """Parse unit number from string, handling '1098/115' → 115 format."""
    if "/" in raw:
        raw = raw.split("/")[-1].strip()
    try:
        return int(float(raw))
    except (ValueError, TypeError):
        return None


def _parse_value_list(raw: str) -> tuple[set[str], list[tuple[str, float]]]:
    """Parse comma-separated value list into exact matches and comparison rules.

    Supports: exact values (1, ANO, YES) and comparisons (>0, <0, >=1, <=0).
    Returns (exact_values_set, comparison_rules_list).
    """
    if not raw:
        return set(), []
    exact = set()
    comparisons = []
    cmp_re = re.compile(r"^([<>]=?)\s*(-?\d+(?:\.\d+)?)$")
    for v in raw.split(","):
        v = v.strip()
        if not v:
            continue
        m = cmp_re.match(v)
        if m:
            comparisons.append((m.group(1), float(m.group(2))))
        else:
            exact.add(v.upper())
    return exact, comparisons


def _check_comparisons(num: float, comparisons: list[tuple[str, float]]) -> bool:
    """Check if a number matches any comparison rule."""
    for op, threshold in comparisons:
        if op == ">" and num > threshold:
            return True
        if op == "<" and num < threshold:
            return True
        if op == ">=" and num >= threshold:
            return True
        if op == "<=" and num <= threshold:
            return True
    return False


def _match_vote(
    raw: str | None,
    num: float | None,
    for_values: tuple[set[str], list],
    against_values: tuple[set[str], list],
    abstain_values: tuple[set[str], list] = (set(), []),
) -> str | None:
    """Match a cell value against for/against/abstain value sets. Returns 'for', 'against', 'abstain', or None."""
    for_exact, for_cmp = for_values
    against_exact, against_cmp = against_values
    abstain_exact, abstain_cmp = abstain_values

    if raw is not None:
        upper = raw.upper()
        if upper in for_exact:
            return "for"
        if upper in against_exact:
            return "against"
        if upper in abstain_exact:
            return "abstain"
    if num is not None:
        # Check numeric as string (e.g. "1" in for_values, "0" in against_values)
        num_str = str(int(num)) if num == int(num) else str(num)
        if num_str in for_exact:
            return "for"
        if num_str in against_exact:
            return "against"
        if num_str in abstain_exact:
            return "abstain"
        # Check comparison rules (e.g. >0, <0)
        if _check_comparisons(num, for_cmp):
            return "for"
        if _check_comparisons(num, against_cmp):
            return "against"
        if _check_comparisons(num, abstain_cmp):
            return "abstain"
    return None


def validate_mapping(mapping: dict) -> str | None:
    """Validate import mapping dict. Returns error message or None if valid."""
    if not isinstance(mapping, dict):
        return "Mapping musí být objekt"

    # Required integer columns
    for key in ("owner_col", "unit_col"):
        val = mapping.get(key)
        if not isinstance(val, int) or val < 0:
            return f"Chybí nebo neplatný '{key}' (musí být nezáporné celé číslo)"

    # Optional start_row
    start_row = mapping.get("start_row", 2)
    if not isinstance(start_row, int) or start_row < 1:
        return "Neplatný 'start_row' (musí být ≥ 1)"

    # Item mappings
    items = mapping.get("item_mappings") or mapping.get("items", [])
    if not isinstance(items, list) or not items:
        return "Chybí 'item_mappings' (musí být neprázdný seznam)"

    for i, item in enumerate(items):
        if not isinstance(item, dict):
            return f"item_mappings[{i}] musí být objekt"
        if not isinstance(item.get("item_id"), int):
            return f"item_mappings[{i}].item_id musí být celé číslo"
        for_col = item.get("for_col")
        against_col = item.get("against_col")
        if for_col is None and against_col is None:
            return f"item_mappings[{i}] musí mít alespoň 'for_col' nebo 'against_col'"
        if for_col is not None and (not isinstance(for_col, int) or for_col < 0):
            return f"item_mappings[{i}].for_col musí být nezáporné celé číslo"
        if against_col is not None and (not isinstance(against_col, int) or against_col < 0):
            return f"item_mappings[{i}].against_col musí být nezáporné celé číslo"

    return None


def read_excel_headers(file_path: str) -> list[str]:
    """Read first row (headers) from Excel file.

    NOTE: This is a thin wrapper for backward compatibility.
    For new code, prefer import_mapping.read_excel_headers() which supports
    sheet_name and header_row parameters.
    """
    from app.services.import_mapping import read_excel_headers as _read_headers
    return _read_headers(file_path)


def preview_voting_import(file_path: str, mapping: dict, voting: Voting, db: Session) -> dict:
    """Parse Excel with mapping and return preview without modifying DB."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    owner_col = mapping["owner_col"]
    unit_col = mapping["unit_col"]
    start_row = mapping.get("start_row", 2)
    item_mappings = mapping.get("item_mappings") or mapping.get("items", [])

    # Parse user-defined values for PRO/PROTI/ZDRŽEL SE
    for_values = _parse_value_list(mapping.get("for_values", "1, ANO, YES, X, PRO"))
    against_values = _parse_value_list(mapping.get("against_values", "0, NE, NO, PROTI"))
    abstain_values = _parse_value_list(mapping.get("abstain_values", "ZDRŽEL, ZDRŽEL SE, ABSTAIN, Z, 2"))

    matched = []
    unmatched = []
    no_match = []
    errors = []

    # Build lookup: unit_number → list of ballots
    ballot_lookup = {}
    for ballot in voting.ballots:
        seen_unit_ids = set()
        # Primary: owner's current units
        for ou in ballot.owner.current_units:
            unit_num = ou.unit.unit_number
            ballot_lookup.setdefault(unit_num, []).append(ballot)
            seen_unit_ids.add(unit_num)
        # Fallback: units_text (shared SJM ballots contain units from all members)
        if ballot.units_text:
            for part in ballot.units_text.split(","):
                part = part.strip()
                if part.isdigit():
                    unit_num = int(part)
                    if unit_num not in seen_unit_ids:
                        ballot_lookup.setdefault(unit_num, []).append(ballot)
                        seen_unit_ids.add(unit_num)

    # Track which ballots we've already seen (for dedup across rows)
    seen_ballots = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        owner_name = _cell(row, owner_col)
        unit_raw = _cell(row, unit_col)

        if not owner_name and not unit_raw:
            continue

        if not unit_raw:
            unmatched.append({
                "row": row_idx,
                "owner_name": owner_name or "",
                "unit_number": "",
                "reason": "Chybí číslo jednotky",
            })
            continue

        unit_number = _parse_unit_number(unit_raw)
        if unit_number is None:
            unmatched.append({
                "row": row_idx,
                "owner_name": owner_name or "",
                "unit_number": unit_raw,
                "reason": f"Neplatné číslo jednotky: {unit_raw}",
            })
            continue

        # Find ballot via unit number
        ballots_for_unit = ballot_lookup.get(unit_number, [])
        if not ballots_for_unit:
            unmatched.append({
                "row": row_idx,
                "owner_name": owner_name or "",
                "unit_number": unit_number,
                "reason": "Jednotka nenalezena v lístcích",
            })
            continue

        # Parse vote choices for each item (ballot-independent)
        vote_choices = {}
        raw_values = {}  # item_id → raw string from cell (for unrecognized tracking)
        for im in item_mappings:
            item_id = im["item_id"]
            for_col = im.get("for_col")
            against_col = im.get("against_col")

            # Primary column — match value against for/against/abstain sets
            if for_col is not None:
                raw = _cell(row, for_col)
                num = _cell_numeric(row, for_col)
                if raw is not None:
                    raw_values[item_id] = raw
                result = _match_vote(raw, num, for_values, against_values, abstain_values)
                if result:
                    vote_choices[item_id] = result

            # Secondary column (PROTI zvlášť) — only if primary didn't match
            if against_col is not None and item_id not in vote_choices:
                raw = _cell(row, against_col)
                num = _cell_numeric(row, against_col)
                if raw is not None and item_id not in raw_values:
                    raw_values[item_id] = raw
                # In against column, for_values → PROTI, against_values → PRO (inverted)
                result = _match_vote(raw, num, for_values, against_values, abstain_values)
                if result == "for":
                    vote_choices[item_id] = "against"
                elif result == "against":
                    vote_choices[item_id] = "for"
                elif result == "abstain":
                    vote_choices[item_id] = "abstain"

        # Detect unrecognized: items with raw values but no match
        unrecognized = {
            item_id: raw_values[item_id]
            for item_id in raw_values
            if item_id not in vote_choices
        }

        # If ALL items have raw values but NONE matched → no_match
        if raw_values and not vote_choices:
            for ballot in ballots_for_unit:
                if ballot.id not in seen_ballots:
                    no_match.append({
                        "row": row_idx,
                        "owner_name": owner_name or ballot.owner.display_name,
                        "unit_number": unit_number,
                        "ballot_id": ballot.id,
                        "raw_values": raw_values,
                    })
            continue

        # Disambiguate when multiple ballots share the same unit
        if len(ballots_for_unit) > 1 and owner_name:
            owner_norm = _strip_diacritics(owner_name)
            narrowed = []
            for b in ballots_for_unit:
                # Check if ballot owner's name appears IN the Excel name
                name = b.owner.name_normalized or ""
                if name and name in owner_norm:
                    narrowed.append(b)
            if narrowed:
                # Re-add SJM partners excluded by name disambiguation
                # Only on units with exactly 2 SJM owners (clear pair)
                if len(narrowed) < len(ballots_for_unit):
                    sjm_on_unit = [
                        b for b in ballots_for_unit
                        if any(
                            ou.unit.unit_number == unit_number
                            and ou.ownership_type and "SJM" in ou.ownership_type.upper()
                            for ou in b.owner.current_units
                        )
                    ]
                    if len(sjm_on_unit) == 2:
                        for b in sjm_on_unit:
                            if b not in narrowed:
                                narrowed.append(b)
                ballots_for_unit = narrowed

        # Propagate votes to all ballots for this unit (each owner has their own share)
        targets = ballots_for_unit if vote_choices else [ballots_for_unit[0]]
        for ballot in targets:
            votes = {
                item_id: {"vote": vote, "count": ballot.total_votes}
                for item_id, vote in vote_choices.items()
            }

            # Merge votes with previously seen rows for same ballot
            if ballot.id in seen_ballots:
                canonical = seen_ballots[ballot.id]
                for item_id, vote_data in votes.items():
                    canonical["votes"][item_id] = vote_data
                votes = canonical["votes"]

            # Collect existing votes for conflict detection (K2)
            existing_votes = {}
            for bv in ballot.votes:
                if bv.vote is not None:
                    existing_votes[bv.voting_item_id] = bv.vote.value

            entry = {
                "row": row_idx,
                "owner_name": owner_name or ballot.owner.display_name,
                "unit_number": unit_number,
                "ballot_id": ballot.id,
                "votes": votes,
            }
            if existing_votes:
                entry["existing_votes"] = existing_votes
            if unrecognized:
                entry["unrecognized"] = unrecognized
            if ballot.id not in seen_ballots:
                seen_ballots[ballot.id] = entry
            matched.append(entry)

    wb.close()

    # K3: Detect duplicate ballot assignments (SJM risk)
    ballot_row_count: dict[int, int] = {}
    for entry in matched:
        bid = entry["ballot_id"]
        ballot_row_count[bid] = ballot_row_count.get(bid, 0) + 1
    duplicates = [
        {"ballot_id": bid, "count": cnt}
        for bid, cnt in ballot_row_count.items()
        if cnt > 1
    ]

    return {
        "total_rows": len(matched) + len(unmatched) + len(no_match),
        "matched": matched,
        "unmatched": unmatched,
        "no_match": no_match,
        "errors": errors,
        "duplicates": duplicates,
    }


def execute_voting_import(file_path: str, mapping: dict, voting: Voting, db: Session) -> dict:
    """Execute the import: update BallotVote records and ballot statuses."""
    clear_existing = mapping.get("clear_existing", False)
    preview = preview_voting_import(file_path, mapping, voting, db)

    processed_count = 0
    skipped_count = 0
    cleared_count = 0
    errors = []

    # Set of ballot IDs that will be updated from import
    matched_ballot_ids = {entry["ballot_id"] for entry in preview["matched"]}

    # If clear mode: reset all ballots that are NOT in matched set back to unprocessed
    if clear_existing:
        for ballot in voting.ballots:
            if ballot.id not in matched_ballot_ids and ballot.status == BallotStatus.PROCESSED:
                ballot.status = BallotStatus.GENERATED
                ballot.processed_at = None
                for bv in ballot.votes:
                    bv.vote = None
                    # Keep votes_count (the weight) as is
                cleared_count += 1

    # Use already-loaded ballot objects (avoids potential ORM identity issues)
    ballot_by_id = {b.id: b for b in voting.ballots}

    processed_ballot_ids = set()
    for entry in preview["matched"]:
        ballot_id = entry["ballot_id"]
        if ballot_id in processed_ballot_ids:
            continue  # Already processed this ballot from a previous row
        processed_ballot_ids.add(ballot_id)

        ballot = ballot_by_id.get(ballot_id)
        if not ballot:
            skipped_count += 1
            continue

        has_real_votes = False
        for bv in ballot.votes:
            item_id = bv.voting_item_id
            if item_id in entry["votes"]:
                vote_data = entry["votes"][item_id]
                # In append mode: skip if vote already set
                if not clear_existing and bv.vote is not None:
                    continue
                vote_map = {"for": VoteValue.FOR, "against": VoteValue.AGAINST, "abstain": VoteValue.ABSTAIN}
                bv.vote = vote_map.get(vote_data["vote"], VoteValue.FOR)
                bv.votes_count = vote_data["count"]
                has_real_votes = True
            elif clear_existing:
                # Clear mode: reset votes not in import
                bv.vote = None

        if has_real_votes:
            ballot.status = BallotStatus.PROCESSED
            ballot.processed_at = utcnow()
            processed_count += 1
        else:
            skipped_count += 1

    # Save mapping for next time (commit handled by caller)
    voting.import_column_mapping = json.dumps(mapping, ensure_ascii=False)

    return {
        "processed_count": processed_count,
        "skipped_count": skipped_count,
        "cleared_count": cleared_count,
        "unmatched_count": len(preview["unmatched"]),
        "clear_existing": clear_existing,
        "errors": preview["errors"],
    }
