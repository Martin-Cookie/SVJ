from __future__ import annotations

"""Import kontaktních údajů vlastníků z Excelu.

Supports dynamic column mapping via `mapping` dict parameter.
Falls back to DEFAULT_CONTACT_MAPPING when no mapping is provided.

Default Excel format (KontaktyVlastnici):
- Sheet "ZU", data od řádku 7 (hlavičky v řádcích 4-6)
- Sloupce (1-indexed, openpyxl):
  16: Jméno, 17: Příjmení
  19: RČ / IČ
  20-24: Trvalá adresa, 25-29: Korespondenční adresa
  30: GSM, 31: Pevný telefon, 32: Email
"""
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Owner, OwnerType
from app.utils import strip_diacritics


# Default mapping — 0-based column indices (converted to 1-based internally for openpyxl)
DEFAULT_CONTACT_MAPPING = {
    "fields": {
        "match_name": 15,         # col 16 (1-based) = Jméno — used for matching
        "match_birth_number": 18, # col 19 (1-based) = RČ/IČ — fallback matching
        "email": 31,              # col 32
        "email_secondary": 33,    # col 34 (poznámka, may contain email)
        "phone": 29,              # col 30
        "phone_landline": 30,     # col 31
        "birth_number": 18,       # col 19
        "perm_street": 19,        # col 20
        "perm_district": 20,      # col 21
        "perm_city": 21,          # col 22
        "perm_zip": 22,           # col 23
        "perm_country": 23,       # col 24
        "corr_street": 24,        # col 25
        "corr_district": 25,      # col 26
        "corr_city": 26,          # col 27
        "corr_zip": 27,           # col 28
        "corr_country": 28,       # col 29
    },
    "sheet_name": "ZU",
    "start_row": 7,
    # Extra fields used by default layout for name building
    "_title_before_col": 14,   # col 15 (1-based) — titul před
    "_last_name_col": 16,      # col 17 (1-based) — příjmení
    "_title_after_col": 17,    # col 18 (1-based) — titul za
}


# Field labels for UI display
_FIELD_LABELS = {
    "email": "Email",
    "email_secondary": "Email 2",
    "phone": "Telefon (GSM)",
    "phone_landline": "Pevný telefon",
    "birth_number": "Rodné číslo / IČ",
    "perm_street": "Trvalá ulice",
    "perm_district": "Trvalá část obce",
    "perm_city": "Trvalá obec",
    "perm_zip": "Trvalé PSČ",
    "perm_country": "Trvalá země",
    "corr_street": "Koresp. ulice",
    "corr_district": "Koresp. část obce",
    "corr_city": "Koresp. obec",
    "corr_zip": "Koresp. PSČ",
    "corr_country": "Koresp. země",
}

# Fields that are imported (not used for matching)
_IMPORT_FIELDS = [
    "email", "email_secondary", "phone", "phone_landline", "birth_number",
    "perm_street", "perm_district", "perm_city", "perm_zip", "perm_country",
    "corr_street", "corr_district", "corr_city", "corr_zip", "corr_country",
]


def _normalize_phone(phone: str) -> str:
    """Normalize phone for comparison: strip non-digit chars, remove +420/00420 prefix."""
    if not phone:
        return ""
    digits = "".join(c for c in phone if c.isdigit() or c == "+")
    if digits.startswith("+420"):
        digits = digits[4:]
    elif digits.startswith("00420"):
        digits = digits[5:]
    elif digits.startswith("420") and len(digits) > 9:
        digits = digits[3:]
    return digits


def _format_phone_for_db(phone: str) -> str:
    """Format phone for DB storage: add +420 prefix for 9-digit CZ numbers."""
    if not phone:
        return ""
    digits = "".join(c for c in phone if c.isdigit())
    if len(digits) == 9:
        return f"+420{digits}"
    if len(digits) == 12 and digits.startswith("420"):
        return f"+{digits}"
    return phone.strip()


def _build_normalized_name(first_name: str | None, last_name: str | None) -> str:
    """Build normalized name for matching: 'příjmení jméno' → stripped lowercase."""
    parts = []
    if last_name:
        parts.append(last_name.strip())
    if first_name:
        parts.append(first_name.strip())
    return strip_diacritics(" ".join(parts))


def _get_cell(cells: dict, field_key: str, fm: dict) -> str | None:
    """Get cell value from cells dict using 0-based mapping (converted to 1-based for openpyxl cells)."""
    col_0 = fm.get(field_key)
    if col_0 is None:
        return None
    col_1 = col_0 + 1  # openpyxl cells dict is 1-based
    return cells.get(col_1)


def preview_contact_import(
    file_path: str,
    db: Session,
    progress: dict | None = None,
    mapping: dict | None = None,
) -> dict:
    """Parse Excel and return preview of changes.

    Args:
        progress: optional shared dict updated with 'total'/'current' for progress tracking.
        mapping: column mapping dict. If None, uses DEFAULT_CONTACT_MAPPING.

    Returns dict with keys:
    - rows: list of dicts with excel_row, excel_name, owner_id, owner_name, matched, changes
    - stats: total_rows, matched_count, unmatched_count, with_changes, changes_by_field
    """
    m = mapping or DEFAULT_CONTACT_MAPPING
    fm = m.get("fields", DEFAULT_CONTACT_MAPPING["fields"])
    sheet_name = m.get("sheet_name", "ZU")
    start_row = m.get("start_row", 7)

    if progress is not None:
        progress["phase"] = "Načítám Excel..."

    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        return {
            "rows": [],
            "stats": {
                "total_rows": 0,
                "matched_count": 0,
                "unmatched_count": 0,
                "with_changes": 0,
                "changes_by_field": {},
            },
            "error": f"Nepodařilo se otevřít Excel soubor: {e}",
        }

    # Select sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif "ZU" in wb.sheetnames:
        ws = wb["ZU"]
    else:
        ws = wb.active

    # Pre-load all data rows into list of tuples (row_num, cells...)
    data_rows = []
    for row in ws.iter_rows(min_row=start_row, values_only=False):
        cells = {}
        for cell in row:
            if cell.value is not None:
                val = str(cell.value).strip()
                if val:
                    cells[cell.column] = val
        if cells:
            data_rows.append((row[0].row, cells))

    wb.close()

    if progress is not None:
        progress["phase"] = "Porovnávám s evidencí..."
        progress["total"] = len(data_rows)
        progress["current"] = 0

    # Build owner lookup by normalized name + RČ/IČ
    owners = db.query(Owner).filter_by(is_active=True).all()
    owner_by_name = {}
    owner_by_rc = {}
    for o in owners:
        if o.name_normalized:
            key = o.name_normalized.strip()
            if key not in owner_by_name:
                owner_by_name[key] = o
        if o.birth_number:
            rc_key = o.birth_number.replace(" ", "").replace("/", "")
            owner_by_rc[rc_key] = o
        if o.company_id:
            ic_key = o.company_id.replace(" ", "")
            owner_by_rc[ic_key] = o

    # Name building columns — for display of Excel name
    # Default layout has separate first_name(16), last_name(17), title_before(15), title_after(18)
    # Custom mapping: match_name is the primary name column
    title_before_col = m.get("_title_before_col")
    last_name_col = m.get("_last_name_col")
    title_after_col = m.get("_title_after_col")
    match_name_col_0 = fm.get("match_name")
    match_rc_col_0 = fm.get("match_birth_number")

    rows = []
    seen_owners = set()
    changes_by_field = {}

    for idx, (row_num, cells) in enumerate(data_rows):
        if progress is not None:
            progress["current"] = idx + 1

        # Get name for matching
        if match_name_col_0 is not None:
            first_name = cells.get(match_name_col_0 + 1)
        else:
            first_name = None

        # Try to get last_name from separate column (default layout)
        if last_name_col is not None:
            last_name = cells.get(last_name_col + 1)
        else:
            last_name = None

        if not first_name and not last_name:
            continue

        # Build display name
        name_parts = []
        if title_before_col is not None:
            t = cells.get(title_before_col + 1)
            if t:
                name_parts.append(t)
        if last_name:
            name_parts.append(last_name)
        if first_name:
            name_parts.append(first_name)
        if title_after_col is not None:
            t = cells.get(title_after_col + 1)
            if t:
                name_parts.append(t)
        excel_name = " ".join(name_parts) if name_parts else (first_name or "")

        # Match to DB owner
        norm_name = _build_normalized_name(first_name, last_name)
        owner = owner_by_name.get(norm_name)

        # Fallback: try RČ/IČ
        if not owner and match_rc_col_0 is not None:
            rc_raw = cells.get(match_rc_col_0 + 1)
            if rc_raw:
                rc_key = rc_raw.replace(" ", "").replace("/", "")
                owner = owner_by_rc.get(rc_key)

        # Skip duplicate rows for same owner
        if owner and owner.id in seen_owners:
            continue
        if owner:
            seen_owners.add(owner.id)

        # Compare fields
        changes = []
        if owner:
            for field_key in _IMPORT_FIELDS:
                col_0 = fm.get(field_key)
                if col_0 is None:
                    continue
                excel_val = cells.get(col_0 + 1)
                if not excel_val:
                    continue

                label = _FIELD_LABELS.get(field_key, field_key)

                # email_secondary: only use if value looks like email
                if field_key == "email_secondary" and "@" not in excel_val:
                    continue

                field = field_key
                # RČ/IČ: use correct DB field based on owner type
                if field == "birth_number" and owner.owner_type == OwnerType.LEGAL_ENTITY:
                    field = "company_id"
                    label = "IČ"

                current_val = getattr(owner, field, None) or ""

                # Intelligent routing for phone/email → secondary fields
                secondary_map = {
                    "phone": ("phone_secondary", "→ Telefon GSM 2"),
                    "email": ("email_secondary", "→ Email 2"),
                }
                if field in secondary_map:
                    sec_field, sec_label = secondary_map[field]
                    sec_val = getattr(owner, sec_field, None) or ""
                    is_phone = field == "phone"

                    if is_phone:
                        match_primary = _normalize_phone(current_val) == _normalize_phone(excel_val)
                        match_secondary = _normalize_phone(sec_val) == _normalize_phone(excel_val)
                    else:
                        match_primary = current_val.strip().lower() == excel_val.strip().lower()
                        match_secondary = sec_val.strip().lower() == excel_val.strip().lower()

                    if not current_val.strip():
                        pass  # Primary empty → fill primary
                    elif match_primary or match_secondary:
                        continue
                    elif not sec_val.strip():
                        changes.append({
                            "field": sec_field,
                            "label": sec_label,
                            "current": "",
                            "new": excel_val,
                            "is_overwrite": False,
                            "is_secondary": True,
                            "primary_label": label,
                            "primary_current": current_val,
                        })
                        changes_by_field[sec_field] = changes_by_field.get(sec_field, 0) + 1
                        continue
                    else:
                        pass  # Both occupied → overwrite primary
                else:
                    if field in ("phone_landline",):
                        if _normalize_phone(current_val) == _normalize_phone(excel_val):
                            continue
                    elif field == "email_secondary":
                        primary_email = (getattr(owner, "email", None) or "").strip().lower()
                        if excel_val.strip().lower() == primary_email:
                            continue
                        if current_val and current_val.strip().lower() == excel_val.strip().lower():
                            continue
                        if any(c["field"] == "email_secondary" for c in changes):
                            continue
                    elif current_val and current_val.strip() == excel_val.strip():
                        continue

                changes.append({
                    "field": field,
                    "label": label,
                    "current": current_val,
                    "new": excel_val,
                    "is_overwrite": bool(current_val.strip()),
                })
                changes_by_field[field] = changes_by_field.get(field, 0) + 1

        rows.append({
            "excel_row": row_num,
            "excel_name": excel_name,
            "owner_id": owner.id if owner else None,
            "owner_name": owner.display_name if owner else None,
            "matched": owner is not None,
            "changes": changes,
        })

    matched_count = sum(1 for r in rows if r["matched"])
    with_changes = sum(1 for r in rows if r["matched"] and r["changes"])

    return {
        "rows": rows,
        "stats": {
            "total_rows": len(rows),
            "matched_count": matched_count,
            "unmatched_count": len(rows) - matched_count,
            "with_changes": with_changes,
            "changes_by_field": changes_by_field,
        },
    }


def execute_contact_import(
    file_path: str,
    db: Session,
    selected_owner_ids: list[int],
    overwrite_existing: bool = False,
    mapping: dict | None = None,
) -> dict:
    """Execute the import for selected owners.

    Returns dict with owners_updated, fields_updated, per_field counts.
    """
    # Re-run preview to get current data
    preview = preview_contact_import(file_path, db, mapping=mapping)

    selected_set = set(selected_owner_ids)
    owners_updated = 0
    fields_updated = 0
    per_field = {}

    for row in preview["rows"]:
        if not row["matched"] or row["owner_id"] not in selected_set:
            continue
        if not row["changes"]:
            continue

        owner = db.query(Owner).get(row["owner_id"])
        if not owner:
            continue

        updated = False
        for change in row["changes"]:
            # Skip overwrites unless explicitly requested
            if change["is_overwrite"] and not overwrite_existing:
                continue

            value = change["new"]
            # Format phone numbers for DB storage
            if change["field"] in ("phone", "phone_secondary", "phone_landline"):
                value = _format_phone_for_db(value)
            setattr(owner, change["field"], value)
            updated = True
            fields_updated += 1
            per_field[change["field"]] = per_field.get(change["field"], 0) + 1

        if updated:
            owners_updated += 1

    # db.commit() is called by the router
    return {
        "owners_updated": owners_updated,
        "fields_updated": fields_updated,
        "per_field": per_field,
    }
