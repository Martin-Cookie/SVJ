"""Shared utilities for Excel import column mapping (owners + contacts).

Provides field definitions, auto-detection of column mappings based on
header names, and validation of mapping dicts.
"""
from __future__ import annotations

from openpyxl import load_workbook

from app.utils import strip_diacritics


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def read_excel_headers(file_path: str, sheet_name: str | None = None, header_row: int = 1) -> list[str]:
    """Read headers from a specific row in an Excel file.

    Returns list of header strings (empty cells become "Sloupec N").
    Supports both .xlsx (openpyxl) and .xls (xlrd) formats.
    """
    if file_path.lower().endswith(".xls") and not file_path.lower().endswith(".xlsx"):
        return _read_xls_headers(file_path, sheet_name, header_row)

    wb = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    headers = []
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        headers = [
            str(c).strip() if c is not None else f"Sloupec {i + 1}"
            for i, c in enumerate(row)
        ]
    wb.close()
    return headers


def _read_xls_headers(file_path: str, sheet_name: str | None, header_row: int) -> list[str]:
    """Read headers from .xls file using xlrd."""
    import xlrd
    wb = xlrd.open_workbook(file_path)
    if sheet_name and sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
    else:
        ws = wb.sheet_by_index(0)
    row_idx = header_row - 1  # xlrd is 0-based
    if row_idx >= ws.nrows:
        return []
    return [
        str(ws.cell_value(row_idx, c)).strip() if ws.cell_value(row_idx, c) != "" else f"Sloupec {c + 1}"
        for c in range(ws.ncols)
    ]


def read_excel_sheet_names(file_path: str) -> list[str]:
    """Return list of sheet names in an Excel file."""
    if file_path.lower().endswith(".xls") and not file_path.lower().endswith(".xlsx"):
        import xlrd
        wb = xlrd.open_workbook(file_path)
        return wb.sheet_names()
    wb = load_workbook(file_path, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


# ---------------------------------------------------------------------------
# Owner field definitions — 31 fields in 6 groups
# ---------------------------------------------------------------------------

OWNER_FIELD_GROUPS = [
    {
        "key": "unit",
        "label": "Jednotka",
        "color": "blue",
        "fields": [
            "unit_kn", "building_number", "podil_scd", "floor_area",
            "room_count", "space_type", "section", "orientation_number",
            "address", "lv_number",
        ],
    },
    {
        "key": "owner",
        "label": "Vlastník",
        "color": "green",
        "fields": [
            "first_name", "last_name", "title", "birth_or_ic",
            "ownership_type",
        ],
    },
    {
        "key": "perm_address",
        "label": "Trvalá adresa",
        "color": "yellow",
        "fields": [
            "perm_street", "perm_district", "perm_city", "perm_zip",
            "perm_country",
        ],
    },
    {
        "key": "corr_address",
        "label": "Korespondenční adresa",
        "color": "purple",
        "fields": [
            "corr_street", "corr_district", "corr_city", "corr_zip",
            "corr_country",
        ],
    },
    {
        "key": "contacts",
        "label": "Kontakty",
        "color": "orange",
        "fields": [
            "phone_gsm", "phone_landline", "email_evidence", "email_contacts",
        ],
    },
    {
        "key": "other",
        "label": "Ostatní",
        "color": "gray",
        "fields": ["owner_since", "note"],
    },
]

OWNER_FIELD_DEFS: dict[str, dict] = {
    # --- Jednotka ---
    "unit_kn": {
        "label": "Číslo jednotky (KN)",
        "required": True,
        "description": "Číslo jednotky z katastru nemovitostí — povinné pro párování",
        "candidates": [
            "cislo jednotky", "cislo jednotky kn", "jednotka kn", "jednotka",
            "unit number", "unit kn", "c.j.", "cj", "kn",
            "cislo jednotky (kn)", "c. jednotky",
        ],
    },
    "building_number": {
        "label": "Číslo prostoru (stavební)",
        "required": False,
        "description": "Stavební číslo prostoru (nemusí se shodovat s číslem jednotky v KN)",
        "candidates": [
            "cislo prostoru", "stavebni cislo", "cislo prostoru stavebni",
            "building number", "prostor", "c.p.", "cp",
        ],
    },
    "podil_scd": {
        "label": "Podíl na SČD",
        "required": False,
        "description": "Spoluvlastnický podíl na společných částech domu (např. 685/71584)",
        "candidates": [
            "podil", "podil na scd", "podil scd", "scd", "share",
            "podil na spolecnych castech domu",
        ],
    },
    "floor_area": {
        "label": "Podlahová plocha (m²)",
        "required": False,
        "candidates": [
            "podlahova plocha", "plocha", "floor area", "m2", "plocha m2",
            "podlahova plocha (m2)", "plocha (m2)",
        ],
    },
    "room_count": {
        "label": "Počet místností",
        "required": False,
        "candidates": [
            "pocet mistnosti", "mistnosti", "room count", "rooms",
            "dispozice", "velikost",
        ],
    },
    "space_type": {
        "label": "Druh prostoru",
        "required": False,
        "candidates": [
            "druh prostoru", "typ prostoru", "space type", "typ", "druh",
            "ucel", "ucel uzivani",
        ],
    },
    "section": {
        "label": "Sekce domu",
        "required": False,
        "candidates": [
            "sekce", "sekce domu", "section", "vchod", "blok",
        ],
    },
    "orientation_number": {
        "label": "Číslo orientační",
        "required": False,
        "candidates": [
            "cislo orientacni", "orientacni cislo", "orientation number",
            "c.o.", "co", "orientacni",
        ],
    },
    "address": {
        "label": "Adresa jednotky",
        "required": False,
        "candidates": [
            "adresa jednotky", "adresa", "ulice jednotky", "address",
        ],
    },
    "lv_number": {
        "label": "LV číslo",
        "required": False,
        "description": "Číslo listu vlastnictví z katastru nemovitostí",
        "candidates": [
            "lv", "lv cislo", "list vlastnictvi", "lv number",
        ],
    },
    # --- Vlastník ---
    "first_name": {
        "label": "Jméno",
        "required": True,
        "description": "Křestní jméno vlastníka — povinné",
        "candidates": [
            "jmeno", "krestni jmeno", "first name", "name", "firstname",
        ],
    },
    "last_name": {
        "label": "Příjmení",
        "required": False,
        "candidates": [
            "prijmeni", "last name", "surname", "lastname", "prijmeni / nazev",
        ],
    },
    "title": {
        "label": "Titul",
        "required": False,
        "candidates": [
            "titul", "titul pred", "title", "titul pred jmenem",
            "akademicky titul",
        ],
    },
    "birth_or_ic": {
        "label": "Rodné číslo / IČ",
        "required": False,
        "description": "Rodné číslo fyzické osoby nebo IČO právnické osoby",
        "candidates": [
            "rodne cislo", "rc", "ic", "ico", "rodne cislo / ic",
            "rc/ic", "rc / ic", "birth number", "company id",
            "rodne cislo/ic",
        ],
    },
    "ownership_type": {
        "label": "Typ vlastnictví",
        "required": False,
        "description": "Typ vlastnictví: výlučné, SJM (spoluvlastnictví manželů), podílové",
        "candidates": [
            "typ vlastnictvi", "vlastnictvi", "ownership type", "ownership",
            "spoluvlastnictvi", "sjm",
        ],
    },
    # --- Trvalá adresa ---
    "perm_street": {
        "label": "Trvalá adresa – ulice",
        "required": False,
        "candidates": [
            "trvala adresa ulice", "trvala ulice", "ulice trvala",
            "perm street", "trvala adresa - ulice", "t. adresa ulice",
        ],
    },
    "perm_district": {
        "label": "Trvalá adresa – část obce",
        "required": False,
        "candidates": [
            "trvala adresa cast obce", "trvala cast obce", "cast obce trvala",
            "perm district", "trvala adresa - cast obce",
        ],
    },
    "perm_city": {
        "label": "Trvalá adresa – město",
        "required": False,
        "candidates": [
            "trvala adresa mesto", "trvala mesto", "mesto trvala", "trvala obec",
            "perm city", "trvala adresa - mesto", "trvala adresa obec",
        ],
    },
    "perm_zip": {
        "label": "Trvalá adresa – PSČ",
        "required": False,
        "candidates": [
            "trvala adresa psc", "trvala psc", "psc trvala",
            "perm zip", "trvala adresa - psc",
        ],
    },
    "perm_country": {
        "label": "Trvalá adresa – stát",
        "required": False,
        "candidates": [
            "trvala adresa stat", "trvala stat", "stat trvala", "trvala zeme",
            "perm country", "trvala adresa - stat",
        ],
    },
    # --- Korespondenční adresa ---
    "corr_street": {
        "label": "Koresp. adresa – ulice",
        "required": False,
        "candidates": [
            "korespondencni adresa ulice", "koresp ulice", "koresp. ulice",
            "korespondencni ulice", "corr street",
            "korespondencni adresa - ulice", "k. adresa ulice",
        ],
    },
    "corr_district": {
        "label": "Koresp. adresa – část obce",
        "required": False,
        "candidates": [
            "korespondencni adresa cast obce", "koresp cast obce",
            "koresp. cast obce", "corr district",
            "korespondencni adresa - cast obce",
        ],
    },
    "corr_city": {
        "label": "Koresp. adresa – město",
        "required": False,
        "candidates": [
            "korespondencni adresa mesto", "koresp mesto", "koresp. mesto",
            "korespondencni mesto", "koresp obec", "koresp. obec",
            "corr city", "korespondencni adresa - mesto",
        ],
    },
    "corr_zip": {
        "label": "Koresp. adresa – PSČ",
        "required": False,
        "candidates": [
            "korespondencni adresa psc", "koresp psc", "koresp. psc",
            "corr zip", "korespondencni adresa - psc",
        ],
    },
    "corr_country": {
        "label": "Koresp. adresa – stát",
        "required": False,
        "candidates": [
            "korespondencni adresa stat", "koresp stat", "koresp. stat",
            "koresp zeme", "koresp. zeme", "corr country",
            "korespondencni adresa - stat",
        ],
    },
    # --- Kontakty ---
    "phone_gsm": {
        "label": "Telefon GSM",
        "required": False,
        "candidates": [
            "telefon", "telefon gsm", "gsm", "mobil", "phone",
            "tel", "mobilni telefon",
        ],
    },
    "phone_landline": {
        "label": "Telefon pevný",
        "required": False,
        "candidates": [
            "telefon pevny", "pevny telefon", "pevna linka", "phone landline",
            "pevny", "tel. pevny",
        ],
    },
    "email_evidence": {
        "label": "Email (Evidence)",
        "required": False,
        "candidates": [
            "email", "e-mail", "email evidence", "email (evidence)",
            "hlavni email", "primarni email",
        ],
    },
    "email_contacts": {
        "label": "Email (Kontakty)",
        "required": False,
        "candidates": [
            "email kontakty", "email (kontakty)", "druhy email",
            "sekundarni email", "email 2", "dalsi email",
        ],
    },
    # --- Ostatní ---
    "owner_since": {
        "label": "Vlastník od",
        "required": False,
        "candidates": [
            "vlastnik od", "od", "owner since", "datum nabytí",
            "datum nabyti", "nabytí vlastnictví", "nabyti vlastnictvi",
        ],
    },
    "note": {
        "label": "Poznámka",
        "required": False,
        "candidates": [
            "poznamka", "note", "notes", "komentar", "popis",
        ],
    },
}


# ---------------------------------------------------------------------------
# Contact field definitions — 17 fields in 5 groups
# ---------------------------------------------------------------------------

CONTACT_FIELD_GROUPS = [
    {
        "key": "matching",
        "label": "Párování",
        "color": "red",
        "fields": ["match_name", "match_birth_number"],
    },
    {
        "key": "contacts",
        "label": "Kontakty",
        "color": "orange",
        "fields": ["email", "email_secondary", "phone", "phone_landline"],
    },
    {
        "key": "identification",
        "label": "Identifikace",
        "color": "green",
        "fields": ["birth_number"],
    },
    {
        "key": "perm_address",
        "label": "Trvalá adresa",
        "color": "yellow",
        "fields": [
            "perm_street", "perm_district", "perm_city", "perm_zip",
            "perm_country",
        ],
    },
    {
        "key": "corr_address",
        "label": "Korespondenční adresa",
        "color": "purple",
        "fields": [
            "corr_street", "corr_district", "corr_city", "corr_zip",
            "corr_country",
        ],
    },
]

CONTACT_FIELD_DEFS: dict[str, dict] = {
    # --- Párování ---
    "match_name": {
        "label": "Jméno (pro párování)",
        "required": True,
        "description": "Jméno vlastníka pro párování s evidencí — povinné",
        "candidates": [
            "jmeno", "prijmeni", "jmeno a prijmeni", "jmeno prijmeni",
            "name", "vlastnik", "osoba",
        ],
    },
    "match_birth_number": {
        "label": "RČ/IČ (fallback párování)",
        "required": False,
        "description": "Rodné číslo nebo IČO — použije se pro párování pokud jméno nestačí",
        "candidates": [
            "rodne cislo", "rc", "ic", "ico", "rc/ic", "rc / ic",
            "rodne cislo / ic", "rodne cislo/ic",
        ],
    },
    # --- Kontakty ---
    "email": {
        "label": "Email",
        "required": False,
        "candidates": [
            "email", "e-mail", "hlavni email", "primarni email",
        ],
    },
    "email_secondary": {
        "label": "Email 2 / Poznámka",
        "required": False,
        "candidates": [
            "email 2", "druhy email", "sekundarni email", "dalsi email",
            "poznamka", "note",
        ],
    },
    "phone": {
        "label": "Telefon (GSM)",
        "required": False,
        "candidates": [
            "telefon", "gsm", "mobil", "phone", "tel", "mobilni telefon",
            "telefon gsm",
        ],
    },
    "phone_landline": {
        "label": "Pevný telefon",
        "required": False,
        "candidates": [
            "telefon pevny", "pevny telefon", "pevna linka", "pevny",
            "tel. pevny", "phone landline",
        ],
    },
    # --- Identifikace ---
    "birth_number": {
        "label": "Rodné číslo / IČ",
        "required": False,
        "candidates": [
            "rodne cislo", "rc", "ic", "ico", "rc/ic", "rc / ic",
            "rodne cislo / ic", "birth number",
        ],
    },
    # --- Trvalá adresa ---
    "perm_street": {
        "label": "Trvalá ulice",
        "required": False,
        "candidates": [
            "trvala adresa ulice", "trvala ulice", "ulice",
            "trvala adresa - ulice",
        ],
    },
    "perm_district": {
        "label": "Trvalá část obce",
        "required": False,
        "candidates": [
            "trvala adresa cast obce", "trvala cast obce", "cast obce",
            "trvala adresa - cast obce",
        ],
    },
    "perm_city": {
        "label": "Trvalá obec",
        "required": False,
        "candidates": [
            "trvala adresa mesto", "trvala mesto", "mesto", "obec",
            "trvala obec", "trvala adresa - mesto",
        ],
    },
    "perm_zip": {
        "label": "Trvalé PSČ",
        "required": False,
        "candidates": [
            "trvala adresa psc", "trvala psc", "psc",
            "trvala adresa - psc",
        ],
    },
    "perm_country": {
        "label": "Trvalá země",
        "required": False,
        "candidates": [
            "trvala adresa stat", "trvala stat", "stat", "zeme",
            "trvala adresa - stat",
        ],
    },
    # --- Korespondenční adresa ---
    "corr_street": {
        "label": "Koresp. ulice",
        "required": False,
        "candidates": [
            "korespondencni adresa ulice", "koresp ulice", "koresp. ulice",
        ],
    },
    "corr_district": {
        "label": "Koresp. část obce",
        "required": False,
        "candidates": [
            "korespondencni adresa cast obce", "koresp cast obce",
            "koresp. cast obce",
        ],
    },
    "corr_city": {
        "label": "Koresp. obec",
        "required": False,
        "candidates": [
            "korespondencni adresa mesto", "koresp mesto", "koresp. mesto",
            "koresp obec", "koresp. obec",
        ],
    },
    "corr_zip": {
        "label": "Koresp. PSČ",
        "required": False,
        "candidates": [
            "korespondencni adresa psc", "koresp psc", "koresp. psc",
        ],
    },
    "corr_country": {
        "label": "Koresp. země",
        "required": False,
        "candidates": [
            "korespondencni adresa stat", "koresp stat", "koresp. stat",
            "koresp zeme", "koresp. zeme",
        ],
    },
}


# ---------------------------------------------------------------------------
# Auto-detection logic
# ---------------------------------------------------------------------------

def _normalize_header(text: str) -> str:
    """Normalize header text for matching: strip diacritics, lowercase, strip punctuation."""
    result = strip_diacritics(text)
    # Remove common punctuation that varies between files
    for ch in "()[]{}.-_/":
        result = result.replace(ch, " ")
    return " ".join(result.split())


def auto_detect_mapping(
    headers: list[str],
    field_defs: dict[str, dict],
    saved_mapping: dict | None = None,
) -> dict[str, dict]:
    """Auto-detect column mapping from Excel headers.

    Returns dict: field_key → {col: int|None, status: str, required: bool, label: str}
    Status values: 'saved', 'auto', 'unmatched'
    """
    normalized_headers = [_normalize_header(h) for h in headers]
    result = {}
    used_cols: set[int] = set()

    # Phase 1: Apply saved mapping (if exists and columns still match)
    if saved_mapping and "fields" in saved_mapping:
        for field_key, col_idx in saved_mapping["fields"].items():
            if field_key not in field_defs:
                continue
            if isinstance(col_idx, int) and 0 <= col_idx < len(headers):
                fdef = field_defs[field_key]
                result[field_key] = {
                    "col": col_idx,
                    "status": "saved",
                    "required": fdef.get("required", False),
                    "label": fdef["label"],
                }
                used_cols.add(col_idx)

    # Phase 2: Auto-detect remaining fields from header names
    for field_key, fdef in field_defs.items():
        if field_key in result:
            continue

        candidates = fdef.get("candidates", [])
        best_col = None
        best_score = 0

        for candidate in candidates:
            norm_candidate = _normalize_header(candidate)
            for col_idx, norm_header in enumerate(normalized_headers):
                if col_idx in used_cols:
                    continue
                # Exact match
                if norm_header == norm_candidate:
                    score = 100
                # Header contains candidate
                elif norm_candidate in norm_header:
                    score = 80
                # Candidate contains header (for short headers like "email")
                elif norm_header in norm_candidate and len(norm_header) >= 3:
                    score = 60
                else:
                    continue

                if score > best_score:
                    best_score = score
                    best_col = col_idx

        if best_col is not None:
            result[field_key] = {
                "col": best_col,
                "status": "auto",
                "required": fdef.get("required", False),
                "label": fdef["label"],
            }
            used_cols.add(best_col)
        else:
            result[field_key] = {
                "col": None,
                "status": "unmatched",
                "required": fdef.get("required", False),
                "label": fdef["label"],
            }

    return result


# ---------------------------------------------------------------------------
# Balance (zůstatky) field definitions — 9 fields in 2 groups
# ---------------------------------------------------------------------------

BALANCE_FIELD_GROUPS = [
    {
        "key": "required",
        "label": "Povinná pole",
        "color": "blue",
        "fields": ["unit_number", "owner_name", "amount"],
    },
    {
        "key": "optional",
        "label": "Nepovinná pole",
        "color": "gray",
        "fields": [
            "variable_symbol", "deposits", "settlement",
            "paid", "paid_date", "status",
        ],
    },
]

BALANCE_FIELD_DEFS: dict[str, dict] = {
    "unit_number": {
        "label": "Katastrální číslo / č. jednotky",
        "required": True,
        "description": "Číslo jednotky pro párování s evidencí",
        "candidates": [
            "katastrální číslo", "katastralni cislo", "číslo jednotky",
            "cislo jednotky", "č. jednotky", "č jednotky", "jednotka",
            "byt", "unit", "unit_number", "kn", "č.j.",
        ],
    },
    "owner_name": {
        "label": "Vlastník",
        "required": True,
        "description": "Jméno vlastníka / dlužníka",
        "candidates": [
            "vlastník", "vlastnik", "jméno", "jmeno", "name", "owner",
            "dlužník", "dluznik", "majitel",
        ],
    },
    "amount": {
        "label": "Nedoplatek / částka",
        "required": True,
        "description": "Výše nedoplatku (kladné=dluh, záporné=přeplatek)",
        "candidates": [
            "nedoplatek", "nedoplatky", "částka", "castka", "dluh",
            "zůstatek", "zustatek", "amount", "balance", "saldo",
        ],
    },
    "variable_symbol": {
        "label": "Variabilní symbol",
        "required": False,
        "candidates": [
            "variabilní symbol", "variabilni symbol", "vs",
            "variable symbol", "var. symbol",
        ],
    },
    "deposits": {
        "label": "Zálohy",
        "required": False,
        "description": "Nedoplatek na zálohách",
        "candidates": [
            "zálohy", "zalohy", "záloha", "zaloha", "deposit", "deposits",
        ],
    },
    "settlement": {
        "label": "Vyúčtování",
        "required": False,
        "description": "Nedoplatek na vyúčtování",
        "candidates": [
            "vyúčtování", "vyuctovani", "settlement", "vyúčt",
        ],
    },
    "paid": {
        "label": "Uhrazeno",
        "required": False,
        "candidates": [
            "uhrazeno", "zaplaceno", "paid", "úhrada", "uhrada",
        ],
    },
    "paid_date": {
        "label": "Datum platby",
        "required": False,
        "candidates": [
            "datum platby", "datum", "date", "datum úhrady", "datum uhrady",
        ],
    },
    "status": {
        "label": "Stav",
        "required": False,
        "candidates": [
            "stav", "status", "poznámka", "poznamka", "note",
        ],
    },
}


# ---------------------------------------------------------------------------
# Space field definitions — 10 fields in 3 groups
# ---------------------------------------------------------------------------

SPACE_FIELD_GROUPS = [
    {
        "key": "space",
        "label": "Prostor",
        "color": "purple",
        "fields": [
            "space_number", "designation", "section", "floor", "area",
        ],
    },
    {
        "key": "tenant",
        "label": "Nájemce",
        "color": "green",
        "fields": [
            "tenant_name", "phone", "email",
        ],
    },
    {
        "key": "contract",
        "label": "Smlouva",
        "color": "orange",
        "fields": [
            "contract_number", "contract_start", "monthly_rent", "variable_symbol",
        ],
    },
]

SPACE_FIELD_DEFS: dict[str, dict] = {
    # --- Prostor ---
    "space_number": {
        "label": "Číslo prostoru",
        "required": True,
        "description": "Identifikátor prostoru — povinné",
        "candidates": [
            "cislo prostoru", "prostor", "cislo", "c.p.", "cp", "space number",
            "oznaceni prostoru", "id prostoru", "mistnost", "cislo mistnosti",
        ],
    },
    "designation": {
        "label": "Označení / účel",
        "required": False,
        "description": "Název nebo účel prostoru (sklad, dílna, nebytový prostor)",
        "candidates": [
            "oznaceni", "ucel", "nazev", "popis", "designation", "typ",
            "ucel uzivani", "druh", "typ prostoru", "popis prostoru",
        ],
    },
    "section": {
        "label": "Sekce / vchod",
        "required": False,
        "candidates": [
            "sekce", "vchod", "blok", "section", "cast domu",
        ],
    },
    "floor": {
        "label": "Podlaží",
        "required": False,
        "candidates": [
            "podlazi", "patro", "floor", "np", "poschodie",
        ],
    },
    "area": {
        "label": "Výměra (m²)",
        "required": False,
        "candidates": [
            "vymera", "plocha", "m2", "area", "vymera m2", "podlahova plocha",
        ],
    },
    # --- Nájemce ---
    "tenant_name": {
        "label": "Jméno nájemce",
        "required": False,
        "description": "Jméno nájemce — pokud vyplněno, automaticky se vytvoří nájemce",
        "candidates": [
            "najemce", "jmeno najemce", "nazev najemce", "najemnik",
            "tenant", "tenant name", "firma", "nazev firmy",
            "prijmeni jmeno", "jmeno prijmeni",
        ],
    },
    "phone": {
        "label": "Telefon nájemce",
        "required": False,
        "candidates": [
            "telefon", "tel", "phone", "mobil", "gsm", "kontakt",
        ],
    },
    "email": {
        "label": "Email nájemce",
        "required": False,
        "candidates": [
            "email", "e-mail", "mail", "email najemce",
        ],
    },
    # --- Smlouva ---
    "contract_number": {
        "label": "Číslo smlouvy",
        "required": False,
        "candidates": [
            "cislo smlouvy", "smlouva", "contract", "contract number",
            "c. smlouvy", "c.s.",
        ],
    },
    "contract_start": {
        "label": "Začátek smlouvy",
        "required": False,
        "candidates": [
            "zacatek smlouvy", "platnost od", "od", "datum smlouvy",
            "contract start", "zacatek", "od data",
        ],
    },
    "monthly_rent": {
        "label": "Měsíční nájemné (Kč)",
        "required": False,
        "candidates": [
            "najemne", "mesicni najemne", "rent", "monthly rent",
            "castka", "najemne kc", "mesicni castka",
        ],
    },
    "variable_symbol": {
        "label": "Variabilní symbol",
        "required": False,
        "candidates": [
            "variabilni symbol", "vs", "variable symbol", "var. symbol",
            "var symbol",
        ],
    },
}


def validate_space_mapping(mapping: dict):
    """Validate space import mapping. Returns error message or None."""
    if not isinstance(mapping, dict) or "fields" not in mapping:
        return "Neplatný formát mapování"

    fields = mapping["fields"]
    if not isinstance(fields, dict):
        return "Neplatný formát mapování polí"

    for field_key, fdef in SPACE_FIELD_DEFS.items():
        if fdef.get("required") and field_key not in fields:
            return f"Chybí povinné pole: {fdef['label']}"

    return None


def validate_balance_mapping(mapping: dict) -> str | None:
    """Validate balance import mapping. Returns error message or None."""
    if not isinstance(mapping, dict) or "fields" not in mapping:
        return "Neplatný formát mapování"

    fields = mapping["fields"]
    if not isinstance(fields, dict):
        return "Neplatný formát mapování polí"

    for field_key, fdef in BALANCE_FIELD_DEFS.items():
        if fdef.get("required") and field_key not in fields:
            return f"Chybí povinné pole: {fdef['label']}"

    return None


def validate_owner_mapping(mapping: dict) -> str | None:
    """Validate owner import mapping. Returns error message or None."""
    if not isinstance(mapping, dict) or "fields" not in mapping:
        return "Neplatný formát mapování"

    fields = mapping["fields"]
    if not isinstance(fields, dict):
        return "Neplatný formát mapování polí"

    # Check required fields
    for field_key, fdef in OWNER_FIELD_DEFS.items():
        if fdef.get("required") and field_key not in fields:
            return f"Chybí povinné pole: {fdef['label']}"

    return None


def validate_contact_mapping(mapping: dict) -> str | None:
    """Validate contact import mapping. Returns error message or None."""
    if not isinstance(mapping, dict) or "fields" not in mapping:
        return "Neplatný formát mapování"

    fields = mapping["fields"]
    if not isinstance(fields, dict):
        return "Neplatný formát mapování polí"

    # Check required fields
    for field_key, fdef in CONTACT_FIELD_DEFS.items():
        if fdef.get("required") and field_key not in fields:
            return f"Chybí povinné pole: {fdef['label']}"

    return None


def build_mapping_context(
    headers: list[str],
    field_defs: dict[str, dict],
    groups: list[dict],
    saved_mapping: dict | None = None,
) -> dict:
    """Build template context for the mapping page.

    Returns dict with keys: headers, detected, groups_data, stats
    """
    detected = auto_detect_mapping(headers, field_defs, saved_mapping)

    # Build groups_data for template
    groups_data = []
    for group in groups:
        group_fields = []
        for field_key in group["fields"]:
            det = detected.get(field_key, {})
            group_fields.append({
                "key": field_key,
                "label": field_defs[field_key]["label"],
                "required": field_defs[field_key].get("required", False),
                "description": field_defs[field_key].get("description", ""),
                "col": det.get("col"),
                "status": det.get("status", "unmatched"),
            })
        groups_data.append({
            "key": group["key"],
            "label": group["label"],
            "color": group["color"],
            "fields": group_fields,
        })

    # Stats
    total_fields = len(field_defs)
    matched_count = sum(1 for d in detected.values() if d["col"] is not None)
    required_missing = sum(
        1 for d in detected.values()
        if d["required"] and d["col"] is None
    )

    return {
        "headers": headers,
        "groups_data": groups_data,
        "stats": {
            "total": total_fields,
            "matched": matched_count,
            "required_missing": required_missing,
        },
    }
