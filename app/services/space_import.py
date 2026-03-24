"""Excel import service for parsing SVJ space (prostor) data.

Supports dynamic column mapping via `mapping` dict parameter.
Creates Space + Tenant + SpaceTenant records.
Auto-detects blocked spaces (kočárkárna, ústředna, trezor etc.).
Optionally auto-creates VariableSymbolMapping + Prescription.
"""
from __future__ import annotations

import logging
from datetime import date, datetime

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import (
    Owner, Prescription, PrescriptionYear, Space, SpaceStatus,
    SpaceTenant, SymbolSource, Tenant, VariableSymbolMapping,
)
from app.utils import build_name_with_titles, strip_diacritics, utcnow

logger = logging.getLogger(__name__)

# Keywords that indicate a blocked/utility space (not rentable)
BLOCKED_KEYWORDS = [
    "kocarkarna", "ustredna", "trezor", "kotelna", "strojovna",
    "sklad odpadu", "komora", "rozvodna", "chodba", "schodiste",
    "vytah", "zasedaci", "spolecna", "technick", "uklid",
]


def _build_field_map(mapping: dict) -> dict:
    """Extract field→column_index dict from mapping, filtering out None values."""
    if not mapping or "fields" not in mapping:
        return {}
    return {k: v for k, v in mapping["fields"].items() if v is not None}


def _cell(row: tuple, idx) -> str:
    """Safely get cell value as stripped string, or empty string."""
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    val = str(row[idx]).strip()
    return val


def _cell_int(row: tuple, idx) -> int:
    """Safely get cell value as integer, or None."""
    raw = _cell(row, idx)
    if not raw:
        return None
    try:
        return int(float(raw))
    except (ValueError, TypeError):
        return None


def _cell_float(row: tuple, idx) -> float:
    """Safely get cell value as float, or None."""
    raw = _cell(row, idx)
    if not raw:
        return None
    try:
        return float(raw.replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def _cell_date(row: tuple, idx):
    """Safely get cell value as date."""
    if idx is None or idx >= len(row) or row[idx] is None:
        return None
    val = row[idx]
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # Try parsing common formats
    raw = str(val).strip()
    if not raw:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d. %m. %Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _detect_blocked(designation: str) -> bool:
    """Auto-detect if space is blocked based on designation keywords."""
    if not designation:
        return False
    norm = strip_diacritics(designation)
    return any(kw in norm for kw in BLOCKED_KEYWORDS)


def _match_owner(db: Session, tenant_name: str):
    """Try to find matching Owner by normalized name. Returns Owner or None."""
    if not tenant_name:
        return None
    name_norm = strip_diacritics(tenant_name)
    # Exact match on name_normalized
    owner = db.query(Owner).filter(Owner.name_normalized == name_norm).first()
    if owner:
        return owner
    # Partial match — last name match
    parts = name_norm.split()
    if parts:
        # Try matching just the last name (first word in normalized = příjmení)
        surname = parts[0]
        if len(surname) >= 3:
            candidates = db.query(Owner).filter(
                Owner.name_normalized.like(f"{surname}%")
            ).all()
            if len(candidates) == 1:
                return candidates[0]
    return None


def preview_spaces_from_excel(file_path: str, mapping: dict, db: Session = None):
    """Parse Excel and return preview dict without saving to DB.

    Returns dict with keys:
        rows_processed, spaces_count, preview_rows, errors,
        blocked_count, with_tenant_count
    """
    fm = _build_field_map(mapping)
    if not fm.get("space_number") and fm.get("space_number") != 0:
        return {
            "rows_processed": 0, "spaces_count": 0, "preview_rows": [],
            "errors": ["Povinné pole 'Číslo prostoru' není namapováno."],
            "blocked_count": 0, "with_tenant_count": 0,
        }

    sheet_name = mapping.get("sheet_name")
    start_row = mapping.get("start_row", 2)

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    except Exception as e:
        logger.error("Failed to open Excel for space preview: %s", e)
        return {
            "rows_processed": 0, "spaces_count": 0, "preview_rows": [],
            "errors": [f"Nelze otevřít soubor: {e}"],
            "blocked_count": 0, "with_tenant_count": 0,
        }

    preview_rows = []
    errors = []
    seen_numbers = set()
    blocked_count = 0
    with_tenant_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        # Skip fully empty rows
        if not any(c is not None for c in row):
            continue

        space_number = _cell_int(row, fm.get("space_number"))
        if space_number is None:
            raw = _cell(row, fm.get("space_number"))
            if raw:
                errors.append(f"Řádek {row_idx}: neplatné číslo prostoru '{raw}'")
            continue

        if space_number in seen_numbers:
            errors.append(f"Řádek {row_idx}: duplicitní číslo prostoru {space_number}")
            continue
        seen_numbers.add(space_number)

        designation = _cell(row, fm.get("designation"))
        section = _cell(row, fm.get("section"))
        floor = _cell_int(row, fm.get("floor"))
        area = _cell_float(row, fm.get("area"))
        tenant_name = _cell(row, fm.get("tenant_name"))
        phone = _cell(row, fm.get("phone"))
        email = _cell(row, fm.get("email"))
        contract_number = _cell(row, fm.get("contract_number"))
        contract_start = _cell_date(row, fm.get("contract_start"))
        monthly_rent = _cell_float(row, fm.get("monthly_rent"))
        vs = _cell(row, fm.get("variable_symbol"))

        is_blocked = _detect_blocked(designation) or _detect_blocked(tenant_name)
        if is_blocked:
            blocked_count += 1
            status = "blocked"
        elif tenant_name:
            status = "rented"
            with_tenant_count += 1
        else:
            status = "vacant"

        # Try owner matching for preview
        owner_match = None
        if db and tenant_name and not is_blocked:
            owner = _match_owner(db, tenant_name)
            if owner:
                owner_match = owner.display_name

        preview_rows.append({
            "row": row_idx,
            "space_number": space_number,
            "designation": designation or "—",
            "section": section or "—",
            "floor": floor,
            "area": area,
            "status": status,
            "tenant_name": tenant_name or "",
            "phone": phone or "",
            "email": email or "",
            "contract_number": contract_number or "",
            "contract_start": contract_start.strftime("%d.%m.%Y") if contract_start else "",
            "monthly_rent": monthly_rent or 0,
            "variable_symbol": vs or "",
            "owner_match": owner_match,
        })

    try:
        wb.close()
    except Exception:
        pass

    return {
        "rows_processed": len(preview_rows) + len(errors),
        "spaces_count": len(preview_rows),
        "preview_rows": preview_rows,
        "errors": errors,
        "blocked_count": blocked_count,
        "with_tenant_count": with_tenant_count,
    }


def import_spaces_from_excel(db: Session, file_path: str, mapping: dict):
    """Parse Excel and save Space + Tenant + SpaceTenant records to DB.

    Returns dict with keys:
        spaces_created, tenants_created, contracts_created,
        vs_mappings_created, rows_processed, errors
    """
    fm = _build_field_map(mapping)
    sheet_name = mapping.get("sheet_name")
    start_row = mapping.get("start_row", 2)

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    except Exception as e:
        logger.error("Failed to open Excel for space import: %s", e)
        return {
            "spaces_created": 0, "tenants_created": 0, "contracts_created": 0,
            "vs_mappings_created": 0, "rows_processed": 0,
            "errors": [f"Nelze otevřít soubor: {e}"],
        }

    spaces_created = 0
    tenants_created = 0
    contracts_created = 0
    vs_mappings_created = 0
    errors = []
    seen_numbers = set()
    now = utcnow()

    # Get latest PrescriptionYear for auto-creating prescriptions
    latest_py = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if not any(c is not None for c in row):
            continue

        space_number = _cell_int(row, fm.get("space_number"))
        if space_number is None:
            raw = _cell(row, fm.get("space_number"))
            if raw:
                errors.append(f"Řádek {row_idx}: neplatné číslo prostoru '{raw}'")
            continue

        if space_number in seen_numbers:
            errors.append(f"Řádek {row_idx}: duplicitní č. prostoru {space_number} — přeskočeno")
            continue
        seen_numbers.add(space_number)

        # Check if space already exists
        existing = db.query(Space).filter_by(space_number=space_number).first()
        if existing:
            errors.append(f"Řádek {row_idx}: prostor č. {space_number} již existuje — přeskočeno")
            continue

        designation = _cell(row, fm.get("designation"))
        section = _cell(row, fm.get("section"))
        floor_val = _cell_int(row, fm.get("floor"))
        area = _cell_float(row, fm.get("area"))
        tenant_name = _cell(row, fm.get("tenant_name"))
        phone = _cell(row, fm.get("phone"))
        email = _cell(row, fm.get("email"))
        contract_number = _cell(row, fm.get("contract_number"))
        contract_start = _cell_date(row, fm.get("contract_start"))
        monthly_rent = _cell_float(row, fm.get("monthly_rent")) or 0.0
        vs = _cell(row, fm.get("variable_symbol"))

        is_blocked = _detect_blocked(designation) or _detect_blocked(tenant_name)

        # Determine status
        if is_blocked:
            status = SpaceStatus.BLOCKED
        elif tenant_name:
            status = SpaceStatus.RENTED
        else:
            status = SpaceStatus.VACANT

        # Create Space
        space = Space(
            space_number=space_number,
            designation=designation or f"Prostor {space_number}",
            section=section or None,
            floor=floor_val,
            area=area,
            status=status,
            blocked_reason="Automaticky detekováno z názvu" if is_blocked else None,
            created_at=now,
            updated_at=now,
        )
        db.add(space)
        db.flush()  # Get space.id
        spaces_created += 1

        # Create Tenant + SpaceTenant if tenant_name provided and not blocked
        if tenant_name and not is_blocked:
            # Try to match to existing Owner
            owner = _match_owner(db, tenant_name)

            # Build name fields
            name_norm = strip_diacritics(tenant_name)
            parts = tenant_name.split()

            tenant = Tenant(
                owner_id=owner.id if owner else None,
                first_name=parts[-1] if len(parts) > 1 and not owner else None,
                last_name=parts[0] if parts and not owner else None,
                name_with_titles=tenant_name if not owner else None,
                name_normalized=name_norm if not owner else None,
                phone=phone or None if not owner else None,
                email=email or None if not owner else None,
                data_source="import",
                created_at=now,
                updated_at=now,
            )
            db.add(tenant)
            db.flush()
            tenants_created += 1

            # Create SpaceTenant (active contract)
            st = SpaceTenant(
                space_id=space.id,
                tenant_id=tenant.id,
                contract_number=contract_number or None,
                contract_start=contract_start,
                monthly_rent=monthly_rent,
                variable_symbol=vs or None,
                is_active=True,
                created_at=now,
                updated_at=now,
            )
            db.add(st)
            db.flush()
            contracts_created += 1

            # Auto-create VariableSymbolMapping
            if vs:
                existing_vs = db.query(VariableSymbolMapping).filter_by(
                    variable_symbol=vs
                ).first()
                if not existing_vs:
                    db.add(VariableSymbolMapping(
                        variable_symbol=vs,
                        space_id=space.id,
                        unit_id=None,
                        source=SymbolSource.AUTO,
                        description=f"Import prostorů — prostor {space_number}",
                        created_at=now,
                    ))
                    vs_mappings_created += 1
                else:
                    errors.append(
                        f"Řádek {row_idx}: VS '{vs}' již existuje — VS mapování nevytvořeno"
                    )

            # Auto-create Prescription
            if monthly_rent > 0 and latest_py:
                existing_presc = db.query(Prescription).filter_by(
                    prescription_year_id=latest_py.id, space_id=space.id
                ).first()
                if not existing_presc:
                    db.add(Prescription(
                        prescription_year_id=latest_py.id,
                        space_id=space.id,
                        unit_id=None,
                        variable_symbol=vs or None,
                        monthly_total=monthly_rent,
                        owner_name=tenant_name,
                        created_at=now,
                        updated_at=now,
                    ))

    try:
        wb.close()
    except Exception:
        pass

    db.commit()

    return {
        "spaces_created": spaces_created,
        "tenants_created": tenants_created,
        "contracts_created": contracts_created,
        "vs_mappings_created": vs_mappings_created,
        "rows_processed": len(seen_numbers) + len(errors),
        "errors": errors,
    }
