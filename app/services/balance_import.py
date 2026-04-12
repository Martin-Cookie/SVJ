"""Import počátečních zůstatků z Excelu.

Parsuje Excel soubor s nedoplatky, páruje na jednotky a vlastníky,
vytváří UnitBalance záznamy.
"""
from __future__ import annotations

import logging
from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Unit, Owner, OwnerUnit, UnitBalance, BalanceSource, VariableSymbolMapping
from app.utils import strip_diacritics

logger = logging.getLogger(__name__)


def preview_balance_import(
    file_path: str,
    mapping: dict,
    year: int,
    db: Session,
) -> dict:
    """Přečte Excel a vrátí náhled pro potvrzení importu.

    Returns dict:
        rows: list[dict] — parsované řádky s párováním
        stats: dict — statistiky (ok, warning, error, skipped, total_amount)
    """
    rows = _parse_excel(file_path, mapping)

    # Lookup mapy
    units_by_number = {u.unit_number: u for u in db.query(Unit).all()}
    vs_map = {}
    for vsm in db.query(VariableSymbolMapping).filter_by(is_active=True).all():
        vs_map[vsm.variable_symbol] = vsm.unit_id

    # Vlastníci per unit pro fuzzy match
    owner_units = (
        db.query(OwnerUnit)
        .filter(OwnerUnit.valid_to.is_(None))
        .all()
    )
    owners_by_unit: dict[int, list] = {}
    for ou in owner_units:
        owners_by_unit.setdefault(ou.unit_id, []).append(ou)

    owner_by_id = {o.id: o for o in db.query(Owner).all()}

    result_rows = []
    stats = {"ok": 0, "warning": 0, "error": 0, "skipped": 0, "total_amount": 0.0}

    for row in rows:
        unit_number = row.get("unit_number")
        owner_name = row.get("owner_name", "")
        amount = row.get("amount", 0.0)

        # Přeskočit řádky CELKEM / prázdné
        if not unit_number or (isinstance(owner_name, str) and owner_name.strip().upper() == "CELKEM"):
            row["status"] = "skipped"
            row["status_label"] = "Přeskočeno"
            stats["skipped"] += 1
            result_rows.append(row)
            continue

        # Párování na jednotku
        unit = units_by_number.get(int(unit_number)) if unit_number else None
        row["unit"] = unit
        row["unit_found"] = unit is not None

        # Párování na vlastníka
        matched_owner = None
        if unit:
            matched_owner = _match_owner(
                owner_name, unit.id, owners_by_unit, owner_by_id
            )
        row["matched_owner"] = matched_owner

        # Vlastníci pro náhled — u SJM jen pár odpovídající Excel jménu
        unit_owners = []
        if unit:
            ou_list = owners_by_unit.get(unit.id, [])
            sjm_ous = [ou for ou in ou_list if (ou.ownership_type or "").strip().upper() == "SJM"]
            if sjm_ous and owner_name:
                excel_words = set(strip_diacritics(owner_name or "").split())
                for ou in sjm_ous:
                    o = owner_by_id.get(ou.owner_id)
                    if o and o.name_normalized and o.name_normalized.split()[0] in excel_words:
                        unit_owners.append(o)
            if not unit_owners and matched_owner:
                unit_owners = [matched_owner]
        row["unit_owners"] = unit_owners

        # VS ověření
        vs = row.get("variable_symbol")
        vs_ok = True
        if vs and unit:
            vs_unit_id = vs_map.get(str(vs).strip())
            vs_ok = vs_unit_id == unit.id if vs_unit_id else True  # neznámé VS = OK
        row["vs_ok"] = vs_ok

        # Status
        if not unit:
            row["status"] = "error"
            row["status_label"] = "Jednotka nenalezena"
            stats["error"] += 1
        elif not matched_owner:
            row["status"] = "warning"
            row["status_label"] = "Vlastník nenalezen v evidenci"
            stats["warning"] += 1
            stats["total_amount"] += amount or 0
        else:
            row["status"] = "ok"
            row["status_label"] = "OK"
            stats["ok"] += 1
            stats["total_amount"] += amount or 0

        result_rows.append(row)

    return {"rows": result_rows, "stats": stats, "year": year}


def execute_balance_import(
    file_path: str,
    mapping: dict,
    year: int,
    db: Session,
) -> dict:
    """Provede import zůstatků — smaže existující pro rok, vytvoří nové.

    Returns dict: imported, skipped, errors
    """
    rows = _parse_excel(file_path, mapping)

    units_by_number = {u.unit_number: u for u in db.query(Unit).all()}
    owner_units = (
        db.query(OwnerUnit)
        .filter(OwnerUnit.valid_to.is_(None))
        .all()
    )
    owners_by_unit: dict[int, list] = {}
    for ou in owner_units:
        owners_by_unit.setdefault(ou.unit_id, []).append(ou)
    owner_by_id = {o.id: o for o in db.query(Owner).all()}

    # Smazat existující zůstatky pro rok
    deleted = db.query(UnitBalance).filter_by(year=year).delete()
    db.flush()
    logger.info("Deleted %d existing balances for year %d", deleted, year)

    imported = 0
    skipped = 0
    errors = 0

    for row in rows:
        unit_number = row.get("unit_number")
        owner_name = row.get("owner_name", "")
        amount = row.get("amount")

        # Přeskočit CELKEM / prázdné
        if not unit_number or (isinstance(owner_name, str) and owner_name.strip().upper() == "CELKEM"):
            skipped += 1
            continue

        unit = units_by_number.get(int(unit_number)) if unit_number else None
        if not unit:
            errors += 1
            continue

        matched_owner = _match_owner(
            owner_name, unit.id, owners_by_unit, owner_by_id
        )

        # Pokud pro tuto jednotku už existuje zůstatek (duplicitní řádek), sečíst
        existing = db.query(UnitBalance).filter_by(unit_id=unit.id, year=year).first()
        if existing:
            existing.opening_amount = round((existing.opening_amount or 0) + (amount or 0), 2)
            # Přidat jméno do poznámky
            if owner_name and existing.owner_name and owner_name.strip() != existing.owner_name.strip():
                existing.note = (existing.note or "") + f"; {owner_name.strip()}: {amount}"
        else:
            # Sestavit poznámku z nepovinných polí
            note_parts = []
            deposits = row.get("deposits")
            settlement = row.get("settlement")
            status = row.get("status")
            if deposits:
                note_parts.append(f"Zálohy: {deposits}")
            if settlement:
                note_parts.append(f"Vyúčtování: {settlement}")
            if status:
                note_parts.append(str(status))

            db.add(UnitBalance(
                unit_id=unit.id,
                year=year,
                opening_amount=round(amount or 0, 2),
                source=BalanceSource.IMPORT,
                owner_id=matched_owner.id if matched_owner else None,
                owner_name=owner_name.strip() if owner_name else None,
                note=", ".join(note_parts) if note_parts else None,
            ))
        imported += 1

    db.commit()
    logger.info("Balance import: %d imported, %d skipped, %d errors", imported, skipped, errors)
    return {"imported": imported, "skipped": skipped, "errors": errors}


def _parse_excel(file_path: str, mapping: dict) -> list[dict]:
    """Parsuje Excel dle mapování, vrací list řádků jako dict."""
    fields = mapping.get("fields", {})
    sheet_name = mapping.get("sheet_name")
    start_row = mapping.get("start_row", 2)

    # Podpora .xls přes xlrd
    if file_path.lower().endswith(".xls"):
        return _parse_xls(file_path, fields, sheet_name, start_row)

    wb = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        parsed = _extract_row(row, fields)
        if parsed:
            parsed["_row_num"] = i
            rows.append(parsed)

    wb.close()
    return rows


def _parse_xls(file_path: str, fields: dict, sheet_name: str | None, start_row: int) -> list[dict]:
    """Parsuje starší .xls formát přes xlrd."""
    import xlrd
    wb = xlrd.open_workbook(file_path)
    if sheet_name and sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
    else:
        ws = wb.sheet_by_index(0)

    rows = []
    for r in range(start_row - 1, ws.nrows):
        row_values = [ws.cell_value(r, c) for c in range(ws.ncols)]
        parsed = _extract_row(row_values, fields)
        if parsed:
            parsed["_row_num"] = r + 1
            rows.append(parsed)

    return rows


def _extract_row(row_values, fields: dict) -> dict | None:
    """Extrahuje hodnoty z řádku dle mapování polí."""
    if not row_values:
        return None

    result = {}
    for field_key, col_idx in fields.items():
        if col_idx is not None and col_idx < len(row_values):
            val = row_values[col_idx]
            result[field_key] = val
        else:
            result[field_key] = None

    # Konverze typů
    for num_field in ("amount", "deposits", "settlement", "paid"):
        val = result.get(num_field)
        if val is not None and val != "":
            try:
                result[num_field] = float(val)
            except (ValueError, TypeError):
                result[num_field] = 0.0
        else:
            result[num_field] = None

    unit_num = result.get("unit_number")
    if unit_num is not None and unit_num != "":
        try:
            result["unit_number"] = int(float(unit_num))
        except (ValueError, TypeError):
            result["unit_number"] = None

    return result


def _match_owner(
    excel_name: str,
    unit_id: int,
    owners_by_unit: dict[int, list],
    owner_by_id: dict[int, "Owner"],
) -> "Owner | None":
    """Fuzzy match vlastníka z Excelu na vlastníka v evidenci dané jednotky."""
    if not excel_name or not excel_name.strip():
        return None

    unit_ous = owners_by_unit.get(unit_id, [])
    if not unit_ous:
        return None

    search = strip_diacritics(excel_name.strip())

    best_owner = None
    best_score = 0

    for ou in unit_ous:
        owner = owner_by_id.get(ou.owner_id)
        if not owner:
            continue
        owner_norm = owner.name_normalized or ""

        # Exact match
        if search == owner_norm:
            return owner

        # Příjmení match (první slovo excelu vs první slovo evidence)
        search_parts = search.split()
        owner_parts = owner_norm.split()
        if search_parts and owner_parts and search_parts[0] == owner_parts[0]:
            score = 80
            if score > best_score:
                best_score = score
                best_owner = owner

        # Partial: excel name contains owner's last name
        if owner_parts and owner_parts[0] in search:
            score = 60
            if score > best_score:
                best_score = score
                best_owner = owner

    return best_owner
