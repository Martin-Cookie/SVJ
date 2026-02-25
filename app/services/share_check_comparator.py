"""
Parse CSV/XLSX files and compare unit share (podíl SČD) values against the database.
"""
from __future__ import annotations

import csv
import re
from io import StringIO

from sqlalchemy.orm import Session

from app.models.share_check import ShareCheckColumnMapping, ShareCheckStatus


# ---- Column detection candidates (case-insensitive) ----

_UNIT_CANDIDATES = [
    "Název jednotky", "Jednotka", "Cislo jednotky",
    "Číslo jednotky", "nazev_jednotky", "cislo_jednotky",
    "Katastrální číslo", "Katastralni cislo",
    "unit_number",
]

_SHARE_CANDIDATES = [
    "Podíl SČD", "Podíl na domu", "podil_scd",
    "Podil na domu", "podil_na_domu", "Podíl na společných částech",
    "Spoluvlastnický podíl", "Spoluvlastnický podíl - 1",
    "Spoluvlastnicke podily",
    "share",
]


def get_file_headers(file_path: str) -> list[str]:
    """Return column headers from a CSV, XLSX, or XLS file."""
    lower = file_path.lower()
    if lower.endswith(".xls") and not lower.endswith(".xlsx"):
        return _get_xls_headers(file_path)
    if lower.endswith(".xlsx"):
        return _get_xlsx_headers(file_path)
    return _get_csv_headers(file_path)


def _get_xls_headers(file_path: str) -> list[str]:
    """Read headers from old .xls format using xlrd."""
    import xlrd
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)
    if ws.nrows == 0:
        return []
    headers = []
    for col_idx in range(ws.ncols):
        val = ws.cell_value(0, col_idx)
        headers.append(str(val).strip() if val else "")
    return [h for h in headers if h]


def _get_xlsx_headers(file_path: str) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value is not None else "")
    wb.close()
    return [h for h in headers if h]


def _get_csv_headers(file_path: str) -> list[str]:
    content = _read_csv_file(file_path)
    first_line = content.split("\n")[0]
    delimiter = ";" if ";" in first_line else ","
    reader = csv.reader(StringIO(first_line), delimiter=delimiter)
    for row in reader:
        return [col.strip() for col in row if col.strip()]
    return []


def get_file_preview(file_path: str, max_rows: int = 3) -> dict[str, list[str]]:
    """Return {header: [sample_value_1, sample_value_2, ...]} for preview in mapping UI."""
    lower = file_path.lower()
    if lower.endswith(".xls") and not lower.endswith(".xlsx"):
        return _get_xls_preview(file_path, max_rows)
    if lower.endswith(".xlsx"):
        return _get_xlsx_preview(file_path, max_rows)
    return _get_csv_preview(file_path, max_rows)


def _get_xls_preview(file_path: str, max_rows: int) -> dict[str, list[str]]:
    import xlrd
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)
    if ws.nrows < 2:
        return {}
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    preview = {}
    for c, h in enumerate(headers):
        if not h:
            continue
        vals = []
        for r in range(1, min(1 + max_rows, ws.nrows)):
            v = ws.cell_value(r, c)
            s = str(int(v)) if isinstance(v, float) and v == int(v) else str(v)
            vals.append(s.strip() if s.strip() else "—")
        preview[h] = vals
    return preview


def _get_xlsx_preview(file_path: str, max_rows: int) -> dict[str, list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=1 + max_rows))
    wb.close()
    if len(rows) < 2:
        return {}
    headers = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
    preview = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        vals = []
        for row in rows[1:]:
            v = row[i].value if i < len(row) else None
            if v is None:
                vals.append("—")
            elif isinstance(v, float) and v == int(v):
                vals.append(str(int(v)))
            else:
                vals.append(str(v).strip() or "—")
        preview[h] = vals
    return preview


def _get_csv_preview(file_path: str, max_rows: int) -> dict[str, list[str]]:
    content = _read_csv_file(file_path)
    if not content:
        return {}
    first_line = content.split("\n")[0]
    delimiter = ";" if ";" in first_line else ","
    reader = csv.DictReader(StringIO(content), delimiter=delimiter)
    rows = []
    for row in reader:
        rows.append(row)
        if len(rows) >= max_rows:
            break
    preview = {}
    for h in (reader.fieldnames or []):
        h = h.strip()
        if not h:
            continue
        preview[h] = [r.get(h, "—").strip() or "—" for r in rows]
    return preview


def _read_csv_file(file_path: str) -> str:
    """Read CSV trying multiple encodings."""
    for encoding in ["utf-8", "cp1250", "latin-1"]:
        try:
            with open(file_path, "r", encoding=encoding) as f:
                content = f.read()
            return content.lstrip("\ufeff")
        except UnicodeDecodeError:
            continue
    return ""


def detect_columns(headers: list[str]) -> tuple[str | None, str | None]:
    """Auto-detect unit and share columns from headers (case-insensitive)."""
    headers_lower = {h.lower(): h for h in headers}

    col_unit = None
    for candidate in _UNIT_CANDIDATES:
        if candidate.lower() in headers_lower:
            col_unit = headers_lower[candidate.lower()]
            break

    col_share = None
    for candidate in _SHARE_CANDIDATES:
        if candidate.lower() in headers_lower:
            col_share = headers_lower[candidate.lower()]
            break

    return col_unit, col_share


def suggest_mapping(headers: list[str], db: Session) -> tuple[str | None, str | None, bool]:
    """
    Suggest column mapping.
    Returns (col_unit, col_share, from_history).
    First checks saved mappings where both columns exist in the current headers.
    Falls back to auto-detection.
    """
    headers_set = set(headers)

    # Check saved mappings (most recently used first)
    saved = (
        db.query(ShareCheckColumnMapping)
        .order_by(ShareCheckColumnMapping.last_used_at.desc())
        .all()
    )
    for mapping in saved:
        if mapping.col_unit in headers_set and mapping.col_share in headers_set:
            return mapping.col_unit, mapping.col_share, True

    # Fallback to auto-detection
    col_unit, col_share = detect_columns(headers)
    return col_unit, col_share, False


def parse_file(file_path: str, col_unit: str, col_share: str) -> list[dict]:
    """
    Parse a CSV, XLSX, or XLS file and return records.
    Returns [{"unit_number": int, "file_share": int}].

    unit: "1098/14" → 14
    share: "12212/4103391" → 12212
    """
    lower = file_path.lower()
    if lower.endswith(".xls") and not lower.endswith(".xlsx"):
        return _parse_xls(file_path, col_unit, col_share)
    if lower.endswith(".xlsx"):
        return _parse_xlsx(file_path, col_unit, col_share)
    return _parse_csv(file_path, col_unit, col_share)


def _parse_unit_number(raw: str) -> int | None:
    """Extract unit number: '1098/14' → 14, '14' → 14, '14.0' → 14."""
    raw = str(raw).strip()
    if "/" in raw:
        raw = raw.split("/")[-1].strip()
    try:
        return int(float(raw))
    except (ValueError, TypeError):
        return None


def _parse_share_value(raw: str) -> int | None:
    """Extract share: '12212/4103391' → 12212, '12212' → 12212."""
    raw = str(raw).strip()
    if "/" in raw:
        raw = raw.split("/")[0].strip()
    try:
        return int(float(raw))
    except (ValueError, TypeError):
        return None


def _parse_csv(file_path: str, col_unit: str, col_share: str) -> list[dict]:
    content = _read_csv_file(file_path)
    if not content:
        return []

    first_line = content.split("\n")[0]
    delimiter = ";" if ";" in first_line else ","
    reader = csv.DictReader(StringIO(content), delimiter=delimiter)

    records = []
    for row in reader:
        unit_raw = row.get(col_unit, "")
        share_raw = row.get(col_share, "")

        unit_number = _parse_unit_number(unit_raw)
        file_share = _parse_share_value(share_raw)

        if unit_number is not None:
            records.append({
                "unit_number": unit_number,
                "file_share": file_share,
            })
    return records


def _parse_xlsx(file_path: str, col_unit: str, col_share: str) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # Find column indices
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    unit_idx = None
    share_idx = None
    for i, h in enumerate(headers):
        if h == col_unit:
            unit_idx = i
        if h == col_share:
            share_idx = i

    if unit_idx is None or share_idx is None:
        wb.close()
        return []

    records = []
    for row in ws.iter_rows(min_row=2):
        unit_raw = str(row[unit_idx].value or "")
        share_raw = str(row[share_idx].value or "")

        unit_number = _parse_unit_number(unit_raw)
        file_share = _parse_share_value(share_raw)

        if unit_number is not None:
            records.append({
                "unit_number": unit_number,
                "file_share": file_share,
            })

    wb.close()
    return records


def _parse_xls(file_path: str, col_unit: str, col_share: str) -> list[dict]:
    """Parse old .xls format using xlrd."""
    import xlrd
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)

    if ws.nrows < 2:
        return []

    # Find column indices from header row
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    unit_idx = None
    share_idx = None
    for i, h in enumerate(headers):
        if h == col_unit:
            unit_idx = i
        if h == col_share:
            share_idx = i

    if unit_idx is None or share_idx is None:
        return []

    records = []
    for row_idx in range(1, ws.nrows):
        unit_raw = str(ws.cell_value(row_idx, unit_idx) or "")
        share_raw = str(ws.cell_value(row_idx, share_idx) or "")

        unit_number = _parse_unit_number(unit_raw)
        file_share = _parse_share_value(share_raw)

        if unit_number is not None:
            records.append({
                "unit_number": unit_number,
                "file_share": file_share,
            })

    return records


def compare_shares(file_records: list[dict], db: Session) -> list[dict]:
    """
    Compare parsed file records against Unit.podil_scd in the database.
    Returns [{"unit_number", "db_share", "file_share", "status"}].
    Sums file_share per unit_number (multiple co-owner rows → one total).
    """
    from app.models import Unit

    # Load all units from DB
    units = db.query(Unit).all()
    db_by_unit: dict[int, int | None] = {}
    for u in units:
        db_by_unit[u.unit_number] = u.podil_scd

    # Aggregate file records — sum shares per unit (multiple co-owner rows)
    summed: dict[int, int] = {}
    for rec in file_records:
        unit_num = rec["unit_number"]
        share = rec["file_share"] or 0
        summed[unit_num] = summed.get(unit_num, 0) + share

    results = []
    file_units_seen = set()

    for unit_num, file_share in summed.items():
        file_units_seen.add(unit_num)

        if unit_num not in db_by_unit:
            results.append({
                "unit_number": unit_num,
                "db_share": None,
                "file_share": file_share,
                "status": ShareCheckStatus.MISSING_DB,
            })
            continue

        db_share = db_by_unit[unit_num]

        if (db_share or 0) == (file_share or 0):
            status = ShareCheckStatus.MATCH
        else:
            status = ShareCheckStatus.DIFFERENCE

        results.append({
            "unit_number": unit_num,
            "db_share": db_share,
            "file_share": file_share,
            "status": status,
        })

    # Units in DB but not in file
    for unit_num, db_share in db_by_unit.items():
        if unit_num not in file_units_seen:
            results.append({
                "unit_number": unit_num,
                "db_share": db_share,
                "file_share": None,
                "status": ShareCheckStatus.MISSING_FILE,
            })

    # Sort: differences first, then missing, then matches
    status_order = {
        ShareCheckStatus.DIFFERENCE: 0,
        ShareCheckStatus.MISSING_DB: 1,
        ShareCheckStatus.MISSING_FILE: 2,
        ShareCheckStatus.MATCH: 3,
    }
    results.sort(key=lambda r: (status_order.get(r["status"], 4), r["unit_number"]))

    return results
