import csv
import io
from datetime import date as date_type, datetime
from io import BytesIO

from fastapi import APIRouter, Depends, Form, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import (
    OwnerType, Space, SpaceStatus, SpaceTenant, Tenant,
    Prescription, PrescriptionYear, SymbolSource, VariableSymbolMapping,
)
from app.utils import (
    build_list_url, build_name_with_titles, excel_auto_width,
    is_htmx_partial, is_valid_email, strip_diacritics, templates, utcnow,
)

from ._helpers import SORT_COLUMNS, _filter_spaces, _space_stats, logger

router = APIRouter()


# ── Create ────────────────────────────────────────────────────────────


@router.get("/novy-formular")
async def space_create_form(request: Request):
    """Formulář pro vytvoření nového prostoru."""
    return templates.TemplateResponse("spaces/partials/_create_form.html", {
        "request": request,
        "form_data": {},
    })


@router.post("/novy")
async def space_create(
    request: Request,
    space_number: str = Form(...),
    designation: str = Form(...),
    section: str = Form(""),
    floor: str = Form(""),
    area: str = Form(""),
    status: str = Form("vacant"),
    blocked_reason: str = Form(""),
    note: str = Form(""),
    tenant_name: str = Form(""),
    tenant_phone: str = Form(""),
    tenant_email: str = Form(""),
    contract_number: str = Form(""),
    contract_start: str = Form(""),
    monthly_rent: str = Form(""),
    variable_symbol: str = Form(""),
    db: Session = Depends(get_db),
):
    """Vytvoření nového prostoru s volitelným nájemcem."""
    form_data = {
        "space_number": space_number, "designation": designation,
        "section": section, "floor": floor, "area": area,
        "status": status, "blocked_reason": blocked_reason, "note": note,
        "tenant_name": tenant_name, "tenant_phone": tenant_phone,
        "tenant_email": tenant_email, "contract_number": contract_number,
        "contract_start": contract_start, "monthly_rent": monthly_rent,
        "variable_symbol": variable_symbol,
    }

    def _err(msg):
        return templates.TemplateResponse("spaces/partials/_create_form.html", {
            "request": request, "error": msg, "form_data": form_data,
        })

    # Parse space_number
    try:
        space_number_int = int(space_number)
    except (ValueError, TypeError):
        return _err("Číslo prostoru musí být celé číslo.")
    if space_number_int < 1 or space_number_int > 99999:
        return _err("Číslo prostoru musí být v rozsahu 1–99999.")

    # Uniqueness
    existing = db.query(Space).filter(Space.space_number == space_number_int).first()
    if existing:
        return _err(f"Prostor s číslem {space_number_int} již existuje.")

    # Parse optional numerics
    floor_int = None
    if floor.strip():
        try:
            floor_int = int(floor.strip())
        except (ValueError, TypeError):
            return _err("Podlaží musí být celé číslo.")

    area_float = None
    if area.strip():
        try:
            area_float = float(area.strip())
        except (ValueError, TypeError):
            return _err("Výměra musí být číslo.")

    # Validate tenant email if provided
    if tenant_email.strip() and not is_valid_email(tenant_email.strip()):
        return _err("Neplatný formát emailu nájemce.")

    # If tenant_name provided, auto-set status to rented
    tenant_name = tenant_name.strip()
    space_status = SpaceStatus(status) if status in [s.value for s in SpaceStatus] else SpaceStatus.VACANT
    if tenant_name and space_status == SpaceStatus.VACANT:
        space_status = SpaceStatus.RENTED

    now = utcnow()
    space = Space(
        space_number=space_number_int,
        designation=designation.strip(),
        section=section.strip() or None,
        floor=floor_int,
        area=area_float,
        status=space_status,
        blocked_reason=blocked_reason.strip() or None,
        note=note.strip() or None,
        created_at=now,
    )
    db.add(space)
    db.flush()

    # Create tenant + contract if tenant_name provided
    if tenant_name:
        parts = tenant_name.split()
        last_name = parts[0] if parts else None
        first_name = " ".join(parts[1:]) if len(parts) > 1 else None

        tenant = Tenant(
            first_name=first_name,
            last_name=last_name,
            name_with_titles=build_name_with_titles(None, first_name, last_name),
            name_normalized=strip_diacritics(tenant_name),
            tenant_type=OwnerType.PHYSICAL,
            phone=tenant_phone.strip() or None,
            email=tenant_email.strip() or None,
            data_source="manual",
            is_active=True,
            created_at=now,
        )
        db.add(tenant)
        db.flush()

        # Parse rent
        rent_float = 0.0
        if monthly_rent.strip():
            try:
                rent_float = float(monthly_rent.strip().replace(",", "."))
            except (ValueError, TypeError):
                rent_float = 0.0

        # Parse contract start

        start_date = None
        if contract_start.strip():
            try:
                start_date = date_type.fromisoformat(contract_start.strip())
            except ValueError:
                pass

        vs = variable_symbol.strip() or None

        st = SpaceTenant(
            space_id=space.id,
            tenant_id=tenant.id,
            monthly_rent=rent_float,
            variable_symbol=vs,
            contract_number=contract_number.strip() or None,
            contract_start=start_date,
            is_active=True,
            created_at=now,
        )
        db.add(st)
        db.flush()

        # Auto-create VS mapping
        if vs:
            existing_vs = db.query(VariableSymbolMapping).filter_by(variable_symbol=vs).first()
            if not existing_vs:
                db.add(VariableSymbolMapping(
                    variable_symbol=vs,
                    space_id=space.id,
                    unit_id=None,
                    source=SymbolSource.MANUAL,
                    description=f"Ruční vytvoření prostoru {space_number_int}",
                    created_at=now,
                ))

        # Auto-create Prescription
        if rent_float > 0:
            latest_py = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
            if latest_py:
                db.add(Prescription(
                    prescription_year_id=latest_py.id,
                    space_id=space.id,
                    unit_id=None,
                    variable_symbol=vs,
                    monthly_total=rent_float,
                    owner_name=tenant_name,
                    created_at=now,
                    updated_at=now,
                ))

    db.commit()

    if request.headers.get("HX-Request"):
        return HTMLResponse(
            content=f'<p class="text-sm text-green-600 p-4">Prostor {space_number_int} vytvořen. '
                    f'<a href="/prostory/{space.id}" class="text-blue-600 hover:underline">Zobrazit</a></p>',
        )
    return RedirectResponse(f"/prostory/{space.id}", status_code=302)


# ── Detail ────────────────────────────────────────────────────────────


@router.get("/{space_id}/upravit-formular")
async def space_edit_form(
    space_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Inline edit formulář pro prostor."""
    space = db.query(Space).get(space_id)
    if not space:
        return RedirectResponse("/prostory", status_code=302)
    return templates.TemplateResponse("spaces/partials/_space_info.html", {
        "request": request,
        "space": space,
        "edit_mode": True,
    })


@router.get("/{space_id}/info")
async def space_info(
    space_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Display-only info sekce prostoru."""
    space = db.query(Space).get(space_id)
    if not space:
        return RedirectResponse("/prostory", status_code=302)
    return templates.TemplateResponse("spaces/partials/_space_info.html", {
        "request": request,
        "space": space,
        "edit_mode": False,
    })


@router.post("/{space_id}/upravit")
async def space_update(
    space_id: int,
    request: Request,
    designation: str = Form(...),
    section: str = Form(""),
    floor: str = Form(""),
    area: str = Form(""),
    status: str = Form("vacant"),
    blocked_reason: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    """Uložení úprav prostoru."""
    space = db.query(Space).get(space_id)
    if not space:
        return RedirectResponse("/prostory", status_code=302)

    floor_int = None
    if floor.strip():
        try:
            floor_int = int(floor.strip())
        except (ValueError, TypeError):
            return templates.TemplateResponse("spaces/partials/_space_info.html", {
                "request": request, "space": space, "edit_mode": True,
                "error": "Podlaží musí být celé číslo.",
            })

    area_float = None
    if area.strip():
        try:
            area_float = float(area.strip())
        except (ValueError, TypeError):
            return templates.TemplateResponse("spaces/partials/_space_info.html", {
                "request": request, "space": space, "edit_mode": True,
                "error": "Výměra musí být číslo.",
            })

    space_status = SpaceStatus(status) if status in [s.value for s in SpaceStatus] else space.status

    space.designation = designation.strip()
    space.section = section.strip() or None
    space.floor = floor_int
    space.area = area_float
    space.status = space_status
    space.blocked_reason = blocked_reason.strip() or None
    space.note = note.strip() or None
    space.updated_at = utcnow()
    db.commit()

    if request.headers.get("HX-Request"):
        return templates.TemplateResponse("spaces/partials/_space_info.html", {
            "request": request,
            "space": space,
            "edit_mode": False,
            "saved": True,
        })
    return RedirectResponse(f"/prostory/{space_id}", status_code=302)


@router.post("/{space_id}/smazat")
async def space_delete(
    space_id: int,
    db: Session = Depends(get_db),
):
    """Smazání prostoru."""
    space = db.query(Space).get(space_id)
    if space:
        db.delete(space)
        db.commit()
    return RedirectResponse("/prostory?flash=deleted", status_code=302)


# ── Správa nájmů ──────────────────────────────────────────────────────


@router.post("/{space_id}/pridat-najemce")
async def space_assign_tenant(
    space_id: int,
    request: Request,
    tenant_id: int = Form(...),
    monthly_rent: str = Form("0"),
    variable_symbol: str = Form(""),
    contract_number: str = Form(""),
    contract_start: str = Form(""),
    contract_end: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    """Přiřadit nájemce k prostoru — vytvoří SpaceTenant + VS mapping + Prescription."""
    space = db.query(Space).get(space_id)
    if not space:
        return RedirectResponse("/prostory", status_code=302)

    tenant = db.query(Tenant).get(tenant_id)
    if not tenant:
        return RedirectResponse(f"/prostory/{space_id}?flash=tenant_not_found", status_code=302)

    # Deactivate current tenant if any
    for st in db.query(SpaceTenant).filter_by(space_id=space_id, is_active=True).all():
        st.is_active = False
        st.updated_at = utcnow()

    # Parse rent
    rent_float = 0.0
    if monthly_rent.strip():
        try:
            rent_float = float(monthly_rent.strip().replace(",", "."))
        except (ValueError, TypeError):
            rent_float = 0.0

    # Parse dates
    from datetime import date as date_type
    start_date = None
    end_date = None
    if contract_start.strip():
        try:
            start_date = date_type.fromisoformat(contract_start.strip())
        except ValueError:
            pass
    if contract_end.strip():
        try:
            end_date = date_type.fromisoformat(contract_end.strip())
        except ValueError:
            pass

    vs = variable_symbol.strip() or None

    # Create SpaceTenant
    st = SpaceTenant(
        space_id=space.id,
        tenant_id=tenant.id,
        monthly_rent=rent_float,
        variable_symbol=vs,
        contract_number=contract_number.strip() or None,
        contract_start=start_date,
        contract_end=end_date,
        note=note.strip() or None,
        is_active=True,
        created_at=utcnow(),
    )
    db.add(st)

    # Update space status
    space.status = SpaceStatus.RENTED
    space.updated_at = utcnow()

    # Auto-create VariableSymbolMapping for VS
    if vs:
        existing_vs = db.query(VariableSymbolMapping).filter_by(variable_symbol=vs).first()
        if not existing_vs:
            db.add(VariableSymbolMapping(
                variable_symbol=vs,
                space_id=space.id,
                unit_id=None,
                source=SymbolSource.AUTO,
                description=f"Prostor {space.space_number} — {space.designation}",
                is_active=True,
                created_at=utcnow(),
            ))

    # Auto-create Prescription if PrescriptionYear exists
    if rent_float > 0:
        latest_py = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        if latest_py:
            existing_presc = db.query(Prescription).filter_by(
                prescription_year_id=latest_py.id, space_id=space.id
            ).first()
            if not existing_presc:
                db.add(Prescription(
                    prescription_year_id=latest_py.id,
                    space_id=space.id,
                    unit_id=None,
                    variable_symbol=vs,
                    space_number=space.space_number,
                    owner_name=tenant.display_name,
                    monthly_total=rent_float,
                    created_at=utcnow(),
                ))
            else:
                # Update existing prescription
                existing_presc.monthly_total = rent_float
                existing_presc.variable_symbol = vs
                existing_presc.owner_name = tenant.display_name
                existing_presc.updated_at = utcnow()

    db.commit()
    return RedirectResponse(f"/prostory/{space_id}?flash=tenant_assigned", status_code=302)


@router.post("/{space_id}/ukoncit-najem")
async def space_terminate_tenant(
    space_id: int,
    db: Session = Depends(get_db),
):
    """Ukončit aktivní nájem na prostoru."""
    space = db.query(Space).get(space_id)
    if not space:
        return RedirectResponse("/prostory", status_code=302)

    for st in db.query(SpaceTenant).filter_by(space_id=space_id, is_active=True).all():
        st.is_active = False
        st.updated_at = utcnow()

    space.status = SpaceStatus.VACANT
    space.updated_at = utcnow()
    db.commit()

    return RedirectResponse(f"/prostory/{space_id}?flash=tenant_terminated", status_code=302)


# ── List ──────────────────────────────────────────────────────────────


@router.get("/")
async def space_list(
    request: Request,
    q: str = Query("", alias="q"),
    stav: str = Query("", alias="stav"),
    sekce: str = Query("", alias="sekce"),
    sort: str = Query("space_number", alias="sort"),
    order: str = Query("asc", alias="order"),
    back: str = Query("", alias="back"),
    flash: str = Query("", alias="flash"),
    db: Session = Depends(get_db),
):
    """Seznam prostorů s filtry, hledáním a řazením."""
    spaces = _filter_spaces(db, q, stav, sekce, sort, order)

    list_url = build_list_url(request)

    # HTMX partial
    if is_htmx_partial(request):
        return templates.TemplateResponse("spaces/partials/_tbody.html", {
            "request": request,
            "spaces": spaces,
            "list_url": list_url,
        })

    stats = _space_stats(db)

    flash_message = None
    flash_type = None
    if flash == "deleted":
        flash_message = "Prostor byl smazán."
    elif flash == "import_ok":
        imported = request.query_params.get("imported", "?")
        tenants = request.query_params.get("tenants", "?")
        flash_message = f"Import dokončen: {imported} prostorů, {tenants} nájemců."
        flash_type = "success"

    return templates.TemplateResponse("spaces/list.html", {
        "request": request,
        "active_nav": "spaces",
        "spaces": spaces,
        "list_url": list_url,
        "back_url": back,
        "q": q,
        "stav": stav,
        "sekce": sekce,
        "sort": sort,
        "order": order,
        "stats": stats,
        "flash_message": flash_message,
        "flash_type": flash_type,
    })


# ── Export ─────────────────────────────────────────────────────────────


@router.get("/exportovat/{fmt}")
async def space_export(
    fmt: str,
    q: str = Query("", alias="q"),
    stav: str = Query("", alias="stav"),
    sekce: str = Query("", alias="sekce"),
    sort: str = Query("space_number", alias="sort"),
    order: str = Query("asc", alias="order"),
    db: Session = Depends(get_db),
):
    """Export filtered spaces to Excel or CSV."""
    if fmt not in ("xlsx", "csv"):
        return RedirectResponse("/prostory", status_code=302)

    spaces = _filter_spaces(db, q, stav, sekce, sort, order)

    headers = ["Č. prostoru", "Označení", "Sekce", "Podlaží", "Výměra", "Stav", "Nájemce", "Měs. nájemné", "VS"]
    status_labels = {"rented": "Pronajato", "vacant": "Volné", "blocked": "Blokované"}

    def _row(s):
        at = s.active_tenant_rel
        tenant_name = at.tenant.display_name if at else ""
        rent = at.monthly_rent if at else ""
        vs = at.variable_symbol if at else ""
        return [
            s.space_number,
            s.designation or "",
            s.section or "",
            s.floor or "",
            s.area or "",
            status_labels.get(s.status.value, s.status.value) if s.status else "",
            tenant_name,
            rent,
            vs,
        ]

    timestamp = datetime.now().strftime("%Y%m%d")

    stav_labels = {"rented": "pronajate", "vacant": "volne", "blocked": "blokovane"}
    if stav and stav in stav_labels:
        suffix = f"_{stav_labels[stav]}"
    elif sekce:
        suffix = f"_sekce_{sekce}"
    elif q:
        suffix = "_hledani"
    else:
        suffix = "_vse"
    filename = f"prostory{suffix}_{timestamp}"

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Prostory"
        bold = Font(bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold

        for row_idx, s in enumerate(spaces, 2):
            for col_idx, val in enumerate(_row(s), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        excel_auto_width(ws)

        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    else:
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(headers)
        for s in spaces:
            writer.writerow(_row(s))
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )


# ── Detail (catch-all — must be last) ─────────────────────────────────


@router.get("/{space_id}")
async def space_detail(
    space_id: int,
    request: Request,
    back: str = Query("", alias="back"),
    flash: str = Query("", alias="flash"),
    db: Session = Depends(get_db),
):
    """Detail prostoru s nájemcem a historií."""
    space = db.query(Space).options(
        joinedload(Space.tenants).joinedload(SpaceTenant.tenant).joinedload(Tenant.owner)
    ).get(space_id)
    if not space:
        return RedirectResponse("/prostory", status_code=302)

    # Back label
    if "/platby/prehled" in back:
        back_label = "Zpět na matici plateb"
    elif "/platby/dluznici" in back:
        back_label = "Zpět na dlužníky"
    elif "/platby/prostor" in back:
        back_label = "Zpět na platby prostoru"
    elif "/platby" in back:
        back_label = "Zpět na platby"
    elif "/najemci/" in back:
        back_label = "Zpět na detail nájemce"
    elif back.startswith("/najemci"):
        back_label = "Zpět na seznam nájemců"
    elif back:
        back_label = "Zpět"
    else:
        back_label = "Zpět na seznam prostorů"

    # Active tenant + history
    active_rel = space.active_tenant_rel
    history = [st for st in space.tenants if not st.is_active]
    history.sort(key=lambda st: st.contract_end or st.created_at, reverse=True)

    # All tenants for assignment dropdown
    all_tenants = db.query(Tenant).options(
        joinedload(Tenant.owner)
    ).filter(Tenant.is_active == True).order_by(Tenant.name_normalized).all()  # noqa: E712

    flash_message = None
    flash_type = None
    if flash == "tenant_assigned":
        flash_message = "Nájemce přiřazen."
    elif flash == "tenant_terminated":
        flash_message = "Nájem ukončen."
    elif flash == "tenant_not_found":
        flash_message = "Nájemce nenalezen."
        flash_type = "error"

    return templates.TemplateResponse("spaces/detail.html", {
        "request": request,
        "active_nav": "spaces",
        "space": space,
        "active_rel": active_rel,
        "history": history,
        "all_tenants": all_tenants,
        "back_url": back or "/prostory",
        "back_label": back_label,
        "flash_message": flash_message,
        "flash_type": flash_type,
    })
