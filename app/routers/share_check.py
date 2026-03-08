import logging
import shutil
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, Query, Request, UploadFile
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.config import settings
from app.database import get_db

logger = logging.getLogger(__name__)
from app.models import (
    ImportLog, Owner, OwnerUnit, ShareCheckColumnMapping, ShareCheckRecord,
    ShareCheckResolution, ShareCheckSession, ShareCheckStatus, Unit,
)
from app.services.share_check_comparator import (
    compare_shares, get_file_headers, get_file_preview, parse_file, suggest_mapping,
)
from app.utils import build_list_url, excel_auto_width, is_htmx_partial, is_safe_path, setup_jinja_filters, strip_diacritics, validate_upload

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")
setup_jinja_filters(templates)


SORT_COLUMNS = {
    "unit": ShareCheckRecord.unit_number,
    "db_share": ShareCheckRecord.db_share,
    "file_share": ShareCheckRecord.file_share,
    "status": ShareCheckRecord.status,
}


@router.get("/")
async def share_check_list(
    request: Request,
    q: str = Query("", alias="sc_q"),
    sort: str = Query("date", alias="sc_sort"),
    order: str = Query("desc", alias="sc_order"),
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    # HTMX partial search — return just tbody rows
    if is_htmx_partial(request):
        sessions = db.query(ShareCheckSession).order_by(ShareCheckSession.created_at.desc()).all()

        if q:
            q_lower = q.lower()
            q_ascii = strip_diacritics(q)
            sessions = [
                s for s in sessions
                if q_lower in (s.filename or "").lower()
                or q_ascii in strip_diacritics(s.filename or "")
                or q_lower in s.created_at.strftime("%d.%m.%Y %H:%M")
            ]

        SORT_KEYS = {
            "filename": lambda s: (s.filename or "").lower(),
            "date": lambda s: s.created_at,
            "matches": lambda s: s.total_matches or 0,
            "differences": lambda s: (s.total_differences or 0) + (s.total_missing_db or 0) + (s.total_missing_file or 0),
        }
        sort_fn = SORT_KEYS.get(sort, SORT_KEYS["date"])
        sessions.sort(key=sort_fn, reverse=(order == "desc"))

        list_url = build_list_url(request)
        ctx = {
            "request": request,
            "sessions": sessions,
            "list_url": list_url,
            "q": q,
        }
        return templates.TemplateResponse("partials/share_check_list_body.html", ctx)

    # Full page → redirect to combined page
    return RedirectResponse("/synchronizace#kontrola-podilu", status_code=302)


@router.post("/nova")
async def share_check_upload(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename:
        return RedirectResponse("/synchronizace#kontrola-podilu", status_code=302)

    err = await validate_upload(file, max_size_mb=50, allowed_extensions=[".csv", ".xlsx", ".xls"])
    if err:
        return RedirectResponse("/synchronizace?chyba=upload#kontrola-podilu", status_code=302)

    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = settings.upload_dir / "share_check" / f"{timestamp}_{file.filename}"
    dest.parent.mkdir(parents=True, exist_ok=True)
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)

    # Redirect to mapping page (PRG pattern)
    from urllib.parse import urlencode
    params = urlencode({"file_path": str(dest), "filename": file.filename})
    return RedirectResponse(f"/kontrola-podilu/mapovani?{params}", status_code=302)


@router.get("/mapovani")
async def share_check_mapping(
    request: Request,
    file_path: str = Query(...),
    filename: str = Query(...),
    db: Session = Depends(get_db),
):
    if not is_safe_path(Path(file_path), settings.upload_dir):
        return RedirectResponse("/synchronizace#kontrola-podilu", status_code=302)
    if not Path(file_path).is_file():
        return RedirectResponse("/synchronizace#kontrola-podilu", status_code=302)

    try:
        headers = get_file_headers(file_path)
    except Exception:
        return RedirectResponse("/synchronizace#kontrola-podilu", status_code=302)

    if not headers:
        return RedirectResponse("/synchronizace#kontrola-podilu", status_code=302)

    col_unit, col_share, from_history = suggest_mapping(headers, db)

    try:
        preview = get_file_preview(file_path)
    except Exception:
        preview = {}

    return templates.TemplateResponse("share_check/mapping.html", {
        "request": request,
        "active_nav": "kontroly",
        "headers": headers,
        "col_unit": col_unit,
        "col_share": col_share,
        "from_history": from_history,
        "file_path": file_path,
        "filename": filename,
        "preview": preview,
    })


@router.post("/potvrdit-mapovani")
async def share_check_confirm_mapping(
    request: Request,
    file_path: str = Form(...),
    filename: str = Form(...),
    col_unit: str = Form(...),
    col_share: str = Form(...),
    db: Session = Depends(get_db),
):
    # Validate file path and existence
    if not is_safe_path(Path(file_path), settings.upload_dir):
        return RedirectResponse("/synchronizace#kontrola-podilu", status_code=302)
    if not Path(file_path).is_file():
        return RedirectResponse("/synchronizace#kontrola-podilu", status_code=302)

    # Parse file
    file_records = parse_file(file_path, col_unit, col_share)
    if not file_records:
        return RedirectResponse("/synchronizace#kontrola-podilu", status_code=302)

    # Compare with DB
    comparison = compare_shares(file_records, db)

    # Save/update column mapping
    existing = (
        db.query(ShareCheckColumnMapping)
        .filter_by(col_unit=col_unit, col_share=col_share)
        .first()
    )
    if existing:
        existing.used_count += 1
        existing.last_used_at = datetime.utcnow()
    else:
        mapping = ShareCheckColumnMapping(
            col_unit=col_unit,
            col_share=col_share,
        )
        db.add(mapping)

    # Create session
    session = ShareCheckSession(
        filename=filename,
        file_path=file_path,
        col_unit=col_unit,
        col_share=col_share,
        total_records=len(comparison),
        total_matches=sum(1 for c in comparison if c["status"] == ShareCheckStatus.MATCH),
        total_differences=sum(1 for c in comparison if c["status"] == ShareCheckStatus.DIFFERENCE),
        total_missing_db=sum(1 for c in comparison if c["status"] == ShareCheckStatus.MISSING_DB),
        total_missing_file=sum(1 for c in comparison if c["status"] == ShareCheckStatus.MISSING_FILE),
    )
    db.add(session)
    db.flush()

    for comp in comparison:
        record = ShareCheckRecord(
            session_id=session.id,
            unit_number=comp["unit_number"],
            db_share=comp["db_share"],
            file_share=comp["file_share"],
            status=comp["status"],
            resolution=(
                ShareCheckResolution.PENDING
                if comp["status"] != ShareCheckStatus.MATCH
                else ShareCheckResolution.SKIPPED
            ),
        )
        db.add(record)

    db.commit()
    return RedirectResponse(f"/kontrola-podilu/{session.id}", status_code=302)


@router.get("/{session_id}")
async def share_check_detail(
    session_id: int,
    request: Request,
    q: str = Query(""),
    filtr: str = Query("", alias="filtr"),
    sort: str = Query("unit", alias="sort"),
    order: str = Query("asc", alias="order"),
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    session = db.query(ShareCheckSession).get(session_id)
    if not session:
        return RedirectResponse("/synchronizace#kontrola-podilu", status_code=302)

    base = db.query(ShareCheckRecord).filter_by(session_id=session_id)

    # Totals for bubbles
    total_match = session.total_matches
    total_diff = session.total_differences
    total_missing_db = session.total_missing_db
    total_missing_file = session.total_missing_file

    # Sum of shares per category
    def _share_sums(flt):
        r = db.query(
            func.sum(ShareCheckRecord.db_share),
            func.sum(ShareCheckRecord.file_share),
        ).filter_by(session_id=session_id).filter(flt).one()
        return r[0] or 0, r[1] or 0

    match_db, match_file = _share_sums(ShareCheckRecord.status == ShareCheckStatus.MATCH)
    diff_db, diff_file = _share_sums(ShareCheckRecord.status == ShareCheckStatus.DIFFERENCE)
    missing_db_db, missing_db_file = _share_sums(ShareCheckRecord.status == ShareCheckStatus.MISSING_DB)
    missing_file_db, missing_file_file = _share_sums(ShareCheckRecord.status == ShareCheckStatus.MISSING_FILE)

    total_db_share = (match_db + diff_db + missing_file_db)
    total_file_share = (match_file + diff_file + missing_db_file)

    # Filter
    query = base
    if filtr == "match":
        query = query.filter(ShareCheckRecord.status == ShareCheckStatus.MATCH)
    elif filtr == "rozdil":
        query = query.filter(ShareCheckRecord.status == ShareCheckStatus.DIFFERENCE)
    elif filtr == "chybi_db":
        query = query.filter(ShareCheckRecord.status == ShareCheckStatus.MISSING_DB)
    elif filtr == "chybi_soubor":
        query = query.filter(ShareCheckRecord.status == ShareCheckStatus.MISSING_FILE)

    # Sorting
    sort_col = SORT_COLUMNS.get(sort)
    if sort_col is not None:
        if order == "desc":
            query = query.order_by(sort_col.desc().nulls_last())
        else:
            query = query.order_by(sort_col.asc().nulls_last())
    else:
        query = query.order_by(ShareCheckRecord.unit_number.asc())

    records = query.all()

    # Search filtering
    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)

        # Pre-build owner_map for search (before the main owner_map build)
        _owner_names: dict[int, list[str]] = {}
        _ou = (
            db.query(OwnerUnit.owner_id, Unit.unit_number, Owner.name_with_titles)
            .join(OwnerUnit.unit)
            .join(Owner, OwnerUnit.owner_id == Owner.id)
            .filter(OwnerUnit.valid_to.is_(None))
            .all()
        )
        for oid, unit_num, oname in _ou:
            _owner_names.setdefault(unit_num, []).append(oname or "")

        records = [
            r for r in records
            if q_lower in str(r.unit_number or "")
            or any(
                q_ascii in strip_diacritics(n)
                for n in _owner_names.get(r.unit_number, [])
            )
        ]

    # Build unit_number → unit_id and unit_number → [(owner_id, name)] mappings
    unit_map = {}
    units = db.query(Unit.unit_number, Unit.id).all()
    for unit_num, unit_id in units:
        unit_map[unit_num] = unit_id

    owner_map: dict[int, list[tuple[int, str]]] = {}
    owner_units = (
        db.query(OwnerUnit.owner_id, Unit.unit_number, Owner.name_with_titles)
        .join(OwnerUnit.unit)
        .join(Owner, OwnerUnit.owner_id == Owner.id)
        .filter(OwnerUnit.valid_to.is_(None))
        .all()
    )
    for oid, unit_num, oname in owner_units:
        owner_map.setdefault(unit_num, []).append((oid, oname))

    back_url = back or "/synchronizace#kontrola-podilu"
    back_label = "Zpět na přehled" if back == "/" else "Zpět na kontroly" if "/synchronizace" in (back or "") else "Zpět"

    return templates.TemplateResponse("share_check/compare.html", {
        "request": request,
        "active_nav": "kontroly",
        "session": session,
        "records": records,
        "filtr": filtr,
        "sort": sort,
        "order": order,
        "q": q,
        "back_url": back_url,
        "back_label": back_label,
        "total_match": total_match,
        "total_diff": total_diff,
        "total_missing_db": total_missing_db,
        "total_missing_file": total_missing_file,
        "match_db": match_db,
        "match_file": match_file,
        "diff_db": diff_db,
        "diff_file": diff_file,
        "missing_db_db": missing_db_db,
        "missing_db_file": missing_db_file,
        "missing_file_db": missing_file_db,
        "missing_file_file": missing_file_file,
        "total_db_share": total_db_share,
        "total_file_share": total_file_share,
        "unit_map": unit_map,
        "owner_map": owner_map,
    })


@router.post("/{session_id}/exportovat")
async def share_check_export(
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Export current filtered view to Excel."""
    form_data = await request.form()
    filtr = form_data.get("filtr", "")

    session = db.query(ShareCheckSession).get(session_id)
    if not session:
        return RedirectResponse("/synchronizace#kontrola-podilu", status_code=302)

    query = db.query(ShareCheckRecord).filter_by(session_id=session_id)
    if filtr == "match":
        query = query.filter(ShareCheckRecord.status == ShareCheckStatus.MATCH)
    elif filtr == "rozdil":
        query = query.filter(ShareCheckRecord.status == ShareCheckStatus.DIFFERENCE)
    elif filtr == "chybi_db":
        query = query.filter(ShareCheckRecord.status == ShareCheckStatus.MISSING_DB)
    elif filtr == "chybi_soubor":
        query = query.filter(ShareCheckRecord.status == ShareCheckStatus.MISSING_FILE)

    records = query.order_by(ShareCheckRecord.unit_number.asc()).all()

    # Build owner_map for owner names in export
    owner_map: dict[int, list[str]] = {}
    owner_units = (
        db.query(OwnerUnit.owner_id, Unit.unit_number, Owner.name_with_titles)
        .join(OwnerUnit.unit)
        .join(Owner, OwnerUnit.owner_id == Owner.id)
        .filter(OwnerUnit.valid_to.is_(None))
        .all()
    )
    for _, unit_num, oname in owner_units:
        owner_map.setdefault(unit_num, []).append(oname)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Kontrola podílu"

    headers = ["Jednotka", "Vlastník", "Podíl DB", "Podíl soubor", "Rozdíl", "Stav"]
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    _STATUS_LABELS = {
        "match": "Shoda",
        "difference": "Rozdíl",
        "missing_db": "Chybí v DB",
        "missing_file": "Chybí v souboru",
    }

    diff_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    for r in records:
        owners = "; ".join(owner_map.get(r.unit_number, []))
        db_share = r.db_share
        file_share = r.file_share
        diff_val = (db_share or 0) - (file_share or 0) if db_share is not None and file_share is not None else None
        row = [
            r.unit_number,
            owners or "—",
            db_share,
            file_share,
            diff_val,
            _STATUS_LABELS.get(r.status.value, r.status.value),
        ]
        ws.append(row)

        if r.status == ShareCheckStatus.DIFFERENCE:
            row_num = ws.max_row
            ws.cell(row=row_num, column=3).fill = diff_fill
            ws.cell(row=row_num, column=4).fill = diff_fill
            ws.cell(row=row_num, column=5).fill = diff_fill

    excel_auto_width(ws)

    _FILTR_LABELS = {
        "": "vse", "match": "shoda", "rozdil": "rozdily",
        "chybi_db": "chybi_db", "chybi_soubor": "chybi_soubor",
    }
    suffix = _FILTR_LABELS.get(filtr, filtr or "vse")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"kontrola_podilu_{suffix}_{timestamp}.xlsx"
    output_path = settings.generated_dir / "exports" / filename
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return FileResponse(
        str(output_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/{session_id}/smazat")
async def share_check_delete(session_id: int, db: Session = Depends(get_db)):
    session = db.query(ShareCheckSession).get(session_id)
    if session:
        try:
            p = Path(session.file_path)
            if p.exists():
                p.unlink()
        except Exception:
            logger.debug("Failed to clean up file: %s", session.file_path)
        db.delete(session)
        db.commit()
    return RedirectResponse("/synchronizace#kontrola-podilu", status_code=302)


@router.post("/{session_id}/aktualizovat")
async def share_check_apply_updates(
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Batch update Unit.podil_scd from selected file values."""
    form = await request.form()

    session = db.query(ShareCheckSession).get(session_id)
    if not session:
        return RedirectResponse("/synchronizace#kontrola-podilu", status_code=302)

    # Collect record IDs from checkboxes: update__{record_id}
    record_ids = []
    for key in form.keys():
        if key.startswith("update__"):
            try:
                rid = int(key.split("__")[1])
                record_ids.append(rid)
            except (ValueError, IndexError):
                continue

    if not record_ids:
        filtr = form.get("filtr", "")
        url = f"/kontrola-podilu/{session_id}"
        if filtr:
            url += f"?filtr={filtr}"
        return RedirectResponse(url, status_code=302)

    success_count = 0
    change_details = []
    now = datetime.now().strftime("%d.%m.%Y %H:%M")

    for rid in record_ids:
        record = db.query(ShareCheckRecord).filter_by(
            id=rid, session_id=session_id,
        ).first()
        if not record or record.file_share is None:
            continue

        unit = db.query(Unit).filter_by(unit_number=record.unit_number).first()
        if not unit:
            continue

        old_val = unit.podil_scd
        unit.podil_scd = record.file_share
        record.db_share = record.file_share
        record.status = ShareCheckStatus.MATCH
        record.resolution = ShareCheckResolution.UPDATED
        record.admin_note = f"Aktualizováno {now}: {old_val} → {record.file_share}"

        success_count += 1
        change_details.append(f"J. {record.unit_number}: {old_val} → {record.file_share}")

    if record_ids:
        log = ImportLog(
            filename=f"share_check_{session_id}",
            file_path=f"share_check_update/{session_id}",
            import_type="share_check_update",
            rows_total=len(record_ids),
            rows_imported=success_count,
            rows_skipped=len(record_ids) - success_count,
            errors="\n".join(change_details) if change_details else None,
        )
        db.add(log)

        # Recalculate session totals
        all_records = db.query(ShareCheckRecord).filter_by(session_id=session_id).all()
        session.total_matches = sum(1 for r in all_records if r.status == ShareCheckStatus.MATCH)
        session.total_differences = sum(1 for r in all_records if r.status == ShareCheckStatus.DIFFERENCE)
        session.total_missing_db = sum(1 for r in all_records if r.status == ShareCheckStatus.MISSING_DB)
        session.total_missing_file = sum(1 for r in all_records if r.status == ShareCheckStatus.MISSING_FILE)

    db.commit()
    filtr = form.get("filtr", "")
    url = f"/kontrola-podilu/{session_id}"
    if filtr:
        url += f"?filtr={filtr}"
    return RedirectResponse(url, status_code=302)
