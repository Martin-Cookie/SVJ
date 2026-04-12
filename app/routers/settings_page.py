import csv
import io
import logging
import smtplib
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, Form, Query, Request
from fastapi.responses import FileResponse, RedirectResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from sqlalchemy import or_

from app.config import settings
from app.database import get_db
from app.models import ActivityAction, EmailLog, Owner, SmtpProfile, SvjInfo, log_activity
from app.utils import (
    build_list_url, decode_smtp_password, encode_smtp_password,
    excel_auto_width, is_htmx_partial, is_safe_path, strip_diacritics, templates,
)

logger = logging.getLogger(__name__)

router = APIRouter()



def _parse_attachments(raw: Optional[str]) -> list:
    """Parse attachment_paths into list of {name, path, exists}.

    Supports both old format (just filenames) and new format (full paths).
    """
    if not raw:
        return []
    result = []
    for part in raw.split(", "):
        part = part.strip()
        if not part:
            continue
        p = Path(part)
        if p.is_absolute():
            result.append({"name": p.name, "path": str(p), "exists": p.exists()})
        else:
            # Old format — just filename, no path available
            result.append({"name": part, "path": "", "exists": False})
    return result


SORT_COLUMNS = {
    "date": EmailLog.created_at,
    "module": EmailLog.module,
    "recipient": EmailLog.recipient_name,
    "subject": EmailLog.subject,
    "status": EmailLog.status,
}


@router.get("/")
async def settings_view(
    request: Request,
    db: Session = Depends(get_db),
    q: str = Query(""),
    sort: str = Query("date"),
    order: str = Query("desc"),
):
    """Stránka nastavení s historií odeslaných emailů a SMTP konfigurací."""
    # Build query
    query = db.query(EmailLog)

    # Search — SQL filter for email/subject/module + diacritics-insensitive name via name_normalized
    if q:
        q_pattern = f"%{q}%"
        q_ascii = f"%{strip_diacritics(q)}%"
        query = query.filter(
            or_(
                EmailLog.recipient_email.ilike(q_pattern),
                EmailLog.recipient_name.ilike(q_pattern),
                EmailLog.name_normalized.like(q_ascii),
                EmailLog.subject.ilike(q_pattern),
                EmailLog.module.ilike(q_pattern),
            )
        )

    # SQL sort + limit
    col = SORT_COLUMNS.get(sort, EmailLog.created_at)
    if order == "asc":
        query = query.order_by(col.asc().nulls_last())
    else:
        query = query.order_by(col.desc().nulls_last())
    email_logs = query.all()

    # Build email → owner_id lookup for clickable recipients
    emails_in_log = {e.recipient_email for e in email_logs if e.recipient_email}
    owner_by_email = {}
    if emails_in_log:
        owners = db.query(Owner.id, Owner.email, Owner.email_secondary).filter(
            Owner.is_active == True
        ).all()
        for o in owners:
            if o.email:
                owner_by_email[o.email.lower()] = o.id
            if o.email_secondary:
                owner_by_email[o.email_secondary.lower()] = o.id

    # Parse attachments for each email log
    attachments_by_id = {e.id: _parse_attachments(e.attachment_paths) for e in email_logs}

    # HTMX partial
    # Build list_url for back navigation
    list_url = build_list_url(request)

    # Globální nastavení odesílání
    svj = db.query(SvjInfo).first()

    # SMTP profily
    smtp_profiles = db.query(SmtpProfile).order_by(SmtpProfile.is_default.desc(), SmtpProfile.id).all()
    profile_count = len(smtp_profiles)

    ctx = {
        "request": request,
        "active_nav": "settings",
        "settings": settings,
        "svj": svj,
        "smtp_profiles": smtp_profiles,
        "profile_count": profile_count,
        "email_logs": email_logs,
        "owner_by_email": owner_by_email,
        "attachments_by_id": attachments_by_id,
        "list_url": list_url,
        "q": q,
        "sort": sort,
        "order": order,
    }

    # Flash zprávy
    flash = request.query_params.get("flash", "")
    if flash == "send_ok":
        ctx["flash_message"] = "Výchozí nastavení odesílání uloženo"

    if is_htmx_partial(request):
        return templates.TemplateResponse(request, "partials/settings_email_tbody.html", ctx)
    return templates.TemplateResponse(request, "settings.html", ctx)


@router.get("/exportovat/{fmt}")
async def email_log_export(
    fmt: str,
    q: str = Query(""),
    sort: str = Query("date"),
    order: str = Query("desc"),
    db: Session = Depends(get_db),
):
    """Export historie odeslaných emailů do Excelu nebo CSV."""
    if fmt not in ("xlsx", "csv"):
        return RedirectResponse("/nastaveni", status_code=302)

    query = db.query(EmailLog)
    if q:
        q_pattern = f"%{q}%"
        q_ascii = f"%{strip_diacritics(q)}%"
        query = query.filter(
            or_(
                EmailLog.recipient_email.ilike(q_pattern),
                EmailLog.recipient_name.ilike(q_pattern),
                EmailLog.name_normalized.like(q_ascii),
                EmailLog.subject.ilike(q_pattern),
                EmailLog.module.ilike(q_pattern),
            )
        )

    col = SORT_COLUMNS.get(sort, EmailLog.created_at)
    if order == "asc":
        query = query.order_by(col.asc().nulls_last())
    else:
        query = query.order_by(col.desc().nulls_last())
    email_logs = query.all()

    headers = ["Datum", "Modul", "Prijemce", "Email", "Predmet", "Prilohy", "Stav"]
    status_labels = {"sent": "OK", "failed": "Chyba", "pending": "Čeká"}
    rows = []
    for e in email_logs:
        attachments = ", ".join(a["name"] for a in _parse_attachments(e.attachment_paths))
        rows.append([
            e.created_at.strftime("%d.%m.%Y %H:%M") if e.created_at else "",
            e.module or "",
            e.recipient_name or "",
            e.recipient_email or "",
            e.subject or "",
            attachments,
            status_labels.get(e.status.value if e.status else "", e.status.value if e.status else ""),
        ])

    timestamp = datetime.now().strftime("%Y%m%d")
    suffix = "_hledani" if q else "_vse"
    filename = f"emaily{suffix}_{timestamp}.{fmt}"

    if fmt == "csv":
        buf = io.StringIO()
        buf.write("\ufeff")
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(headers)
        writer.writerows(rows)
        return Response(
            content=buf.getvalue().encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Historie emailu"
    bold = Font(bold=True)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h).font = bold
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    excel_auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/odesilani")
async def save_send_settings(
    request: Request,
    send_batch_size: int = Form(10),
    send_batch_interval: int = Form(5),
    send_confirm_each_batch: Optional[str] = Form(None),
    send_test_email_address: str = Form(""),
    db: Session = Depends(get_db),
):
    """Uložit výchozí globální nastavení odesílání."""
    svj = db.query(SvjInfo).first()
    if svj:
        svj.send_batch_size = max(1, min(100, send_batch_size))
        svj.send_batch_interval = max(1, min(60, send_batch_interval))
        svj.send_confirm_each_batch = send_confirm_each_batch == "true"
        svj.send_test_email_address = send_test_email_address.strip() or svj.send_test_email_address
        db.commit()
    return RedirectResponse("/nastaveni?flash=send_ok", status_code=302)


def _smtp_profiles_ctx(db: Session, **extra) -> dict:
    """Sdílený kontext pro smtp_info partial."""
    profiles = db.query(SmtpProfile).order_by(SmtpProfile.is_default.desc(), SmtpProfile.id).all()
    return {"smtp_profiles": profiles, "profile_count": len(profiles), **extra}


@router.get("/smtp/profily")
async def smtp_profiles_list(request: Request, db: Session = Depends(get_db)):
    """HTMX partial: seznam SMTP profilů."""
    return templates.TemplateResponse(request, "partials/smtp_info.html", _smtp_profiles_ctx(db))


@router.get("/smtp/novy-formular")
async def smtp_new_form(request: Request, db: Session = Depends(get_db)):
    """HTMX partial: formulář pro nový profil."""
    profiles = db.query(SmtpProfile).all()
    if len(profiles) >= 3:
        return templates.TemplateResponse(request, "partials/smtp_info.html",
            _smtp_profiles_ctx(db, error="Maximum 3 SMTP profily."))
    # Zobrazit existující profily + nový formulář na konci
    return templates.TemplateResponse(request, "partials/smtp_info_with_new.html", {
        **_smtp_profiles_ctx(db),
        "profile": None,
    })


@router.get("/smtp/{profile_id}/formular")
async def smtp_edit_form(request: Request, profile_id: int, db: Session = Depends(get_db)):
    """HTMX partial: editační formulář jednoho profilu (swap za kartu)."""
    profile = db.query(SmtpProfile).get(profile_id)
    if not profile:
        return templates.TemplateResponse(request, "partials/smtp_info.html", _smtp_profiles_ctx(db))
    profiles = db.query(SmtpProfile).all()
    return templates.TemplateResponse(request, "partials/smtp_profile_form.html", {
        "profile": profile,
        "profile_count": len(profiles),
    })


@router.get("/smtp/{profile_id}/karta")
async def smtp_profile_card(request: Request, profile_id: int, db: Session = Depends(get_db)):
    """HTMX partial: read-only karta jednoho profilu."""
    profile = db.query(SmtpProfile).get(profile_id)
    if not profile:
        return templates.TemplateResponse(request, "partials/smtp_info.html", _smtp_profiles_ctx(db))
    profile_count = db.query(SmtpProfile).count()
    return templates.TemplateResponse(request, "partials/smtp_profile_card.html", {
        "profile": profile,
        "profile_count": profile_count,
    })


@router.post("/smtp/novy")
async def smtp_create(
    request: Request,
    name: str = Form(""),
    smtp_host: str = Form(""),
    smtp_port: int = Form(465),
    smtp_user: str = Form(""),
    smtp_password: str = Form(""),
    smtp_from_name: str = Form(""),
    smtp_from_email: str = Form(""),
    smtp_use_tls: Optional[str] = Form(None),
    imap_save_sent: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """Vytvoření nového SMTP profilu."""
    count = db.query(SmtpProfile).count()
    if count >= 3:
        return templates.TemplateResponse(request, "partials/smtp_info.html",
            _smtp_profiles_ctx(db, error="Maximum 3 SMTP profily."))

    if not name.strip() or not smtp_host.strip() or not smtp_from_email.strip():
        return templates.TemplateResponse(request, "partials/smtp_info_with_new.html", {
            **_smtp_profiles_ctx(db),
            "profile": None,
            "error": "Název, SMTP server a email odesílatele jsou povinné.",
        })

    # Placeholder •••• není skutečné heslo
    if smtp_password and all(c == "•" for c in smtp_password):
        smtp_password = ""

    if not smtp_password:
        return templates.TemplateResponse(request, "partials/smtp_info_with_new.html", {
            **_smtp_profiles_ctx(db),
            "profile": None,
            "error": "Heslo je povinné pro nový profil.",
        })

    is_default = count == 0  # první profil je automaticky výchozí
    profile = SmtpProfile(
        name=name.strip(),
        smtp_host=smtp_host.strip(),
        smtp_port=smtp_port,
        smtp_user=smtp_user.strip(),
        smtp_password_b64=encode_smtp_password(smtp_password),
        smtp_from_name=smtp_from_name.strip(),
        smtp_from_email=smtp_from_email.strip(),
        smtp_use_tls=smtp_use_tls == "true",
        imap_save_sent=imap_save_sent == "true",
        is_default=is_default,
    )
    db.add(profile)
    log_activity(db, ActivityAction.CREATED, "smtp_profile", str(profile.id),
                 entity_name=name.strip(), description=smtp_host)
    db.commit()

    return templates.TemplateResponse(request, "partials/smtp_info.html",
        _smtp_profiles_ctx(db, saved=True))


@router.post("/smtp/{profile_id}")
async def smtp_update(
    request: Request,
    profile_id: int,
    name: str = Form(""),
    smtp_host: str = Form(""),
    smtp_port: int = Form(465),
    smtp_user: str = Form(""),
    smtp_password: str = Form(""),
    smtp_from_name: str = Form(""),
    smtp_from_email: str = Form(""),
    smtp_use_tls: Optional[str] = Form(None),
    imap_save_sent: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """Aktualizace existujícího SMTP profilu."""
    profile = db.query(SmtpProfile).get(profile_id)
    if not profile:
        return templates.TemplateResponse(request, "partials/smtp_info.html", _smtp_profiles_ctx(db))

    if not name.strip() or not smtp_host.strip() or not smtp_from_email.strip():
        return templates.TemplateResponse(request, "partials/smtp_profile_form.html", {
            "profile": profile,
            "profile_count": db.query(SmtpProfile).count(),
            "error": "Název, SMTP server a email odesílatele jsou povinné.",
        })

    profile.name = name.strip()
    profile.smtp_host = smtp_host.strip()
    profile.smtp_port = smtp_port
    profile.smtp_user = smtp_user.strip()
    if smtp_password and not all(c == "•" for c in smtp_password):
        profile.smtp_password_b64 = encode_smtp_password(smtp_password)
    profile.smtp_from_name = smtp_from_name.strip()
    profile.smtp_from_email = smtp_from_email.strip()
    profile.smtp_use_tls = smtp_use_tls == "true"
    profile.imap_save_sent = imap_save_sent == "true"

    log_activity(db, ActivityAction.UPDATED, "smtp_profile", str(profile_id),
                 entity_name=name.strip(), description=smtp_host)
    db.commit()

    return templates.TemplateResponse(request, "partials/smtp_info.html",
        _smtp_profiles_ctx(db, saved=True))


@router.post("/smtp/{profile_id}/test")
async def test_smtp_profile(request: Request, profile_id: int, db: Session = Depends(get_db)):
    """Test SMTP připojení pro konkrétní profil."""
    profile = db.query(SmtpProfile).get(profile_id)
    if not profile:
        return templates.TemplateResponse(request, "partials/smtp_info.html", _smtp_profiles_ctx(db))

    profile_count = db.query(SmtpProfile).count()
    try:
        from app.services.email_service import _create_smtp
        password = decode_smtp_password(profile.smtp_password_b64) if profile.smtp_password_b64 else ""
        server = _create_smtp(profile.smtp_host, profile.smtp_port, profile.smtp_use_tls, timeout=10)
        if profile.smtp_user:
            server.login(profile.smtp_user, password)
        server.quit()
        return templates.TemplateResponse(request, "partials/smtp_profile_card.html", {
            "profile": profile, "profile_count": profile_count, "smtp_test_ok": True,
        })
    except smtplib.SMTPAuthenticationError:
        return templates.TemplateResponse(request, "partials/smtp_profile_card.html", {
            "profile": profile, "profile_count": profile_count,
            "smtp_test_error": "Přihlášení selhalo — zkontrolujte uživatele a heslo.",
        })
    except Exception as e:
        logger.warning("SMTP test failed for profile %s: %s", profile_id, e)
        return templates.TemplateResponse(request, "partials/smtp_profile_card.html", {
            "profile": profile, "profile_count": profile_count,
            "smtp_test_error": f"Připojení selhalo: {e}",
        })


@router.post("/smtp/{profile_id}/vychozi")
async def smtp_set_default(request: Request, profile_id: int, db: Session = Depends(get_db)):
    """Nastavit profil jako výchozí."""
    profile = db.query(SmtpProfile).get(profile_id)
    if not profile:
        return templates.TemplateResponse(request, "partials/smtp_info.html", _smtp_profiles_ctx(db))

    # Reset is_default na všech
    db.query(SmtpProfile).update({"is_default": False})
    profile.is_default = True
    db.commit()

    return templates.TemplateResponse(request, "partials/smtp_info.html",
        _smtp_profiles_ctx(db, saved=True))


@router.post("/smtp/{profile_id}/smazat")
async def smtp_delete(request: Request, profile_id: int, db: Session = Depends(get_db)):
    """Smazat SMTP profil (nelze smazat jediný)."""
    count = db.query(SmtpProfile).count()
    if count <= 1:
        return templates.TemplateResponse(request, "partials/smtp_info.html",
            _smtp_profiles_ctx(db, error="Nelze smazat jediný SMTP profil."))

    profile = db.query(SmtpProfile).get(profile_id)
    if not profile:
        return templates.TemplateResponse(request, "partials/smtp_info.html", _smtp_profiles_ctx(db))

    was_default = profile.is_default
    log_activity(db, ActivityAction.DELETED, "smtp_profile", str(profile_id),
                 entity_name=profile.name)
    db.delete(profile)
    db.commit()

    # Pokud smazaný byl výchozí, nastavit první zbývající
    if was_default:
        first = db.query(SmtpProfile).first()
        if first:
            first.is_default = True
            db.commit()

    return templates.TemplateResponse(request, "partials/smtp_info.html",
        _smtp_profiles_ctx(db, saved=True))


@router.get("/priloha/{log_id}/{filename}")
async def serve_attachment(
    log_id: int,
    filename: str,
    db: Session = Depends(get_db),
):
    """Serve an email attachment file for in-browser preview."""
    log = db.query(EmailLog).get(log_id)
    if not log or not log.attachment_paths:
        return RedirectResponse("/nastaveni", status_code=302)

    # Find matching path in attachment_paths
    for part in log.attachment_paths.split(", "):
        part = part.strip()
        p = Path(part)
        if p.name == filename and p.is_absolute() and p.exists():
            # Validate path is within allowed directories
            if is_safe_path(p, settings.upload_dir, settings.generated_dir):
                suffix = p.suffix.lower()
                media_types = {
                    ".pdf": "application/pdf",
                    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ".xls": "application/vnd.ms-excel",
                    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    ".csv": "text/csv",
                }
                media_type = media_types.get(suffix, "application/octet-stream")
                return FileResponse(str(p), media_type=media_type, filename=p.name)

    return RedirectResponse("/nastaveni", status_code=302)
