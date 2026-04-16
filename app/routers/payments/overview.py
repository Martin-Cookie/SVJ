"""Přehled plateb — matice, dlužníci, detail jednotky."""

import io
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, Form, Query, Request
from fastapi.responses import RedirectResponse, Response, StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import PrescriptionYear, Unit, Space
from app.services.payment_overview import (
    compute_debtor_list,
    compute_payment_matrix,
    compute_unit_payment_detail,
    compute_space_debtor_list,
    compute_space_payment_matrix,
    compute_space_payment_detail,
)
from app.utils import build_list_url, excel_auto_width, is_htmx_partial, strip_diacritics, utcnow

from ._helpers import templates, compute_nav_stats, MONTH_NAMES_SHORT

router = APIRouter()

MONTH_NAMES = MONTH_NAMES_SHORT


# ── Matice plateb ────────────────────────────────────────────────────


SORT_COLUMNS_MATRIX = {
    "cislo": None,     # Python sort
    "sekce": None,
    "vlastnik": None,
    "predpis": None,
    "prevod": None,
    "celkem": None,
    "saldo": None,
    # Měsíční sloupce m1–m12
    **{f"m{i}": None for i in range(1, 13)},
}


def _matrix_sort_fns(entita: str) -> dict:
    """Sort fn mapa pro matici plateb — sdílená mezi GET view a exportem."""
    if entita == "prostory":
        fns = {
            "cislo": lambda r: r["space"].space_number or "",
            "sekce": lambda r: "",
            "vlastnik": lambda r: strip_diacritics(
                r["tenant_rel"].tenant.display_name if r.get("tenant_rel") and r["tenant_rel"].tenant else ""
            ),
            "predpis": lambda r: r["monthly"],
            "prevod": lambda r: r.get("opening", 0),
            "celkem": lambda r: r["total_paid"],
            "saldo": lambda r: r["saldo"],
        }
    else:
        fns = {
            "cislo": lambda r: r["unit"].unit_number or 0,
            "sekce": lambda r: (r["prescription"].section or "").lower(),
            "vlastnik": lambda r: strip_diacritics(r["owner"].display_name if r["owner"] else r["prescription"].owner_name or ""),
            "predpis": lambda r: r["monthly"],
            "prevod": lambda r: r.get("opening", 0),
            "celkem": lambda r: r["total_paid"],
            "saldo": lambda r: r["saldo"],
        }
    for mi in range(1, 13):
        fns[f"m{mi}"] = (lambda m: lambda r: r["months"].get(m, {}).get("paid", 0))(mi)
    return fns


def _sorted_owners_names(owners) -> str:
    """Stabilní pořadí spoluvlastníků pro export — podle name_normalized."""
    if not owners:
        return ""
    return ", ".join(o.display_name for o in sorted(owners, key=lambda o: (o.name_normalized or "").lower()))


@router.get("/prehled")
async def platby_prehled(
    request: Request,
    rok: int = Query(0),
    typ: str = Query(""),
    q: str = Query(""),
    sort: str = Query("cislo"),
    order: str = Query("asc"),
    vse_mesice: int = Query(0, alias="vse_mesice"),
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    """Matice plateb — jednotky/prostory × měsíce."""
    entita = request.query_params.get("entita", "")
    show_all_months = bool(vse_mesice)

    # Výchozí rok = nejnovější PrescriptionYear
    if not rok:
        latest = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        rok = latest.year if latest else utcnow().year

    years = [y.year for y in db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).all()]

    if entita == "prostory":
        matrix = compute_space_payment_matrix(db, rok)
        rows = matrix["rows"]

        if q:
            q_ascii = strip_diacritics(q)
            rows = [
                r for r in rows
                if q_ascii in strip_diacritics(str(r["space"].space_number))
                or q_ascii in strip_diacritics(r["space"].designation or "")
                or q_ascii in strip_diacritics(
                    r["tenant_rel"].tenant.display_name if r.get("tenant_rel") and r["tenant_rel"].tenant else ""
                )
            ]

        total_units = len(matrix["rows"])
        space_types = []
        # Counts pro stat-card bubliny Jednotky/Prostory
        if not is_htmx_partial(request):
            units_matrix_all = compute_payment_matrix(db, rok, space_type="")
            units_count_total = len(units_matrix_all["units"])
        else:
            units_count_total = 0
        spaces_count_total = len(matrix["rows"])
    else:
        matrix = compute_payment_matrix(db, rok, space_type=typ)
        rows = matrix["units"]

        if q:
            q_ascii = strip_diacritics(q)
            rows = [
                r for r in rows
                if q_ascii in strip_diacritics(str(r["unit"].unit_number))
                or q_ascii in strip_diacritics(r["owner"].display_name if r["owner"] else "")
                or q_ascii in strip_diacritics(r["prescription"].owner_name or "")
                or q_ascii in strip_diacritics(r["prescription"].variable_symbol or "")
                or q_ascii in strip_diacritics(r["prescription"].section or "")
            ]

        total_units = len(matrix["units"])
        space_types = matrix["space_types"]
        # Counts pro stat-card bubliny
        if not is_htmx_partial(request):
            if typ:
                units_all = compute_payment_matrix(db, rok, space_type="")
                units_count_total = len(units_all["units"])
            else:
                units_count_total = total_units
            spaces_all = compute_space_payment_matrix(db, rok)
            spaces_count_total = len(spaces_all["rows"])
        else:
            units_count_total = 0
            spaces_count_total = 0

    sort_key = sort if sort in SORT_COLUMNS_MATRIX else "cislo"
    reverse = order == "desc"
    sort_fns = _matrix_sort_fns(entita)
    rows.sort(key=sort_fns.get(sort_key, sort_fns["cislo"]), reverse=reverse)

    list_url = build_list_url(request)

    ctx = {
        "request": request,
        "active_nav": "platby",
        "rows": rows,
        "rok": rok,
        "years": years,
        "typ": typ,
        "q": q,
        "sort": sort_key,
        "order": order,
        "entita": entita,
        "back_url": back,
        "list_url": list_url,
        "month_names": MONTH_NAMES,
        "months_with_data": matrix["months_with_data"],
        "show_all_months": show_all_months,
        "space_types": space_types,
        "total_prescribed": matrix["total_prescribed"],
        "total_paid": matrix["total_paid"],
        "total_units": total_units,
        "units_count": units_count_total,
        "spaces_count": spaces_count_total,
        "active_tab": "prehled",
        **(compute_nav_stats(db) if not is_htmx_partial(request) else {}),
    }

    if is_htmx_partial(request):
        return templates.TemplateResponse(request, "payments/partials/prehled_tbody.html", ctx)

    return templates.TemplateResponse(request, "payments/prehled.html", ctx)


# ── Export matice plateb ──────────────────────────────────────────────


@router.get("/prehled/exportovat/{fmt}")
async def matice_export(
    request: Request,
    fmt: str,
    rok: int = Query(0),
    typ: str = Query(""),
    q: str = Query(""),
    sort: str = Query("cislo"),
    order: str = Query("asc"),
    entita: str = Query(""),
    db: Session = Depends(get_db),
):
    """Export matice plateb do Excelu nebo CSV — jednotky i prostory."""
    import csv as csv_mod
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    if fmt not in ("xlsx", "csv"):
        return RedirectResponse("/platby/prehled", status_code=302)

    if not rok:
        latest = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        rok = latest.year if latest else utcnow().year

    if entita == "prostory":
        matrix = compute_space_payment_matrix(db, rok)
        rows = matrix["rows"]
        if q:
            q_ascii = strip_diacritics(q)
            rows = [
                r for r in rows
                if q_ascii in strip_diacritics(str(r["space"].space_number))
                or q_ascii in strip_diacritics(r["space"].designation or "")
                or q_ascii in strip_diacritics(
                    r["tenant_rel"].tenant.display_name if r.get("tenant_rel") and r["tenant_rel"].tenant else ""
                )
            ]
    else:
        matrix = compute_payment_matrix(db, rok, space_type=typ)
        rows = matrix["units"]
        if q:
            q_ascii = strip_diacritics(q)
            rows = [
                r for r in rows
                if q_ascii in strip_diacritics(str(r["unit"].unit_number))
                or q_ascii in strip_diacritics(r["owner"].display_name if r["owner"] else "")
                or q_ascii in strip_diacritics(r["prescription"].owner_name or "")
                or q_ascii in strip_diacritics(r["prescription"].variable_symbol or "")
                or q_ascii in strip_diacritics(r["prescription"].section or "")
            ]

    months_with_data = sorted(matrix["months_with_data"])

    sort_fns = _matrix_sort_fns(entita)
    rows.sort(key=sort_fns.get(sort, sort_fns["cislo"]), reverse=(order == "desc"))

    month_labels = ["Led", "Uno", "Bre", "Dub", "Kve", "Cer", "Cvc", "Srp", "Zar", "Rij", "Lis", "Pro"]
    if entita == "prostory":
        headers = ["Č. prost.", "Označení", "Nájemce", "Předpis/měs", "Převod"]
    else:
        headers = ["Č. jedn.", "Sekce", "Vlastník", "Předpis/měs", "Převod"]
    for m in months_with_data:
        headers.append(month_labels[m - 1])
    headers += ["Celkem", "Saldo"]

    typ_suffix = f"_{strip_diacritics(typ)}" if typ else ""
    q_suffix = f"_hledani" if q else ""
    suffix = typ_suffix or q_suffix or "_vse"
    date_str = datetime.now().strftime("%Y%m%d")
    base_name = "matice_prostor" if entita == "prostory" else "matice_plateb"

    def _row_values(r):
        """Vrátí (cislo, popis, jmeno) podle entity."""
        if entita == "prostory":
            space = r["space"]
            tenant = r.get("tenant_rel")
            name = tenant.tenant.display_name if tenant and tenant.tenant else ""
            return (space.space_number, space.designation or "", name)
        else:
            owner_name = _sorted_owners_names(r["owners"]) or (r["prescription"].owner_name or "")
            return (r["unit"].unit_number, r["prescription"].section or "", owner_name)

    if fmt == "csv":
        buf_csv = io.StringIO()
        buf_csv.write("\ufeff")
        writer = csv_mod.writer(buf_csv, delimiter=";")
        writer.writerow(headers)
        for r in rows:
            cislo, popis, jmeno = _row_values(r)
            row_out = [cislo, popis, jmeno, r["monthly"], r.get("opening", 0)]
            for m in months_with_data:
                row_out.append(r["months"].get(m, {}).get("paid", 0))
            row_out += [r["total_paid"], r["saldo"]]
            writer.writerow(row_out)
        filename = f"{base_name}_{rok}{suffix}_{date_str}.csv"
        return Response(
            content=buf_csv.getvalue().encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Matice {rok}"

    bold = Font(bold=True)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold

    for i, r in enumerate(rows, 2):
        cislo, popis, jmeno = _row_values(r)
        ws.cell(row=i, column=1, value=cislo)
        ws.cell(row=i, column=2, value=popis)
        ws.cell(row=i, column=3, value=jmeno)
        ws.cell(row=i, column=4, value=r["monthly"])
        ws.cell(row=i, column=5, value=r.get("opening", 0))
        col = 6
        for m in months_with_data:
            paid = r["months"].get(m, {}).get("paid", 0)
            cell = ws.cell(row=i, column=col, value=paid)
            if paid < r["monthly"] and r["monthly"] > 0:
                cell.fill = red_fill
            col += 1
        ws.cell(row=i, column=col, value=r["total_paid"])
        saldo_cell = ws.cell(row=i, column=col + 1, value=r["saldo"])
        if r["saldo"] < 0:
            saldo_cell.fill = red_fill

    excel_auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{base_name}_{rok}{suffix}_{date_str}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Dlužníci ─────────────────────────────────────────────────────────


SORT_COLUMNS_DEBTORS = {
    "cislo": None,
    "vlastnik": None,
    "predpis": None,
    "zaplaceno": None,
    "saldo": None,
    "mesice": None,
}


def _compute_debtors_filtered(
    db: Session,
    rok: int,
    q: str = "",
    sort: str = "saldo",
    order: str = "desc",
    entita: str = "",
):
    """Sdílená logika — výpočet + filtrování + řazení dlužníků (jednotky nebo prostory)."""
    if entita == "prostory":
        debtors, months_with_data = compute_space_debtor_list(db, rok)

        for r in debtors:
            r["months_unpaid"] = sum(
                1 for m in range(1, 13)
                if r["months"].get(m, {}).get("status") in ("unpaid", "partial")
                and m in months_with_data
            )

        if q:
            q_ascii = strip_diacritics(q)
            debtors = [
                r for r in debtors
                if q_ascii in strip_diacritics(str(r["space"].space_number))
                or q_ascii in strip_diacritics(r["space"].designation or "")
                or q_ascii in strip_diacritics(
                    r["tenant_rel"].tenant.display_name if r.get("tenant_rel") and r["tenant_rel"].tenant else ""
                )
            ]

        sort_key = sort if sort in SORT_COLUMNS_DEBTORS else "saldo"
        reverse = order == "desc"
        sort_fns = {
            "cislo": lambda r: r["space"].space_number or "",
            "vlastnik": lambda r: strip_diacritics(
                r["tenant_rel"].tenant.display_name if r.get("tenant_rel") and r["tenant_rel"].tenant else ""
            ),
            "predpis": lambda r: r["monthly"],
            "zaplaceno": lambda r: r["total_paid"],
            "saldo": lambda r: r["saldo"],
            "mesice": lambda r: r["months_unpaid"],
        }
        debtors.sort(key=sort_fns.get(sort_key, sort_fns["saldo"]), reverse=reverse)
    else:
        debtors, months_with_data = compute_debtor_list(db, rok)

        for r in debtors:
            r["months_unpaid"] = sum(
                1 for m in range(1, 13)
                if r["months"].get(m, {}).get("status") in ("unpaid", "partial")
                and m in months_with_data
            )

        if q:
            q_ascii = strip_diacritics(q)
            debtors = [
                r for r in debtors
                if q_ascii in strip_diacritics(str(r["unit"].unit_number))
                or q_ascii in strip_diacritics(r["owner"].display_name if r["owner"] else "")
                or q_ascii in strip_diacritics(r["prescription"].owner_name or "")
            ]

        sort_key = sort if sort in SORT_COLUMNS_DEBTORS else "saldo"
        reverse = order == "desc"
        sort_fns = {
            "cislo": lambda r: r["unit"].unit_number or 0,
            "vlastnik": lambda r: strip_diacritics(r["owner"].display_name if r["owner"] else r["prescription"].owner_name or ""),
            "predpis": lambda r: r["monthly"],
            "zaplaceno": lambda r: r["total_paid"],
            "saldo": lambda r: r["saldo"],
            "mesice": lambda r: r["months_unpaid"],
        }
        debtors.sort(key=sort_fns.get(sort_key, sort_fns["saldo"]), reverse=reverse)

    return debtors, months_with_data, sort_key


@router.get("/dluznici")
async def platby_dluznici(
    request: Request,
    rok: int = Query(0),
    q: str = Query(""),
    sort: str = Query("saldo"),
    order: str = Query("desc"),
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    """Seznam dlužníků — jednotky nebo prostory."""
    entita = request.query_params.get("entita", "")

    if not rok:
        latest = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        rok = latest.year if latest else utcnow().year

    years = [y.year for y in db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).all()]

    debtors, months_with_data, sort_key = _compute_debtors_filtered(db, rok, q, sort, order, entita)

    # Counts pro stat-card bubliny (obě entity)
    units_count = 0
    spaces_count = 0
    if not is_htmx_partial(request):
        u_list, _ = compute_debtor_list(db, rok)
        units_count = len(u_list)
        s_list, _ = compute_space_debtor_list(db, rok)
        spaces_count = len(s_list)

    list_url = build_list_url(request)

    ctx = {
        "request": request,
        "active_nav": "platby",
        "debtors": debtors,
        "rok": rok,
        "years": years,
        "q": q,
        "sort": sort_key,
        "order": order,
        "entita": entita,
        "back_url": back,
        "list_url": list_url,
        "months_with_data": months_with_data,
        "total_debt": sum(-r["saldo"] for r in debtors),
        "units_count": units_count,
        "spaces_count": spaces_count,
        "active_tab": "dluznici",
        **(compute_nav_stats(db) if not is_htmx_partial(request) else {}),
    }

    if is_htmx_partial(request):
        return templates.TemplateResponse(request, "payments/partials/dluznici_tbody.html", ctx)

    return templates.TemplateResponse(request, "payments/dluznici.html", ctx)


# ── Export dlužníků ──────────────────────────────────────────────────


@router.get("/dluznici/exportovat/{fmt}")
async def dluznici_export(
    fmt: str,
    rok: int = Query(0),
    q: str = Query(""),
    sort: str = Query("saldo"),
    order: str = Query("desc"),
    entita: str = Query(""),
    db: Session = Depends(get_db),
):
    """Export dlužníků (xlsx/csv) — jednotky nebo prostory, respektuje filtry."""
    import csv as csv_mod
    import io as io_mod
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if fmt not in ("xlsx", "csv"):
        return RedirectResponse("/platby/dluznici", status_code=302)

    if not rok:
        latest = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        rok = latest.year if latest else utcnow().year

    debtors, _months_with_data, _sort_key = _compute_debtors_filtered(db, rok, q, sort, order, entita)

    is_spaces = entita == "prostory"
    if is_spaces:
        headers = ["Č. prostoru", "Označení", "Nájemce", "Předpis/měs", "Zaplaceno", "Saldo", "Měsíce"]

        def _row(r):
            tenant_name = ""
            if r.get("tenant_rel") and r["tenant_rel"].tenant:
                tenant_name = r["tenant_rel"].tenant.display_name or ""
            return [
                r["space"].space_number,
                r["space"].designation or "",
                tenant_name,
                r["monthly"],
                r["total_paid"],
                r["saldo"],
                r["months_unpaid"],
            ]
    else:
        headers = ["Katastr. č.", "Vlastník", "Předpis/měs", "Zaplaceno", "Saldo", "Měsíce"]

        def _row(r):
            owner_name = ", ".join(o.display_name for o in r["owners"]) if r["owners"] else (r["prescription"].owner_name or "")
            return [
                r["unit"].unit_number,
                owner_name,
                r["monthly"],
                r["total_paid"],
                r["saldo"],
                r["months_unpaid"],
            ]

    # Suffix dle filtru
    if entita == "prostory":
        suffix = "_prostory"
    elif q:
        suffix = "_hledani"
    else:
        suffix = "_vsichni"
    date_str = datetime.now().strftime("%Y%m%d")
    filename = f"dluznici_{rok}{suffix}_{date_str}"

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = f"Dlužníci {rok}"

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        for i, r in enumerate(debtors, 2):
            for col_idx, val in enumerate(_row(r), 1):
                ws.cell(row=i, column=col_idx, value=val)

        excel_auto_width(ws)

        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    else:
        buf = io_mod.StringIO()
        writer = csv_mod.writer(buf, delimiter=";")
        writer.writerow(headers)
        for r in debtors:
            writer.writerow(_row(r))
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )


# ── Detail plateb jednotky ───────────────────────────────────────────


@router.get("/jednotka/{unit_id}")
async def platby_jednotka(
    unit_id: int,
    request: Request,
    rok: int = Query(0),
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    """Platební detail jedné jednotky."""
    unit = db.query(Unit).get(unit_id)
    if not unit:
        return RedirectResponse("/platby/prehled", status_code=302)

    if not rok:
        latest = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        rok = latest.year if latest else utcnow().year

    years = [y.year for y in db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).all()]

    detail = compute_unit_payment_detail(db, unit_id, rok)
    if not detail:
        return RedirectResponse("/platby/prehled", status_code=302)

    back_url = back or "/platby/prehled"
    if "/platby/prehled" in back_url:
        back_label = "Zpět na matici plateb"
    elif "/platby/dluznici" in back_url:
        back_label = "Zpět na dlužníky"
    elif "/jednotky/" in back_url:
        back_label = "Zpět na detail jednotky"
    else:
        back_label = "Zpět na platby"

    return templates.TemplateResponse(request, "payments/jednotka_platby.html", {
        "active_nav": "platby",
        "active_tab": "prehled",
        "unit": unit,
        "detail": detail,
        "rok": rok,
        "years": years,
        "back_url": back_url,
        "back_label": back_label,
        "hide_nav_back": True,
        "month_names": MONTH_NAMES,
        **(compute_nav_stats(db) if not is_htmx_partial(request) else {}),
    })


# ── Detail plateb prostoru ──────────────────────────────────────────


@router.get("/prostor/{space_id}")
async def platby_prostor(
    space_id: int,
    request: Request,
    rok: int = Query(0),
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    """Platební detail jednoho prostoru."""
    space = db.query(Space).get(space_id)
    if not space:
        return RedirectResponse("/platby/prehled?entita=prostory", status_code=302)

    if not rok:
        latest = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        rok = latest.year if latest else utcnow().year

    years = [y.year for y in db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).all()]

    detail = compute_space_payment_detail(db, space_id, rok)
    if not detail:
        return RedirectResponse("/platby/prehled?entita=prostory", status_code=302)

    back_url = back or "/platby/prehled?entita=prostory"
    if "/platby/prehled" in back_url:
        back_label = "Zpět na matici plateb"
    elif "/platby/dluznici" in back_url:
        back_label = "Zpět na dlužníky"
    elif "/prostory/" in back_url:
        back_label = "Zpět na detail prostoru"
    else:
        back_label = "Zpět na platby"

    return templates.TemplateResponse(request, "payments/prostor_platby.html", {
        "active_nav": "platby",
        "active_tab": "prehled",
        "space": space,
        "detail": detail,
        "rok": rok,
        "years": years,
        "back_url": back_url,
        "back_label": back_label,
        "hide_nav_back": True,
        "month_names": MONTH_NAMES,
        **(compute_nav_stats(db) if not is_htmx_partial(request) else {}),
    })
