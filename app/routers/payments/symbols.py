"""Router pro správu variabilních symbolů (VS → jednotka mapování)."""

import csv
import io
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, Form, Query, Request
from fastapi.responses import RedirectResponse, Response
from sqlalchemy import asc as sa_asc, desc as sa_desc
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import ActivityAction, VariableSymbolMapping, Unit, Space, SymbolSource, log_activity
from app.utils import build_list_url, excel_auto_width, is_htmx_partial, strip_diacritics
from ._helpers import templates, logger, compute_nav_stats

router = APIRouter()


SORT_COLUMNS = {
    "vs": VariableSymbolMapping.variable_symbol,
    "jednotka": None,  # Python-side sort
    "zdroj": VariableSymbolMapping.source,
    "popis": VariableSymbolMapping.description,
}


def _filter_symbols(db: Session, q: str, zdroj: str, entita: str, sort: str, order: str):
    """Sdílená filtrační logika pro seznam i export."""
    query = db.query(VariableSymbolMapping).options(
        joinedload(VariableSymbolMapping.unit),
        joinedload(VariableSymbolMapping.space),
    )

    if q:
        q_escaped = q.replace("%", r"\%").replace("_", r"\_")
        q_ascii = strip_diacritics(q)
        q_ascii_escaped = q_ascii.replace("%", r"\%").replace("_", r"\_")
        query = query.filter(
            VariableSymbolMapping.variable_symbol.like(f"%{q_escaped}%", escape="\\")
            | VariableSymbolMapping.description.like(f"%{q_ascii_escaped}%", escape="\\")
        )

    if zdroj:
        query = query.filter(VariableSymbolMapping.source == zdroj)

    if entita == "jednotky":
        query = query.filter(VariableSymbolMapping.unit_id.isnot(None))
    elif entita == "prostory":
        query = query.filter(VariableSymbolMapping.space_id.isnot(None))

    col = SORT_COLUMNS.get(sort)
    if col is not None:
        order_fn = sa_desc if order == "desc" else sa_asc
        query = query.order_by(order_fn(col).nulls_last())
    else:
        query = query.order_by(VariableSymbolMapping.variable_symbol)

    mappings = query.all()

    if sort == "jednotka":
        mappings.sort(
            key=lambda m: m.unit.unit_number if m.unit else 0,
            reverse=(order == "desc"),
        )

    return mappings


@router.get("/symboly")
async def symboly_seznam(
    request: Request,
    sort: str = "vs",
    order: str = "asc",
    q: str = "",
    zdroj: str = "",
    db: Session = Depends(get_db),
):
    """Seznam VS mapování."""
    entity = request.query_params.get("entita", "")
    mappings = _filter_symbols(db, q, zdroj, entity, sort, order)

    # Jednotky a prostory pro formulář přidání
    units = db.query(Unit).order_by(Unit.unit_number).all()
    spaces = db.query(Space).order_by(Space.space_number).all()

    # Zdroje pro bubliny
    sources = db.query(VariableSymbolMapping.source).distinct().all()
    sources = sorted(set(s[0].value for s in sources if s[0]))

    list_url = build_list_url(request)
    back_url = request.query_params.get("back", "")

    # Flash zprávy
    flash_message = ""
    flash_param = request.query_params.get("flash", "")
    chyba = request.query_params.get("chyba", "")
    if flash_param == "ok":
        flash_message = "Variabilní symbol přidán."
    elif flash_param == "upraveno":
        flash_message = "Variabilní symbol upraven."
    elif flash_param == "smazano":
        flash_message = "Variabilní symbol smazán."
    elif chyba == "duplicita":
        flash_message = "Variabilní symbol již existuje."
    elif chyba == "prazdny":
        flash_message = "Variabilní symbol nesmí být prázdný."

    # Count by entity type for bubbles
    total_vs_count = db.query(VariableSymbolMapping).count()
    unit_vs_count = db.query(VariableSymbolMapping).filter(VariableSymbolMapping.unit_id.isnot(None)).count()
    space_vs_count = db.query(VariableSymbolMapping).filter(VariableSymbolMapping.space_id.isnot(None)).count()

    ctx = {
        "request": request,
        "active_nav": "platby",
        "mappings": mappings,
        "units": units,
        "spaces": spaces,
        "sources": sources,
        "sort": sort,
        "order": order,
        "q": q,
        "zdroj": zdroj,
        "entita": entity,
        "total_vs_count": total_vs_count,
        "unit_vs_count": unit_vs_count,
        "space_vs_count": space_vs_count,
        "list_url": list_url,
        "back_url": back_url,
        "flash_message": flash_message,
        "flash_type": "error" if chyba else "",
        "active_tab": "symboly",
        **(compute_nav_stats(db) if not is_htmx_partial(request) else {}),
    }

    if is_htmx_partial(request):
        return templates.TemplateResponse(request, "payments/partials/symboly_tbody.html", ctx)

    return templates.TemplateResponse(request, "payments/symboly.html", ctx)


@router.get("/symboly/exportovat/{fmt}")
async def symboly_export(
    fmt: str,
    q: str = Query(""),
    zdroj: str = Query(""),
    entita: str = Query(""),
    sort: str = Query("vs"),
    order: str = Query("asc"),
    db: Session = Depends(get_db),
):
    """Export filtrovaných variabilních symbolů."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if fmt not in ("xlsx", "csv"):
        return RedirectResponse("/platby/symboly", status_code=302)

    mappings = _filter_symbols(db, q, zdroj, entita, sort, order)

    headers = ["Variabilní symbol", "Entita", "Číslo", "Označení", "Zdroj", "Popis"]
    source_labels = {"auto": "Automaticky", "manual": "Ručně", "legacy": "Historicky"}

    def _row(m):
        if m.unit:
            entity_type = "Jednotka"
            number = str(m.unit.unit_number)
            designation = ""
        elif m.space:
            entity_type = "Prostor"
            number = str(m.space.space_number)
            designation = m.space.designation or ""
        else:
            entity_type = ""
            number = ""
            designation = ""
        return [
            m.variable_symbol or "",
            entity_type,
            number,
            designation,
            source_labels.get(m.source.value if m.source else "", ""),
            m.description or "",
        ]

    if entita:
        suffix = f"_{entita}"
    elif zdroj:
        suffix = f"_{zdroj}"
    elif q:
        suffix = "_hledani"
    else:
        suffix = "_vsechny"
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"symboly{suffix}_{timestamp}"

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Variabilní symboly"
        bold = Font(bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold

        for row_idx, m in enumerate(mappings, 2):
            for col_idx, val in enumerate(_row(m), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        excel_auto_width(ws)

        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    else:
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(headers)
        for m in mappings:
            writer.writerow(_row(m))
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )


def _symboly_redirect_url(form_data, flash: str = "", chyba: str = "") -> str:
    """Sestaví redirect URL zpět na symboly se zachováním filtrů."""
    params = []
    if flash:
        params.append(f"flash={flash}")
    if chyba:
        params.append(f"chyba={chyba}")
    for key in ("q", "sort", "order", "zdroj", "entita", "back"):
        val = form_data.get(key, "")
        if val:
            params.append(f"{key}={val}")
    qs = "&".join(params)
    return f"/platby/symboly?{qs}" if qs else "/platby/symboly"


@router.post("/symboly/pridat")
async def symbol_pridat(
    request: Request,
    variable_symbol: str = Form(...),
    entity_type: str = Form("unit"),
    unit_id: int = Form(0),
    space_id: int = Form(0),
    description: str = Form(""),
    db: Session = Depends(get_db),
):
    """Přidat nové VS mapování."""
    form_data = await request.form()

    # Validace VS — pouze číslice a alfanumerické znaky
    vs_clean = variable_symbol.strip()
    if not vs_clean:
        return RedirectResponse(_symboly_redirect_url(form_data, chyba="prazdny"), status_code=302)

    # Kontrola duplicity
    existing = db.query(VariableSymbolMapping).filter_by(variable_symbol=vs_clean).first()
    if existing:
        return RedirectResponse(_symboly_redirect_url(form_data, chyba="duplicita"), status_code=302)

    mapping = VariableSymbolMapping(
        variable_symbol=vs_clean,
        source=SymbolSource.MANUAL,
        description=description.strip() or None,
    )
    if entity_type == "space" and space_id:
        mapping.space_id = space_id
        mapping.unit_id = None
    else:
        mapping.unit_id = unit_id if unit_id else None
        mapping.space_id = None

    db.add(mapping)
    db.flush()
    log_activity(
        db, ActivityAction.CREATED, "vs_mapping", "platby",
        entity_id=mapping.id,
        entity_name=f"VS {mapping.variable_symbol}",
    )
    db.commit()
    return RedirectResponse(_symboly_redirect_url(form_data, flash="ok"), status_code=302)


@router.get("/symboly/{mapping_id}/upravit-formular")
async def symbol_edit_form(
    request: Request,
    mapping_id: int,
    db: Session = Depends(get_db),
):
    """Vrátí HTMX partial s editačním řádkem."""
    mapping = db.query(VariableSymbolMapping).options(
        joinedload(VariableSymbolMapping.unit),
        joinedload(VariableSymbolMapping.space),
    ).get(mapping_id)
    if not mapping:
        return RedirectResponse("/platby/symboly", status_code=302)

    units = db.query(Unit).order_by(Unit.unit_number).all()
    spaces = db.query(Space).order_by(Space.space_number).all()

    return templates.TemplateResponse(request, "payments/partials/_symboly_edit_row.html", {
        "m": mapping,
        "units": units,
        "spaces": spaces,
        "q": request.query_params.get("q", ""),
        "sort": request.query_params.get("sort", "vs"),
        "order": request.query_params.get("order", "asc"),
        "zdroj": request.query_params.get("zdroj", ""),
        "entita": request.query_params.get("entita", ""),
        "back_url": request.query_params.get("back", ""),
        "list_url": request.query_params.get("list_url", "/platby/symboly"),
    })


@router.get("/symboly/{mapping_id}/info")
async def symbol_info_row(
    request: Request,
    mapping_id: int,
    db: Session = Depends(get_db),
):
    """Vrátí HTMX partial se zobrazovacím řádkem (pro cancel)."""
    mapping = db.query(VariableSymbolMapping).options(
        joinedload(VariableSymbolMapping.unit),
        joinedload(VariableSymbolMapping.space),
    ).get(mapping_id)
    if not mapping:
        return RedirectResponse("/platby/symboly", status_code=302)

    list_url = request.query_params.get("list_url", "/platby/symboly")

    return templates.TemplateResponse(request, "payments/partials/_symboly_view_row.html", {
        "m": mapping,
        "q": request.query_params.get("q", ""),
        "sort": request.query_params.get("sort", "vs"),
        "order": request.query_params.get("order", "asc"),
        "zdroj": request.query_params.get("zdroj", ""),
        "entita": request.query_params.get("entita", ""),
        "back_url": request.query_params.get("back", ""),
        "list_url": list_url,
    })


@router.post("/symboly/{mapping_id}/upravit")
async def symbol_upravit(
    request: Request,
    mapping_id: int,
    variable_symbol: str = Form(""),
    entity_type: str = Form("unit"),
    unit_id: int = Form(0),
    space_id: int = Form(0),
    description: str = Form(""),
    db: Session = Depends(get_db),
):
    """Upravit existující VS mapování."""
    form_data = await request.form()
    mapping = db.query(VariableSymbolMapping).get(mapping_id)
    if not mapping:
        return RedirectResponse(_symboly_redirect_url(form_data, chyba="nenalezeno"), status_code=302)

    # Aktualizace VS pokud se změnil
    vs_clean = variable_symbol.strip()
    if not vs_clean:
        return RedirectResponse(_symboly_redirect_url(form_data, chyba="prazdny"), status_code=302)
    if vs_clean != mapping.variable_symbol:
        existing = db.query(VariableSymbolMapping).filter_by(variable_symbol=vs_clean).first()
        if existing:
            return RedirectResponse(_symboly_redirect_url(form_data, chyba="duplicita"), status_code=302)
        mapping.variable_symbol = vs_clean

    if entity_type == "space" and space_id:
        mapping.space_id = space_id
        mapping.unit_id = None
    else:
        mapping.unit_id = unit_id if unit_id else None
        mapping.space_id = None
    mapping.description = description.strip() or None
    log_activity(
        db, ActivityAction.UPDATED, "vs_mapping", "platby",
        entity_id=mapping.id,
        entity_name=f"VS {mapping.variable_symbol}",
    )
    db.commit()
    return RedirectResponse(_symboly_redirect_url(form_data, flash="upraveno"), status_code=302)


@router.post("/symboly/{mapping_id}/smazat")
async def symbol_smazat(
    request: Request,
    mapping_id: int,
    db: Session = Depends(get_db),
):
    """Smazat VS mapování."""
    form_data = await request.form()
    mapping = db.query(VariableSymbolMapping).get(mapping_id)
    if mapping:
        log_activity(
            db, ActivityAction.DELETED, "vs_mapping", "platby",
            entity_id=mapping.id,
            entity_name=f"VS {mapping.variable_symbol}",
        )
        db.delete(mapping)
        db.commit()
    return RedirectResponse(_symboly_redirect_url(form_data, flash="smazano"), status_code=302)
