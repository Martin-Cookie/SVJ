"""Router pro bankovní výpisy — import CSV, seznam, detail, párování."""

import csv
import io
import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, Form, Query, Request, UploadFile, File
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response
from sqlalchemy import asc as sa_asc, desc as sa_desc, func
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.config import settings
from app.models import (
    ActivityAction, BankStatement, Owner, OwnerUnit, Payment, PaymentAllocation,
    PaymentDirection, PaymentMatchStatus, Space, SpaceTenant, Tenant, Unit,
    log_activity,
)
from app.utils import (
    build_list_url, excel_auto_width, is_htmx_partial, is_safe_path,
    strip_diacritics, utcnow, validate_upload, UPLOAD_LIMITS,
)
from ._helpers import templates, compute_nav_stats, MONTH_NAMES_LONG

router = APIRouter()


def _build_suggest_map(
    payments: list,
    name_index: list[tuple[set, int]],
) -> dict[int, int]:
    """Vybudovat mapu payment_id → entity_id z name_index (words_set, entity_id).

    Pro každou UNMATCHED příjmovou platbu najde nejlepší shodu dle společných slov
    v counter_account_name + note + message.
    """
    suggest_map: dict[int, int] = {}
    for p in payments:
        if p.match_status != PaymentMatchStatus.UNMATCHED or p.direction != PaymentDirection.INCOME:
            continue
        text_parts = [p.counter_account_name or "", p.note or "", p.message or ""]
        sender_words = {w for w in strip_diacritics(" ".join(text_parts)).split() if len(w) > 2}
        if not sender_words:
            continue
        best_id = None
        best_score = 0
        for name_words, entity_id in name_index:
            common = sender_words & name_words
            if len(common) >= 1 and len(common) > best_score:
                best_score = len(common)
                best_id = entity_id
        if best_id and best_score >= 1:
            suggest_map[p.id] = best_id
    return suggest_map


# ── Seznam výpisů ──────────────────────────────────────────────────────


def _filter_statements(
    db: Session,
    q: str = "",
    rok: int = 0,
    stav: str = "",
) -> list:
    """Sdílená filtrační logika pro výpisy (seznam, HTMX partial, export)."""
    query = db.query(BankStatement).order_by(BankStatement.period_from.desc())

    if rok:
        query = query.filter(func.strftime("%Y", BankStatement.period_from) == str(rok))

    if stav == "zamcene":
        query = query.filter(BankStatement.locked_at.isnot(None))
    elif stav == "otevrene":
        query = query.filter(BankStatement.locked_at.is_(None))

    statements = query.all()

    if q:
        q_ascii = strip_diacritics(q)
        def _match(s):
            period = ""
            if s.period_from:
                period = f"{MONTH_NAMES_LONG.get(s.period_from.month, '')} {s.period_from.year}"
            haystack = strip_diacritics(f"{period} {s.filename or ''} {s.bank_account or ''}")
            return q_ascii in haystack
        statements = [s for s in statements if _match(s)]

    return statements


def _attach_statement_stats(db: Session, statements: list) -> None:
    """Doplní _stats dict každému výpisu (jedním dotazem)."""
    if not statements:
        return
    stmt_ids = [s.id for s in statements]
    stats_rows = (
        db.query(
            Payment.statement_id,
            func.count(Payment.id).label("total"),
            func.count(Payment.id).filter(Payment.match_status == PaymentMatchStatus.AUTO_MATCHED).label("matched"),
            func.count(Payment.id).filter(Payment.match_status == PaymentMatchStatus.SUGGESTED).label("suggested"),
            func.count(Payment.id).filter(Payment.match_status == PaymentMatchStatus.MANUAL).label("manual"),
            func.count(Payment.id).filter(Payment.match_status == PaymentMatchStatus.UNMATCHED).label("unmatched"),
        )
        .filter(Payment.statement_id.in_(stmt_ids))
        .group_by(Payment.statement_id)
        .all()
    )
    stats_map = {r.statement_id: r for r in stats_rows}
    for stmt in statements:
        r = stats_map.get(stmt.id)
        stmt._stats = {
            "total": r.total if r else 0,
            "matched": (r.matched + r.manual) if r else 0,
            "suggested": r.suggested if r else 0,
            "unmatched": r.unmatched if r else 0,
        }


@router.get("/vypisy")
async def vypisy_seznam(
    request: Request,
    q: str = "",
    rok: int = 0,
    stav: str = "",
    db: Session = Depends(get_db),
):
    """Seznam importovaných bankovních výpisů."""
    statements = _filter_statements(db, q, rok, stav)
    _attach_statement_stats(db, statements)

    # Roky pro bubliny — ze všech (nefiltrovaných) výpisů
    all_years = (
        db.query(func.strftime("%Y", BankStatement.period_from))
        .filter(BankStatement.period_from.isnot(None))
        .distinct()
        .all()
    )
    years = sorted({int(y[0]) for y in all_years if y[0]}, reverse=True)

    # Celkový počet pro bubliny
    total_count = db.query(BankStatement).count()
    locked_count = db.query(BankStatement).filter(BankStatement.locked_at.isnot(None)).count()
    open_count = total_count - locked_count

    list_url = build_list_url(request)
    back_url = request.query_params.get("back", "")

    ctx = {
        "request": request,
        "active_nav": "platby",
        "active_tab": "vypisy",
        "statements": statements,
        "years": years,
        "total_count": total_count,
        "locked_count": locked_count,
        "open_count": open_count,
        "q": q,
        "rok": rok,
        "stav": stav,
        "list_url": list_url,
        "back_url": back_url,
        "month_names": MONTH_NAMES_LONG,
        **(compute_nav_stats(db) if not is_htmx_partial(request) else {}),
    }

    if is_htmx_partial(request):
        return templates.TemplateResponse(request, "payments/partials/vypisy_list.html", ctx)

    return templates.TemplateResponse(request, "payments/vypisy.html", ctx)


@router.get("/vypisy/exportovat/{fmt}")
async def vypisy_export(
    fmt: str,
    q: str = Query(""),
    rok: int = Query(0),
    stav: str = Query(""),
    db: Session = Depends(get_db),
):
    """Export filtrovaných bankovních výpisů."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if fmt not in ("xlsx", "csv"):
        return RedirectResponse("/platby/vypisy", status_code=302)

    statements = _filter_statements(db, q, rok, stav)
    _attach_statement_stats(db, statements)

    headers = [
        "Období", "Od", "Do", "Soubor", "Účet",
        "Počet plateb", "Příjem (Kč)", "Výdaj (Kč)",
        "Napárováno", "Návrhů", "Nenapárováno", "Zamčeno",
    ]

    def _row(s):
        period = ""
        if s.period_from:
            period = f"{MONTH_NAMES_LONG.get(s.period_from.month, '')} {s.period_from.year}"
        stats = getattr(s, "_stats", {}) or {}
        return [
            period,
            s.period_from.strftime("%d.%m.%Y") if s.period_from else "",
            s.period_to.strftime("%d.%m.%Y") if s.period_to else "",
            s.filename or "",
            s.bank_account or "",
            s.transaction_count or 0,
            round(s.total_income or 0),
            round(s.total_expense or 0),
            stats.get("matched", 0),
            stats.get("suggested", 0),
            stats.get("unmatched", 0),
            s.locked_at.strftime("%d.%m.%Y") if s.locked_at else "",
        ]

    if rok:
        suffix = f"_{rok}"
    elif stav:
        suffix = f"_{strip_diacritics(stav)}"
    elif q:
        suffix = "_hledani"
    else:
        suffix = "_vse"
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"vypisy{suffix}_{timestamp}"

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Bankovní výpisy"
        bold = Font(bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold

        for row_idx, s in enumerate(statements, 2):
            for col_idx, val in enumerate(_row(s), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        excel_auto_width(ws)

        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    else:
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(headers)
        for s in statements:
            writer.writerow(_row(s))
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )


# ── Import CSV ─────────────────────────────────────────────────────────


@router.get("/vypisy/import")
async def vypis_import_form(request: Request, db: Session = Depends(get_db)):
    """Formulář pro import bankovního výpisu z CSV."""
    back_url = request.query_params.get("back", "")
    return templates.TemplateResponse(request, "payments/vypis_import.html", {
        "active_nav": "platby",
        "active_tab": "vypisy",
        "back_url": back_url,
        **compute_nav_stats(db),
    })


@router.post("/vypisy/import")
async def vypis_import_upload(
    request: Request,
    file: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    """Zpracování importu bankovního výpisu z CSV."""
    from app.services.bank_import import parse_fio_csv
    from app.services.payment_matching import match_payments

    form_data = await request.form()
    force = form_data.get("force_overwrite")
    saved_path = form_data.get("saved_path", "")

    # Při force_overwrite použít uložený soubor
    if force and saved_path:
        saved_file = Path(saved_path)
        if not is_safe_path(saved_file, Path(settings.upload_dir) / "temp"):
            return templates.TemplateResponse(request, "payments/vypis_import.html", {
                "active_nav": "platby",
                "active_tab": "vypisy",
                "error": "Neplatná cesta k souboru.",
                **compute_nav_stats(db),
            })
        if not saved_file.is_file():
            return templates.TemplateResponse(request, "payments/vypis_import.html", {
                "active_nav": "platby",
                "active_tab": "vypisy",
                "error": "Uložený soubor expiroval. Nahrajte soubor znovu.",
                **compute_nav_stats(db),
            })
        file_content = saved_file.read_bytes()
        original_filename = saved_file.name.split("_", 2)[-1] if "_" in saved_file.name else saved_file.name
    else:
        # Nový upload — validace
        if not file or not file.filename:
            return templates.TemplateResponse(request, "payments/vypis_import.html", {
                "active_nav": "platby",
                "active_tab": "vypisy",
                "error": "Vyberte soubor CSV.",
                **compute_nav_stats(db),
            })
        error = await validate_upload(file, **UPLOAD_LIMITS["csv"])
        if error:
            return templates.TemplateResponse(request, "payments/vypis_import.html", {
                "active_nav": "platby",
                "active_tab": "vypisy",
                "error": error,
                **compute_nav_stats(db),
            })
        file_content = await file.read()
        original_filename = file.filename

    # Čtení a parsování
    try:
        result = parse_fio_csv(file_content, original_filename)
    except Exception as e:
        logger.error("CSV parse error: %s", e)
        return templates.TemplateResponse(request, "payments/vypis_import.html", {
            "active_nav": "platby",
            "active_tab": "vypisy",
            "error": f"Chyba při čtení CSV: {e}",
            **compute_nav_stats(db),
        })

    if result["errors"]:
        return templates.TemplateResponse(request, "payments/vypis_import.html", {
            "active_nav": "platby",
            "active_tab": "vypisy",
            "error": "Chyby při parsování: " + "; ".join(result["errors"][:5]),
            **compute_nav_stats(db),
        })

    if not result["transactions"]:
        return templates.TemplateResponse(request, "payments/vypis_import.html", {
            "active_nav": "platby",
            "active_tab": "vypisy",
            "error": "CSV neobsahuje žádné transakce.",
            **compute_nav_stats(db),
        })

    meta = result["metadata"]

    # Kontrola duplicity výpisu (podle období)
    existing = (
        db.query(BankStatement)
        .filter_by(period_from=meta.get("period_from"), period_to=meta.get("period_to"))
        .first()
    )
    if existing and not force:
        # Uložit soubor na disk pro pozdější použití
        temp_dir = Path(settings.upload_dir) / "temp"
        temp_dir.mkdir(parents=True, exist_ok=True)
        timestamp = utcnow().strftime("%Y%m%d_%H%M%S")
        temp_path = temp_dir / f"{timestamp}_{original_filename}"
        temp_path.write_bytes(file_content)

        period_label = ""
        if meta.get("period_from"):
            period_label = f"{meta['period_from'].strftime('%d.%m.%Y')} – {meta['period_to'].strftime('%d.%m.%Y')}"
        # Počet ručně přiřazených plateb
        manual_count = (
            db.query(Payment)
            .filter(
                Payment.statement_id == existing.id,
                Payment.match_status == PaymentMatchStatus.MANUAL,
            )
            .count()
        )
        return templates.TemplateResponse(request, "payments/vypis_import.html", {
            "active_nav": "platby",
            "active_tab": "vypisy",
            "confirm_overwrite": True,
            "existing": existing,
            "period_label": period_label,
            "filename": original_filename,
            "saved_path": str(temp_path),
            "transaction_count": len(result["transactions"]),
            "manual_count": manual_count,
            **compute_nav_stats(db),
        })

    # Zachování ručních přiřazení
    manual_map = {}
    if existing:
        preserve = form_data.get("preserve_manual")
        if preserve:
            manual_payments = (
                db.query(Payment.operation_id, Payment.unit_id, Payment.match_status)
                .filter(
                    Payment.statement_id == existing.id,
                    Payment.match_status == PaymentMatchStatus.MANUAL,
                    Payment.unit_id.isnot(None),
                    Payment.operation_id.isnot(None),
                )
                .all()
            )
            manual_map = {p.operation_id: p.unit_id for p in manual_payments}
        db.delete(existing)
        db.flush()

    # Uložit soubor
    dest_dir = settings.upload_dir / "csv"
    dest_dir.mkdir(parents=True, exist_ok=True)
    ts = utcnow().strftime("%Y%m%d_%H%M%S")
    dest_path = dest_dir / f"{ts}_{original_filename}"
    with open(dest_path, "wb") as f:
        f.write(file_content)

    # Vytvořit BankStatement
    statement = BankStatement(
        filename=original_filename,
        file_path=str(dest_path),
        bank_account=meta.get("bank_account"),
        period_from=meta.get("period_from"),
        period_to=meta.get("period_to"),
        opening_balance=meta.get("opening_balance"),
        closing_balance=meta.get("closing_balance"),
        total_income=meta.get("total_income", 0),
        total_expense=meta.get("total_expense", 0),
        transaction_count=len(result["transactions"]),
    )
    db.add(statement)
    db.flush()

    # Vložit transakce (s deduplikací přes operation_id)
    existing_ops = set(
        r[0] for r in db.query(Payment.operation_id).filter(
            Payment.operation_id.isnot(None)
        ).all()
    )

    inserted = 0
    skipped = 0
    for t in result["transactions"]:
        if t["operation_id"] in existing_ops:
            skipped += 1
            continue

        db.add(Payment(
            statement_id=statement.id,
            operation_id=t["operation_id"],
            date=t["date"],
            amount=t["amount"],
            direction=PaymentDirection.INCOME if t["direction"] == "income" else PaymentDirection.EXPENSE,
            counter_account=t["counter_account"],
            counter_account_name=t["counter_account_name"],
            bank_code=t["bank_code"],
            bank_name=t["bank_name"],
            ks=t["ks"],
            vs=t["vs"],
            ss=t["ss"],
            note=t["note"],
            message=t["message"],
            payment_type=t["payment_type"],
            match_status=PaymentMatchStatus.UNMATCHED,
        ))
        existing_ops.add(t["operation_id"])
        inserted += 1

    db.flush()

    # Automatické párování
    pf = meta.get("period_from")
    year = pf.year if pf else utcnow().year
    try:
        match_result = match_payments(db, statement.id, year)
    except Exception as e:
        logger.error("Matching failed for statement %d: %s", statement.id, e)
        match_result = {"matched": 0, "suggested": 0, "unmatched": inserted, "total": inserted}

    # Obnovit ruční přiřazení z předchozího importu
    restored = 0
    if manual_map:
        for payment in db.query(Payment).filter(
            Payment.statement_id == statement.id,
            Payment.operation_id.in_(list(manual_map.keys())),
        ).all():
            payment.unit_id = manual_map[payment.operation_id]
            payment.match_status = PaymentMatchStatus.MANUAL
            # Dual-write: vytvořit alokaci
            db.add(PaymentAllocation(
                payment_id=payment.id,
                unit_id=payment.unit_id,
                owner_id=payment.owner_id,
                prescription_id=payment.prescription_id,
                amount=payment.amount,
            ))
            restored += 1
        match_result["matched"] += restored

    statement.matched_count = match_result["matched"]
    log_activity(
        db, ActivityAction.IMPORTED, "bank_statement", "platby",
        entity_id=statement.id,
        entity_name=statement.filename or f"Výpis #{statement.id}",
        description=f"{match_result.get('total', 0)} plateb",
    )
    db.commit()

    # Cleanup temp souboru
    if saved_path:
        try:
            Path(saved_path).unlink(missing_ok=True)
        except OSError:
            pass

    return RedirectResponse(
        f"/platby/vypisy/{statement.id}?flash=import_ok"
        f"&inserted={inserted}&skipped={skipped}"
        f"&matched={match_result['matched']}"
        f"&suggested={match_result.get('suggested', 0)}"
        f"&unmatched={match_result['unmatched']}",
        status_code=302,
    )


# ── Detail výpisu ──────────────────────────────────────────────────────


SORT_COLUMNS_PAYMENTS = {
    "datum": Payment.date,
    "castka": Payment.amount,
    "vs": Payment.vs,
    "protiucet": Payment.counter_account_name,
    "stav": Payment.match_status,
    "jednotka": None,  # Python-side sort
    "prostor": None,  # Python-side sort
}


def _filter_payments(
    db: Session,
    statement_id: int,
    q: str = "",
    stav: str = "",
    smer: str = "",
    typ: str = "",
    sort: str = "datum",
    order: str = "asc",
) -> list:
    """Sdílená filtrační logika pro detail výpisu (seznam, HTMX, export)."""
    from app.models import PaymentAllocation as PA
    query = (
        db.query(Payment)
        .filter_by(statement_id=statement_id)
        .options(
            joinedload(Payment.unit),
            joinedload(Payment.space),
            joinedload(Payment.owner),
            joinedload(Payment.allocations).joinedload(PA.unit),
            joinedload(Payment.allocations).joinedload(PA.space),
        )
    )

    if stav:
        query = query.filter(Payment.match_status == stav)

    if smer == "prijem":
        query = query.filter(Payment.direction == PaymentDirection.INCOME)
    elif smer == "vydej":
        query = query.filter(Payment.direction == PaymentDirection.EXPENSE)

    if typ == "jednotky":
        query = query.filter(Payment.unit_id.isnot(None))
    elif typ == "prostory":
        query = query.filter(Payment.space_id.isnot(None))

    col = SORT_COLUMNS_PAYMENTS.get(sort, Payment.date)
    if col is not None:
        order_fn = sa_desc if order == "desc" else sa_asc
        query = query.order_by(order_fn(col).nulls_last())

    payments = query.all()

    if sort in ("jednotka", "prostor"):
        def _unit_key(p):
            if p.unit:
                return p.unit.unit_number or 0
            for a in (p.allocations or []):
                if a.unit:
                    return a.unit.unit_number or 0
            return 0

        def _space_key(p):
            if p.space:
                return p.space.space_number or 0
            for a in (p.allocations or []):
                if a.space:
                    return a.space.space_number or 0
            return 0

        key_fn = _unit_key if sort == "jednotka" else _space_key
        payments.sort(key=key_fn, reverse=(order == "desc"))

    if q:
        q_ascii = strip_diacritics(q)
        payments = [
            p for p in payments
            if q in (p.vs or "")
            or q_ascii in strip_diacritics(p.counter_account_name or "")
            or q_ascii in strip_diacritics(p.note or "")
            or q_ascii in strip_diacritics(p.message or "")
        ]

    return payments


@router.get("/vypisy/{statement_id}")
async def vypis_detail(
    request: Request,
    statement_id: int,
    sort: str = "datum",
    order: str = "asc",
    q: str = "",
    stav: str = "",
    smer: str = "",
    typ: str = "",
    db: Session = Depends(get_db),
):
    """Detail bankovního výpisu — seznam plateb."""
    statement = db.query(BankStatement).get(statement_id)
    if not statement:
        return RedirectResponse("/platby/vypisy", status_code=302)

    payments = _filter_payments(db, statement_id, q, stav, smer, typ, sort, order)

    # Statistiky
    total_income = sum(p.amount for p in payments if p.direction == PaymentDirection.INCOME)
    total_expense = sum(p.amount for p in payments if p.direction == PaymentDirection.EXPENSE)
    matched_count = sum(1 for p in payments if p.match_status != PaymentMatchStatus.UNMATCHED)

    # Bubble counts (celkové počty bez filtrů — ale respektují typ filtr)
    typ_filter = []
    if typ == "jednotky":
        typ_filter.append(Payment.unit_id.isnot(None))
    elif typ == "prostory":
        typ_filter.append(Payment.space_id.isnot(None))

    all_payments_for_counts = (
        db.query(Payment.match_status, Payment.direction, func.count())
        .filter_by(statement_id=statement_id)
        .filter(*typ_filter)
        .group_by(Payment.match_status, Payment.direction)
        .all()
    )
    bubble_counts = {
        "vse": 0,
        "auto_matched": 0,
        "suggested": 0,
        "manual": 0,
        "unmatched": 0,
        "prijem": 0,
        "vydej": 0,
    }
    for status, direction, cnt in all_payments_for_counts:
        bubble_counts["vse"] += cnt
        bubble_counts[status.value] += cnt
        if direction == PaymentDirection.INCOME:
            bubble_counts["prijem"] += cnt
        else:
            bubble_counts["vydej"] += cnt

    # Typ bubble counts (celkové počty bez stav/smer filtrů)
    all_count = db.query(func.count()).filter(Payment.statement_id == statement_id).scalar() or 0
    unit_count = db.query(func.count()).filter(
        Payment.statement_id == statement_id, Payment.unit_id.isnot(None)
    ).scalar() or 0
    space_count = db.query(func.count()).filter(
        Payment.statement_id == statement_id, Payment.space_id.isnot(None)
    ).scalar() or 0
    typ_counts = {"vse": all_count, "jednotky": unit_count, "prostory": space_count}

    # Flash zprávy
    flash_message = ""
    flash_type = ""
    flash = request.query_params.get("flash", "")
    if flash == "import_ok":
        inserted = request.query_params.get("inserted", "0")
        skipped = request.query_params.get("skipped", "0")
        matched = request.query_params.get("matched", "0")
        suggested_count = request.query_params.get("suggested", "0")
        unmatched = request.query_params.get("unmatched", "0")
        flash_message = (
            f"Import dokončen: {inserted} plateb vloženo"
            + (f", {skipped} duplicit přeskočeno" if int(skipped) > 0 else "")
            + f", {matched} napárováno"
            + (f", {suggested_count} návrhů" if int(suggested_count) > 0 else "")
            + f", {unmatched} nenapárováno."
        )
    elif flash == "match_ok":
        flash_message = "Ruční přiřazení uloženo."
    elif flash == "match_fail":
        flash_message = "Jednotka s tímto číslem nebyla nalezena."
        flash_type = "error"
    elif flash == "confirmed":
        flash_message = "Návrh potvrzen."
    elif flash == "bulk_confirmed":
        count = request.query_params.get("count", "0")
        flash_message = f"Potvrzeno {count} návrhů."
    elif flash == "rejected":
        flash_message = "Návrh odmítnut."
    elif flash == "rematch_ok":
        matched = request.query_params.get("matched", "0")
        suggested_count = request.query_params.get("suggested", "0")
        flash_message = (
            f"Přepárování dokončeno: {matched} plateb napárováno"
            + (f", {suggested_count} návrhů" if int(suggested_count) > 0 else "")
            + "."
        )
    elif flash == "locked":
        flash_message = "Výpis je zamčený — párování nelze měnit."
        flash_type = "warning"
    elif flash == "lock_ok":
        flash_message = "Párování zamčeno."
    elif flash == "unlock_ok":
        flash_message = "Párování odemčeno."

    # Nesrovnalosti pro upozornění
    from app.services.payment_discrepancy import detect_discrepancies, DISCREPANCY_LABELS
    discrepancies = detect_discrepancies(db, statement_id)

    # Kandidáti pro nenapárované platby
    from app.services.payment_matching import compute_candidates
    from app.models import Prescription, PrescriptionYear
    pf = statement.period_from
    cand_year = pf.year if pf else utcnow().year
    candidates_map = compute_candidates(db, payments, cand_year, statement_id=statement.id)

    # Mapa unit_id → měsíční předpis + VS (pro tooltipy)
    unit_monthly = {}
    unit_vs = {}
    py = db.query(PrescriptionYear).filter_by(year=cand_year).first()
    if py:
        for presc in db.query(Prescription).filter_by(prescription_year_id=py.id).all():
            if presc.unit_id:
                if presc.monthly_total:
                    unit_monthly[presc.unit_id] = presc.monthly_total
                if presc.variable_symbol:
                    unit_vs[presc.unit_id] = presc.variable_symbol

    # Units + owner names for assignment dropdown (1 JOIN místo 3 dotazů)
    unit_owner_names = {}  # unit_id → owner display_name
    unit_name_index = []  # list of (words_set, unit_id)
    units_by_id = {}
    unit_owner_rows = (
        db.query(Unit, Owner)
        .outerjoin(OwnerUnit, (OwnerUnit.unit_id == Unit.id) & (OwnerUnit.valid_to.is_(None)))
        .outerjoin(Owner, Owner.id == OwnerUnit.owner_id)
        .all()
    )
    for unit, owner in unit_owner_rows:
        units_by_id[unit.id] = unit
        if owner:
            if owner.display_name:
                existing = unit_owner_names.get(unit.id)
                if existing:
                    unit_owner_names[unit.id] = existing + ", " + owner.display_name
                else:
                    unit_owner_names[unit.id] = owner.display_name
            if owner.name_normalized:
                words = {w for w in strip_diacritics(owner.name_normalized).split() if len(w) > 2}
                if words:
                    unit_name_index.append((words, unit.id))
    # Řadit podle jména vlastníka (bez diakritiky), jednotky bez vlastníka na konec
    all_units_list = sorted(
        units_by_id.values(),
        key=lambda u: strip_diacritics(unit_owner_names.get(u.id, "zzz")),
    )

    # Suggest map: payment_id → unit_id (pre-select based on counterparty name / message)
    unit_suggest_map = _build_suggest_map(payments, unit_name_index)

    # Spaces + tenant names for assignment dropdown
    all_spaces = db.query(Space).order_by(Space.space_number).all()
    space_tenant_names = {}
    active_sts = (
        db.query(SpaceTenant)
        .filter_by(is_active=True)
        .options(joinedload(SpaceTenant.tenant).joinedload(Tenant.owner))
        .all()
    )
    space_monthly = {}  # space_id → monthly_rent
    space_vs = {}  # space_id → variable_symbol
    space_name_index = []  # (words_set, space_id)
    for st in active_sts:
        if st.monthly_rent:
            space_monthly[st.space_id] = st.monthly_rent
        if st.variable_symbol:
            space_vs[st.space_id] = st.variable_symbol
        if st.tenant:
            name = st.tenant.display_name
            if name:
                space_tenant_names[st.space_id] = name
                norm = strip_diacritics(name)
                words = {w for w in norm.split() if len(w) > 2}
                if words:
                    space_name_index.append((words, st.space_id))
    # Řadit podle jména nájemce (bez diakritiky), prostory bez nájemce na konec
    all_spaces.sort(key=lambda s: strip_diacritics(space_tenant_names.get(s.id, "zzz")))

    # Suggest map: payment_id → space_id (pre-select based on counterparty/message)
    space_suggest_map = _build_suggest_map(payments, space_name_index)

    list_url = build_list_url(request)
    back_url = request.query_params.get("back", "")

    ctx = {
        "request": request,
        "active_nav": "platby",
        "active_tab": "vypisy",
        "statement": statement,
        "payments": payments,
        "total_income": total_income,
        "total_expense": total_expense,
        "matched_count": matched_count,
        "candidates_map": candidates_map,
        "unit_monthly": unit_monthly,
        "unit_vs": unit_vs,
        "all_units_list": all_units_list,
        "unit_owner_names": unit_owner_names,
        "unit_suggest_map": unit_suggest_map,
        "all_spaces": all_spaces,
        "space_tenant_names": space_tenant_names,
        "space_monthly": space_monthly,
        "space_vs": space_vs,
        "space_suggest_map": space_suggest_map,
        "sort": sort,
        "order": order,
        "q": q,
        "stav": stav,
        "smer": smer,
        "typ": typ,
        "bubble_counts": bubble_counts,
        "typ_counts": typ_counts,
        "list_url": list_url,
        "back_url": back_url,
        "discrepancies": discrepancies,
        "discrepancy_labels": DISCREPANCY_LABELS,
        "flash_message": flash_message,
        "flash_type": flash_type,
        "month_names": MONTH_NAMES_LONG,
        **(compute_nav_stats(db) if not is_htmx_partial(request) else {}),
    }

    if is_htmx_partial(request):
        return templates.TemplateResponse(request, "payments/partials/vypis_tbody.html", ctx)

    return templates.TemplateResponse(request, "payments/vypis_detail.html", ctx)


# ── Download zdrojového CSV souboru výpisu ────────────────────────────


@router.get("/vypisy/{statement_id}/soubor")
async def vypis_download_source(statement_id: int, db: Session = Depends(get_db)):
    """Stažení původního CSV souboru bankovního výpisu."""
    statement = db.query(BankStatement).get(statement_id)
    if not statement or not statement.file_path:
        return RedirectResponse("/platby/vypisy", status_code=302)

    file_path = Path(statement.file_path)
    allowed_root = Path(settings.upload_dir)
    if not file_path.exists() or not is_safe_path(file_path, allowed_root):
        return RedirectResponse(f"/platby/vypisy/{statement_id}", status_code=302)

    return FileResponse(
        path=str(file_path),
        filename=statement.filename,
        media_type="text/csv",
    )


# ── Export plateb z výpisu ─────────────────────────────────────────────


@router.get("/vypisy/{statement_id}/exportovat/{fmt}")
async def vypis_detail_export(
    statement_id: int,
    fmt: str,
    q: str = Query(""),
    stav: str = Query(""),
    smer: str = Query(""),
    typ: str = Query(""),
    sort: str = Query("datum"),
    order: str = Query("asc"),
    db: Session = Depends(get_db),
):
    """Export filtrovaných plateb z detailu výpisu."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if fmt not in ("xlsx", "csv"):
        return RedirectResponse(f"/platby/vypisy/{statement_id}", status_code=302)

    statement = db.query(BankStatement).get(statement_id)
    if not statement:
        return RedirectResponse("/platby/vypisy", status_code=302)

    payments = _filter_payments(db, statement_id, q, stav, smer, typ, sort, order)

    status_labels = {
        "auto_matched": "Napárováno",
        "suggested": "Návrh",
        "manual": "Ručně",
        "unmatched": "Nenapárováno",
    }

    headers = [
        "Datum", "Směr", "Částka (Kč)", "VS", "KS", "SS",
        "Protiúčet", "Název protistrany", "Zpráva", "Poznámka",
        "Jednotka", "Prostor", "Vlastník / Nájemce", "Stav",
    ]

    def _row(p):
        entity_name = ""
        if p.owner and p.owner.display_name:
            entity_name = p.owner.display_name
        unit_num = ""
        if p.unit:
            unit_num = str(p.unit.unit_number)
        else:
            for a in (p.allocations or []):
                if a.unit:
                    unit_num = str(a.unit.unit_number)
                    break
        space_num = ""
        if p.space:
            space_num = str(p.space.space_number)
        else:
            for a in (p.allocations or []):
                if a.space:
                    space_num = str(a.space.space_number)
                    break
        direction = "Příjem" if p.direction == PaymentDirection.INCOME else "Výdej"
        status_val = p.match_status.value if p.match_status else ""
        return [
            p.date.strftime("%d.%m.%Y") if p.date else "",
            direction,
            round(p.amount or 0, 2),
            p.vs or "",
            p.ks or "",
            p.ss or "",
            p.counter_account or "",
            p.counter_account_name or "",
            p.message or "",
            p.note or "",
            unit_num,
            space_num,
            entity_name,
            status_labels.get(status_val, status_val),
        ]

    # Suffix dle filtru (priorita: stav → smer → typ → hledani → vse)
    stav_labels = {
        "auto_matched": "naparovano",
        "suggested": "navrh",
        "manual": "rucne",
        "unmatched": "nenaparovano",
    }
    smer_labels = {"prijem": "prijem", "vydej": "vydej"}
    typ_labels = {"jednotky": "jednotky", "prostory": "prostory"}
    if stav:
        suffix = f"_{stav_labels.get(stav, strip_diacritics(stav))}"
    elif smer:
        suffix = f"_{smer_labels.get(smer, strip_diacritics(smer))}"
    elif typ:
        suffix = f"_{typ_labels.get(typ, strip_diacritics(typ))}"
    elif q:
        suffix = "_hledani"
    else:
        suffix = "_vse"
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"vypis_{statement_id}{suffix}_{timestamp}"

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Platby"
        bold = Font(bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold

        for row_idx, p in enumerate(payments, 2):
            for col_idx, val in enumerate(_row(p), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        excel_auto_width(ws)

        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    else:
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(headers)
        for p in payments:
            writer.writerow(_row(p))
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )


# ── Ruční přiřazení platby ─────────────────────────────────────────────


def _detail_redirect_url(statement_id: int, form_data, flash: str = "", anchor: str = "") -> str:
    """Sestaví redirect URL zpět na detail výpisu se zachováním filtrů."""
    params = []
    if flash:
        params.append(f"flash={flash}")
    for key in ("q", "sort", "order", "stav", "smer", "typ", "back"):
        val = form_data.get(key, "")
        if val:
            params.append(f"{key}={val}")
    qs = "&".join(params)
    url = f"/platby/vypisy/{statement_id}?{qs}" if qs else f"/platby/vypisy/{statement_id}"
    if anchor:
        url += f"#{anchor}"
    return url


@router.post("/vypisy/{statement_id}/prirazeni/{payment_id}")
async def platba_prirazeni(
    request: Request,
    statement_id: int,
    payment_id: int,
    db: Session = Depends(get_db),
):
    """Ručně přiřadit platbu k jednotce/jednotkám (čísla oddělená čárkou)."""
    form_data = await request.form()

    # Zamčený výpis — nelze měnit
    statement = db.query(BankStatement).get(statement_id)
    if statement and statement.locked_at:
        return RedirectResponse(_detail_redirect_url(statement_id, form_data, "locked"), status_code=302)

    unit_id_raw = form_data.get("unit_id", "").strip()

    payment = db.query(Payment).filter_by(id=payment_id, statement_id=statement_id).first()
    if not payment:
        return RedirectResponse(_detail_redirect_url(statement_id, form_data), status_code=302)

    # Parsovat čísla jednotek (může být "5" nebo "5, 425")
    unit_numbers = []
    for part in unit_id_raw.replace(",", " ").split():
        part = part.strip()
        if part.isdigit():
            unit_numbers.append(int(part))
    if not unit_numbers:
        return RedirectResponse(
            _detail_redirect_url(statement_id, form_data, "match_fail", anchor=f"p-{payment_id}"),
            status_code=302,
        )

    # Najít jednotky
    units = db.query(Unit).filter(Unit.unit_number.in_(unit_numbers)).all()
    if not units or len(units) != len(set(unit_numbers)):
        return RedirectResponse(
            _detail_redirect_url(statement_id, form_data, "match_fail", anchor=f"p-{payment_id}"),
            status_code=302,
        )

    from app.models import OwnerUnit, Prescription, PrescriptionYear

    # Smazat staré alokace
    db.query(PaymentAllocation).filter_by(payment_id=payment.id).delete()

    if len(units) == 1:
        # Single-unit přiřazení
        unit = units[0]
        payment.unit_id = unit.id
        payment.match_status = PaymentMatchStatus.MANUAL

        ou = db.query(OwnerUnit).filter_by(unit_id=unit.id).filter(OwnerUnit.valid_to.is_(None)).first()
        if ou:
            payment.owner_id = ou.owner_id

        db.add(PaymentAllocation(
            payment_id=payment.id,
            unit_id=unit.id,
            owner_id=payment.owner_id,
            prescription_id=payment.prescription_id,
            amount=payment.amount,
        ))
    else:
        # Multi-unit přiřazení — rozdělit částku podle předpisů
        payment.unit_id = None
        payment.match_status = PaymentMatchStatus.MANUAL

        # Najít předpisy pro rozložení částky
        latest_py = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        prescriptions_map = {}
        if latest_py:
            for presc in db.query(Prescription).filter_by(prescription_year_id=latest_py.id).all():
                if presc.unit_id:
                    prescriptions_map[presc.unit_id] = presc

        total_monthly = sum(
            prescriptions_map[u.id].monthly_total
            for u in units if u.id in prescriptions_map and prescriptions_map[u.id].monthly_total
        ) or 1.0

        for unit in units:
            presc = prescriptions_map.get(unit.id)
            # Rozdělit částku proporcionálně podle předpisů
            if presc and presc.monthly_total and total_monthly > 0:
                alloc_amount = round(payment.amount * presc.monthly_total / total_monthly, 2)
            else:
                alloc_amount = round(payment.amount / len(units), 2)

            ou = db.query(OwnerUnit).filter_by(unit_id=unit.id).filter(OwnerUnit.valid_to.is_(None)).first()
            db.add(PaymentAllocation(
                payment_id=payment.id,
                unit_id=unit.id,
                owner_id=ou.owner_id if ou else None,
                prescription_id=presc.id if presc else None,
                amount=alloc_amount,
            ))

    db.commit()

    return RedirectResponse(
        _detail_redirect_url(statement_id, form_data, "match_ok", anchor=f"p-{payment_id}"),
        status_code=302,
    )


@router.post("/vypisy/{statement_id}/prirazeni-prostor/{payment_id}")
async def platba_prirazeni_prostor(
    request: Request,
    statement_id: int,
    payment_id: int,
    db: Session = Depends(get_db),
):
    """Ručně přiřadit platbu k prostoru (číslo prostoru)."""
    form_data = await request.form()

    statement = db.query(BankStatement).get(statement_id)
    if statement and statement.locked_at:
        return RedirectResponse(_detail_redirect_url(statement_id, form_data, "locked"), status_code=302)

    space_num_raw = form_data.get("space_id", "").strip()
    payment = db.query(Payment).filter_by(id=payment_id, statement_id=statement_id).first()
    if not payment:
        return RedirectResponse(_detail_redirect_url(statement_id, form_data), status_code=302)

    # Najít prostor podle čísla
    space = db.query(Space).filter_by(space_number=space_num_raw).first()
    if not space:
        return RedirectResponse(
            _detail_redirect_url(statement_id, form_data, "match_fail", anchor=f"p-{payment_id}"),
            status_code=302,
        )

    # Smazat staré alokace
    db.query(PaymentAllocation).filter_by(payment_id=payment.id).delete()

    payment.space_id = space.id
    payment.unit_id = None
    payment.match_status = PaymentMatchStatus.MANUAL

    db.add(PaymentAllocation(
        payment_id=payment.id,
        space_id=space.id,
        owner_id=payment.owner_id,
        amount=payment.amount,
    ))

    db.commit()

    return RedirectResponse(
        _detail_redirect_url(statement_id, form_data, "match_ok", anchor=f"p-{payment_id}"),
        status_code=302,
    )


# ── Potvrzení / odmítnutí návrhu ──────────────────────────────────────


@router.post("/vypisy/{statement_id}/potvrdit-vse")
async def platba_potvrdit_vse(
    request: Request,
    statement_id: int,
    db: Session = Depends(get_db),
):
    """Potvrdit všechny SUGGESTED přiřazení najednou (→ MANUAL)."""
    form_data = await request.form()

    statement = db.query(BankStatement).get(statement_id)
    if statement and statement.locked_at:
        return RedirectResponse(_detail_redirect_url(statement_id, form_data, "locked"), status_code=302)

    count = db.query(Payment).filter(
        Payment.statement_id == statement_id,
        Payment.match_status == PaymentMatchStatus.SUGGESTED,
    ).update({Payment.match_status: PaymentMatchStatus.MANUAL})
    db.commit()

    return RedirectResponse(
        _detail_redirect_url(statement_id, form_data, f"bulk_confirmed&count={count}"),
        status_code=302,
    )


@router.post("/vypisy/{statement_id}/potvrdit/{payment_id}")
async def platba_potvrdit(
    request: Request,
    statement_id: int,
    payment_id: int,
    db: Session = Depends(get_db),
):
    """Potvrdit navržené přiřazení (SUGGESTED → MANUAL)."""
    form_data = await request.form()

    # Zamčený výpis — nelze měnit
    statement = db.query(BankStatement).get(statement_id)
    if statement and statement.locked_at:
        return RedirectResponse(_detail_redirect_url(statement_id, form_data, "locked"), status_code=302)

    payment = db.query(Payment).filter_by(id=payment_id, statement_id=statement_id).first()
    if payment and payment.match_status == PaymentMatchStatus.SUGGESTED:
        payment.match_status = PaymentMatchStatus.MANUAL
        db.commit()

    return RedirectResponse(
        _detail_redirect_url(statement_id, form_data, "confirmed", anchor=f"p-{payment_id}"),
        status_code=302,
    )


@router.post("/vypisy/{statement_id}/odmitnout/{payment_id}")
async def platba_odmitnout(
    request: Request,
    statement_id: int,
    payment_id: int,
    db: Session = Depends(get_db),
):
    """Odmítnout navržené přiřazení (SUGGESTED → UNMATCHED)."""
    form_data = await request.form()

    # Zamčený výpis — nelze měnit
    statement = db.query(BankStatement).get(statement_id)
    if statement and statement.locked_at:
        return RedirectResponse(_detail_redirect_url(statement_id, form_data, "locked"), status_code=302)

    payment = db.query(Payment).filter_by(id=payment_id, statement_id=statement_id).first()
    if payment and payment.match_status == PaymentMatchStatus.SUGGESTED:
        payment.unit_id = None
        payment.space_id = None
        payment.owner_id = None
        payment.prescription_id = None
        payment.match_status = PaymentMatchStatus.UNMATCHED
        # Smazat alokace
        db.query(PaymentAllocation).filter_by(payment_id=payment.id).delete()
        db.commit()

    return RedirectResponse(
        _detail_redirect_url(statement_id, form_data, "rejected", anchor=f"p-{payment_id}"),
        status_code=302,
    )


# ── Přepárování výpisu ─────────────────────────────────────────────────


@router.post("/vypisy/{statement_id}/preparovat")
async def vypis_preparovat(
    request: Request,
    statement_id: int,
    db: Session = Depends(get_db),
):
    """Znovu spustit automatické párování pro výpis."""
    from app.services.payment_matching import match_payments

    form_data = await request.form()

    statement = db.query(BankStatement).get(statement_id)
    if not statement:
        return RedirectResponse("/platby/vypisy", status_code=302)

    # Zamčený výpis — nelze přepárovat
    if statement.locked_at:
        url = _detail_redirect_url(statement_id, form_data, "locked")
        return RedirectResponse(url, status_code=302)

    # Smazat alokace pro auto + suggested platby
    reset_payment_ids = [
        pid for (pid,) in db.query(Payment.id).filter(
            Payment.statement_id == statement_id,
            Payment.match_status.in_([PaymentMatchStatus.AUTO_MATCHED, PaymentMatchStatus.SUGGESTED]),
        ).all()
    ]
    if reset_payment_ids:
        db.query(PaymentAllocation).filter(
            PaymentAllocation.payment_id.in_(reset_payment_ids)
        ).delete(synchronize_session="fetch")

    # Reset auto + suggested (ponech ruční)
    db.query(Payment).filter(
        Payment.statement_id == statement_id,
        Payment.match_status.in_([PaymentMatchStatus.AUTO_MATCHED, PaymentMatchStatus.SUGGESTED]),
    ).update({
        Payment.unit_id: None,
        Payment.space_id: None,
        Payment.owner_id: None,
        Payment.prescription_id: None,
        Payment.match_status: PaymentMatchStatus.UNMATCHED,
    }, synchronize_session="fetch")
    db.flush()

    year = statement.period_from.year if statement.period_from else utcnow().year
    result = match_payments(db, statement_id, year)
    statement.matched_count = result["matched"]
    db.commit()

    url = _detail_redirect_url(
        statement_id, form_data,
        f"rematch_ok&matched={result['matched']}&suggested={result.get('suggested', 0)}",
    )
    return RedirectResponse(url, status_code=302)


# ── Smazání výpisu ─────────────────────────────────────────────────────


@router.post("/vypisy/{statement_id}/smazat")
async def vypis_smazat(
    request: Request,
    statement_id: int,
    db: Session = Depends(get_db),
):
    """Smazat bankovní výpis a jeho platby."""
    statement = db.query(BankStatement).get(statement_id)
    if statement and statement.locked_at:
        return RedirectResponse(f"/platby/vypisy/{statement_id}", status_code=302)
    if statement:
        # Smazat soubor
        if statement.file_path:
            try:
                Path(statement.file_path).unlink()
            except OSError:
                logger.debug("Nelze smazat soubor %s", statement.file_path, exc_info=True)
        log_activity(
            db, ActivityAction.DELETED, "bank_statement", "platby",
            entity_id=statement.id,
            entity_name=statement.filename or f"Výpis #{statement.id}",
        )
        db.delete(statement)
        db.commit()
    return RedirectResponse("/platby/vypisy", status_code=302)


# ── Zamčení / odemčení párování ──────────────────────────────────────


@router.post("/vypisy/{statement_id}/zamknout")
async def vypis_zamknout(
    request: Request,
    statement_id: int,
    db: Session = Depends(get_db),
):
    """Zamknout/odemknout párování výpisu."""
    form_data = await request.form()

    statement = db.query(BankStatement).get(statement_id)
    if not statement:
        return RedirectResponse("/platby/vypisy", status_code=302)

    if statement.locked_at:
        statement.locked_at = None
        flash = "unlock_ok"
    else:
        statement.locked_at = utcnow()
        flash = "lock_ok"
    db.commit()

    url = _detail_redirect_url(statement_id, form_data, flash)
    return RedirectResponse(url, status_code=302)
