"""Vyúčtování — seznam, detail, generování, stav, mazání, export."""

import csv
import io
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, Form, Query, Request
from fastapi.responses import RedirectResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import (
    OwnerUnit, PrescriptionYear, Settlement, SettlementItem, SettlementStatus, Unit,
)
from app.services.settlement_service import (
    generate_settlements,
    get_settlement_detail,
    update_settlement_status,
)
from app.utils import build_list_url, excel_auto_width, is_htmx_partial, strip_diacritics, utcnow

from ._helpers import templates, compute_nav_stats

router = APIRouter()

# ── Pomocné ──────────────────────────────────────────────────────────

STATUS_LABELS = {
    "": "Vše",
    "generated": "Vygenerováno",
    "sent": "Odesláno",
    "paid": "Zaplaceno",
    "overdue": "Po splatnosti",
}

STATUS_COLORS = {
    "generated": ("bg-blue-100 text-blue-800", "dark:bg-blue-900 dark:text-blue-200"),
    "sent": ("bg-yellow-100 text-yellow-800", "dark:bg-yellow-900 dark:text-yellow-200"),
    "paid": ("bg-green-100 text-green-800", "dark:bg-green-900 dark:text-green-200"),
    "overdue": ("bg-red-100 text-red-800", "dark:bg-red-900 dark:text-red-200"),
}

SORT_COLUMNS = {
    "cislo": None,      # Python sort
    "vlastnik": None,
    "predpis": None,
    "zaplaceno": None,
    "vysledek": None,
    "stav": None,
}


# ── Seznam vyúčtování ────────────────────────────────────────────────


@router.get("/vyuctovani")
async def vyuctovani_seznam(
    request: Request,
    rok: int = Query(0),
    stav: str = Query(""),
    q: str = Query(""),
    sort: str = Query("cislo"),
    order: str = Query("asc"),
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    """Seznam vyúčtování s filtry, hledáním a řazením."""
    # Výchozí rok
    if not rok:
        latest = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        rok = latest.year if latest else utcnow().year

    years = [y.year for y in db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).all()]

    # Dotaz
    query = (
        db.query(Settlement)
        .filter_by(year=rok)
        .options(
            joinedload(Settlement.unit).joinedload(Unit.owners).joinedload(OwnerUnit.owner),
            joinedload(Settlement.owner),
            joinedload(Settlement.items),
        )
    )

    # Načteme VŠECHNA vyúčtování roku (pro bubliny + souhrnné statistiky)
    all_settlements = query.all()

    # Statistiky bublin (před filtrováním)
    bubble_counts = {"": len(all_settlements)}
    for s_enum in SettlementStatus:
        bubble_counts[s_enum.value] = sum(1 for s in all_settlements if s.status == s_enum)

    # Souhrnné statistiky
    total_overpay = sum(abs(s.result_amount) for s in all_settlements if s.result_amount < 0)
    total_underpay = sum(s.result_amount for s in all_settlements if s.result_amount > 0)

    # Stav filtr (po statistikách, aby bubliny ukazovaly celkové počty)
    if stav and stav in STATUS_LABELS:
        try:
            status_filter = SettlementStatus(stav)
            settlements = [s for s in all_settlements if s.status == status_filter]
        except ValueError:
            settlements = all_settlements
    else:
        settlements = all_settlements

    # Hledání
    if q:
        q_ascii = strip_diacritics(q)
        def _owners_text(s):
            if s.unit and s.unit.current_owners:
                return " ".join(ou.owner.display_name for ou in s.unit.current_owners)
            return s.owner.display_name if s.owner else ""
        settlements = [
            s for s in settlements
            if q_ascii in strip_diacritics(str(s.unit.unit_number) if s.unit else "")
            or q_ascii in strip_diacritics(_owners_text(s))
            or q_ascii in strip_diacritics(s.variable_symbol or "")
        ]

    # Řazení
    sort_key = sort if sort in SORT_COLUMNS else "cislo"
    reverse = order == "desc"
    sort_fns = {
        "cislo": lambda s: s.unit.unit_number if s.unit else 0,
        "vlastnik": lambda s: strip_diacritics(" ".join(ou.owner.display_name for ou in s.unit.current_owners) if s.unit and s.unit.current_owners else (s.owner.display_name if s.owner else "")),
        "predpis": lambda s: _annual_prescription(s),
        "zaplaceno": lambda s: _total_paid(s),
        "vysledek": lambda s: s.result_amount or 0,
        "stav": lambda s: s.status.value if s.status else "",
    }
    settlements.sort(key=sort_fns.get(sort_key, sort_fns["cislo"]), reverse=reverse)

    list_url = build_list_url(request)

    # Flash zprávy
    flash_message = ""
    flash_param = request.query_params.get("flash", "")
    if flash_param == "generated":
        created = request.query_params.get("created", "0")
        updated = request.query_params.get("updated", "0")
        flash_message = f"Vygenerováno {created} nových"
        if int(updated or 0) > 0:
            flash_message += f", aktualizováno {updated}"
        flash_message += " vyúčtování."
    elif flash_param == "deleted":
        flash_message = "Vyúčtování roku smazána."
    elif flash_param == "bulk_stav":
        count = request.query_params.get("count", "0")
        stav_label = request.query_params.get("stav_label", "")
        flash_message = f"Stav {count} vyúčtování změněn na \u201E{stav_label}\u201C."

    ctx = {
        "request": request,
        "active_nav": "platby",
        "active_tab": "vyuctovani",
        "settlements": settlements,
        "rok": rok,
        "years": years,
        "stav": stav,
        "q": q,
        "sort": sort_key,
        "order": order,
        "back_url": back,
        "list_url": list_url,
        "status_labels": STATUS_LABELS,
        "status_colors": STATUS_COLORS,
        "bubble_counts": bubble_counts,
        "total_overpay": total_overpay,
        "total_underpay": total_underpay,
        "flash_message": flash_message,
        **(compute_nav_stats(db) if not is_htmx_partial(request) else {}),
    }

    if is_htmx_partial(request):
        return templates.TemplateResponse("payments/partials/vyuctovani_tbody.html", ctx)

    return templates.TemplateResponse("payments/vyuctovani.html", ctx)


def _annual_prescription(s: Settlement) -> float:
    """Celkový roční předpis = result + paid (pro sort)."""
    paid = _total_paid(s)
    return (s.result_amount or 0) + paid


def _total_paid(s: Settlement) -> float:
    """Celkem zaplaceno = sum(items.paid)."""
    return sum(item.paid or 0 for item in s.items)


# ── Detail vyúčtování ────────────────────────────────────────────────


@router.get("/vyuctovani/{settlement_id}")
async def vyuctovani_detail(
    settlement_id: int,
    request: Request,
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    """Detail jednoho vyúčtování."""
    detail = get_settlement_detail(db, settlement_id)
    if not detail:
        return RedirectResponse("/platby/vyuctovani", status_code=302)

    settlement = detail["settlement"]
    back_url = back or "/platby/vyuctovani"
    if "/platby/vyuctovani" in back_url and back_url != "/platby/vyuctovani":
        back_label = "Zpět na vyúčtování"
    elif "/jednotky/" in back_url:
        back_label = "Zpět na detail jednotky"
    else:
        back_label = "Zpět na vyúčtování"

    # Flash z query parametru → globální toast
    flash_param = request.query_params.get("flash", "")
    flash_message = ""
    if flash_param == "stav_ok":
        stav_label = request.query_params.get("stav_label", "")
        flash_message = f"Stav změněn na {stav_label}."

    return templates.TemplateResponse("payments/vyuctovani_detail.html", {
        "request": request,
        "active_nav": "platby",
        "active_tab": "vyuctovani",
        "settlement": settlement,
        "detail": detail,
        "back_url": back_url,
        "back_label": back_label,
        "hide_nav_back": True,
        "status_labels": STATUS_LABELS,
        "status_colors": STATUS_COLORS,
        "flash_message": flash_message,
        **compute_nav_stats(db),
    })


# ── Generování vyúčtování ────────────────────────────────────────────


def _vyuctovani_redirect_url(form_data, flash_params: str = "") -> str:
    """Sestaví redirect URL zpět na vyúčtování se zachováním filtrů."""
    params = []
    if flash_params:
        params.append(flash_params)
    for key in ("rok", "q", "stav", "sort", "order", "back"):
        val = form_data.get(key, "")
        if val:
            params.append(f"{key}={val}")
    qs = "&".join(params)
    return f"/platby/vyuctovani?{qs}" if qs else "/platby/vyuctovani"


@router.post("/vyuctovani/generovat")
async def vyuctovani_generovat(
    request: Request,
    rok: int = Form(...),
    back: str = Form(""),
    db: Session = Depends(get_db),
):
    """Generování vyúčtování pro rok."""
    form_data = await request.form()
    result = generate_settlements(db, rok)
    db.commit()
    return RedirectResponse(
        _vyuctovani_redirect_url(form_data, f"flash=generated&created={result['created']}&updated={result['updated']}"),
        status_code=302,
    )


# ── Změna stavu ──────────────────────────────────────────────────────


@router.post("/vyuctovani/{settlement_id}/stav")
async def vyuctovani_zmena_stavu(
    settlement_id: int,
    request: Request,
    novy_stav: str = Form(...),
    back: str = Form(""),
    db: Session = Depends(get_db),
):
    """Změna stavu vyúčtování."""
    settlement = update_settlement_status(db, settlement_id, novy_stav)
    if not settlement:
        return RedirectResponse("/platby/vyuctovani", status_code=302)
    db.commit()

    stav_label = STATUS_LABELS.get(novy_stav, novy_stav)
    redirect_url = f"/platby/vyuctovani/{settlement_id}?flash=stav_ok&stav_label={stav_label}"
    if back:
        redirect_url += f"&back={back}"
    return RedirectResponse(redirect_url, status_code=302)


# ── Smazání roku ─────────────────────────────────────────────────────


@router.post("/vyuctovani/smazat-rok")
async def vyuctovani_smazat_rok(
    request: Request,
    rok: int = Form(...),
    back: str = Form(""),
    db: Session = Depends(get_db),
):
    """Smazání všech vyúčtování roku."""
    form_data = await request.form()
    settlements = db.query(Settlement).filter_by(year=rok).all()
    for s in settlements:
        db.delete(s)
    db.commit()

    return RedirectResponse(
        _vyuctovani_redirect_url(form_data, "flash=deleted"),
        status_code=302,
    )


# ── Hromadná změna stavu ──────────────────────────────────────────


@router.post("/vyuctovani/hromadny-stav")
async def vyuctovani_hromadny_stav(
    request: Request,
    db: Session = Depends(get_db),
):
    """Hromadná změna stavu vybraných vyúčtování."""
    form = await request.form()
    ids = form.getlist("settlement_ids")
    novy_stav = form.get("novy_stav", "")
    rok = int(form.get("rok", 0))
    back = form.get("back", "")

    try:
        status_enum = SettlementStatus(novy_stav)
    except ValueError:
        return RedirectResponse(f"/platby/vyuctovani?rok={rok}", status_code=302)

    count = 0
    for sid in ids:
        settlement = db.query(Settlement).get(int(sid))
        if settlement:
            settlement.status = status_enum
            settlement.updated_at = utcnow()
            count += 1
    db.commit()

    stav_label = STATUS_LABELS.get(novy_stav, novy_stav)
    return RedirectResponse(
        _vyuctovani_redirect_url(form, f"flash=bulk_stav&count={count}&stav_label={stav_label}"),
        status_code=302,
    )


@router.post("/vyuctovani/hromadny-stav-filtrovane")
async def vyuctovani_hromadny_stav_filtrovane(
    request: Request,
    db: Session = Depends(get_db),
):
    """Hromadná změna stavu všech vyúčtování odpovídajících filtru."""
    form = await request.form()
    novy_stav = form.get("novy_stav", "")
    rok = int(form.get("rok", 0))
    stav_filtr = form.get("stav", "")

    try:
        status_enum = SettlementStatus(novy_stav)
    except ValueError:
        return RedirectResponse(f"/platby/vyuctovani?rok={rok}", status_code=302)

    query = db.query(Settlement).filter_by(year=rok)
    if stav_filtr:
        try:
            query = query.filter_by(status=SettlementStatus(stav_filtr))
        except ValueError:
            pass

    count = 0
    for s in query.all():
        s.status = status_enum
        s.updated_at = utcnow()
        count += 1
    db.commit()

    stav_label = STATUS_LABELS.get(novy_stav, novy_stav)
    return RedirectResponse(
        _vyuctovani_redirect_url(form, f"flash=bulk_stav&count={count}&stav_label={stav_label}"),
        status_code=302,
    )


# ── Export vyúčtování ─────────────────────────────────────────────


STAV_SUFFIX = {
    "generated": "vygenerovano",
    "sent": "odeslano",
    "paid": "zaplaceno",
    "overdue": "po_splatnosti",
}


def _get_filtered_settlements(db: Session, rok: int, stav: str, q: str):
    """Společná filtrace pro seznam i export."""
    query = (
        db.query(Settlement)
        .filter_by(year=rok)
        .options(
            joinedload(Settlement.unit).joinedload(Unit.owners).joinedload(OwnerUnit.owner),
            joinedload(Settlement.owner),
            joinedload(Settlement.items),
        )
    )
    if stav and stav in STATUS_LABELS:
        try:
            query = query.filter(Settlement.status == SettlementStatus(stav))
        except ValueError:
            pass

    settlements = query.all()

    if q:
        q_ascii = strip_diacritics(q)
        def _owners_text_export(s):
            if s.unit and s.unit.current_owners:
                return " ".join(ou.owner.display_name for ou in s.unit.current_owners)
            return s.owner.display_name if s.owner else ""
        settlements = [
            s for s in settlements
            if q_ascii in strip_diacritics(str(s.unit.unit_number) if s.unit else "")
            or q_ascii in strip_diacritics(_owners_text_export(s))
            or q_ascii in strip_diacritics(s.variable_symbol or "")
        ]

    settlements.sort(key=lambda s: s.unit.unit_number if s.unit else 0)
    return settlements


@router.get("/vyuctovani/exportovat/{fmt}")
async def vyuctovani_export(
    fmt: str,
    rok: int = Query(0),
    stav: str = Query(""),
    q: str = Query(""),
    detail: str = Query(""),
    db: Session = Depends(get_db),
):
    """Export vyúčtování do Excelu nebo CSV."""
    if fmt not in ("xlsx", "csv"):
        return RedirectResponse("/platby/vyuctovani", status_code=302)

    if not rok:
        latest = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        rok = latest.year if latest else utcnow().year

    settlements = _get_filtered_settlements(db, rok, stav, q)

    # Suffix pro název souboru
    suffix = STAV_SUFFIX.get(stav, "vse")
    if q:
        suffix = "hledani"
    timestamp = datetime.now().strftime("%Y%m%d")
    is_detailed = detail == "1"
    detail_suffix = "_polozky" if is_detailed else ""
    filename = f"vyuctovani_{rok}_{suffix}{detail_suffix}_{timestamp}"

    headers_summary = [
        "Č. jednotky", "Vlastník", "VS", "Předpis měsíční",
        "Předpis roční", "Zaplaceno", "Výsledek", "Stav",
    ]

    if fmt == "xlsx":
        return _export_xlsx(settlements, headers_summary, filename, is_detailed)
    else:
        return _export_csv(settlements, headers_summary, filename, is_detailed)


def _settlement_row(s: Settlement) -> list:
    """Jeden řádek souhrnu vyúčtování."""
    paid = sum(item.paid or 0 for item in s.items)
    annual = (s.result_amount or 0) + paid
    monthly = round(annual / 12, 2) if annual else 0
    result = s.result_amount or 0
    stav = STATUS_LABELS.get(s.status.value, s.status.value) if s.status else ""
    return [
        s.unit.unit_number if s.unit else "",
        ", ".join(ou.owner.display_name for ou in s.unit.current_owners) if s.unit and s.unit.current_owners else (s.owner.display_name if s.owner else ""),
        s.variable_symbol or "",
        monthly,
        annual,
        paid,
        result,
        stav,
    ]


def _export_xlsx(settlements, headers, filename, is_detailed):
    """Generování Excel souboru."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vyúčtování"
    bold = Font(bold=True)

    if not is_detailed:
        # Souhrn
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold
        for row_idx, s in enumerate(settlements, 2):
            for col_idx, val in enumerate(_settlement_row(s), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
    else:
        # Detailní — hlavní řádek + podřádky
        detail_headers = headers + ["Položka", "Kategorie", "Měsíčně", "Ročně", "Zapl. položky", "Výsl. položky"]
        for col, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold
        row_idx = 2
        for s in settlements:
            base = _settlement_row(s)
            if s.items:
                for item in s.items:
                    row_data = base + [
                        item.name or "",
                        item.distribution_key or "",
                        item.cost_unit or 0,
                        item.cost_building or 0,
                        item.paid or 0,
                        item.result or 0,
                    ]
                    for col_idx, val in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=val)
                    row_idx += 1
                    base = [""] * len(headers)  # další položky bez hlavních dat
            else:
                for col_idx, val in enumerate(base + ["", "", 0, 0, 0, 0], 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
                row_idx += 1

    excel_auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


def _export_csv(settlements, headers, filename, is_detailed):
    """Generování CSV souboru."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_ALL)

    if not is_detailed:
        writer.writerow(headers)
        for s in settlements:
            writer.writerow(_settlement_row(s))
    else:
        detail_headers = headers + ["Položka", "Kategorie", "Měsíčně", "Ročně", "Zapl. položky", "Výsl. položky"]
        writer.writerow(detail_headers)
        for s in settlements:
            base = _settlement_row(s)
            if s.items:
                for item in s.items:
                    writer.writerow(base + [
                        item.name or "",
                        item.distribution_key or "",
                        item.cost_unit or 0,
                        item.cost_building or 0,
                        item.paid or 0,
                        item.result or 0,
                    ])
                    base = [""] * len(headers)
            else:
                writer.writerow(base + ["", "", 0, 0, 0, 0])

    return Response(
        content=buf.getvalue().encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )
