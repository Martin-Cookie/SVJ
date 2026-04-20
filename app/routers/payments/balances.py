"""Router pro počáteční zůstatky jednotek — seznam, ruční správa, import z Excelu."""
from __future__ import annotations

import csv
import io
import json
import shutil
from datetime import datetime
from io import BytesIO
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, Query, Request, UploadFile
from fastapi.responses import RedirectResponse, Response
from sqlalchemy.orm import Session, joinedload

from app.config import settings
from app.database import get_db
from app.models import (
    ActivityAction, BalanceSource, Owner, OwnerUnit, SvjInfo, Unit, UnitBalance,
    VariableSymbolMapping, log_activity,
)
from app.utils import (
    build_list_url, build_import_wizard, excel_auto_width, flash_from_params,
    is_htmx_partial, is_safe_path, strip_diacritics, utcnow, validate_upload,
    UPLOAD_LIMITS,
)
from ._helpers import templates, logger, compute_nav_stats

router = APIRouter()


# ── Helpers ──────────────────────────────────────────────────────────────

def _load_balance_mapping(db: Session) -> dict | None:
    info = db.query(SvjInfo).first()
    if info and info.balance_import_mapping:
        try:
            return json.loads(info.balance_import_mapping)
        except (json.JSONDecodeError, TypeError):
            pass
    return None


def _save_balance_mapping(db: Session, mapping: dict):
    info = db.query(SvjInfo).first()
    if info:
        info.balance_import_mapping = json.dumps(mapping, ensure_ascii=False)
        db.commit()


def _owners_for_units(db: Session) -> dict[int, list]:
    """Vrátí dict unit_id → list[Owner] (aktivní vlastníci)."""
    owner_units = (
        db.query(OwnerUnit)
        .filter(OwnerUnit.valid_to.is_(None))
        .options(joinedload(OwnerUnit.owner))
        .all()
    )
    result: dict[int, list] = {}
    for ou in owner_units:
        result.setdefault(ou.unit_id, []).append(ou.owner)
    return result


# ── Seznam zůstatků ──────────────────────────────────────────────────────


def _unit_vs_map(db: Session) -> dict[int, list[str]]:
    """unit_id → list variabilních symbolů (pro vyhledávání)."""
    mappings = db.query(VariableSymbolMapping).filter(VariableSymbolMapping.unit_id.isnot(None)).all()
    result: dict[int, list[str]] = {}
    for m in mappings:
        if m.unit_id and m.variable_symbol:
            result.setdefault(m.unit_id, []).append(m.variable_symbol)
    return result


def _filter_balances(
    db: Session,
    rok: int,
    q: str = "",
    sort: str = "jednotka",
    order: str = "asc",
) -> list:
    """Sdílená filtrační logika pro seznam i export."""
    query = (
        db.query(UnitBalance)
        .options(joinedload(UnitBalance.unit), joinedload(UnitBalance.owner))
    )
    if rok:
        query = query.filter(UnitBalance.year == rok)

    balances = query.all()

    if q:
        q_ascii = strip_diacritics(q)
        vs_map = _unit_vs_map(db)

        def _match(b):
            unit_num = str(b.unit.unit_number) if b.unit else ""
            owner_name = b.owner.name_normalized if b.owner else (b.owner_name or "")
            haystack = strip_diacritics(f"{unit_num} {owner_name}")
            if q_ascii in haystack:
                return True
            # VS match (ponecháváme přesně)
            if b.unit_id:
                for vs in vs_map.get(b.unit_id, []):
                    if q in vs:
                        return True
            return False

        balances = [b for b in balances if _match(b)]

    if sort == "jednotka":
        balances.sort(key=lambda b: b.unit.unit_number if b.unit else 0, reverse=(order == "desc"))
    elif sort == "castka":
        balances.sort(key=lambda b: b.opening_amount, reverse=(order == "desc"))
    elif sort == "vlastnik":
        balances.sort(
            key=lambda b: (b.owner.name_normalized if b.owner else b.owner_name or ""),
            reverse=(order == "desc"),
        )

    return balances


@router.get("/zustatky")
async def zustatky_seznam(
    request: Request,
    rok: int = 0,
    q: str = "",
    sort: str = "jednotka",
    order: str = "asc",
    db: Session = Depends(get_db),
):
    """Seznam počátečních zůstatků."""
    years = (
        db.query(UnitBalance.year)
        .distinct()
        .order_by(UnitBalance.year.desc())
        .all()
    )
    years = [y[0] for y in years]

    if rok == 0:
        rok = years[0] if years else utcnow().year

    balances = _filter_balances(db, rok, q, sort, order)
    total_count = db.query(UnitBalance).filter(UnitBalance.year == rok).count() if rok else 0

    units = db.query(Unit).order_by(Unit.unit_number).all()
    existing_unit_ids = {b.unit_id for b in balances}

    total_preplatek = sum(b.opening_amount for b in balances if b.opening_amount > 0)
    total_nedoplatek = sum(b.opening_amount for b in balances if b.opening_amount < 0)

    # Vlastníci per unit pro formulář
    owners_by_unit = _owners_for_units(db)

    list_url = build_list_url(request)
    back_url = request.query_params.get("back", "")

    flash_message, flash_type = flash_from_params(request, {
        "ok": ("Zůstatek uložen.", "success"),
        "smazano": ("Zůstatek smazán.", "success"),
        "chyba_rok": ("Rok musí být mezi 2020 a 2040.", "error"),
        "import_ok": ("Zůstatky úspěšně naimportovány.", "success"),
        "import_chyba": ("Chyba při importu zůstatků.", "error"),
    })

    ctx = {
        "request": request,
        "active_nav": "platby",
        "balances": balances,
        "units": units,
        "years": years,
        "rok": rok,
        "q": q,
        "total_count": total_count,
        "sort": sort,
        "order": order,
        "existing_unit_ids": existing_unit_ids,
        "total_preplatek": total_preplatek,
        "total_nedoplatek": total_nedoplatek,
        "owners_by_unit": owners_by_unit,
        "list_url": list_url,
        "back_url": back_url,
        "flash_message": flash_message,
        "flash_type": flash_type,
        "active_tab": "zustatky",
        **(compute_nav_stats(db) if not is_htmx_partial(request) else {}),
    }

    if is_htmx_partial(request):
        return templates.TemplateResponse(request, "payments/partials/zustatky_tbody.html", ctx)

    return templates.TemplateResponse(request, "payments/zustatky.html", ctx)


# ── Export zůstatků ──────────────────────────────────────────────────────


@router.get("/zustatky/exportovat/{fmt}")
async def zustatky_export(
    fmt: str,
    rok: int = Query(0),
    q: str = Query(""),
    sort: str = Query("jednotka"),
    order: str = Query("asc"),
    db: Session = Depends(get_db),
):
    """Export počátečních zůstatků (xlsx/csv) respektující rok a hledání."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if fmt not in ("xlsx", "csv"):
        return RedirectResponse("/platby/zustatky", status_code=302)

    if rok == 0:
        latest = db.query(UnitBalance.year).order_by(UnitBalance.year.desc()).first()
        rok = latest[0] if latest else utcnow().year

    balances = _filter_balances(db, rok, q, sort, order)

    headers = ["Rok", "Katastr. č.", "Vlastník", "Částka (Kč)", "Typ", "Zdroj", "Poznámka"]

    source_labels = {
        "manual": "Ručně",
        "import": "Import",
        "legacy": "Historicky",
    }

    def _row(b):
        owner_name = ""
        if b.owner and b.owner.display_name:
            owner_name = b.owner.display_name
        elif b.owner_name:
            owner_name = b.owner_name
        amount = b.opening_amount or 0
        typ = "Přeplatek" if amount > 0 else ("Nedoplatek" if amount < 0 else "Vyrovnáno")
        source_val = b.source.value if b.source else ""
        return [
            b.year,
            b.unit.unit_number if b.unit else "",
            owner_name,
            round(amount, 2),
            typ,
            source_labels.get(source_val, source_val),
            b.note or "",
        ]

    suffix = "_hledani" if q else f"_{rok}"
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"zustatky{suffix}_{timestamp}"

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = f"Zůstatky {rok}"
        bold = Font(bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold

        for row_idx, b in enumerate(balances, 2):
            for col_idx, val in enumerate(_row(b), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        excel_auto_width(ws)

        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    else:
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(headers)
        for b in balances:
            writer.writerow(_row(b))
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )


# ── Ruční správa ─────────────────────────────────────────────────────────

@router.post("/zustatky/pridat")
async def zustatek_pridat(
    request: Request,
    unit_id: int = Form(...),
    year: int = Form(...),
    opening_amount: float = Form(...),
    owner_id: int = Form(0),
    owner_name: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    """Přidat/aktualizovat počáteční zůstatek."""
    if year < 2020 or year > 2040:
        return RedirectResponse("/platby/zustatky?flash=chyba_rok", status_code=302)

    _owner_id = owner_id if owner_id else None
    _owner_name = owner_name.strip() or None
    # Pokud owner_id je zadáno, vyplnit owner_name automaticky
    if _owner_id and not _owner_name:
        owner = db.query(Owner).get(_owner_id)
        if owner:
            _owner_name = owner.display_name

    existing = db.query(UnitBalance).filter_by(unit_id=unit_id, year=year).first()
    if existing:
        existing.opening_amount = opening_amount
        existing.owner_id = _owner_id
        existing.owner_name = _owner_name
        existing.note = note.strip() or None
        existing.source = BalanceSource.MANUAL
    else:
        db.add(UnitBalance(
            unit_id=unit_id,
            year=year,
            opening_amount=opening_amount,
            source=BalanceSource.MANUAL,
            owner_id=_owner_id,
            owner_name=_owner_name,
            note=note.strip() or None,
        ))
    log_activity(
        db, ActivityAction.UPDATED if existing else ActivityAction.CREATED,
        "unit_balance", "platby",
        entity_name=f"Zůstatek jednotky {unit_id} / {year}",
    )
    db.commit()
    return RedirectResponse(f"/platby/zustatky?rok={year}&flash=ok", status_code=302)


@router.get("/zustatky/{balance_id}/upravit-formular")
async def zustatek_edit_form(
    request: Request,
    balance_id: int,
    db: Session = Depends(get_db),
):
    """Vrátí HTMX partial s editačním řádkem."""
    balance = db.query(UnitBalance).options(
        joinedload(UnitBalance.unit), joinedload(UnitBalance.owner),
    ).get(balance_id)
    if not balance:
        return RedirectResponse("/platby/zustatky", status_code=302)

    # Vlastníci pro dropdown
    unit_owners = []
    if balance.unit_id:
        ous = (
            db.query(OwnerUnit)
            .filter(OwnerUnit.unit_id == balance.unit_id, OwnerUnit.valid_to.is_(None))
            .options(joinedload(OwnerUnit.owner))
            .all()
        )
        unit_owners = [ou.owner for ou in ous]

    rok = request.query_params.get("rok", str(balance.year))
    return templates.TemplateResponse(request, "payments/partials/_zustatek_edit_row.html", {
        "b": balance,
        "unit_owners": unit_owners,
        "rok": rok,
        "q": request.query_params.get("q", ""),
        "sort": request.query_params.get("sort", "jednotka"),
        "order": request.query_params.get("order", "asc"),
        "back_url": request.query_params.get("back", ""),
        "list_url": request.query_params.get("list_url", "/platby/zustatky"),
    })


@router.get("/zustatky/{balance_id}/info")
async def zustatek_info_row(
    request: Request,
    balance_id: int,
    db: Session = Depends(get_db),
):
    """Vrátí HTMX partial se zobrazovacím řádkem (pro cancel)."""
    balance = db.query(UnitBalance).options(
        joinedload(UnitBalance.unit), joinedload(UnitBalance.owner),
    ).get(balance_id)
    if not balance:
        return RedirectResponse("/platby/zustatky", status_code=302)

    return templates.TemplateResponse(request, "payments/partials/_zustatek_view_row.html", {
        "b": balance,
        "rok": request.query_params.get("rok", str(balance.year)),
        "q": request.query_params.get("q", ""),
        "sort": request.query_params.get("sort", "jednotka"),
        "order": request.query_params.get("order", "asc"),
        "back_url": request.query_params.get("back", ""),
        "list_url": request.query_params.get("list_url", "/platby/zustatky"),
    })


@router.post("/zustatky/{balance_id}/upravit")
async def zustatek_upravit(
    request: Request,
    balance_id: int,
    opening_amount: float = Form(...),
    owner_id: int = Form(0),
    owner_name: str = Form(""),
    note: str = Form(""),
    rok: int = Form(0),
    q: str = Form(""),
    sort: str = Form("jednotka"),
    order: str = Form("asc"),
    back: str = Form(""),
    db: Session = Depends(get_db),
):
    """Upravit existující zůstatek."""
    balance = db.query(UnitBalance).get(balance_id)
    if not balance:
        return RedirectResponse("/platby/zustatky", status_code=302)

    _owner_id = owner_id if owner_id else None
    _owner_name = owner_name.strip() or None
    if _owner_id and not _owner_name:
        owner = db.query(Owner).get(_owner_id)
        if owner:
            _owner_name = owner.display_name

    balance.opening_amount = opening_amount
    balance.owner_id = _owner_id
    balance.owner_name = _owner_name
    balance.note = note.strip() or None
    balance.source = BalanceSource.MANUAL
    log_activity(
        db, ActivityAction.UPDATED, "unit_balance", "platby",
        entity_id=balance.id,
        entity_name=f"Zůstatek jednotky {balance.unit_id} / {balance.year}",
    )
    db.commit()
    _rok = rok or balance.year
    qs = f"rok={_rok}&flash=ok"
    if q:
        qs += f"&q={q}"
    if sort != "jednotka":
        qs += f"&sort={sort}"
    if order != "asc":
        qs += f"&order={order}"
    if back:
        qs += f"&back={back}"
    return RedirectResponse(f"/platby/zustatky?{qs}#bal-{balance.id}", status_code=302)


@router.post("/zustatky/{balance_id}/smazat")
async def zustatek_smazat(
    request: Request,
    balance_id: int,
    db: Session = Depends(get_db),
):
    """Smazat zůstatek."""
    balance = db.query(UnitBalance).get(balance_id)
    rok = balance.year if balance else 0
    if balance:
        log_activity(
            db, ActivityAction.DELETED, "unit_balance", "platby",
            entity_id=balance.id,
            entity_name=f"Zůstatek jednotky {balance.unit_id} / {balance.year}",
        )
        db.delete(balance)
        db.commit()
    return RedirectResponse(f"/platby/zustatky?rok={rok}&flash=smazano", status_code=302)


# ── API pro owners per unit (JS dropdown) ────────────────────────────────

@router.get("/zustatky/vlastnici/{unit_id}")
async def zustatky_owners_for_unit(
    unit_id: int,
    db: Session = Depends(get_db),
):
    """Vrátí JSON seznam vlastníků dané jednotky."""
    from fastapi.responses import JSONResponse
    owner_units = (
        db.query(OwnerUnit)
        .filter(OwnerUnit.unit_id == unit_id, OwnerUnit.valid_to.is_(None))
        .options(joinedload(OwnerUnit.owner))
        .all()
    )
    owners = [{"id": ou.owner.id, "name": ou.owner.display_name} for ou in owner_units]
    return JSONResponse(owners)


# ── Import z Excelu ──────────────────────────────────────────────────────

@router.get("/zustatky/import")
async def zustatky_import_upload(
    request: Request,
    db: Session = Depends(get_db),
):
    """Upload formulář pro import zůstatků."""
    return templates.TemplateResponse(request, "payments/zustatky_import.html", {
        "active_nav": "platby",
        "active_tab": "zustatky",
        **build_import_wizard(1),
        "default_year": utcnow().year,
        **(compute_nav_stats(db) if not is_htmx_partial(request) else {}),
    })


@router.post("/zustatky/import")
async def zustatky_import_upload_post(
    request: Request,
    file: UploadFile = File(...),
    year: int = Form(...),
    db: Session = Depends(get_db),
):
    """Zpracování uploadu — uložit soubor, redirect na mapování."""
    err = await validate_upload(file, **UPLOAD_LIMITS["excel"])
    if err:
        return templates.TemplateResponse(request, "payments/zustatky_import.html", {
            "active_nav": "platby",
            "active_tab": "zustatky",
            **build_import_wizard(1),
            "default_year": year,
            "error": err,
            **(compute_nav_stats(db) if not is_htmx_partial(request) else {}),
        })

    # Uložit soubor
    upload_dir = settings.upload_dir / "excel"
    upload_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = upload_dir / f"{ts}_{file.filename}"
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)

    return _balance_mapping_page(request, str(dest), file.filename, year, db)


@router.get("/zustatky/import/mapovani")
async def zustatky_import_mapping_get(
    request: Request,
    file_path: str = "",
    filename: str = "",
    year: int = 0,
    db: Session = Depends(get_db),
):
    """Mapování sloupců — GET (po reload)."""
    if not file_path or not is_safe_path(Path(file_path), settings.upload_dir):
        return RedirectResponse("/platby/zustatky/import", status_code=302)
    return _balance_mapping_page(request, file_path, filename, year or utcnow().year, db)


@router.post("/zustatky/import/mapovani")
async def zustatky_import_mapping_post(
    request: Request,
    file_path: str = Form(""),
    filename: str = Form(""),
    year: int = Form(0),
    sheet_name: str = Form(""),
    start_row: int = Form(1),
    db: Session = Depends(get_db),
):
    """Reload mapování s jiným sheet/řádkem."""
    if not file_path or not is_safe_path(Path(file_path), settings.upload_dir):
        return RedirectResponse("/platby/zustatky/import", status_code=302)
    return _balance_mapping_page(
        request, file_path, filename, year or utcnow().year, db,
        sheet_name=sheet_name or None, start_row=start_row,
    )


def _balance_mapping_page(
    request, file_path, filename, year, db,
    sheet_name=None, start_row=None,
):
    from app.services.import_mapping import (
        read_excel_headers, read_excel_sheet_names,
        BALANCE_FIELD_DEFS, BALANCE_FIELD_GROUPS,
        build_mapping_context,
    )
    saved_mapping = _load_balance_mapping(db)
    sheets = read_excel_sheet_names(file_path)
    if start_row is None:
        start_row = saved_mapping.get("start_row", 1) if saved_mapping else 1
    if sheet_name is None and saved_mapping:
        sheet_name = saved_mapping.get("sheet_name")

    headers = read_excel_headers(file_path, sheet_name=sheet_name, header_row=start_row)
    ctx = build_mapping_context(headers, BALANCE_FIELD_DEFS, BALANCE_FIELD_GROUPS, saved_mapping)

    return templates.TemplateResponse(request, "payments/zustatky_mapping.html", {
        "active_nav": "platby",
        "active_tab": "zustatky",
        **build_import_wizard(2),
        "file_path": file_path,
        "filename": filename,
        "year": year,
        "sheets": sheets,
        "current_sheet": sheet_name or (sheets[0] if sheets else ""),
        "start_row": start_row,
        **ctx,
        **(compute_nav_stats(db) if not is_htmx_partial(request) else {}),
    })


@router.post("/zustatky/import/nahled")
async def zustatky_import_preview(
    request: Request,
    file_path: str = Form(""),
    filename: str = Form(""),
    year: int = Form(0),
    mapping_json: str = Form(""),
    db: Session = Depends(get_db),
):
    """Náhled importu s párováním."""
    if not file_path or not is_safe_path(Path(file_path), settings.upload_dir):
        return RedirectResponse("/platby/zustatky/import", status_code=302)

    import json as _json
    try:
        mapping = _json.loads(mapping_json)
    except (ValueError, TypeError):
        return RedirectResponse("/platby/zustatky/import", status_code=302)

    from app.services.import_mapping import validate_balance_mapping, BALANCE_FIELD_DEFS, BALANCE_FIELD_GROUPS
    err = validate_balance_mapping(mapping)
    if err:
        return _balance_mapping_page(request, file_path, filename, year, db)

    # Uložit mapování pokud uživatel zaškrtl
    if mapping.pop("save", False):
        _save_balance_mapping(db, mapping)

    from app.services.balance_import import preview_balance_import
    result = preview_balance_import(file_path, mapping, year, db)

    return templates.TemplateResponse(request, "payments/zustatky_preview.html", {
        "active_nav": "platby",
        "active_tab": "zustatky",
        **build_import_wizard(3),
        "file_path": file_path,
        "filename": filename,
        "year": year,
        "mapping_json": _json.dumps(mapping, ensure_ascii=False),
        "preview": result,
        **(compute_nav_stats(db) if not is_htmx_partial(request) else {}),
    })


@router.post("/zustatky/import/potvrdit")
async def zustatky_import_confirm(
    request: Request,
    file_path: str = Form(""),
    year: int = Form(0),
    mapping_json: str = Form(""),
    db: Session = Depends(get_db),
):
    """Provedení importu."""
    if not file_path or not is_safe_path(Path(file_path), settings.upload_dir):
        return RedirectResponse("/platby/zustatky?flash=import_chyba", status_code=302)

    import json as _json
    try:
        mapping = _json.loads(mapping_json)
    except (ValueError, TypeError):
        return RedirectResponse("/platby/zustatky?flash=import_chyba", status_code=302)

    from app.services.balance_import import execute_balance_import
    try:
        result = execute_balance_import(file_path, mapping, year, db)
        logger.info("Balance import done: %s", result)
        log_activity(
            db, ActivityAction.IMPORTED, "unit_balance", "platby",
            entity_name=f"Zůstatky {year}",
            description=f"{result.get('created', 0) + result.get('updated', 0)} jednotek",
        )
        db.commit()
    except Exception as e:
        logger.error("Balance import failed: %s", e)
        return RedirectResponse("/platby/zustatky?flash=import_chyba", status_code=302)

    return RedirectResponse(f"/platby/zustatky?rok={year}&flash=import_ok", status_code=302)
