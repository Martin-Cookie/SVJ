import csv
import io
import logging
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import RedirectResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import case, func
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import (
    ActivityLog, Ballot, BallotStatus, BallotVote, BankStatement,
    EmailDeliveryStatus, EmailLog, Owner, OwnerUnit, Payment,
    PaymentDirection, PaymentMatchStatus, PrescriptionYear,
    Space, SpaceStatus, SpaceTenant, SvjInfo,
    TaxDistribution, TaxDocument, TaxSession,
    Tenant, Unit, Voting,
)
from app.utils import excel_auto_width, strip_diacritics, templates, utcnow

logger = logging.getLogger(__name__)
router = APIRouter()

# Normalizace raw module stringů na kanonické klíče — sjednocuje EmailLog vs ActivityLog,
# aby `tax` a `dane` splynuly v jednu bublinu "Rozesílání".
_MODULE_CANONICAL = {
    "tax": "dane",
    "voting": "hlasovani",
    "tenants": "najemci",
    "water_meter": "vodometry",
    "water_notice": "vodometry",
    "water_meters": "vodometry",
}


_KNOWN_MODULES = {
    "vlastnici", "jednotky", "najemci", "prostory",
    "hlasovani", "dane", "sync", "platby", "payment_notice",
    "sprava", "nastaveni", "water_meters", "vodometry",
}


def _norm_module(m: str) -> str:
    canonical = _MODULE_CANONICAL.get(m or "", m or "")
    # Neznámé moduly (test, None, číselné) → spadnou do "sprava"
    if canonical not in _KNOWN_MODULES:
        return "sprava"
    return canonical


@router.get("/prehled/rozdil-podilu")
async def shares_breakdown(request: Request, vse: int = 0, db: Session = Depends(get_db)):
    """Porovnání podílů dle prohlášení vs. evidence vlastníků."""
    svj_info = db.query(SvjInfo).first()
    declared_shares = svj_info.total_shares if svj_info and svj_info.total_shares else 0

    # Per-unit: sum of active owner votes
    owner_votes_subq = (
        db.query(
            OwnerUnit.unit_id,
            func.sum(OwnerUnit.votes).label("owners_votes"),
        )
        .filter(OwnerUnit.valid_to.is_(None))
        .group_by(OwnerUnit.unit_id)
        .subquery()
    )

    rows = (
        db.query(
            Unit.id,
            Unit.unit_number,
            Unit.podil_scd,
            owner_votes_subq.c.owners_votes,
        )
        .outerjoin(owner_votes_subq, Unit.id == owner_votes_subq.c.unit_id)
        .order_by(Unit.unit_number)
        .all()
    )

    items = []
    total_units_scd = 0
    total_owners_votes = 0
    for unit_id, unit_number, podil_scd, owners_votes in rows:
        p = podil_scd or 0
        o = owners_votes or 0
        total_units_scd += p
        total_owners_votes += o
        items.append({
            "unit_id": unit_id,
            "unit_number": unit_number,
            "podil_scd": p,
            "owners_votes": o,
            "diff": o - p,
        })

    show_all = bool(vse)
    filtered_items = items if show_all else [i for i in items if i["diff"] != 0]

    return templates.TemplateResponse(request, "dashboard_shares.html", {
        "active_nav": "dashboard",
        "declared_shares": declared_shares,
        "total_units_scd": total_units_scd,
        "total_owners_votes": total_owners_votes,
        "diff_owners": total_owners_votes - declared_shares,
        "diff_units": total_units_scd - declared_shares,
        "items": filtered_items,
        "show_all": show_all,
        "total_count": len(items),
        "diff_count": len([i for i in items if i["diff"] != 0]),
    })


@router.get("/")
async def home(
    request: Request,
    q: str = Query(""),
    modul: str = Query(""),
    sort: str = Query("date"),
    order: str = Query("desc"),
    db: Session = Depends(get_db),
):
    """Hlavní přehledová stránka se statistikami a poslední aktivitou."""
    owners_count = db.query(Owner).filter_by(is_active=True).count()
    units_count = db.query(Unit).count()
    # Voting stats: count per status (lightweight)
    svj_info = db.query(SvjInfo).first()
    declared_shares = svj_info.total_shares if svj_info and svj_info.total_shares else 0

    voting_counts = (
        db.query(Voting.status, func.count(Voting.id))
        .group_by(Voting.status)
        .all()
    )
    total_votings = sum(c for _, c in voting_counts)

    # Nejnovější hlasování per status — jedna agregace přes max(updated_at)
    voting_by_status = {}
    if voting_counts:
        latest_per_status_subq = (
            db.query(
                Voting.status.label("status"),
                func.max(Voting.updated_at).label("max_updated"),
            )
            .group_by(Voting.status)
            .subquery()
        )
        latest_votings = (
            db.query(Voting)
            .join(
                latest_per_status_subq,
                (Voting.status == latest_per_status_subq.c.status)
                & (Voting.updated_at == latest_per_status_subq.c.max_updated),
            )
            .all()
        )
        latest_ids = [v.id for v in latest_votings]

        # Počet lístků per voting (1 dotaz)
        ballot_counts = dict(
            db.query(Ballot.voting_id, func.count(Ballot.id))
            .filter(Ballot.voting_id.in_(latest_ids))
            .group_by(Ballot.voting_id)
            .all()
        ) if latest_ids else {}

        # Kvórum: sum(total_votes) distinct PROCESSED lístků s ≥1 vote (1 dotaz, group by)
        processed_votes_map: dict = {}
        if latest_ids:
            voted_sub = (
                db.query(Ballot.voting_id, Ballot.id, Ballot.total_votes)
                .join(BallotVote, BallotVote.ballot_id == Ballot.id)
                .filter(
                    Ballot.voting_id.in_(latest_ids),
                    Ballot.status == BallotStatus.PROCESSED,
                    BallotVote.vote.isnot(None),
                )
                .distinct()
                .subquery()
            )
            for vid, total in (
                db.query(voted_sub.c.voting_id, func.coalesce(func.sum(voted_sub.c.total_votes), 0))
                .group_by(voted_sub.c.voting_id)
                .all()
            ):
                processed_votes_map[vid] = total or 0

        status_counts_map = {s: c for s, c in voting_counts}
        for latest in latest_votings:
            processed_votes = processed_votes_map.get(latest.id, 0)
            quorum_pct = round(processed_votes / declared_shares * 100, 2) if declared_shares else 0
            voting_by_status[latest.status.value] = {
                "count": status_counts_map.get(latest.status, 0),
                "latest": latest,
                "date": latest.start_date or latest.created_at,
                "total_ballots": ballot_counts.get(latest.id, 0),
                "quorum_pct": quorum_pct,
            }

    # Tax sessions — count per status in one query
    tax_status_counts = (
        db.query(TaxSession.send_status, func.count(TaxSession.id))
        .group_by(TaxSession.send_status)
        .all()
    )

    # Latest session per status — jedna agregace přes max(created_at)
    latest_session_ids = {}
    tax_by_status = {}
    if tax_status_counts:
        latest_tax_subq = (
            db.query(
                TaxSession.send_status.label("status"),
                func.max(TaxSession.created_at).label("max_created"),
            )
            .group_by(TaxSession.send_status)
            .subquery()
        )
        latest_tax_sessions = (
            db.query(TaxSession)
            .join(
                latest_tax_subq,
                (TaxSession.send_status == latest_tax_subq.c.status)
                & (TaxSession.created_at == latest_tax_subq.c.max_created),
            )
            .all()
        )
        tax_status_map = {s: c for s, c in tax_status_counts}
        for latest in latest_tax_sessions:
            s = latest.send_status.value
            latest_session_ids[latest.id] = s
            tax_by_status[s] = {
                "count": tax_status_map.get(latest.send_status, 0),
                "latest": latest,
                "date": latest.created_at,
                "sent": 0,
                "total_dists": 0,
            }

    # Single query for distribution stats of latest sessions per status
    if latest_session_ids:
        dist_stats = (
            db.query(
                TaxDocument.session_id,
                func.count(TaxDistribution.id),
                func.sum(case(
                    (TaxDistribution.email_status == EmailDeliveryStatus.SENT, 1),
                    else_=0,
                )),
            )
            .join(TaxDistribution.document)
            .filter(TaxDocument.session_id.in_(latest_session_ids.keys()))
            .group_by(TaxDocument.session_id)
            .all()
        )
        for session_id, total, sent in dist_stats:
            s = latest_session_ids[session_id]
            tax_by_status[s]["total_dists"] = total
            tax_by_status[s]["sent"] = sent or 0

    # Unified activity: EmailLog + ActivityLog (limit 200 pro výkon)
    recent_emails = db.query(EmailLog).order_by(EmailLog.created_at.desc()).limit(200).all()
    recent_activity = db.query(ActivityLog).order_by(ActivityLog.created_at.desc()).limit(200).all()

    unified = []
    # URL mapování pro entity_type → detail stránka
    _entity_urls = {
        "voting": "/hlasovani/{id}",
        "tax_session": "/rozesilani/{id}",
        "owner": "/vlastnici/{id}",
    }

    # Seskupení payment_notice emailů (a jiných hromadných emailů) dle den+modul+subject
    from collections import defaultdict
    grouped_emails = defaultdict(list)
    for e in recent_emails:
        if e.module == "payment_notice":
            day = e.created_at.strftime("%Y-%m-%d") if e.created_at else "?"
            key = (day, e.module, e.subject or "")
            grouped_emails[key].append(e)
        else:
            url = f"/rozesilani/{e.reference_id}" if e.reference_id else ""
            unified.append({
                "type": "email",
                "created_at": e.created_at,
                "module": _norm_module(e.module),
                "description": e.subject or "",
                "detail": e.recipient_name or e.recipient_email or "",
                "status": e.status.value if e.status else "",
                "url": url,
                "entity": e,
            })

    # Seskupené payment_notice → jeden řádek za den+subject
    for (day, mod, subject), emails in grouped_emails.items():
        latest = max(emails, key=lambda e: e.created_at or datetime.min)
        sent_count = sum(1 for e in emails if e.status and e.status.value == "sent")
        failed_count = sum(1 for e in emails if e.status and e.status.value == "failed")
        count = len(emails)
        detail = f"{count}× odesláno"
        if failed_count:
            detail = f"{sent_count}× odesláno, {failed_count}× chyba"
        ref_id = latest.reference_id
        url = f"/platby/vypisy/{ref_id}" if ref_id else ""
        unified.append({
            "type": "email",
            "created_at": latest.created_at,
            "module": _norm_module(mod),
            "description": f"{subject}" if count == 1 else f"{subject} ({count}×)",
            "detail": detail,
            "status": "sent" if failed_count == 0 else "failed",
            "url": url,
            "entity": latest,
            "grouped_count": count,
        })

    for a in recent_activity:
        url_tpl = _entity_urls.get(a.entity_type, "")
        url = url_tpl.format(id=a.entity_id) if url_tpl and a.entity_id else ""
        unified.append({
            "type": "activity",
            "created_at": a.created_at,
            "module": _norm_module(a.module),
            "description": a.entity_name or "",
            "detail": a.description or "",
            "status": a.action.value if a.action else "",
            "url": url,
            "entity": a,
        })

    # Sort combined
    unified.sort(key=lambda x: x["created_at"] or datetime.min, reverse=True)

    # Počty per modul (pro bubliny) — před filtrováním
    module_counts = defaultdict(int)
    for item in unified:
        module_counts[item["module"]] += 1

    # Fixní pořadí bublin shodné se sidebarem (skryje se položka s 0 záznamy)
    _MODULE_ORDER = [
        "vlastnici", "jednotky", "najemci", "prostory",
        "hlasovani", "dane", "sync", "platby", "payment_notice",
        "sprava", "nastaveni",
    ]
    module_counts_ordered = [
        (key, module_counts[key]) for key in _MODULE_ORDER if module_counts.get(key, 0) > 0
    ]
    # Neznámé moduly přemapovány na "sprava" v _norm_module()

    # Filtr dle modulu (akceptuje i legacy raw klíče jako `tax`, `voting`)
    if modul:
        modul_canonical = _norm_module(modul)
        unified = [item for item in unified if item["module"] == modul_canonical]

    # Search filtering
    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        unified = [
            item for item in unified
            if q_lower in item["description"].lower()
            or q_ascii in strip_diacritics(item["description"])
            or q_lower in item["detail"].lower()
            or q_ascii in strip_diacritics(item["detail"])
            or q_lower in item["module"].lower()
        ]

    # Sorting
    SORT_KEYS = {
        "date": lambda x: x["created_at"],
        "module": lambda x: x["module"].lower(),
        "description": lambda x: strip_diacritics(x["description"]),
        "detail": lambda x: strip_diacritics(x["detail"]),
        "status": lambda x: x["status"],
    }
    sort_fn = SORT_KEYS.get(sort, SORT_KEYS["date"])
    unified.sort(key=sort_fn, reverse=(order == "desc"))

    # Share statistics (svj_info + declared_shares already loaded above)
    owners_scd = db.query(func.sum(OwnerUnit.votes)).filter(OwnerUnit.valid_to.is_(None)).scalar() or 0
    units_scd = db.query(func.sum(Unit.podil_scd)).scalar() or 0

    # Payment stats — jedna agregace místo 4 samostatných dotazů
    statement_count = db.query(BankStatement).count()
    confirmed_statuses = [PaymentMatchStatus.AUTO_MATCHED, PaymentMatchStatus.MANUAL]
    payment_stats = db.query(
        func.sum(case((Payment.match_status.in_(confirmed_statuses), 1), else_=0)).label("matched"),
        func.sum(case((Payment.match_status == PaymentMatchStatus.UNMATCHED, 1), else_=0)).label("unmatched"),
        func.coalesce(
            func.sum(case((Payment.match_status.in_(confirmed_statuses), Payment.amount), else_=0)),
            0,
        ).label("income"),
    ).filter(Payment.direction == PaymentDirection.INCOME).one()
    matched_payments = payment_stats.matched or 0
    unmatched_payments = payment_stats.unmatched or 0
    total_income = payment_stats.income or 0

    # Debtor count (imported from payments helpers if available)
    debtor_count = 0
    try:
        from app.routers.payments._helpers import _count_debtors_fast
        latest_py = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        if latest_py:
            debtor_count = _count_debtors_fast(db, latest_py.year)
    except Exception as e:
        logger.warning("Debtor count failed: %s", e)

    # Tenant stats
    tenants_count = db.query(Tenant).filter_by(is_active=True).count()
    tenants_with_contract = (
        db.query(func.count(func.distinct(SpaceTenant.tenant_id)))
        .filter(SpaceTenant.is_active == True)
        .scalar() or 0
    )
    tenants_linked = (
        db.query(func.count(Tenant.id))
        .filter(Tenant.is_active == True, Tenant.owner_id.isnot(None))
        .scalar() or 0
    )

    # Space stats
    space_total = db.query(Space).count()
    space_status_counts = dict(
        db.query(Space.status, func.count(Space.id))
        .group_by(Space.status).all()
    )
    space_rented = space_status_counts.get(SpaceStatus.RENTED, 0)
    space_vacant = space_status_counts.get(SpaceStatus.VACANT, 0)
    space_blocked = space_status_counts.get(SpaceStatus.BLOCKED, 0)

    # Contracts expiring within 3 months
    from datetime import timedelta
    expiry_cutoff = utcnow().date() + timedelta(days=90)
    expiring_contracts = (
        db.query(SpaceTenant)
        .filter(
            SpaceTenant.is_active == True,
            SpaceTenant.contract_end.isnot(None),
            SpaceTenant.contract_end <= expiry_cutoff,
        )
        .count()
    )

    ctx = {
        "request": request,
        "active_nav": "dashboard",
        "owners_count": owners_count,
        "units_count": units_count,
        "active_votings": total_votings,
        "voting_by_status": voting_by_status,
        "active_tax_count": sum(c for _, c in tax_status_counts),
        "tax_by_status": tax_by_status,
        "recent_activity": unified,
        "modul": modul,
        "module_counts": dict(module_counts),
        "module_counts_ordered": module_counts_ordered,
        "declared_shares": declared_shares,
        "owners_scd": owners_scd,
        "units_scd": units_scd,
        "q": q,
        "sort": sort,
        "order": order,
        "statement_count": statement_count,
        "matched_payments": matched_payments,
        "unmatched_payments": unmatched_payments,
        "total_income": total_income,
        "debtor_count": debtor_count,
        "tenants_count": tenants_count,
        "tenants_with_contract": tenants_with_contract,
        "tenants_linked": tenants_linked,
        "space_total": space_total,
        "space_rented": space_rented,
        "space_vacant": space_vacant,
        "space_blocked": space_blocked,
        "expiring_contracts": expiring_contracts,
    }

    if request.headers.get("HX-Request") and not request.headers.get("HX-Boosted"):
        return templates.TemplateResponse(request, "partials/dashboard_activity_body.html", ctx)

    return templates.TemplateResponse(request, "dashboard.html", ctx)


# Module labels for export — klíče jsou kanonické (po _norm_module)
_MODULE_LABELS = {
    "dane": "Rozesílání", "hlasovani": "Hlasování", "prostory": "Prostory",
    "najemci": "Nájemci", "vlastnici": "Vlastníci", "jednotky": "Jednotky",
    "sprava": "Administrace", "nastaveni": "Nastavení",
    "sync": "Kontroly", "platby": "Platby", "import": "Import",
    "backup": "Zálohy", "share_check": "Kontrola podílů",
    "payment_notice": "Platební upozornění",
}

_STATUS_LABELS = {
    "sent": "Odesláno", "failed": "Chyba", "pending": "Čeká",
    "created": "Vytvořeno", "updated": "Aktualizováno", "deleted": "Smazáno",
    "imported": "Importováno", "exported": "Exportováno", "restored": "Obnoveno",
    "state_change": "Změna stavu", "status_changed": "Změna stavu",
    "confirmed": "Potvrzeno", "purged": "Smazáno",
}


@router.get("/exportovat/{fmt}")
async def dashboard_export(
    fmt: str,
    q: str = Query(""),
    sort: str = Query("date"),
    order: str = Query("desc"),
    db: Session = Depends(get_db),
):
    """Export přehledu aktivity do Excelu nebo CSV."""
    if fmt not in ("xlsx", "csv"):
        return RedirectResponse("/", status_code=302)

    # Build unified activity (same as dashboard view)
    recent_emails = db.query(EmailLog).order_by(EmailLog.created_at.desc()).all()
    recent_acts = db.query(ActivityLog).order_by(ActivityLog.created_at.desc()).all()

    unified = []
    for e in recent_emails:
        unified.append({
            "created_at": e.created_at,
            "module": _norm_module(e.module),
            "description": e.subject or "",
            "detail": e.recipient_name or e.recipient_email or "",
            "status": e.status.value if e.status else "",
        })
    for a in recent_acts:
        unified.append({
            "created_at": a.created_at,
            "module": _norm_module(a.module),
            "description": a.entity_name or "",
            "detail": a.description or "",
            "status": a.action.value if a.action else "",
        })

    unified.sort(key=lambda x: x["created_at"] or datetime.min, reverse=True)

    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        unified = [
            item for item in unified
            if q_lower in item["description"].lower()
            or q_ascii in strip_diacritics(item["description"])
            or q_lower in item["detail"].lower()
            or q_ascii in strip_diacritics(item["detail"])
            or q_lower in item["module"].lower()
        ]

    SORT_KEYS = {
        "date": lambda x: x["created_at"],
        "module": lambda x: x["module"].lower(),
        "description": lambda x: strip_diacritics(x["description"]),
        "detail": lambda x: strip_diacritics(x["detail"]),
        "status": lambda x: x["status"],
    }
    sort_fn = SORT_KEYS.get(sort, SORT_KEYS["date"])
    unified.sort(key=sort_fn, reverse=(order == "desc"))

    headers = ["Datum", "Modul", "Popis", "Detail", "Stav"]
    rows = []
    for item in unified:
        rows.append([
            item["created_at"].strftime("%d.%m.%Y %H:%M") if item["created_at"] else "",
            _MODULE_LABELS.get(item["module"], item["module"]),
            item["description"],
            item["detail"],
            _STATUS_LABELS.get(item["status"], item["status"]),
        ])

    timestamp = datetime.now().strftime("%Y%m%d")
    suffix = "_hledani" if q else "_vse"
    filename = f"aktivita{suffix}_{timestamp}.{fmt}"

    if fmt == "csv":
        buf = io.StringIO()
        buf.write("\ufeff")
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(headers)
        writer.writerows(rows)
        return Response(
            content=buf.getvalue().encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Aktivita"
    bold = Font(bold=True)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h).font = bold
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    excel_auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
