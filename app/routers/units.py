import csv
import io
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, Form, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from markupsafe import escape
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import cast, func, String
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import ActivityAction, Owner, OwnerUnit, SvjInfo, Unit, UnitBalance, Payment, PaymentAllocation, PaymentDirection, PaymentMatchStatus, Prescription, PrescriptionYear, log_activity
from app.routers.payments._helpers import compute_debt_map
from app.services.code_list_service import get_all_code_lists
from app.services.owner_exchange import recalculate_unit_votes
from app.utils import build_list_url, excel_auto_width, is_htmx_partial, strip_diacritics, templates, utcnow

router = APIRouter()


def _parse_numeric_fields(floor_area: str, podil_scd: str):
    """Parse floor_area and podil_scd strings, return (float|None, float|None, list[str]).

    Pokud uživatel zadal neprázdnou hodnotu a nelze ji převést na číslo,
    vrátí chybu (ne warning). Prázdný vstup = legitimní NULL.
    """
    errors = []
    floor_area_val = floor_area.strip() if floor_area else ""
    podil_scd_val = podil_scd.strip() if podil_scd else ""

    # Plocha
    if not floor_area_val:
        floor_area_float = None
    else:
        try:
            floor_area_float = float(floor_area_val)
        except (ValueError, TypeError):
            floor_area_float = None
            errors.append(f"Plocha '{escape(floor_area_val)}' není platné číslo.")
        if floor_area_float is not None and (floor_area_float < 0 or floor_area_float > 9999):
            errors.append(f"Plocha {floor_area_float} mimo rozsah 0–9999.")
            floor_area_float = None

    # Podíl SČD
    if not podil_scd_val:
        podil_scd_float = None
    else:
        try:
            podil_scd_float = float(podil_scd_val)
        except (ValueError, TypeError):
            podil_scd_float = None
            errors.append(f"Podíl SČD '{escape(podil_scd_val)}' není platné číslo.")
        if podil_scd_float is not None and (podil_scd_float < 0 or podil_scd_float > 99999999):
            errors.append(f"Podíl SČD {podil_scd_float} mimo rozsah 0–99999999.")
            podil_scd_float = None

    return floor_area_float, podil_scd_float, errors


def _build_warn_html(warnings: list) -> str:
    """Build warning HTML from list of warning strings."""
    if not warnings:
        return ""
    warn_items = "".join(f"<li>{w}</li>" for w in warnings)
    return (
        '<div class="bg-yellow-50 border border-yellow-200 rounded-lg p-3 mb-2">'
        '<p class="text-sm font-medium text-yellow-800 mb-1">Upozornění</p>'
        f'<ul class="text-sm text-yellow-700 list-disc list-inside">{warn_items}</ul></div>'
    )


SORT_COLUMNS = {
    "unit_number": Unit.unit_number,
    "building": Unit.building_number,
    "space_type": Unit.space_type,
    "section": Unit.section,
    "address": Unit.address,
    "lv": Unit.lv_number,
    "room_count": Unit.room_count,
    "floor_area": Unit.floor_area,
    "podil": Unit.podil_scd,
}


@router.get("/nova-formular")
async def unit_create_form(request: Request, db: Session = Depends(get_db)):
    """Formulář pro vytvoření nové jednotky."""
    return templates.TemplateResponse(request, "partials/unit_create_form.html", {
        "code_lists": get_all_code_lists(db),
    })


@router.post("/nova")
async def unit_create(
    request: Request,
    unit_number: str = Form(...),
    building_number: str = Form(""),
    space_type: str = Form(""),
    section: str = Form(""),
    address: str = Form(""),
    lv_number: str = Form(""),
    room_count: str = Form(""),
    floor_area: str = Form(""),
    podil_scd: str = Form(""),
    db: Session = Depends(get_db),
):
    """Vytvoření nové jednotky z formuláře."""

    # Parse unit_number
    try:
        unit_number_int = int(unit_number)
    except (ValueError, TypeError):
        return templates.TemplateResponse(request, "partials/unit_create_form.html", {
            "error": "Číslo jednotky musí být celé číslo.",
            "code_lists": get_all_code_lists(db),
        })
    if unit_number_int < 1 or unit_number_int > 99999:
        return templates.TemplateResponse(request, "partials/unit_create_form.html", {
            "error": "Číslo jednotky musí být v rozsahu 1–99999.",
            "code_lists": get_all_code_lists(db),
        })

    # Check uniqueness
    existing = db.query(Unit).filter(Unit.unit_number == unit_number_int).first()
    if existing:
        return templates.TemplateResponse(request, "partials/unit_create_form.html", {
            "error": f"Jednotka s číslem {unit_number_int} již existuje.",
            "code_lists": get_all_code_lists(db),
        })

    # Validate building_number range
    bn_clean = building_number.strip()
    if bn_clean:
        try:
            bn_int = int(bn_clean)
            if bn_int < 1 or bn_int > 99999:
                return templates.TemplateResponse(request, "partials/unit_create_form.html", {
                    "error": f"Číslo budovy {bn_int} mimo rozsah 1–99999.",
                    "code_lists": get_all_code_lists(db),
                })
        except (ValueError, TypeError):
            pass  # alphanumeric building numbers are OK

    try:
        lv_number_int = int(lv_number.strip()) if lv_number.strip() else None
    except (ValueError, TypeError):
        lv_number_int = None
    if lv_number_int is not None and (lv_number_int < 1 or lv_number_int > 99999):
        lv_number_int = None
    floor_area_float, podil_scd_float, parse_errors = _parse_numeric_fields(floor_area, podil_scd)

    # Validace: pokud uživatel zadal neplatné číselné hodnoty, vrátit formulář s chybou
    if parse_errors:
        return templates.TemplateResponse(request, "partials/unit_create_form.html", {
            "error": " ".join(parse_errors),
            "code_lists": get_all_code_lists(db),
        })

    unit = Unit(
        unit_number=unit_number_int,
        building_number=building_number.strip() or None,
        space_type=space_type.strip() or None,
        section=section.strip() or None,
        address=address.strip() or None,
        lv_number=lv_number_int,
        room_count=room_count.strip() or None,
        floor_area=floor_area_float,
        podil_scd=podil_scd_float,
        created_at=utcnow(),
    )
    db.add(unit)
    db.flush()
    log_activity(db, ActivityAction.CREATED, "unit", "jednotky",
                 entity_id=unit.id, entity_name=f"Jednotka {unit_number_int}")
    db.commit()

    if request.headers.get("HX-Request"):
        return HTMLResponse(
            content=f'<p class="text-sm text-green-600 p-4">Jednotka {unit_number_int} vytvořena. <a href="/jednotky/{unit.id}" class="text-blue-600 hover:underline">Zobrazit</a></p>',
        )
    return RedirectResponse(f"/jednotky/{unit.id}", status_code=302)


@router.get("/{unit_id}/vlastnici-sekce")
async def unit_owners_section(
    unit_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Sekce vlastníků jednotky pro HTMX obnovení."""
    unit = db.query(Unit).options(
        joinedload(Unit.owners).joinedload(OwnerUnit.owner)
    ).get(unit_id)
    if not unit:
        return HTMLResponse("<p class='text-sm text-red-600'>Jednotka nenalezena.</p>")
    return templates.TemplateResponse(request, "partials/unit_owners.html", {
        "unit": unit,
    })


@router.get("/{unit_id}/vlastnik/{ou_id}/upravit-formular")
async def owner_unit_edit_form(
    unit_id: int,
    ou_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Inline editační řádek vazby vlastník–jednotka."""
    unit = db.query(Unit).get(unit_id)
    ou = db.query(OwnerUnit).options(joinedload(OwnerUnit.owner)).get(ou_id)
    if not unit or not ou or ou.unit_id != unit_id:
        return HTMLResponse("<tr><td colspan='5' class='text-sm text-red-600 px-3 py-2'>Záznam nenalezen.</td></tr>")
    return templates.TemplateResponse(request, "partials/unit_owner_edit_row.html", {
        "unit": unit,
        "ou": ou,
    })


@router.post("/{unit_id}/vlastnik/{ou_id}/upravit")
async def owner_unit_update(
    unit_id: int,
    ou_id: int,
    request: Request,
    share: str = Form(...),
    ownership_type: str = Form(""),
    db: Session = Depends(get_db),
):
    """Uložení změn podílu a typu vlastnictví na jednotce."""
    unit = db.query(Unit).options(
        joinedload(Unit.owners).joinedload(OwnerUnit.owner)
    ).get(unit_id)
    ou = db.query(OwnerUnit).get(ou_id)
    if not unit or not ou or ou.unit_id != unit_id:
        return HTMLResponse("<p class='text-sm text-red-600'>Záznam nenalezen.</p>")

    # Parse share — accept both float (0.5) and fraction (1/2)
    try:
        if "/" in share:
            parts = share.split("/")
            share_val = float(parts[0].strip()) / float(parts[1].strip())
        else:
            share_val = float(share)
    except (ValueError, ZeroDivisionError):
        share_val = ou.share  # keep original on parse error
    if share_val < 0 or share_val > 1:
        share_val = ou.share  # keep original on out-of-range

    ou.share = share_val
    ou.ownership_type = ownership_type.strip() or None

    # Recalculate votes for all owners of the unit
    recalculate_unit_votes(unit, db)

    log_activity(db, ActivityAction.UPDATED, "unit", "jednotky",
                 entity_id=unit.id, entity_name=f"Jednotka {unit.unit_number}",
                 description=f"Upraven podíl vlastníka {ou.owner.display_name if ou.owner else ''}")
    db.commit()

    # Refresh relationships
    db.refresh(unit)

    return templates.TemplateResponse(request, "partials/unit_owners.html", {
        "unit": unit,
    })


@router.get("/{unit_id}/upravit-formular")
async def unit_edit_form(
    unit_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Formulář pro inline editaci údajů jednotky."""
    unit = db.query(Unit).get(unit_id)
    if not unit:
        return RedirectResponse("/jednotky", status_code=302)
    return templates.TemplateResponse(request, "partials/unit_edit_form.html", {
        "unit": unit,
        "code_lists": get_all_code_lists(db),
    })


@router.get("/{unit_id}/info")
async def unit_info(
    unit_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Zobrazení údajů jednotky po zrušení editace."""
    unit = db.query(Unit).get(unit_id)
    if not unit:
        return RedirectResponse("/jednotky", status_code=302)
    return templates.TemplateResponse(request, "partials/unit_info.html", {
        "unit": unit,
    })


@router.post("/{unit_id}/upravit")
async def unit_update(
    unit_id: int,
    request: Request,
    unit_number: str = Form(...),
    building_number: str = Form(""),
    space_type: str = Form(""),
    section: str = Form(""),
    orientation_number: str = Form(""),
    address: str = Form(""),
    lv_number: str = Form(""),
    room_count: str = Form(""),
    floor_area: str = Form(""),
    podil_scd: str = Form(""),
    db: Session = Depends(get_db),
):
    """Uložení změn údajů jednotky."""

    unit = db.query(Unit).get(unit_id)
    if not unit:
        return RedirectResponse("/jednotky", status_code=302)

    # Parse and validate unit_number
    try:
        unit_number_int = int(unit_number)
    except (ValueError, TypeError):
        return templates.TemplateResponse(request, "partials/unit_edit_form.html", {
            "unit": unit,
            "error": "Číslo jednotky musí být celé číslo.",
            "code_lists": get_all_code_lists(db),
        })
    if unit_number_int < 1 or unit_number_int > 99999:
        return templates.TemplateResponse(request, "partials/unit_edit_form.html", {
            "unit": unit,
            "error": "Číslo jednotky musí být v rozsahu 1–99999.",
            "code_lists": get_all_code_lists(db),
        })

    # Check uniqueness (exclude self)
    existing = db.query(Unit).filter(
        Unit.unit_number == unit_number_int, Unit.id != unit_id
    ).first()
    if existing:
        return templates.TemplateResponse(request, "partials/unit_edit_form.html", {
            "unit": unit,
            "error": f"Jednotka s číslem {unit_number_int} již existuje.",
            "code_lists": get_all_code_lists(db),
        })

    try:
        orientation_number_int = int(orientation_number.strip()) if orientation_number.strip() else None
    except (ValueError, TypeError):
        orientation_number_int = None
    if orientation_number_int is not None and (orientation_number_int < 1 or orientation_number_int > 9999):
        orientation_number_int = None
    try:
        lv_number_int = int(lv_number.strip()) if lv_number.strip() else None
    except (ValueError, TypeError):
        lv_number_int = None
    if lv_number_int is not None and (lv_number_int < 1 or lv_number_int > 99999):
        lv_number_int = None
    floor_area_float, podil_scd_float, parse_errors = _parse_numeric_fields(floor_area, podil_scd)

    # Validace: pokud uživatel zadal neplatné číselné hodnoty, vrátit formulář s chybou
    if parse_errors:
        return templates.TemplateResponse(request, "partials/unit_edit_form.html", {
            "unit": unit,
            "error": " ".join(parse_errors),
            "code_lists": get_all_code_lists(db),
        })

    unit.unit_number = unit_number_int
    unit.building_number = building_number.strip() or None
    unit.space_type = space_type.strip() or None
    unit.section = section.strip() or None
    unit.orientation_number = orientation_number_int
    unit.address = address.strip() or None
    unit.lv_number = lv_number_int
    unit.room_count = room_count.strip() or None
    unit.floor_area = floor_area_float
    unit.podil_scd = podil_scd_float

    recalculate_unit_votes(unit, db)

    log_activity(db, ActivityAction.UPDATED, "unit", "jednotky",
                 entity_id=unit.id, entity_name=f"Jednotka {unit.unit_number}")
    db.commit()

    if request.headers.get("HX-Request"):
        # Refresh unit + owners (recalculate changed owner votes)
        db.expire(unit, ["owners"])
        info_html = templates.TemplateResponse(request, "partials/unit_info.html", {
            "unit": unit,
            "saved": True,
        }).body.decode()
        owners_html = templates.TemplateResponse(request, "partials/unit_owners.html", {
            "unit": unit,
        }).body.decode()
        # OOB swap: main target gets unit_info, owners section updates out-of-band
        # unit_owners.html already has <div id="unit-owners">, add hx-swap-oob attribute
        owners_oob = owners_html.replace('<div id="unit-owners">', '<div id="unit-owners" hx-swap-oob="true">', 1)
        return HTMLResponse(content=info_html + owners_oob)
    return RedirectResponse(f"/jednotky/{unit_id}", status_code=302)


def _filter_units(db: Session, q="", typ="", sekce="", sort="unit_number", order="asc"):
    """Filter and sort units. Returns list[Unit] with eager-loaded owners."""
    query = db.query(Unit).options(
        joinedload(Unit.owners).joinedload(OwnerUnit.owner)
    )

    if q:
        search = f"%{q}%"
        search_ascii = f"%{strip_diacritics(q)}%"
        query = query.filter(
            cast(Unit.unit_number, String).ilike(search)
            | Unit.building_number.ilike(search)
            | Unit.space_type.ilike(search)
            | Unit.section.ilike(search)
            | Unit.address.ilike(search)
            | Unit.owners.any(OwnerUnit.owner.has(Owner.name_normalized.like(search_ascii)))
        )
    if typ:
        query = query.filter(Unit.space_type == typ)
    if sekce:
        query = query.filter(Unit.section == sekce)

    # Sorting
    if sort == "owners":
        # SQL subquery: min owner name_normalized per unit
        owner_sub = (
            db.query(OwnerUnit.unit_id, func.min(Owner.name_normalized).label("min_owner"))
            .join(Owner, OwnerUnit.owner_id == Owner.id)
            .filter(OwnerUnit.valid_to.is_(None))
            .group_by(OwnerUnit.unit_id)
            .subquery()
        )
        query = query.outerjoin(owner_sub, Unit.id == owner_sub.c.unit_id)
        col = owner_sub.c.min_owner
        query = query.order_by(col.desc().nulls_last() if order == "desc" else col.asc().nulls_last())
        units = query.all()
    else:
        sort_col = SORT_COLUMNS.get(sort, Unit.unit_number)
        if order == "desc":
            query = query.order_by(sort_col.desc().nulls_last())
        else:
            query = query.order_by(sort_col.asc().nulls_last())
        units = query.all()

    return units


@router.get("/")
async def unit_list(
    request: Request,
    q: str = Query("", alias="q"),
    typ: str = Query("", alias="typ"),
    sekce: str = Query("", alias="sekce"),
    sort: str = Query("unit_number", alias="sort"),
    order: str = Query("asc", alias="order"),
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    """Seznam jednotek s filtry, hledáním a řazením."""
    units = _filter_units(db, q, typ, sekce, sort, order)

    # Debt map — platební dluh (předpis × měsíce + zůstatky - zaplaceno)
    debt_map = {}
    latest_py = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
    if latest_py:
        debt_map = compute_debt_map(db, latest_py.year)

    # Python-side sort by debt
    if sort == "dluh":
        units.sort(key=lambda u: debt_map.get(u.id, 0), reverse=(order == "desc"))

    # Current list URL for back navigation
    list_url = build_list_url(request)

    # HTMX partial
    if is_htmx_partial(request):
        return templates.TemplateResponse(request, "partials/unit_table_body.html", {
            "units": units,
            "list_url": list_url,
            "debt_map": debt_map,
        })

    # Stats
    total_units = db.query(Unit).count()
    total_scd = db.query(func.sum(Unit.podil_scd)).scalar() or 0
    svj_info = db.query(SvjInfo).first()
    declared_shares = svj_info.total_shares if svj_info and svj_info.total_shares else 0

    type_counts_raw = (
        db.query(Unit.space_type, func.count(Unit.id))
        .filter(Unit.space_type.isnot(None))
        .group_by(Unit.space_type)
        .all()
    )
    type_counts = {st: cnt for st, cnt in type_counts_raw}

    sections = [
        r[0] for r in
        db.query(Unit.section).filter(Unit.section.isnot(None)).distinct().order_by(Unit.section).all()
    ]

    return templates.TemplateResponse(request, "units/list.html", {
        "active_nav": "units",
        "units": units,
        "list_url": list_url,
        "back_url": back,
        "q": q,
        "typ": typ,
        "sekce": sekce,
        "sort": sort,
        "order": order,
        "debt_map": debt_map,
        "stats": {
            "total_units": total_units,
            "total_scd": total_scd,
            "declared_shares": declared_shares,
            "type_counts": type_counts,
            "sections": sections,
        },
    })


@router.get("/exportovat/{fmt}")
async def unit_export(
    fmt: str,
    q: str = Query("", alias="q"),
    typ: str = Query("", alias="typ"),
    sekce: str = Query("", alias="sekce"),
    sort: str = Query("unit_number", alias="sort"),
    order: str = Query("asc", alias="order"),
    db: Session = Depends(get_db),
):
    """Export filtered units to Excel or CSV."""
    if fmt not in ("xlsx", "csv"):
        return RedirectResponse("/jednotky", status_code=302)

    units = _filter_units(db, q, typ, sekce, sort, order)

    headers = ["Č. jednotky", "Budova", "Typ prostoru", "Sekce", "Adresa", "LV", "Místnosti", "Plocha", "Podíl SČD", "Vlastníci"]

    def _row(u):
        owners = ", ".join(ou.owner.display_name for ou in u.current_owners)
        return [
            u.unit_number,
            u.building_number or "",
            u.space_type or "",
            u.section or "",
            u.address or "",
            u.lv_number or "",
            u.room_count or "",
            u.floor_area or "",
            u.podil_scd or "",
            owners,
        ]

    timestamp = datetime.now().strftime("%Y%m%d")

    # Suffix podle aktivního filtru
    typ_labels = {"byt": "byt", "garáž": "garaz", "jiný nebytový prostor": "nebytovy"}
    if typ and typ in typ_labels:
        suffix = f"_{typ_labels[typ]}"
    elif sekce:
        suffix = f"_sekce_{strip_diacritics(sekce)}"
    elif q:
        suffix = "_hledani"
    else:
        suffix = "_vsechny"
    filename = f"jednotky{suffix}_{timestamp}"

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Jednotky"
        bold = Font(bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold

        for row_idx, u in enumerate(units, 2):
            for col_idx, val in enumerate(_row(u), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        excel_auto_width(ws)

        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    else:
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(headers)
        for u in units:
            writer.writerow(_row(u))
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )


@router.get("/{unit_id}")
async def unit_detail(
    unit_id: int,
    request: Request,
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    """Detail jednotky s vlastníky a editací."""
    unit = db.query(Unit).options(
        joinedload(Unit.owners).joinedload(OwnerUnit.owner)
    ).get(unit_id)
    if not unit:
        return RedirectResponse("/jednotky", status_code=302)

    if "/sprava/hromadne" in back:
        back_label = "Zpět na hromadné úpravy"
    elif "/vlastnici/" in back:
        back_label = "Zpět na detail vlastníka"
    elif back.startswith("/vlastnici"):
        back_label = "Zpět na seznam vlastníků"
    elif "/synchronizace/" in back:
        back_label = "Zpět na porovnání"
    elif "/platby/predpisy/" in back:
        back_label = "Zpět na předpisy"
    elif "/platby/vypisy/" in back:
        back_label = "Zpět na výpis"
    elif "/platby/symboly" in back:
        back_label = "Zpět na symboly"
    elif "/platby/zustatky" in back:
        back_label = "Zpět na zůstatky"
    elif "/platby/prehled" in back:
        back_label = "Zpět na matici plateb"
    elif "/platby/dluznici" in back:
        back_label = "Zpět na dlužníky"
    elif "/platby/jednotka" in back:
        back_label = "Zpět na platby jednotky"
    elif "/platby" in back:
        back_label = "Zpět na platby"
    elif back:
        back_label = "Zpět na seznam jednotek"
    else:
        back_label = "Zpět na seznam jednotek"

    # Počáteční zůstatek (dluh z importu)
    opening_balance = 0
    balance_year = None
    bal_py = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
    if bal_py:
        balance_year = bal_py.year
        ob = db.query(func.sum(UnitBalance.opening_amount)).filter(
            UnitBalance.unit_id == unit.id,
            UnitBalance.year == bal_py.year,
        ).scalar()
        opening_balance = ob or 0

    # Platební stav — aktuální rok
    latest_py = bal_py
    payment_status = None
    payment_debt = 0
    payment_year = None
    if latest_py:
        payment_year = latest_py.year
        presc = db.query(Prescription).filter_by(
            prescription_year_id=latest_py.id, unit_id=unit.id
        ).first()
        if presc:
            monthly = presc.monthly_total or 0
            confirmed_statuses = [PaymentMatchStatus.AUTO_MATCHED, PaymentMatchStatus.MANUAL]
            total_paid = db.query(
                func.coalesce(func.sum(PaymentAllocation.amount), 0)
            ).join(Payment).filter(
                PaymentAllocation.unit_id == unit.id,
                Payment.direction == PaymentDirection.INCOME,
                Payment.match_status.in_(confirmed_statuses),
                func.extract("year", Payment.date) == latest_py.year,
            ).scalar() or 0
            # Počet měsíců s daty
            months_count = db.query(
                func.distinct(func.extract("month", Payment.date))
            ).filter(
                Payment.direction == PaymentDirection.INCOME,
                Payment.match_status.in_(confirmed_statuses),
                func.extract("year", Payment.date) == latest_py.year,
            ).count()
            expected = monthly * months_count
            if total_paid >= expected and expected > 0:
                payment_status = "ok"
            elif total_paid > 0:
                payment_status = "partial"
                payment_debt = expected - total_paid
            elif months_count > 0:
                payment_status = "unpaid"
                payment_debt = expected

    return templates.TemplateResponse(request, "units/detail.html", {
        "active_nav": "units",
        "unit": unit,
        "back_url": back or "/jednotky",
        "back_label": back_label,
        "payment_status": payment_status,
        "payment_debt": payment_debt,
        "payment_year": payment_year,
        "opening_balance": opening_balance,
        "balance_year": balance_year,
    })
