import logging
import shutil
import threading
import time as _time
from datetime import date, datetime

logger = logging.getLogger(__name__)
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import cast, func, String
from sqlalchemy.orm import Session, joinedload

from app.config import settings
from app.database import SessionLocal, get_db
from app.models import ImportLog, Owner, OwnerType, OwnerUnit, SvjInfo, Unit, ActivityAction, log_activity
from app.services.excel_import import import_owners_from_excel, preview_owners_from_excel
from app.utils import build_list_url, excel_auto_width, is_htmx_partial, is_safe_path, is_valid_email, setup_jinja_filters, strip_diacritics, validate_upload

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")
setup_jinja_filters(templates)

# In-memory progress tracker for contact import background processing
_contact_import_progress: dict[str, dict] = {}


SORT_COLUMNS = {
    "name": Owner.name_normalized,
    "type": Owner.owner_type,
    "email": Owner.email,
    "phone": Owner.phone,
    "podil": None,  # handled in Python — needs sum across units
    "jednotky": None,  # handled in Python
    "sekce": None,  # handled in Python
}


@router.get("/novy-formular")
async def owner_create_form(request: Request):
    return templates.TemplateResponse("partials/owner_create_form.html", {
        "request": request,
    })


@router.post("/novy")
async def owner_create(
    request: Request,
    first_name: str = Form(...),
    last_name: str = Form(...),
    title: str = Form(""),
    owner_type: str = Form("physical"),
    email: str = Form(""),
    phone: str = Form(""),
    birth_number: str = Form(""),
    db: Session = Depends(get_db),
):
    # Build name_with_titles and name_normalized
    parts_wt = []
    if title:
        parts_wt.append(title)
    if last_name:
        parts_wt.append(last_name)
    if first_name:
        parts_wt.append(first_name)
    name_with_titles = " ".join(parts_wt)

    parts_norm = []
    if last_name:
        parts_norm.append(last_name)
    if first_name:
        parts_norm.append(first_name)
    name_normalized = strip_diacritics(" ".join(parts_norm))

    owner = Owner(
        first_name=first_name.strip(),
        last_name=last_name.strip() or None,
        title=title.strip() or None,
        owner_type=OwnerType(owner_type),
        name_with_titles=name_with_titles,
        name_normalized=name_normalized,
        email=(email.strip() if email.strip() and is_valid_email(email.strip()) else None),
        phone=phone.strip() or None,
        birth_number=birth_number.strip() or None,
        data_source="manual",
        is_active=True,
        created_at=datetime.utcnow(),
    )
    db.add(owner)
    db.commit()

    return RedirectResponse(f"/vlastnici/{owner.id}", status_code=302)


def _filter_owners(db: Session, q="", owner_type="", vlastnictvi="", kontakt="", stav="", sekce="", sort="name", order="asc"):
    """Filter and sort owners. Returns list[Owner] with eager-loaded units."""
    query = db.query(Owner).filter_by(is_active=True).options(
        joinedload(Owner.units).joinedload(OwnerUnit.unit)
    )
    if q:
        search = f"%{q}%"
        search_ascii = f"%{strip_diacritics(q)}%"
        query = query.filter(
            Owner.name_normalized.like(search_ascii)
            | Owner.name_with_titles.ilike(search)
            | Owner.first_name.ilike(search)
            | Owner.last_name.ilike(search)
            | Owner.email.ilike(search)
            | Owner.phone.ilike(search)
            | Owner.birth_number.ilike(search)
            | Owner.company_id.ilike(search)
            | Owner.units.any((OwnerUnit.valid_to.is_(None)) & OwnerUnit.unit.has(cast(Unit.unit_number, String).ilike(search)))
        )
    if owner_type:
        query = query.filter(Owner.owner_type == owner_type)
    if vlastnictvi == "_empty":
        query = query.filter(
            Owner.units.any(
                (OwnerUnit.valid_to.is_(None)) & ((OwnerUnit.ownership_type.is_(None)) | (OwnerUnit.ownership_type == ""))
            )
        )
    elif vlastnictvi:
        query = query.filter(
            Owner.units.any((OwnerUnit.valid_to.is_(None)) & (OwnerUnit.ownership_type == vlastnictvi))
        )
    if kontakt == "s_emailem":
        query = query.filter(Owner.email.isnot(None), Owner.email != "")
    elif kontakt == "bez_emailu":
        query = query.filter((Owner.email.is_(None)) | (Owner.email == ""))
    elif kontakt == "s_telefonem":
        query = query.filter(Owner.phone.isnot(None), Owner.phone != "")
    elif kontakt == "bez_telefonu":
        query = query.filter((Owner.phone.is_(None)) | (Owner.phone == ""))
    if stav == "bez_jednotky":
        query = query.filter(~Owner.units.any(OwnerUnit.valid_to.is_(None)))
    if sekce:
        query = query.filter(
            Owner.units.any((OwnerUnit.valid_to.is_(None)) & OwnerUnit.unit.has(Unit.section == sekce))
        )

    # Sorting
    sort_col = SORT_COLUMNS.get(sort)
    if sort == "podil":
        owners = query.all()
        owners.sort(
            key=lambda o: sum(ou.votes for ou in o.current_units),
            reverse=(order == "desc"),
        )
    elif sort == "jednotky":
        owners = query.all()
        owners.sort(
            key=lambda o: (o.current_units[0].unit.unit_number if o.current_units else 0),
            reverse=(order == "desc"),
        )
    elif sort == "sekce":
        owners = query.all()
        owners.sort(
            key=lambda o: (o.current_units[0].unit.section or "") if o.current_units else "",
            reverse=(order == "desc"),
        )
    elif sort_col is not None:
        if order == "desc":
            query = query.order_by(sort_col.desc().nulls_last())
        else:
            query = query.order_by(sort_col.asc().nulls_last())
        owners = query.all()
    else:
        owners = query.order_by(Owner.name_normalized).all()

    return owners


@router.get("/")
async def owner_list(
    request: Request,
    q: str = Query("", alias="q"),
    owner_type: str = Query("", alias="typ"),
    vlastnictvi: str = Query("", alias="vlastnictvi"),
    kontakt: str = Query("", alias="kontakt"),
    stav: str = Query("", alias="stav"),
    sekce: str = Query("", alias="sekce"),
    sort: str = Query("name", alias="sort"),
    order: str = Query("asc", alias="order"),
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    owners = _filter_owners(db, q, owner_type, vlastnictvi, kontakt, stav, sekce, sort, order)

    # Current list URL for back navigation
    list_url = build_list_url(request)

    # Return partial only for targeted HTMX requests (search/filter), not boosted navigation
    if is_htmx_partial(request):
        return templates.TemplateResponse("partials/owner_table_body.html", {
            "request": request,
            "owners": owners,
            "list_url": list_url,
        })

    # Stats for header
    all_owners = db.query(Owner).filter_by(is_active=True).count()
    total_units = db.query(Unit).count()
    type_counts_raw = (
        db.query(Owner.owner_type, func.count(Owner.id))
        .filter_by(is_active=True)
        .group_by(Owner.owner_type)
        .all()
    )
    type_counts = {ot.value: cnt for ot, cnt in type_counts_raw}
    sections = [
        r[0] for r in
        db.query(Unit.section).filter(Unit.section.isnot(None)).distinct().order_by(Unit.section).all()
    ]
    emails_count = db.query(Owner).filter(
        Owner.is_active == True,
        Owner.email.isnot(None),
        Owner.email != "",
    ).count()
    phones_count = db.query(Owner).filter(
        Owner.is_active == True,
        Owner.phone.isnot(None),
        Owner.phone != "",
    ).count()

    no_units_count = db.query(Owner).filter(
        Owner.is_active == True,
        ~Owner.units.any(OwnerUnit.valid_to.is_(None)),
    ).count()

    total_scd = db.query(func.sum(OwnerUnit.votes)).filter(OwnerUnit.valid_to.is_(None)).scalar() or 0
    svj_info = db.query(SvjInfo).first()
    declared_shares = svj_info.total_shares if svj_info and svj_info.total_shares else 0

    # Ownership type counts
    ownership_counts_raw = (
        db.query(
            func.coalesce(OwnerUnit.ownership_type, ""),
            func.count(func.distinct(OwnerUnit.owner_id)),
        )
        .filter(OwnerUnit.valid_to.is_(None))
        .group_by(func.coalesce(OwnerUnit.ownership_type, ""))
        .all()
    )
    ownership_counts = {ot or "": cnt for ot, cnt in ownership_counts_raw}

    return templates.TemplateResponse("owners/list.html", {
        "request": request,
        "active_nav": "owners",
        "owners": owners,
        "list_url": list_url,
        "back_url": back,
        "q": q,
        "owner_type": owner_type,
        "vlastnictvi": vlastnictvi,
        "kontakt": kontakt,
        "stav": stav,
        "sekce": sekce,
        "sort": sort,
        "order": order,
        "owner_types": OwnerType,
        "stats": {
            "total_owners": all_owners,
            "total_units": total_units,
            "type_counts": type_counts,
            "sections": sections,
            "emails_count": emails_count,
            "phones_count": phones_count,
            "no_units_count": no_units_count,
            "total_scd": total_scd,
            "declared_shares": declared_shares,
            "ownership_counts": ownership_counts,
        },
    })


def _format_address(owner, prefix):
    """Format address fields into a single string."""
    parts = []
    street = getattr(owner, f"{prefix}_street")
    district = getattr(owner, f"{prefix}_district")
    city = getattr(owner, f"{prefix}_city")
    zip_code = getattr(owner, f"{prefix}_zip")
    country = getattr(owner, f"{prefix}_country")
    if street:
        parts.append(street)
    if district:
        parts.append(district)
    if city and zip_code:
        parts.append(f"{zip_code} {city}")
    elif city:
        parts.append(city)
    elif zip_code:
        parts.append(zip_code)
    if country:
        parts.append(country)
    return ", ".join(parts)


@router.get("/exportovat/{fmt}")
async def owner_export(
    fmt: str,
    q: str = Query("", alias="q"),
    owner_type: str = Query("", alias="typ"),
    vlastnictvi: str = Query("", alias="vlastnictvi"),
    kontakt: str = Query("", alias="kontakt"),
    stav: str = Query("", alias="stav"),
    sekce: str = Query("", alias="sekce"),
    sort: str = Query("name", alias="sort"),
    order: str = Query("asc", alias="order"),
    db: Session = Depends(get_db),
):
    """Export filtered owners to Excel or CSV."""
    if fmt not in ("xlsx", "csv"):
        return RedirectResponse("/vlastnici", status_code=302)

    owners = _filter_owners(db, q, owner_type, vlastnictvi, kontakt, stav, sekce, sort, order)

    headers = ["Vlastník", "Typ", "Jednotky", "Sekce", "Email", "Email 2", "Telefon", "Podíl SČD", "RČ/IČ", "Trvalá adresa", "Korespondenční adresa"]

    def _row(o):
        units = ", ".join(str(ou.unit.unit_number) for ou in o.current_units)
        sections = ", ".join(sorted(set(ou.unit.section for ou in o.current_units if ou.unit.section)))
        typ = "Právnická os." if o.owner_type == OwnerType.LEGAL_ENTITY else "Fyzická os."
        podil = sum(ou.votes for ou in o.current_units)
        rc_ic = o.company_id or o.birth_number or ""
        return [
            o.display_name,
            typ,
            units,
            sections,
            o.email or "",
            o.email_secondary or "",
            o.phone or "",
            podil,
            rc_ic,
            _format_address(o, "perm"),
            _format_address(o, "corr"),
        ]

    timestamp = datetime.now().strftime("%Y%m%d")
    from fastapi.responses import Response

    if fmt == "xlsx":
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font


        wb = Workbook()
        ws = wb.active
        ws.title = "Vlastníci"
        bold = Font(bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold

        for row_idx, o in enumerate(owners, 2):
            for col_idx, val in enumerate(_row(o), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        excel_auto_width(ws)

        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="vlastnici_{timestamp}.xlsx"'},
        )
    else:
        import csv
        import io
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        for o in owners:
            writer.writerow(_row(o))
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="vlastnici_{timestamp}.csv"'},
        )


@router.get("/import-kontaktu")
async def contact_import_page():
    """Redirect to unified import page (contacts section)."""
    return RedirectResponse("/vlastnici/import#kontakty", status_code=302)


@router.post("/import-kontaktu")
async def contact_import_upload(
    request: Request,
    file: UploadFile = File(...),
):
    if not file.filename:
        return RedirectResponse("/vlastnici/import?chyba_kontakty=format#kontakty", status_code=302)

    err = await validate_upload(file, max_size_mb=50, allowed_extensions=[".xlsx", ".xls"])
    if err:
        return RedirectResponse("/vlastnici/import?chyba_kontakty=format#kontakty", status_code=302)

    from datetime import datetime as _dt
    timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
    dest = settings.upload_dir / "excel" / f"{timestamp}_{file.filename}"
    dest.parent.mkdir(parents=True, exist_ok=True)
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)

    file_key = dest.name  # unique key for progress tracking

    # Initialize progress tracker
    _contact_import_progress[file_key] = {
        "done": False,
        "error": None,
        "result": None,
        "file_path": str(dest),
        "filename": file.filename,
        "started_at": _time.monotonic(),
        "total": 0,
        "current": 0,
        "phase": "Připravuji...",
    }

    # Start background processing thread
    thread = threading.Thread(
        target=_run_contact_preview,
        args=(file_key, str(dest)),
        daemon=True,
    )
    thread.start()

    from urllib.parse import quote
    return RedirectResponse(f"/vlastnici/import-kontaktu/zpracovani?soubor={quote(file_key)}", status_code=302)


def _run_contact_preview(file_key: str, file_path: str):
    """Background thread: parse Excel and compare with DB."""
    db = SessionLocal()
    try:
        from app.services.contact_import import preview_contact_import
        progress = _contact_import_progress[file_key]
        result = preview_contact_import(file_path, db, progress=progress)
        progress["result"] = result
    except Exception as e:
        logger.exception("Contact import failed for %s", file_key)
        _contact_import_progress[file_key]["error"] = str(e)
    finally:
        _contact_import_progress[file_key]["done"] = True
        db.close()


def _contact_progress_ctx(progress: dict) -> dict:
    """Compute progress context for contact import templates."""
    total = progress.get("total", 0)
    current = progress.get("current", 0)
    pct = int(current / total * 100) if total > 0 else 0
    elapsed = _time.monotonic() - progress["started_at"]

    eta_text = ""
    if current > 0 and total > 0:
        per_row = elapsed / current
        remaining = (total - current) * per_row
        if remaining >= 60:
            eta_text = f"{int(remaining // 60)} min {int(remaining % 60)} s"
        elif remaining >= 1:
            eta_text = f"{int(remaining)} s"

    elapsed_text = f"{int(elapsed // 60)} min {int(elapsed % 60)} s" if elapsed >= 60 else f"{int(elapsed)} s"

    return {
        "total": total,
        "current": current,
        "pct": pct,
        "elapsed": elapsed_text,
        "eta": eta_text,
        "phase": progress.get("phase", "Připravuji..."),
    }


@router.get("/import-kontaktu/zpracovani")
async def contact_import_processing(
    request: Request,
    soubor: str = Query("", alias="soubor"),
):
    """Progress page — HTMX polls /import-kontaktu/zpracovani-stav."""
    progress = _contact_import_progress.get(soubor)
    if not progress:
        return RedirectResponse("/vlastnici/import#kontakty", status_code=302)
    if progress.get("done"):
        from urllib.parse import quote
        return RedirectResponse(f"/vlastnici/import-kontaktu/nahled?soubor={quote(soubor)}", status_code=302)

    return templates.TemplateResponse("owners/contact_import_processing.html", {
        "request": request,
        "active_nav": "owners",
        "file_key": soubor,
        **_contact_progress_ctx(progress),
    })


@router.get("/import-kontaktu/zpracovani-stav")
async def contact_import_status(
    request: Request,
    soubor: str = Query("", alias="soubor"),
):
    """HTMX polling endpoint — returns progress partial or HX-Redirect when done."""
    progress = _contact_import_progress.get(soubor)
    if not progress:
        response = HTMLResponse("")
        response.headers["HX-Redirect"] = "/vlastnici/import#kontakty"
        return response

    if progress.get("done"):
        from urllib.parse import quote
        response = HTMLResponse("")
        response.headers["HX-Redirect"] = f"/vlastnici/import-kontaktu/nahled?soubor={quote(soubor)}"
        return response

    return templates.TemplateResponse("partials/contact_import_progress.html", {
        "request": request,
        **_contact_progress_ctx(progress),
    })


@router.get("/import-kontaktu/nahled")
async def contact_import_preview_page(
    request: Request,
    soubor: str = Query("", alias="soubor"),
):
    """Show preview from cached result after background processing."""
    data = _contact_import_progress.pop(soubor, None)
    if not data or not data.get("result"):
        # Check for error
        if data and data.get("error"):
            return RedirectResponse(f"/vlastnici/import?chyba_kontakty=zpracovani#kontakty", status_code=302)
        return RedirectResponse("/vlastnici/import#kontakty", status_code=302)

    return templates.TemplateResponse("owners/contact_import_preview.html", {
        "request": request,
        "active_nav": "owners",
        "preview": data["result"],
        "file_path": data["file_path"],
        "filename": data.get("filename", ""),
    })


@router.get("/import-kontaktu/znovu")
async def contact_import_rerun(
    request: Request,
    soubor: str = Query("", alias="soubor"),
):
    """Re-run preview for an already uploaded file."""
    from pathlib import Path
    if not soubor or not is_safe_path(Path(soubor), settings.upload_dir) or not Path(soubor).is_file():
        return RedirectResponse("/vlastnici/import#kontakty", status_code=302)

    file_key = Path(soubor).name

    _contact_import_progress[file_key] = {
        "done": False,
        "error": None,
        "result": None,
        "file_path": soubor,
        "filename": Path(soubor).name,
        "started_at": _time.monotonic(),
        "total": 0,
        "current": 0,
        "phase": "Připravuji...",
    }

    thread = threading.Thread(
        target=_run_contact_preview,
        args=(file_key, soubor),
        daemon=True,
    )
    thread.start()

    from urllib.parse import quote
    return RedirectResponse(f"/vlastnici/import-kontaktu/zpracovani?soubor={quote(file_key)}", status_code=302)


@router.post("/import-kontaktu/potvrdit")
async def contact_import_confirm(
    request: Request,
    file_path: str = Form(...),
    overwrite: str = Form(""),
    db: Session = Depends(get_db),
):
    if not is_safe_path(Path(file_path), settings.upload_dir):
        return RedirectResponse("/vlastnici/import#kontakty", status_code=302)

    form_data = await request.form()
    selected = [int(v) for v in form_data.getlist("selected_owners")]

    if not selected:
        return RedirectResponse("/vlastnici/import#kontakty", status_code=302)

    from app.services.contact_import import execute_contact_import
    result = execute_contact_import(file_path, db, selected, overwrite_existing=bool(overwrite))

    # Log the import
    log = ImportLog(
        filename=Path(file_path).name,
        file_path=file_path,
        import_type="contacts_excel",
        rows_total=result["owners_updated"] + (len(selected) - result["owners_updated"]),
        rows_imported=result["owners_updated"],
        rows_skipped=len(selected) - result["owners_updated"],
    )
    db.add(log)
    log_activity(db, ActivityAction.IMPORTED, "import", "vlastnici",
                 entity_name="Import kontaktů",
                 description=f"{result['owners_updated']} vlastníků, {result['fields_updated']} polí")
    db.commit()

    return templates.TemplateResponse("owners/contact_import_result.html", {
        "request": request,
        "active_nav": "owners",
        "result": result,
        "file_path": file_path,
    })


@router.get("/import")
async def import_page(
    request: Request,
    chyba_kontakty: str = Query("", alias="chyba_kontakty"),
    db: Session = Depends(get_db),
):
    imports = db.query(ImportLog).filter_by(import_type="owners_excel").order_by(ImportLog.created_at.desc()).all()
    contact_imports = db.query(ImportLog).filter_by(import_type="contacts_excel").order_by(ImportLog.created_at.desc()).all()

    contact_flash = None
    if chyba_kontakty == "format":
        contact_flash = "Nahrajte soubor ve formátu .xlsx"
    elif chyba_kontakty == "zpracovani":
        contact_flash = "Chyba při zpracování souboru"

    return templates.TemplateResponse("owners/import.html", {
        "request": request,
        "active_nav": "import",
        "imports": imports,
        "contact_imports": contact_imports,
        "contact_flash": contact_flash,
    })


@router.post("/import")
async def import_excel_preview(
    request: Request,
    file: UploadFile = File(...),
):
    """Step 1: Upload Excel, show preview of parsed data."""
    err = await validate_upload(file, max_size_mb=50, allowed_extensions=[".xlsx", ".xls"]) if file.filename else "Nahrajte prosím soubor ve formátu .xlsx"
    if err:
        return templates.TemplateResponse("owners/import.html", {
            "request": request,
            "active_nav": "import",
            "flash_message": err,
            "flash_type": "error",
        })

    # Save uploaded file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = settings.upload_dir / "excel" / f"{timestamp}_{file.filename}"
    dest.parent.mkdir(parents=True, exist_ok=True)
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)

    # Parse without saving to DB
    preview = preview_owners_from_excel(str(dest))

    return templates.TemplateResponse("owners/import_preview.html", {
        "request": request,
        "active_nav": "import",
        "preview": preview,
        "file_path": str(dest),
        "filename": file.filename,
    })


@router.post("/import/potvrdit")
async def import_excel_confirm(
    request: Request,
    file_path: str = Form(...),
    filename: str = Form(""),
    db: Session = Depends(get_db),
):
    """Step 2: Confirm preview and save to DB."""
    if not is_safe_path(Path(file_path), settings.upload_dir):
        return RedirectResponse("/vlastnici/import", status_code=302)

    from app.models.owner import OwnerUnit

    # Clear existing owners
    db.query(OwnerUnit).delete()
    db.query(Owner).delete()
    db.query(Unit).delete()
    db.commit()

    # Import
    result = import_owners_from_excel(db, file_path)

    # Log the import
    log = ImportLog(
        filename=filename,
        file_path=file_path,
        import_type="owners_excel",
        rows_total=result["rows_processed"],
        rows_imported=result["owners_created"],
        rows_skipped=len(result["errors"]),
        errors="\n".join(result["errors"]) if result["errors"] else None,
    )
    db.add(log)
    log_activity(db, ActivityAction.IMPORTED, "import", "vlastnici",
                 entity_name="Import vlastníků",
                 description=f"{result['owners_created']} vlastníků, {result.get('units_created', 0)} jednotek")
    db.commit()

    return templates.TemplateResponse("owners/import_result.html", {
        "request": request,
        "active_nav": "import",
        "result": result,
    })


@router.post("/import/{log_id}/smazat")
async def import_delete(
    log_id: int,
    db: Session = Depends(get_db),
):
    """Delete an import log entry and its uploaded file (data remain intact)."""
    from pathlib import Path

    log = db.query(ImportLog).filter_by(id=log_id, import_type="owners_excel").first()
    if not log:
        return RedirectResponse("/vlastnici/import", status_code=302)

    # Remove uploaded file
    try:
        p = Path(log.file_path)
        if p.exists():
            p.unlink()
    except Exception:
        logger.debug("Failed to clean up file: %s", log.file_path)

    # Remove log entry only
    db.delete(log)
    db.commit()

    return RedirectResponse("/vlastnici/import", status_code=302)


@router.post("/import-kontaktu/{log_id}/smazat")
async def contact_import_delete(
    log_id: int,
    db: Session = Depends(get_db),
):
    """Delete a contact import log entry and its uploaded file."""
    from pathlib import Path

    log = db.query(ImportLog).filter_by(id=log_id, import_type="contacts_excel").first()
    if not log:
        return RedirectResponse("/vlastnici/import", status_code=302)

    try:
        p = Path(log.file_path)
        if p.exists():
            p.unlink()
    except Exception:
        logger.debug("Failed to clean up file: %s", log.file_path)

    db.delete(log)
    db.commit()

    return RedirectResponse("/vlastnici/import", status_code=302)


@router.get("/{owner_id}")
async def owner_detail(
    owner_id: int,
    request: Request,
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    from app.services.code_list_service import get_all_code_lists

    owner = db.query(Owner).options(
        joinedload(Owner.units).joinedload(OwnerUnit.unit)
    ).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)

    # Units not yet assigned to this owner (current only)
    assigned_unit_ids = [ou.unit_id for ou in owner.current_units]
    if assigned_unit_ids:
        available_units = db.query(Unit).filter(
            Unit.id.notin_(assigned_unit_ids)
        ).order_by(Unit.unit_number).all()
    else:
        available_units = db.query(Unit).order_by(Unit.unit_number).all()

    svj_info = db.query(SvjInfo).first()
    declared_shares = svj_info.total_shares if svj_info and svj_info.total_shares else 0

    return templates.TemplateResponse("owners/detail.html", {
        "request": request,
        "active_nav": "owners",
        "owner": owner,
        "available_units": available_units,
        "declared_shares": declared_shares,
        "back_url": back or "/vlastnici",
        "back_label": (
            "Zpět na hromadné úpravy" if "/sprava/hromadne" in back
            else "Zpět na detail jednotky" if "/jednotky/" in back
            else "Zpět na seznam jednotek" if back.startswith("/jednotky")
            else "Zpět na porovnání" if "/synchronizace/" in back
            else "Zpět na hlasovací lístek" if "/hlasovani/" in back
            else "Zpět na nastavení" if back.startswith("/nastaveni")
            else "Zpět na seznam vlastníků"
        ),
        "code_lists": get_all_code_lists(db),
    })


@router.get("/{owner_id}/identita-formular")
async def owner_identity_edit_form(
    owner_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)
    return templates.TemplateResponse("partials/owner_identity_form.html", {
        "request": request,
        "owner": owner,
    })


@router.get("/{owner_id}/identita-info")
async def owner_identity_info(
    owner_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)
    return templates.TemplateResponse("partials/owner_identity_info.html", {
        "request": request,
        "owner": owner,
    })


def _rebuild_owner_name(owner: Owner) -> None:
    """Rebuild name_with_titles and name_normalized from identity fields."""
    parts_wt = []
    if owner.title:
        parts_wt.append(owner.title)
    if owner.last_name:
        parts_wt.append(owner.last_name)
    if owner.first_name:
        parts_wt.append(owner.first_name)
    owner.name_with_titles = " ".join(parts_wt)

    parts_norm = []
    if owner.last_name:
        parts_norm.append(owner.last_name)
    if owner.first_name:
        parts_norm.append(owner.first_name)
    owner.name_normalized = strip_diacritics(" ".join(parts_norm))


def _find_duplicate_owners(db: Session, owner: Owner) -> list[Owner]:
    """Find other active owners with the same name_normalized (potential duplicates)."""
    if not owner.name_normalized:
        return []
    return (
        db.query(Owner)
        .filter(
            Owner.id != owner.id,
            Owner.is_active == True,
            Owner.name_normalized == owner.name_normalized,
        )
        .options(joinedload(Owner.units).joinedload(OwnerUnit.unit))
        .all()
    )


def _header_oob_html(owner: Owner) -> str:
    """Build OOB swap HTML for owner display name + badges in page header."""
    from markupsafe import escape
    name_html = (
        f'<h1 id="owner-display-name" hx-swap-oob="true"'
        f' class="text-2xl font-bold text-gray-800">{escape(owner.display_name)}</h1>'
    )
    type_badge = (
        '<span class="px-2 py-1 text-xs font-medium bg-blue-100 text-blue-800 rounded-full">Právnická osoba</span>'
        if owner.owner_type == OwnerType.LEGAL_ENTITY
        else '<span class="px-2 py-1 text-xs font-medium bg-gray-100 text-gray-800 rounded-full">Fyzická osoba</span>'
    )
    extra_badges = ""
    if owner.birth_number:
        extra_badges += f'<span class="px-2 py-1 text-xs font-medium bg-gray-50 text-gray-700 rounded border border-gray-200">RČ: {escape(owner.birth_number)}</span>'
    if owner.company_id:
        extra_badges += f'<span class="px-2 py-1 text-xs font-medium bg-gray-50 text-gray-700 rounded border border-gray-200">IČ: {escape(owner.company_id)}</span>'
    badges_html = f'<div id="owner-badges" hx-swap-oob="true" class="mt-1 flex items-center gap-2">{type_badge}{extra_badges}</div>'
    return name_html + badges_html


@router.post("/{owner_id}/identita-upravit")
async def owner_identity_update(
    owner_id: int,
    request: Request,
    owner_type: str = Form("physical"),
    title: str = Form(""),
    last_name: str = Form(""),
    first_name_physical: str = Form(""),
    first_name_legal: str = Form(""),
    birth_number: str = Form(""),
    company_id: str = Form(""),
    db: Session = Depends(get_db),
):
    owner = db.query(Owner).options(
        joinedload(Owner.units).joinedload(OwnerUnit.unit)
    ).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)

    # Remember old normalized name to find duplicates BEFORE changing
    old_normalized = owner.name_normalized

    owner.owner_type = OwnerType(owner_type)

    if owner_type == "legal":
        owner.first_name = first_name_legal.strip()
        owner.last_name = None
        owner.title = None
        owner.birth_number = None
        owner.company_id = company_id.strip() or None
    else:
        owner.first_name = first_name_physical.strip()
        owner.last_name = last_name.strip() or None
        owner.title = title.strip() or None
        owner.birth_number = birth_number.strip() or None
        owner.company_id = None

    _rebuild_owner_name(owner)
    owner.updated_at = datetime.utcnow()
    db.commit()

    # Find duplicates by OLD name (before edit)
    duplicates = []
    if old_normalized:
        duplicates = (
            db.query(Owner)
            .filter(
                Owner.id != owner.id,
                Owner.is_active == True,
                Owner.name_normalized == old_normalized,
            )
            .options(joinedload(Owner.units).joinedload(OwnerUnit.unit))
            .all()
        )

    if request.headers.get("HX-Request"):
        from fastapi.responses import HTMLResponse
        identity_html = templates.TemplateResponse("partials/owner_identity_info.html", {
            "request": request,
            "owner": owner,
            "saved": True,
            "duplicates": duplicates,
        })
        body = identity_html.body.decode() + _header_oob_html(owner)
        return HTMLResponse(body)
    return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)


@router.post("/{owner_id}/sloucit")
async def owner_merge(
    owner_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Merge duplicate owners into this one. Move their units, deactivate them."""
    form = await request.form()
    merge_ids = [int(v) for v in form.getlist("merge_ids")]
    if not merge_ids:
        return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)

    owner = db.query(Owner).options(
        joinedload(Owner.units).joinedload(OwnerUnit.unit)
    ).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)

    duplicates = []
    for dup_id in merge_ids:
        dup = db.query(Owner).options(
            joinedload(Owner.units).joinedload(OwnerUnit.unit)
        ).get(dup_id)
        if dup and dup.id != owner.id:
            duplicates.append(dup)

    from app.services.owner_service import merge_owners
    merge_owners(owner, duplicates, db)
    db.commit()
    db.refresh(owner)

    if request.headers.get("HX-Request"):
        from fastapi.responses import HTMLResponse
        from app.services.code_list_service import get_all_code_lists
        # Refresh identity section (no more duplicates)
        identity_html = templates.TemplateResponse("partials/owner_identity_info.html", {
            "request": request,
            "owner": owner,
            "saved": True,
            "duplicates": [],
        })
        # Also refresh units section via OOB (new units were merged in)
        available_units, declared_shares = _owner_units_context(owner, db)
        units_html = templates.TemplateResponse("partials/owner_units_section.html", {
            "request": request,
            "owner": owner,
            "available_units": available_units,
            "declared_shares": declared_shares,
            "code_lists": get_all_code_lists(db),
        })
        units_oob = (
            f'<div id="owner-units-section" hx-swap-oob="true">'
            f'{units_html.body.decode()}</div>'
        )
        body = identity_html.body.decode() + _header_oob_html(owner) + units_oob
        return HTMLResponse(body)
    return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)


@router.get("/{owner_id}/upravit-formular")
async def owner_edit_form(
    owner_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)
    return templates.TemplateResponse("partials/owner_contact_form.html", {
        "request": request,
        "owner": owner,
    })


@router.get("/{owner_id}/info")
async def owner_info(
    owner_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)
    return templates.TemplateResponse("partials/owner_contact_info.html", {
        "request": request,
        "owner": owner,
    })


def _address_context(owner, prefix):
    """Extract address fields for a given prefix (perm/corr)."""
    return {
        "prefix": prefix,
        "address_label": "Trvalá adresa" if prefix == "perm" else "Korespondenční adresa",
        "street": getattr(owner, f"{prefix}_street"),
        "district": getattr(owner, f"{prefix}_district"),
        "city": getattr(owner, f"{prefix}_city"),
        "zip": getattr(owner, f"{prefix}_zip"),
        "country": getattr(owner, f"{prefix}_country"),
    }


@router.get("/{owner_id}/adresa/{prefix}/upravit-formular")
async def owner_address_edit_form(
    owner_id: int,
    prefix: str,
    request: Request,
    db: Session = Depends(get_db),
):
    if prefix not in ("perm", "corr"):
        return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)
    return templates.TemplateResponse("partials/owner_address_form.html", {
        "request": request,
        "owner": owner,
        **_address_context(owner, prefix),
    })


@router.get("/{owner_id}/adresa/{prefix}/info")
async def owner_address_info(
    owner_id: int,
    prefix: str,
    request: Request,
    db: Session = Depends(get_db),
):
    if prefix not in ("perm", "corr"):
        return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)
    return templates.TemplateResponse("partials/owner_address_info.html", {
        "request": request,
        "owner": owner,
        **_address_context(owner, prefix),
    })


@router.post("/{owner_id}/adresa/{prefix}/upravit")
async def owner_address_update(
    owner_id: int,
    prefix: str,
    request: Request,
    street: str = Form(""),
    district: str = Form(""),
    city: str = Form(""),
    zip: str = Form(""),
    country: str = Form(""),
    db: Session = Depends(get_db),
):
    if prefix not in ("perm", "corr"):
        return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)

    setattr(owner, f"{prefix}_street", street.strip() or None)
    setattr(owner, f"{prefix}_district", district.strip() or None)
    setattr(owner, f"{prefix}_city", city.strip() or None)
    setattr(owner, f"{prefix}_zip", zip.strip() or None)
    setattr(owner, f"{prefix}_country", country.strip() or None)
    owner.updated_at = datetime.utcnow()
    db.commit()

    if request.headers.get("HX-Request"):
        return templates.TemplateResponse("partials/owner_address_info.html", {
            "request": request,
            "owner": owner,
            "saved": True,
            **_address_context(owner, prefix),
        })
    return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)


@router.post("/{owner_id}/upravit")
async def owner_update(
    owner_id: int,
    request: Request,
    email: str = Form(""),
    email_secondary: str = Form(""),
    phone: str = Form(""),
    phone_secondary: str = Form(""),
    phone_landline: str = Form(""),
    db: Session = Depends(get_db),
):
    owner = db.query(Owner).get(owner_id)
    if owner:
        owner.email = (email.strip() if email.strip() and is_valid_email(email.strip()) else None)
        owner.email_secondary = (email_secondary.strip() if email_secondary.strip() and is_valid_email(email_secondary.strip()) else None)
        owner.phone = phone.strip() or None
        owner.phone_secondary = phone_secondary.strip() or None
        owner.phone_landline = phone_landline.strip() or None
        owner.updated_at = datetime.utcnow()
        db.commit()

    if request.headers.get("HX-Request"):
        return templates.TemplateResponse("partials/owner_contact_info.html", {
            "request": request,
            "owner": owner,
            "saved": True,
        })
    return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)


def _owner_units_context(owner, db):
    """Helper to build context for owner_units_section partial."""
    assigned_unit_ids = [ou.unit_id for ou in owner.current_units]
    if assigned_unit_ids:
        available_units = db.query(Unit).filter(
            Unit.id.notin_(assigned_unit_ids)
        ).order_by(Unit.unit_number).all()
    else:
        available_units = db.query(Unit).order_by(Unit.unit_number).all()
    svj_info = db.query(SvjInfo).first()
    declared_shares = svj_info.total_shares if svj_info and svj_info.total_shares else 0
    return available_units, declared_shares


@router.post("/{owner_id}/jednotky/pridat")
async def owner_add_unit(
    owner_id: int,
    request: Request,
    unit_id: str = Form(...),
    ownership_type: str = Form(""),
    share: str = Form("1.0"),
    votes: str = Form("0"),
    db: Session = Depends(get_db),
):
    owner = db.query(Owner).options(
        joinedload(Owner.units).joinedload(OwnerUnit.unit)
    ).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)

    # Check for duplicate
    unit_id_int = int(unit_id)
    exists = db.query(OwnerUnit).filter_by(
        owner_id=owner_id, unit_id=unit_id_int
    ).filter(OwnerUnit.valid_to.is_(None)).first()
    if not exists:
        ou = OwnerUnit(
            owner_id=owner_id,
            unit_id=unit_id_int,
            ownership_type=ownership_type or None,
            share=float(share) if share else 1.0,
            votes=int(votes) if votes else 0,
            valid_from=date.today(),
        )
        db.add(ou)
        db.flush()

        from app.services.owner_exchange import recalculate_unit_votes
        unit = db.query(Unit).get(unit_id_int)
        if unit:
            recalculate_unit_votes(unit, db)

        db.commit()
        # Refresh owner to get updated units
        db.refresh(owner)

    if request.headers.get("HX-Request"):
        from app.services.code_list_service import get_all_code_lists
        available_units, declared_shares = _owner_units_context(owner, db)
        return templates.TemplateResponse("partials/owner_units_section.html", {
            "request": request,
            "owner": owner,
            "available_units": available_units,
            "declared_shares": declared_shares,
            "code_lists": get_all_code_lists(db),
        })
    return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)


@router.post("/{owner_id}/jednotky/{ou_id}/odebrat")
async def owner_remove_unit(
    owner_id: int,
    ou_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    ou = db.query(OwnerUnit).filter_by(id=ou_id, owner_id=owner_id).first()
    if ou:
        ou.valid_to = date.today()
        db.commit()

    owner = db.query(Owner).options(
        joinedload(Owner.units).joinedload(OwnerUnit.unit)
    ).get(owner_id)

    if request.headers.get("HX-Request"):
        from app.services.code_list_service import get_all_code_lists
        available_units, declared_shares = _owner_units_context(owner, db)
        return templates.TemplateResponse("partials/owner_units_section.html", {
            "request": request,
            "owner": owner,
            "available_units": available_units,
            "declared_shares": declared_shares,
            "code_lists": get_all_code_lists(db),
        })
    return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)
