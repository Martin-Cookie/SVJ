from __future__ import annotations

import csv
import io
from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Depends, Form, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import ActivityAction, Owner, OwnerType, OwnerUnit, Prescription, PrescriptionYear, SvjInfo, Unit, UnitBalance, log_activity
from app.routers.payments._helpers import compute_debt_map
from app.services.code_list_service import get_all_code_lists
from app.services.owner_exchange import recalculate_unit_votes
from app.services.owner_service import merge_owners
from app.utils import build_list_url, excel_auto_width, is_htmx_partial, is_valid_email, strip_diacritics, utcnow

from ._helpers import (
    SORT_COLUMNS,
    _address_context,
    _filter_owners,
    _format_address,
    _header_oob_html,
    _owner_units_context,
    _rebuild_owner_name,
    logger,
    templates,
)

router = APIRouter()


@router.get("/novy-formular")
async def owner_create_form(request: Request):
    """Formulář pro vytvoření nového vlastníka."""
    return templates.TemplateResponse(request, "partials/owner_create_form.html", {
    })


@router.post("/novy")
async def owner_create(
    request: Request,
    first_name: str = Form(...),
    last_name: str = Form(...),
    title: str = Form(""),
    owner_type: str = Form("physical"),
    email: str = Form(""),
    phone: str = Form(""),
    birth_number: str = Form(""),
    force_create: str = Form(""),
    db: Session = Depends(get_db),
):
    """Vytvoření nového vlastníka s kontrolou duplicit."""
    # Build name_with_titles and name_normalized
    parts_wt = []
    if title:
        parts_wt.append(title)
    if last_name:
        parts_wt.append(last_name)
    if first_name:
        parts_wt.append(first_name)
    name_with_titles = " ".join(parts_wt)

    parts_norm = []
    if last_name:
        parts_norm.append(last_name)
    if first_name:
        parts_norm.append(first_name)
    name_normalized = strip_diacritics(" ".join(parts_norm))

    # Validate email — return form with error if invalid
    email_clean = email.strip() if email else ""
    if email_clean and not is_valid_email(email_clean):
        return templates.TemplateResponse(request, "partials/owner_create_form.html", {
            "error": f"Neplatný formát emailu: {email_clean}",
            "form_data": {
                "first_name": first_name, "last_name": last_name,
                "title": title, "owner_type": owner_type,
                "email": email, "phone": phone,
                "birth_number": birth_number,
            },
        })

    # Check for duplicates — name, birth_number, email
    duplicates = []
    if name_normalized:
        dup_name = db.query(Owner).filter(
            Owner.is_active == True,
            Owner.name_normalized == name_normalized,
        ).first()
        if dup_name:
            duplicates.append(("jméno", dup_name))
    bn_clean = birth_number.strip() if birth_number else ""
    if bn_clean:
        dup_bn = db.query(Owner).filter(
            Owner.is_active == True,
            Owner.birth_number == bn_clean,
        ).first()
        if dup_bn and dup_bn not in [d[1] for d in duplicates]:
            duplicates.append(("RČ", dup_bn))
    if email_clean:
        dup_email = db.query(Owner).filter(
            Owner.is_active == True,
            Owner.email == email_clean,
        ).first()
        if dup_email and dup_email not in [d[1] for d in duplicates]:
            duplicates.append(("email", dup_email))

    # If duplicates found and user hasn't confirmed, show warning
    if duplicates and force_create != "1":
        return templates.TemplateResponse(request, "partials/owner_create_form.html", {
            "duplicates": duplicates,
            "form_data": {
                "first_name": first_name, "last_name": last_name,
                "title": title, "owner_type": owner_type,
                "email": email, "phone": phone,
                "birth_number": birth_number,
            },
        })

    owner = Owner(
        first_name=first_name.strip(),
        last_name=last_name.strip() or None,
        title=title.strip() or None,
        owner_type=OwnerType(owner_type),
        name_with_titles=name_with_titles,
        name_normalized=name_normalized,
        email=email_clean or None,
        phone=phone.strip() or None,
        birth_number=bn_clean or None,
        data_source="manual",
        is_active=True,
        created_at=utcnow(),
    )
    db.add(owner)
    db.flush()
    log_activity(db, ActivityAction.CREATED, "owner", "vlastnici",
                 entity_id=owner.id, entity_name=owner.display_name)
    db.commit()

    redirect_url = f"/vlastnici/{owner.id}?info=vytvoren"
    return RedirectResponse(redirect_url, status_code=302)


@router.get("/")
async def owner_list(
    request: Request,
    q: str = Query("", alias="q"),
    owner_type: str = Query("", alias="typ"),
    vlastnictvi: str = Query("", alias="vlastnictvi"),
    kontakt: str = Query("", alias="kontakt"),
    stav: str = Query("", alias="stav"),
    sekce: str = Query("", alias="sekce"),
    sort: str = Query("name", alias="sort"),
    order: str = Query("asc", alias="order"),
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    """Seznam vlastníků s filtry, hledáním a řazením."""
    owners = _filter_owners(db, q, owner_type, vlastnictvi, kontakt, stav, sekce, sort, order)

    # Debt map — platební dluh per owner (přes unit_id → owner)
    debt_map = {}
    latest_py = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
    if latest_py:
        unit_debt_map = compute_debt_map(db, latest_py.year)
        if unit_debt_map:
            # Mapování unit_id → owner_ids (aktivní vlastníci)
            ou_rows = db.query(OwnerUnit.unit_id, OwnerUnit.owner_id).filter(
                OwnerUnit.unit_id.in_(unit_debt_map.keys()),
                OwnerUnit.valid_to.is_(None),
            ).all()
            for unit_id, owner_id in ou_rows:
                debt_map[owner_id] = debt_map.get(owner_id, 0) + unit_debt_map[unit_id]

    # Python-side sort by debt
    if sort == "dluh":
        owners.sort(key=lambda o: debt_map.get(o.id, 0), reverse=(order == "desc"))

    # Current list URL for back navigation
    list_url = build_list_url(request)

    # Mapa unit_id → VS + měsíční předpis (pro tooltipy u čísel jednotek)
    unit_vs = {}
    unit_monthly = {}
    current_year = utcnow().year
    py = db.query(PrescriptionYear).filter_by(year=current_year).first()
    if py:
        for presc in db.query(Prescription).filter_by(prescription_year_id=py.id).all():
            if presc.unit_id:
                if presc.variable_symbol:
                    unit_vs[presc.unit_id] = presc.variable_symbol
                if presc.monthly_total:
                    unit_monthly[presc.unit_id] = presc.monthly_total

    # Return partial only for targeted HTMX requests (search/filter), not boosted navigation
    if is_htmx_partial(request):
        return templates.TemplateResponse(request, "partials/owner_table_body.html", {
            "owners": owners,
            "list_url": list_url,
            "debt_map": debt_map,
            "unit_vs": unit_vs,
            "unit_monthly": unit_monthly,
        })

    # Stats for header
    all_owners = db.query(Owner).filter_by(is_active=True).count()
    total_units = db.query(Unit).count()
    type_counts_raw = (
        db.query(Owner.owner_type, func.count(Owner.id))
        .filter_by(is_active=True)
        .group_by(Owner.owner_type)
        .all()
    )
    type_counts = {ot.value: cnt for ot, cnt in type_counts_raw}
    sections = [
        r[0] for r in
        db.query(Unit.section).filter(Unit.section.isnot(None)).distinct().order_by(Unit.section).all()
    ]
    emails_count = db.query(Owner).filter(
        Owner.is_active == True,
        or_(
            (Owner.email.isnot(None)) & (Owner.email != ""),
            (Owner.email_secondary.isnot(None)) & (Owner.email_secondary != ""),
        ),
    ).count()
    phones_count = db.query(Owner).filter(
        Owner.is_active == True,
        Owner.phone.isnot(None),
        Owner.phone != "",
    ).count()

    no_units_count = db.query(Owner).filter(
        Owner.is_active == True,
        ~Owner.units.any(OwnerUnit.valid_to.is_(None)),
    ).count()

    total_scd = db.query(func.sum(OwnerUnit.votes)).filter(OwnerUnit.valid_to.is_(None)).scalar() or 0
    svj_info = db.query(SvjInfo).first()
    declared_shares = svj_info.total_shares if svj_info and svj_info.total_shares else 0

    # Ownership type counts
    ownership_counts_raw = (
        db.query(
            func.coalesce(OwnerUnit.ownership_type, ""),
            func.count(func.distinct(OwnerUnit.owner_id)),
        )
        .filter(OwnerUnit.valid_to.is_(None))
        .group_by(func.coalesce(OwnerUnit.ownership_type, ""))
        .all()
    )
    ownership_counts = {ot or "": cnt for ot, cnt in ownership_counts_raw}

    return templates.TemplateResponse(request, "owners/list.html", {
        "active_nav": "owners",
        "owners": owners,
        "list_url": list_url,
        "back_url": back,
        "q": q,
        "owner_type": owner_type,
        "vlastnictvi": vlastnictvi,
        "kontakt": kontakt,
        "stav": stav,
        "sekce": sekce,
        "sort": sort,
        "order": order,
        "debt_map": debt_map,
        "unit_vs": unit_vs,
        "unit_monthly": unit_monthly,
        "owner_types": OwnerType,
        "stats": {
            "total_owners": all_owners,
            "total_units": total_units,
            "type_counts": type_counts,
            "sections": sections,
            "emails_count": emails_count,
            "phones_count": phones_count,
            "no_units_count": no_units_count,
            "total_scd": total_scd,
            "declared_shares": declared_shares,
            "ownership_counts": ownership_counts,
        },
    })


@router.get("/exportovat/{fmt}")
async def owner_export(
    fmt: str,
    q: str = Query("", alias="q"),
    owner_type: str = Query("", alias="typ"),
    vlastnictvi: str = Query("", alias="vlastnictvi"),
    kontakt: str = Query("", alias="kontakt"),
    stav: str = Query("", alias="stav"),
    sekce: str = Query("", alias="sekce"),
    sort: str = Query("name", alias="sort"),
    order: str = Query("asc", alias="order"),
    db: Session = Depends(get_db),
):
    """Export filtered owners to Excel or CSV."""
    if fmt not in ("xlsx", "csv"):
        return RedirectResponse("/vlastnici", status_code=302)

    owners = _filter_owners(db, q, owner_type, vlastnictvi, kontakt, stav, sekce, sort, order)

    headers = ["Vlastník", "Typ", "Jednotky", "Sekce", "Email", "Email 2", "Telefon", "Podíl SČD", "RČ/IČ", "Trvalá adresa", "Korespondenční adresa"]

    def _row(o):
        units = ", ".join(str(ou.unit.unit_number) for ou in o.current_units)
        sections = ", ".join(sorted(set(ou.unit.section for ou in o.current_units if ou.unit.section)))
        typ = "Právnická os." if o.owner_type == OwnerType.LEGAL_ENTITY else "Fyzická os."
        podil = sum(ou.votes for ou in o.current_units)
        rc_ic = o.company_id or o.birth_number or ""
        return [
            o.display_name,
            typ,
            units,
            sections,
            o.email or "",
            o.email_secondary or "",
            o.phone or "",
            podil,
            rc_ic,
            _format_address(o, "perm"),
            _format_address(o, "corr"),
        ]

    timestamp = datetime.now().strftime("%Y%m%d")

    # Suffix podle aktivního filtru
    typ_labels = {"physical": "fyzicke", "legal": "pravnicke"}
    kontakt_labels = {"s_emailem": "s_emailem", "bez_emailu": "bez_emailu", "s_telefonem": "s_telefonem", "bez_telefonu": "bez_telefonu"}
    if owner_type and owner_type in typ_labels:
        suffix = f"_{typ_labels[owner_type]}"
    elif kontakt and kontakt in kontakt_labels:
        suffix = f"_{kontakt_labels[kontakt]}"
    elif stav:
        suffix = f"_{strip_diacritics(stav)}"
    elif vlastnictvi:
        suffix = f"_{strip_diacritics(vlastnictvi)}"
    elif sekce:
        suffix = f"_sekce_{strip_diacritics(sekce)}"
    elif q:
        suffix = "_hledani"
    else:
        suffix = "_vsichni"
    filename = f"vlastnici{suffix}_{timestamp}"

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Vlastníci"
        bold = Font(bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold

        for row_idx, o in enumerate(owners, 2):
            for col_idx, val in enumerate(_row(o), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        excel_auto_width(ws)

        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    else:
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(headers)
        for o in owners:
            writer.writerow(_row(o))
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )


@router.get("/{owner_id}")
async def owner_detail(
    owner_id: int,
    request: Request,
    back: str = Query("", alias="back"),
    info: str = Query(""),
    db: Session = Depends(get_db),
):
    """Detail vlastníka s jednotkami, kontakty a adresami."""

    owner = db.query(Owner).options(
        joinedload(Owner.units).joinedload(OwnerUnit.unit)
    ).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)

    # Units not yet assigned to this owner (current only)
    assigned_unit_ids = [ou.unit_id for ou in owner.current_units]
    if assigned_unit_ids:
        available_units = db.query(Unit).filter(
            Unit.id.notin_(assigned_unit_ids)
        ).order_by(Unit.unit_number).all()
    else:
        available_units = db.query(Unit).order_by(Unit.unit_number).all()

    svj_info = db.query(SvjInfo).first()
    declared_shares = svj_info.total_shares if svj_info and svj_info.total_shares else 0

    # Platební dluh vlastníka (přes jeho jednotky)
    owner_debt = 0
    unit_debt_map = {}
    bal_py = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
    if bal_py:
        unit_debt_map = compute_debt_map(db, bal_py.year)
        for ou in owner.current_units:
            owner_debt += unit_debt_map.get(ou.unit_id, 0)

    ctx = {
        "request": request,
        "active_nav": "owners",
        "owner": owner,
        "available_units": available_units,
        "declared_shares": declared_shares,
        "owner_debt": owner_debt,
        "unit_debt_map": unit_debt_map,
        "back_url": back or "/vlastnici",
        "back_label": (
            "Zpět na hromadné úpravy" if "/sprava/hromadne" in back
            else "Zpět na detail jednotky" if "/jednotky/" in back
            else "Zpět na seznam jednotek" if back.startswith("/jednotky")
            else "Zpět na porovnání" if "/synchronizace/" in back
            else "Zpět na hlasovací lístek" if "/hlasovani/" in back
            else "Zpět na nastavení" if back.startswith("/nastaveni")
            else "Zpět na detail nájemce" if "/najemci/" in back
            else "Zpět na duplicity" if back.startswith("/sprava/duplicity")
            else "Zpět na seznam vlastníků"
        ),
        "code_lists": get_all_code_lists(db),
    }
    if info == "vytvoren":
        ctx["flash_message"] = f"Vlastník {owner.display_name} byl vytvořen."
    elif info == "neplatny-email":
        ctx["flash_message"] = "Vlastník vytvořen, ale zadaný email měl neplatný formát a nebyl uložen."
        ctx["flash_type"] = "warning"
    return templates.TemplateResponse(request, "owners/detail.html", ctx)


@router.get("/{owner_id}/identita-formular")
async def owner_identity_edit_form(
    owner_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Formulář pro inline editaci identity vlastníka."""
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)
    return templates.TemplateResponse(request, "partials/owner_identity_form.html", {
        "owner": owner,
    })


@router.get("/{owner_id}/identita-info")
async def owner_identity_info(
    owner_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Zobrazení identity vlastníka po zrušení editace."""
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)
    return templates.TemplateResponse(request, "partials/owner_identity_info.html", {
        "owner": owner,
    })


@router.post("/{owner_id}/identita-upravit")
async def owner_identity_update(
    owner_id: int,
    request: Request,
    owner_type: str = Form("physical"),
    title: str = Form(""),
    last_name: str = Form(""),
    first_name_physical: str = Form(""),
    first_name_legal: str = Form(""),
    birth_number: str = Form(""),
    company_id: str = Form(""),
    db: Session = Depends(get_db),
):
    """Uložení změn identity vlastníka (jméno, typ, RČ/IČ)."""
    owner = db.query(Owner).options(
        joinedload(Owner.units).joinedload(OwnerUnit.unit)
    ).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)

    # Remember old normalized name to find duplicates BEFORE changing
    old_normalized = owner.name_normalized

    owner.owner_type = OwnerType(owner_type)

    if owner_type == "legal":
        owner.first_name = first_name_legal.strip()
        owner.last_name = None
        owner.title = None
        owner.birth_number = None
        owner.company_id = company_id.strip() or None
    else:
        owner.first_name = first_name_physical.strip()
        owner.last_name = last_name.strip() or None
        owner.title = title.strip() or None
        owner.birth_number = birth_number.strip() or None
        owner.company_id = None

    _rebuild_owner_name(owner)
    owner.updated_at = utcnow()
    log_activity(db, ActivityAction.UPDATED, "owner", "vlastnici",
                 entity_id=owner.id, entity_name=owner.display_name,
                 description="Upravena identita")
    db.commit()

    # Find duplicates by OLD name (before edit)
    duplicates = []
    if old_normalized:
        duplicates = (
            db.query(Owner)
            .filter(
                Owner.id != owner.id,
                Owner.is_active == True,
                Owner.name_normalized == old_normalized,
            )
            .options(joinedload(Owner.units).joinedload(OwnerUnit.unit))
            .all()
        )

    if request.headers.get("HX-Request"):
        identity_html = templates.TemplateResponse(request, "partials/owner_identity_info.html", {
            "owner": owner,
            "saved": True,
            "duplicates": duplicates,
        })
        body = identity_html.body.decode() + _header_oob_html(owner)
        return HTMLResponse(body)
    return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)


@router.post("/{owner_id}/sloucit")
async def owner_merge(
    owner_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Merge duplicate owners into this one. Move their units, deactivate them."""
    form = await request.form()
    merge_ids = [int(v) for v in form.getlist("merge_ids")]
    if not merge_ids:
        return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)

    owner = db.query(Owner).options(
        joinedload(Owner.units).joinedload(OwnerUnit.unit)
    ).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)

    duplicates = []
    for dup_id in merge_ids:
        dup = db.query(Owner).options(
            joinedload(Owner.units).joinedload(OwnerUnit.unit)
        ).get(dup_id)
        if dup and dup.id != owner.id:
            duplicates.append(dup)

    merge_owners(owner, duplicates, db)
    log_activity(db, ActivityAction.UPDATED, "owner", "vlastnici",
                 entity_id=owner.id, entity_name=owner.display_name,
                 description=f"Sloučeno {len(duplicates)} duplicit")
    db.commit()
    db.refresh(owner)

    if request.headers.get("HX-Request"):
        # Refresh identity section (no more duplicates)
        identity_html = templates.TemplateResponse(request, "partials/owner_identity_info.html", {
            "owner": owner,
            "saved": True,
            "duplicates": [],
        })
        # Also refresh units section via OOB (new units were merged in)
        available_units, declared_shares = _owner_units_context(owner, db)
        _bal_py = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        _udm = compute_debt_map(db, _bal_py.year) if _bal_py else {}
        units_html = templates.TemplateResponse(request, "partials/owner_units_section.html", {
            "owner": owner,
            "available_units": available_units,
            "declared_shares": declared_shares,
            "code_lists": get_all_code_lists(db),
            "unit_debt_map": _udm,
        })
        units_oob = (
            f'<div id="owner-units-section" hx-swap-oob="true">'
            f'{units_html.body.decode()}</div>'
        )
        body = identity_html.body.decode() + _header_oob_html(owner) + units_oob
        return HTMLResponse(body)
    return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)


@router.get("/{owner_id}/upravit-formular")
async def owner_edit_form(
    owner_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Formulář pro inline editaci kontaktních údajů vlastníka."""
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)
    return templates.TemplateResponse(request, "partials/owner_contact_form.html", {
        "owner": owner,
    })


@router.get("/{owner_id}/info")
async def owner_info(
    owner_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Zobrazení kontaktních údajů vlastníka po zrušení editace."""
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)
    return templates.TemplateResponse(request, "partials/owner_contact_info.html", {
        "owner": owner,
    })


@router.get("/{owner_id}/adresa/{prefix}/upravit-formular")
async def owner_address_edit_form(
    owner_id: int,
    prefix: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Formulář pro inline editaci adresy vlastníka."""
    if prefix not in ("perm", "corr"):
        return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)
    return templates.TemplateResponse(request, "partials/owner_address_form.html", {
        "owner": owner,
        **_address_context(owner, prefix),
    })


@router.get("/{owner_id}/adresa/{prefix}/info")
async def owner_address_info(
    owner_id: int,
    prefix: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Zobrazení adresy vlastníka po zrušení editace."""
    if prefix not in ("perm", "corr"):
        return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)
    return templates.TemplateResponse(request, "partials/owner_address_info.html", {
        "owner": owner,
        **_address_context(owner, prefix),
    })


@router.post("/{owner_id}/adresa/{prefix}/upravit")
async def owner_address_update(
    owner_id: int,
    prefix: str,
    request: Request,
    street: str = Form(""),
    district: str = Form(""),
    city: str = Form(""),
    zip: str = Form(""),
    country: str = Form(""),
    db: Session = Depends(get_db),
):
    """Uložení změn adresy vlastníka."""
    if prefix not in ("perm", "corr"):
        return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)
    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)

    setattr(owner, f"{prefix}_street", street.strip() or None)
    setattr(owner, f"{prefix}_district", district.strip() or None)
    setattr(owner, f"{prefix}_city", city.strip() or None)
    setattr(owner, f"{prefix}_zip", zip.strip() or None)
    setattr(owner, f"{prefix}_country", country.strip() or None)
    owner.updated_at = utcnow()
    addr_label = "trvalá adresa" if prefix == "perm" else "korespondenční adresa"
    log_activity(db, ActivityAction.UPDATED, "owner", "vlastnici",
                 entity_id=owner.id, entity_name=owner.display_name,
                 description=f"Upravena {addr_label}")
    db.commit()

    if request.headers.get("HX-Request"):
        return templates.TemplateResponse(request, "partials/owner_address_info.html", {
            "owner": owner,
            "saved": True,
            **_address_context(owner, prefix),
        })
    return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)


@router.post("/{owner_id}/upravit")
async def owner_update(
    owner_id: int,
    request: Request,
    email: str = Form(""),
    email_secondary: str = Form(""),
    phone: str = Form(""),
    phone_secondary: str = Form(""),
    phone_landline: str = Form(""),
    db: Session = Depends(get_db),
):
    """Uložení změn kontaktních údajů vlastníka."""
    owner = db.query(Owner).get(owner_id)
    if owner:
        new_email = email.strip() if email.strip() and is_valid_email(email.strip()) else None
        email_changed = new_email != owner.email
        owner.email = new_email
        owner.email_secondary = (email_secondary.strip() if email_secondary.strip() and is_valid_email(email_secondary.strip()) else None)
        owner.phone = phone.strip() or None
        owner.phone_secondary = phone_secondary.strip() or None
        owner.phone_landline = phone_landline.strip() or None
        # Reset bounce flag when primary email is changed to a new non-empty value —
        # nová adresa dostane novou šanci doručit, dokud neselže.
        if email_changed and new_email and owner.email_invalid:
            owner.email_invalid = False
            owner.email_invalid_reason = None
        owner.updated_at = utcnow()
        log_activity(db, ActivityAction.UPDATED, "owner", "vlastnici",
                     entity_id=owner.id, entity_name=owner.display_name,
                     description="Upraveny kontaktní údaje")
        db.commit()

    if request.headers.get("HX-Request"):
        return templates.TemplateResponse(request, "partials/owner_contact_info.html", {
            "owner": owner,
            "saved": True,
        })
    return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)


@router.post("/{owner_id}/jednotky/pridat")
async def owner_add_unit(
    owner_id: int,
    request: Request,
    unit_id: str = Form(...),
    ownership_type: str = Form(""),
    share: str = Form("1.0"),
    votes: str = Form("0"),
    db: Session = Depends(get_db),
):
    """Přiřazení jednotky k vlastníkovi."""
    owner = db.query(Owner).options(
        joinedload(Owner.units).joinedload(OwnerUnit.unit)
    ).get(owner_id)
    if not owner:
        return RedirectResponse("/vlastnici", status_code=302)

    # Check for duplicate
    unit_id_int = int(unit_id)
    exists = db.query(OwnerUnit).filter_by(
        owner_id=owner_id, unit_id=unit_id_int
    ).filter(OwnerUnit.valid_to.is_(None)).first()
    if not exists:
        ou = OwnerUnit(
            owner_id=owner_id,
            unit_id=unit_id_int,
            ownership_type=ownership_type or None,
            share=float(share) if share else 1.0,
            votes=int(votes) if votes else 0,
            valid_from=date.today(),
        )
        db.add(ou)
        db.flush()

        unit = db.query(Unit).get(unit_id_int)
        if unit:
            recalculate_unit_votes(unit, db)

        db.commit()
        # Refresh owner to get updated units
        db.refresh(owner)

    if request.headers.get("HX-Request"):
        available_units, declared_shares = _owner_units_context(owner, db)
        _bal_py = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        _udm = compute_debt_map(db, _bal_py.year) if _bal_py else {}
        return templates.TemplateResponse(request, "partials/owner_units_section.html", {
            "owner": owner,
            "available_units": available_units,
            "declared_shares": declared_shares,
            "code_lists": get_all_code_lists(db),
            "unit_debt_map": _udm,
        })
    return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)


@router.post("/{owner_id}/jednotky/{ou_id}/odebrat")
async def owner_remove_unit(
    owner_id: int,
    ou_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Odebrání jednotky od vlastníka (nastavení valid_to)."""
    ou = db.query(OwnerUnit).filter_by(id=ou_id, owner_id=owner_id).first()
    if ou:
        ou.valid_to = date.today()
        db.commit()

    owner = db.query(Owner).options(
        joinedload(Owner.units).joinedload(OwnerUnit.unit)
    ).get(owner_id)

    if request.headers.get("HX-Request"):
        available_units, declared_shares = _owner_units_context(owner, db)
        _bal_py = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
        _udm = compute_debt_map(db, _bal_py.year) if _bal_py else {}
        return templates.TemplateResponse(request, "partials/owner_units_section.html", {
            "owner": owner,
            "available_units": available_units,
            "declared_shares": declared_shares,
            "code_lists": get_all_code_lists(db),
            "unit_debt_map": _udm,
        })
    return RedirectResponse(f"/vlastnici/{owner_id}", status_code=302)
