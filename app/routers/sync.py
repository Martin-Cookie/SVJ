import logging
import re
import shutil
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, Query, Request, UploadFile
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import and_, cast, func, Integer, or_
from sqlalchemy.orm import Session, joinedload

from app.config import settings
from app.database import get_db

logger = logging.getLogger(__name__)
from app.models import (
    ImportLog, Owner, OwnerUnit, ShareCheckSession, SyncRecord, SyncResolution,
    SyncSession, SyncStatus, Unit,
)
from app.services.csv_comparator import compare_owners, parse_sousede_csv
from app.services.owner_exchange import execute_exchange, prepare_exchange_preview
from app.services.owner_matcher import normalize_for_matching
from app.utils import build_list_url, excel_auto_width, is_htmx_partial, setup_jinja_filters, strip_diacritics, validate_upload


router = APIRouter()
templates = Jinja2Templates(directory="app/templates")
setup_jinja_filters(templates)


@router.get("/")
async def sync_list(
    request: Request,
    sync_q: str = Query("", alias="sync_q"),
    sync_sort: str = Query("date", alias="sync_sort"),
    sync_order: str = Query("desc", alias="sync_order"),
    sc_q: str = Query("", alias="sc_q"),
    sc_sort: str = Query("date", alias="sc_sort"),
    sc_order: str = Query("desc", alias="sc_order"),
    # Legacy params (HTMX search uses these)
    q: str = Query(""),
    sort: str = Query(""),
    order: str = Query(""),
    back: str = Query("", alias="back"),
    chyba: str = Query("", alias="chyba"),
    db: Session = Depends(get_db),
):
    # Legacy fallback: if old q/sort/order params are used (HTMX partial)
    _sync_q = sync_q or q
    _sync_sort = sync_sort if sync_sort != "date" or not sort else sort
    _sync_order = sync_order if sync_order != "desc" or not order else order

    # --- Sync sessions ---
    sync_sessions = db.query(SyncSession).order_by(SyncSession.created_at.desc()).all()
    if _sync_q:
        q_lower = _sync_q.lower()
        q_ascii = strip_diacritics(_sync_q)
        sync_sessions = [
            s for s in sync_sessions
            if q_lower in (s.csv_filename or "").lower()
            or q_ascii in strip_diacritics(s.csv_filename or "")
            or q_lower in s.created_at.strftime("%d.%m.%Y %H:%M")
        ]

    SYNC_SORT_KEYS = {
        "filename": lambda s: (s.csv_filename or "").lower(),
        "date": lambda s: s.created_at,
        "matches": lambda s: s.total_matches or 0,
        "differences": lambda s: (s.total_differences or 0) + (s.total_missing or 0),
    }
    sort_fn = SYNC_SORT_KEYS.get(_sync_sort, SYNC_SORT_KEYS["date"])
    sync_sessions.sort(key=sort_fn, reverse=(_sync_order == "desc"))

    list_url = build_list_url(request)

    # HTMX partial for sync search
    if is_htmx_partial(request):
        ctx = {
            "request": request,
            "sessions": sync_sessions,
            "list_url": list_url,
            "q": _sync_q,
        }
        return templates.TemplateResponse("partials/sync_list_body.html", ctx)

    # --- Share check sessions (for full page only) ---
    sc_sessions = db.query(ShareCheckSession).order_by(ShareCheckSession.created_at.desc()).all()
    if sc_q:
        q_lower = sc_q.lower()
        q_ascii = strip_diacritics(sc_q)
        sc_sessions = [
            s for s in sc_sessions
            if q_lower in (s.filename or "").lower()
            or q_ascii in strip_diacritics(s.filename or "")
            or q_lower in s.created_at.strftime("%d.%m.%Y %H:%M")
        ]

    SC_SORT_KEYS = {
        "filename": lambda s: (s.filename or "").lower(),
        "date": lambda s: s.created_at,
        "matches": lambda s: s.total_matches or 0,
        "differences": lambda s: (s.total_differences or 0) + (s.total_missing_db or 0) + (s.total_missing_file or 0),
    }
    sc_sort_fn = SC_SORT_KEYS.get(sc_sort, SC_SORT_KEYS["date"])
    sc_sessions.sort(key=sc_sort_fn, reverse=(sc_order == "desc"))

    # Flash messages from query params
    _CHYBA_MSG = {
        "upload": "Nepodporovaný formát nebo příliš velký soubor.",
        "soubor": "Nebyl vybrán žádný soubor.",
        "cesta": "Neplatná cesta k souboru.",
        "hlavicky": "Nelze přečíst hlavičky souboru.",
        "prazdny": "Soubor neobsahuje žádná data k porovnání.",
    }
    flash_message = _CHYBA_MSG.get(chyba, "")
    flash_type = "error" if flash_message else ""

    ctx = {
        "request": request,
        "active_nav": "kontroly",
        "sync_sessions": sync_sessions,
        "sc_sessions": sc_sessions,
        "back_url": back,
        "list_url": list_url,
        "sync_q": _sync_q,
        "sync_sort": _sync_sort,
        "sync_order": _sync_order,
        "sc_q": sc_q,
        "sc_sort": sc_sort,
        "sc_order": sc_order,
        # Legacy: partials use "sessions" and "q"
        "sessions": sync_sessions,
        "q": _sync_q,
        "flash_message": flash_message,
        "flash_type": flash_type,
    }

    return templates.TemplateResponse("sync/index.html", ctx)


@router.post("/{session_id}/smazat")
async def sync_delete(session_id: int, db: Session = Depends(get_db)):
    """Delete a sync session, its records, and the uploaded CSV file."""
    session = db.query(SyncSession).get(session_id)
    if session:
        # Remove CSV file
        try:
            p = Path(session.csv_path)
            if p.exists():
                p.unlink()
        except Exception:
            logger.debug("Failed to clean up CSV: %s", session.csv_path)
        # Cascade deletes SyncRecord entries
        db.delete(session)
        db.commit()
    return RedirectResponse("/synchronizace", status_code=302)


@router.get("/nova")
async def sync_create_page():
    return RedirectResponse("/synchronizace", status_code=302)


@router.post("/nova")
async def sync_create(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename:
        return RedirectResponse("/synchronizace/nova", status_code=302)

    err = await validate_upload(file, max_size_mb=50, allowed_extensions=[".csv"])
    if err:
        return RedirectResponse("/synchronizace?chyba=upload", status_code=302)

    # Save CSV
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = settings.upload_dir / "csv" / f"{timestamp}_{file.filename}"
    dest.parent.mkdir(parents=True, exist_ok=True)
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)

    # Read CSV content - try multiple encodings
    csv_content = None
    for encoding in ["utf-8", "cp1250", "latin-1"]:
        try:
            with open(dest, "r", encoding=encoding) as f:
                csv_content = f.read()
            break
        except UnicodeDecodeError:
            continue

    if not csv_content:
        return RedirectResponse("/synchronizace", status_code=302)

    csv_records = parse_sousede_csv(csv_content)

    # Get current Excel data
    owners = db.query(Owner).filter_by(is_active=True).options(
        joinedload(Owner.units)
    ).all()

    excel_data = []
    for owner in owners:
        for ou in owner.current_units:
            unit_num = str(ou.unit.unit_number)
            excel_data.append({
                "unit_number": unit_num,
                "owner_name": owner.display_name,
                "first_name": owner.first_name,
                "last_name": owner.last_name or "",
                "name_normalized": owner.name_normalized,
                "owner_type": owner.owner_type.value,
                "space_type": ou.unit.space_type or "",
                "podil_scd": ou.unit.podil_scd or 0,
                "ownership_type": ou.ownership_type or "",
            })

    # Compare
    comparison = compare_owners(csv_records, excel_data)

    # Create session and records
    session = SyncSession(
        csv_filename=file.filename,
        csv_path=str(dest),
        total_records=len(comparison),
        total_matches=sum(1 for c in comparison if c["status"] == SyncStatus.MATCH),
        total_name_order=sum(1 for c in comparison if c["status"] == SyncStatus.NAME_ORDER),
        total_differences=sum(1 for c in comparison if c["status"] == SyncStatus.DIFFERENCE),
        total_missing=sum(
            1 for c in comparison
            if c["status"] in (SyncStatus.MISSING_CSV, SyncStatus.MISSING_EXCEL)
        ),
    )
    db.add(session)
    db.flush()

    for comp in comparison:
        record = SyncRecord(
            session_id=session.id,
            unit_number=comp["unit_number"],
            csv_owner_name=comp.get("csv_owner_name"),
            excel_owner_name=comp.get("excel_owner_name"),
            csv_ownership_type=comp.get("csv_ownership_type"),
            excel_ownership_type=comp.get("excel_ownership_type"),
            csv_space_type=comp.get("csv_space_type"),
            excel_space_type=comp.get("excel_space_type"),
            excel_podil_scd=comp.get("excel_podil_scd"),
            csv_share=comp.get("csv_share"),
            csv_email=comp.get("csv_email", ""),
            csv_phone=comp.get("csv_phone", ""),
            status=comp["status"],
            match_details=comp.get("match_details"),
            resolution=(
                SyncResolution.ACCEPTED
                if comp["status"] in (SyncStatus.MATCH, SyncStatus.NAME_ORDER)
                else SyncResolution.PENDING
            ),
        )
        db.add(record)

    db.commit()
    return RedirectResponse(f"/synchronizace/{session.id}", status_code=302)


SYNC_SORT_COLUMNS = {
    "unit": cast(SyncRecord.unit_number, Integer),
    "owner": SyncRecord.excel_owner_name,
    "space_type": SyncRecord.excel_space_type,
    "ownership": SyncRecord.excel_ownership_type,
    "podil": SyncRecord.excel_podil_scd,
    "match": SyncRecord.match_details,
    "action": SyncRecord.resolution,
}


@router.get("/{session_id}")
async def sync_detail(
    session_id: int,
    request: Request,
    q: str = Query(""),
    filtr: str = Query("", alias="filtr"),
    sort: str = Query("unit", alias="sort"),
    order: str = Query("asc", alias="order"),
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    session = db.query(SyncSession).get(session_id)
    if not session:
        return RedirectResponse("/synchronizace", status_code=302)

    # Field difference condition (name, type, ownership, or share differs)
    # Use coalesce so that NULL vs "value" counts as a difference
    field_diff = or_(
        func.coalesce(SyncRecord.csv_owner_name, '') != func.coalesce(SyncRecord.excel_owner_name, ''),
        func.coalesce(SyncRecord.csv_space_type, '') != func.coalesce(SyncRecord.excel_space_type, ''),
        func.coalesce(SyncRecord.csv_ownership_type, '') != func.coalesce(SyncRecord.excel_ownership_type, ''),
        func.coalesce(SyncRecord.excel_podil_scd, 0) != func.coalesce(SyncRecord.csv_share, 0),
    )

    # Podíl-only difference condition
    podil_diff = (
        func.coalesce(SyncRecord.excel_podil_scd, 0) != func.coalesce(SyncRecord.csv_share, 0)
    )

    base = db.query(SyncRecord).filter_by(session_id=session_id)
    # Partial match = owner matched (MATCH status) but some field differs
    total_partial = base.filter(SyncRecord.status == SyncStatus.MATCH, field_diff).count()
    # Full match = MATCH status and no field differences
    total_full_match = session.total_matches - total_partial
    # Records where podíl specifically differs
    total_podil_diff = base.filter(podil_diff).count()
    # Sum of podíl values for differing records only
    podil_diff_sums = db.query(
        func.sum(SyncRecord.excel_podil_scd),
        func.sum(SyncRecord.csv_share),
    ).filter_by(session_id=session_id).filter(podil_diff).one()
    podil_diff_excel = podil_diff_sums[0] or 0
    podil_diff_csv = podil_diff_sums[1] or 0
    # Sum of podíl values (total)
    sums = db.query(
        func.sum(SyncRecord.excel_podil_scd),
        func.sum(SyncRecord.csv_share),
    ).filter_by(session_id=session_id).one()
    total_excel_podil = sums[0] or 0
    total_csv_podil = sums[1] or 0

    # Sum of podíl per filter category
    def _podil_sums(flt):
        r = db.query(
            func.sum(SyncRecord.excel_podil_scd),
            func.sum(SyncRecord.csv_share),
        ).filter_by(session_id=session_id).filter(flt).one()
        return r[0] or 0, r[1] or 0

    match_excel, match_csv = _podil_sums(and_(SyncRecord.status == SyncStatus.MATCH, ~field_diff))
    partial_excel, partial_csv = _podil_sums(and_(SyncRecord.status == SyncStatus.MATCH, field_diff))
    name_order_excel, name_order_csv = _podil_sums(SyncRecord.status == SyncStatus.NAME_ORDER)
    diff_excel, diff_csv = _podil_sums(SyncRecord.status == SyncStatus.DIFFERENCE)
    missing_excel, missing_csv = _podil_sums(SyncRecord.status.in_([SyncStatus.MISSING_CSV, SyncStatus.MISSING_EXCEL]))

    query = base
    if filtr == "match":
        query = query.filter(SyncRecord.status == SyncStatus.MATCH, ~field_diff)
    elif filtr == "partial":
        query = query.filter(SyncRecord.status == SyncStatus.MATCH, field_diff)
    elif filtr == "name_order":
        query = query.filter(SyncRecord.status == SyncStatus.NAME_ORDER)
    elif filtr == "difference":
        query = query.filter(SyncRecord.status == SyncStatus.DIFFERENCE)
    elif filtr == "podil_diff":
        query = query.filter(podil_diff)
    elif filtr == "missing":
        query = query.filter(SyncRecord.status.in_([SyncStatus.MISSING_CSV, SyncStatus.MISSING_EXCEL]))

    # Sorting
    sort_col = SYNC_SORT_COLUMNS.get(sort)
    if sort_col is not None:
        if order == "desc":
            query = query.order_by(sort_col.desc().nulls_last())
        else:
            query = query.order_by(sort_col.asc().nulls_last())
    else:
        query = query.order_by(cast(SyncRecord.unit_number, Integer).asc())
    records = query.all()

    # Search filtering
    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        records = [
            r for r in records
            if q_lower in str(r.unit_number or "")
            or q_ascii in strip_diacritics(r.excel_owner_name or "")
            or q_ascii in strip_diacritics(r.csv_owner_name or "")
            or q_lower in (r.excel_space_type or "").lower()
            or q_lower in (r.excel_ownership_type or "").lower()
        ]

    # Build unit_number → [(owner_id, owner_name), ...] mapping for clickable owner links
    # and unit_number → unit_id mapping for clickable unit links
    owner_map = {}
    unit_map = {}
    owner_units = (
        db.query(OwnerUnit.owner_id, Unit.unit_number, Owner.name_with_titles, Unit.id)
        .join(OwnerUnit.unit)
        .join(Owner, OwnerUnit.owner_id == Owner.id)
        .filter(OwnerUnit.valid_to.is_(None))
        .all()
    )
    for oid, unit_num, oname, unit_id in owner_units:
        short = str(unit_num)
        owner_map.setdefault(short, []).append((oid, oname))
        unit_map[short] = unit_id

    back_url = back or "/synchronizace"
    back_label = "Zpět na přehled" if back == "/" else "Zpět na kontroly" if "/synchronizace" in (back or "") else "Zpět"

    return templates.TemplateResponse("sync/compare.html", {
        "request": request,
        "active_nav": "kontroly",
        "session": session,
        "records": records,
        "filtr": filtr,
        "sort": sort,
        "order": order,
        "q": q,
        "back_url": back_url,
        "back_label": back_label,
        "total_full_match": total_full_match,
        "total_partial": total_partial,
        "total_podil_diff": total_podil_diff,
        "podil_diff_excel": podil_diff_excel,
        "podil_diff_csv": podil_diff_csv,
        "total_excel_podil": total_excel_podil,
        "total_csv_podil": total_csv_podil,
        "match_excel": match_excel,
        "match_csv": match_csv,
        "partial_excel": partial_excel,
        "partial_csv": partial_csv,
        "name_order_excel": name_order_excel,
        "name_order_csv": name_order_csv,
        "diff_excel": diff_excel,
        "diff_csv": diff_csv,
        "missing_excel": missing_excel,
        "missing_csv": missing_csv,
        "owner_map": owner_map,
        "unit_map": unit_map,
    })


@router.post("/{session_id}/prijmout/{record_id}")
async def accept_change(
    session_id: int,
    record_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    record = db.query(SyncRecord).get(record_id)
    if record:
        record.resolution = SyncResolution.ACCEPTED
        db.commit()

    if request.headers.get("HX-Request"):
        return templates.TemplateResponse("partials/sync_row.html", {
            "request": request,
            "record": record,
        })
    return RedirectResponse(f"/synchronizace/{session_id}", status_code=302)


@router.post("/{session_id}/odmitnout/{record_id}")
async def reject_change(
    session_id: int,
    record_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    record = db.query(SyncRecord).get(record_id)
    if record:
        record.resolution = SyncResolution.REJECTED
        db.commit()

    if request.headers.get("HX-Request"):
        return templates.TemplateResponse("partials/sync_row.html", {
            "request": request,
            "record": record,
        })
    return RedirectResponse(f"/synchronizace/{session_id}", status_code=302)


@router.post("/{session_id}/upravit/{record_id}")
async def manual_edit(
    session_id: int,
    record_id: int,
    corrected_name: str = Form(...),
    request: Request = None,
    db: Session = Depends(get_db),
):
    record = db.query(SyncRecord).get(record_id)
    if record:
        record.admin_corrected_name = corrected_name
        record.resolution = SyncResolution.MANUAL_EDIT
        db.commit()

    if request and request.headers.get("HX-Request"):
        return templates.TemplateResponse("partials/sync_row.html", {
            "request": request,
            "record": record,
        })
    return RedirectResponse(f"/synchronizace/{session_id}", status_code=302)


@router.post("/{session_id}/exportovat")
async def export_excel(
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    form_data = await request.form()
    filtr = form_data.get("filtr", "")

    session = db.query(SyncSession).get(session_id)
    if not session:
        return RedirectResponse("/synchronizace", status_code=302)

    # Apply same filter logic as sync_detail
    field_diff = or_(
        func.coalesce(SyncRecord.csv_owner_name, '') != func.coalesce(SyncRecord.excel_owner_name, ''),
        func.coalesce(SyncRecord.csv_space_type, '') != func.coalesce(SyncRecord.excel_space_type, ''),
        func.coalesce(SyncRecord.csv_ownership_type, '') != func.coalesce(SyncRecord.excel_ownership_type, ''),
        func.coalesce(SyncRecord.excel_podil_scd, 0) != func.coalesce(SyncRecord.csv_share, 0),
    )
    podil_diff = (
        func.coalesce(SyncRecord.excel_podil_scd, 0) != func.coalesce(SyncRecord.csv_share, 0)
    )

    query = db.query(SyncRecord).filter_by(session_id=session_id)
    if filtr == "match":
        query = query.filter(SyncRecord.status == SyncStatus.MATCH, ~field_diff)
    elif filtr == "partial":
        query = query.filter(SyncRecord.status == SyncStatus.MATCH, field_diff)
    elif filtr == "name_order":
        query = query.filter(SyncRecord.status == SyncStatus.NAME_ORDER)
    elif filtr == "difference":
        query = query.filter(SyncRecord.status == SyncStatus.DIFFERENCE)
    elif filtr == "podil_diff":
        query = query.filter(podil_diff)
    elif filtr == "missing":
        query = query.filter(SyncRecord.status.in_([SyncStatus.MISSING_CSV, SyncStatus.MISSING_EXCEL]))

    records = query.order_by(cast(SyncRecord.unit_number, Integer).asc()).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Porovnání"

    headers = [
        "Jednotka", "Vlastník (Evidence)", "Vlastník (CSV)",
        "Typ prostoru (Ev)", "Typ prostoru (CSV)",
        "Vlastnictví (Ev)", "Vlastnictví (CSV)",
        "Podíl (Ev)", "Podíl (CSV)",
        "Shoda", "Stav", "Akce",
    ]
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    _STATUS_LABELS = {
        "match": "Shoda",
        "name_order": "Přeházená jména",
        "difference": "Rozdíl",
        "missing_csv": "Chybí v CSV",
        "missing_excel": "Chybí v evidenci",
    }
    _RESOLUTION_LABELS = {
        "pending": "Čeká",
        "accepted": "OK",
        "rejected": "Zamítnuto",
        "manual_edit": "Upraveno",
        "exchanged": "Výměna",
    }

    diff_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    for r in records:
        match_pct = (r.match_details.split("|")[0].strip() if r.match_details else "")
        row = [
            r.unit_number,
            r.excel_owner_name or "",
            r.csv_owner_name or "",
            r.excel_space_type or "",
            r.csv_space_type or "",
            r.excel_ownership_type or "",
            r.csv_ownership_type or "",
            r.excel_podil_scd,
            r.csv_share,
            match_pct,
            _STATUS_LABELS.get(r.status.value, r.status.value),
            _RESOLUTION_LABELS.get(r.resolution.value, r.resolution.value),
        ]
        ws.append(row)

        # Highlight differing cells
        row_num = ws.max_row
        if (r.excel_owner_name or "") != (r.csv_owner_name or ""):
            ws.cell(row=row_num, column=2).fill = diff_fill
            ws.cell(row=row_num, column=3).fill = diff_fill
        if (r.excel_space_type or "") != (r.csv_space_type or ""):
            ws.cell(row=row_num, column=4).fill = diff_fill
            ws.cell(row=row_num, column=5).fill = diff_fill
        if (r.excel_ownership_type or "") != (r.csv_ownership_type or ""):
            ws.cell(row=row_num, column=6).fill = diff_fill
            ws.cell(row=row_num, column=7).fill = diff_fill
        if (r.excel_podil_scd or 0) != (r.csv_share or 0):
            ws.cell(row=row_num, column=8).fill = diff_fill
            ws.cell(row=row_num, column=9).fill = diff_fill

    excel_auto_width(ws)

    _FILTR_LABELS = {
        "": "vse", "match": "shoda", "partial": "castecna",
        "name_order": "jmena", "difference": "rozdily",
        "podil_diff": "podily", "missing": "chybi",
    }
    suffix = _FILTR_LABELS.get(filtr, filtr or "vse")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"porovnani_{suffix}_{timestamp}.xlsx"
    output_path = settings.generated_dir / "exports" / filename
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return FileResponse(
        str(output_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


ALLOWED_UPDATE_FIELDS = {"ownership_type", "space_type", "podil_scd", "owner_name"}


@router.post("/{session_id}/aktualizovat")
async def apply_selected_updates(
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Apply selected CSV field updates to the database."""
    form = await request.form()

    # Parse form keys: update__{record_id}__{field} → {record_id: {field: value}}
    updates: dict[int, dict[str, str]] = {}
    for key, value in form.items():
        if not key.startswith("update__"):
            continue
        parts = key.split("__")
        if len(parts) != 3:
            continue
        _, record_id_str, field = parts
        if field not in ALLOWED_UPDATE_FIELDS:
            continue
        try:
            rid = int(record_id_str)
        except ValueError:
            continue
        updates.setdefault(rid, {})[field] = value

    session = db.query(SyncSession).get(session_id)
    if not session:
        return RedirectResponse("/synchronizace", status_code=302)

    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    csv_name = session.csv_filename

    success_count = 0
    error_count = 0
    change_details = []

    for record_id, fields in updates.items():
        record = db.query(SyncRecord).filter_by(
            id=record_id, session_id=session_id,
        ).first()
        if not record or not record.unit_number:
            error_count += 1
            continue

        short_num = record.unit_number
        unit = (
            db.query(Unit)
            .filter(
                Unit.unit_number == int(short_num)
            )
            .first()
        )
        if not unit:
            error_count += 1
            change_details.append(f"Jednotka {short_num}: nenalezena")
            continue

        owner_unit = db.query(OwnerUnit).filter_by(unit_id=unit.id).filter(OwnerUnit.valid_to.is_(None)).first()
        if not owner_unit:
            error_count += 1
            change_details.append(f"Jednotka {short_num}: vlastník nenalezen")
            continue

        changes = []
        for field, new_value in fields.items():
            if field == "owner_name":
                # Split CSV name into individual names
                csv_names = re.split(r'\s*[;,]\s*', new_value.strip())
                csv_names = [n.strip() for n in csv_names if n.strip()]

                all_owner_units = db.query(OwnerUnit).filter_by(unit_id=unit.id).filter(OwnerUnit.valid_to.is_(None)).all()
                all_owners = []
                ou_by_owner = {}
                for aou in all_owner_units:
                    o = db.query(Owner).get(aou.owner_id)
                    if o:
                        all_owners.append(o)
                        ou_by_owner[o.id] = aou

                # Match CSV names to existing owners by fuzzy similarity
                used_db = set()
                matched_pairs = []  # (owner, csv_name, ratio)
                unmatched_csv = []

                for cn in csv_names:
                    csv_norm = normalize_for_matching(cn)
                    best_owner, best_ratio = None, -1
                    for o in all_owners:
                        if o.id in used_db:
                            continue
                        r = SequenceMatcher(None, csv_norm, o.name_normalized or "").ratio()
                        if r > best_ratio:
                            best_ratio = r
                            best_owner = o

                    if best_owner and best_ratio >= 0.75:
                        used_db.add(best_owner.id)
                        matched_pairs.append((best_owner, cn, best_ratio))
                    else:
                        unmatched_csv.append(cn)

                # DB owners not matched to any CSV name → soft-delete
                unmatched_db = [o for o in all_owners if o.id not in used_db]

                # Rename matched owners (name correction)
                for owner, matched_name, _ratio in matched_pairs:
                    old_val = owner.name_with_titles
                    if old_val != matched_name:
                        name_parts = matched_name.split(None, 1)
                        if len(name_parts) == 2:
                            owner.last_name = name_parts[0]
                            owner.first_name = name_parts[1]
                        else:
                            owner.first_name = name_parts[0]
                            owner.last_name = None
                        owner.name_with_titles = matched_name
                        owner.name_normalized = normalize_for_matching(matched_name)
                        changes.append(f"jméno: {old_val} → {matched_name}")

                # Hard-delete OwnerUnits for unmatched DB owners (no history)
                for o in unmatched_db:
                    aou = ou_by_owner.get(o.id)
                    if aou:
                        db.delete(aou)
                        changes.append(f"odebrán: {o.name_with_titles}")
                db.flush()

                # Create new Owner + OwnerUnit for unmatched CSV names
                total_votes = unit.podil_scd or 0
                new_count = len(matched_pairs) + len(unmatched_csv)
                votes_each = total_votes // new_count if new_count > 0 else 0

                from app.models import OwnerType
                for cn in unmatched_csv:
                    # Search entire DB for existing owner before creating new
                    cn_simple = strip_diacritics(cn.strip())
                    existing_global = (
                        db.query(Owner)
                        .filter(Owner.name_normalized == cn_simple, Owner.is_active == True)
                        .first()
                    )
                    if not existing_global:
                        # Also try stemmed match (csv_sync owners use normalize_for_matching)
                        cn_stemmed = normalize_for_matching(cn)
                        existing_global = (
                            db.query(Owner)
                            .filter(Owner.name_normalized == cn_stemmed, Owner.is_active == True)
                            .first()
                        )

                    if existing_global:
                        owner = existing_global
                        changes.append(f"přidán (existující): {cn}")
                    else:
                        name_parts = cn.split(None, 1)
                        is_legal = re.search(
                            r'\b(s\.r\.o\.|a\.s\.|spol\.|z\.s\.|v\.o\.s\.)\b',
                            cn, re.IGNORECASE,
                        )
                        owner = Owner(
                            first_name=name_parts[1] if len(name_parts) == 2 else name_parts[0],
                            last_name=name_parts[0] if len(name_parts) == 2 else None,
                            name_with_titles=cn,
                            name_normalized=strip_diacritics(cn.strip()),
                            owner_type=OwnerType.LEGAL_ENTITY if is_legal else OwnerType.PHYSICAL,
                            data_source="csv_sync",
                            is_active=True,
                        )
                        db.add(owner)
                        db.flush()
                        changes.append(f"přidán: {cn}")

                    new_ou = OwnerUnit(
                        owner_id=owner.id,
                        unit_id=unit.id,
                        ownership_type=record.csv_ownership_type or "",
                        share=1.0 / new_count if new_count > 1 else 1.0,
                        votes=votes_each,
                    )
                    db.add(new_ou)

                # Recalculate votes for remaining owners if count changed
                if unmatched_db or unmatched_csv:
                    db.flush()
                    active_ous = db.query(OwnerUnit).filter_by(unit_id=unit.id).filter(OwnerUnit.valid_to.is_(None)).all()
                    if active_ous:
                        base = total_votes // len(active_ous)
                        remainder = total_votes % len(active_ous)
                        for idx, aou in enumerate(active_ous):
                            aou.votes = base + (1 if idx < remainder else 0)
                            aou.share = 1.0 / len(active_ous)

                record.excel_owner_name = new_value.strip()
            elif field == "ownership_type":
                old_val = record.excel_ownership_type or owner_unit.ownership_type or ""
                # Update ALL active OwnerUnits on this unit (not just the first)
                all_ous = db.query(OwnerUnit).filter_by(unit_id=unit.id).filter(OwnerUnit.valid_to.is_(None)).all()
                for aou in all_ous:
                    aou.ownership_type = new_value
                record.excel_ownership_type = new_value
                changes.append(f"vlastnictví: {old_val} → {new_value}")
            elif field == "space_type":
                old_val = record.excel_space_type or unit.space_type or ""
                unit.space_type = new_value
                record.excel_space_type = new_value
                changes.append(f"typ: {old_val} → {new_value}")
            elif field == "podil_scd":
                old_val = record.excel_podil_scd or unit.podil_scd
                unit.podil_scd = float(new_value)
                record.excel_podil_scd = float(new_value)
                from app.services.owner_exchange import recalculate_unit_votes
                recalculate_unit_votes(unit, db)
                changes.append(
                    f"podíl: {old_val} → {new_value}"
                )

        if changes:
            # Each change as separate log entry with source file and timestamp
            note_entries = [f"{ch} (z {csv_name}, {now})" for ch in changes]
            record.admin_note = (record.admin_note + "\n" if record.admin_note else "") + "\n".join(note_entries)
            record.resolution = SyncResolution.MANUAL_EDIT

            # Recalculate status after update — check all fields
            csv_parts = sorted(p.strip() for p in re.split(r'[;,]', record.csv_owner_name or '') if p.strip())
            excel_parts = sorted(p.strip() for p in re.split(r'[;,]', record.excel_owner_name or '') if p.strip())
            names_ok = csv_parts == excel_parts
            type_ok = (record.csv_space_type or '') == (record.excel_space_type or '')
            own_ok = (record.csv_ownership_type or '') == (record.excel_ownership_type or '')
            podil_ok = (record.excel_podil_scd or 0) == (record.csv_share or 0)
            if names_ok:
                record.status = SyncStatus.MATCH
                if type_ok and own_ok and podil_ok:
                    record.match_details = "100%"

            success_count += 1
            change_details.append(f"J. {short_num}: {', '.join(changes)}")

    # Batch log & recalculate session totals
    if updates:
        log = ImportLog(
            filename=f"sync_session_{session_id}",
            file_path=f"sync_update/{session_id}",
            import_type="sync_update",
            rows_total=len(updates),
            rows_imported=success_count,
            rows_skipped=error_count,
            errors="\n".join(change_details) if change_details else None,
        )
        db.add(log)

        # Recalculate session counts from current record statuses
        all_records = db.query(SyncRecord).filter_by(session_id=session_id).all()
        session.total_matches = sum(1 for r in all_records if r.status == SyncStatus.MATCH)
        session.total_name_order = sum(1 for r in all_records if r.status == SyncStatus.NAME_ORDER)
        session.total_differences = sum(1 for r in all_records if r.status == SyncStatus.DIFFERENCE)
        session.total_missing = sum(1 for r in all_records if r.status in (SyncStatus.MISSING_CSV, SyncStatus.MISSING_EXCEL))

    db.commit()
    filtr = form.get("filtr", "")
    url = f"/synchronizace/{session_id}"
    if filtr:
        url += f"?filtr={filtr}"
    return RedirectResponse(url, status_code=302)


def _build_contact_preview(session_id: int, db: Session) -> list[dict]:
    """Build preview of contact transfers from CSV to owners."""
    records = (
        db.query(SyncRecord)
        .filter_by(session_id=session_id)
        .filter(SyncRecord.status == SyncStatus.MATCH)
        .all()
    )

    preview = []
    seen_owners: set[int] = set()
    for record in records:
        if not record.unit_number:
            continue
        if not record.csv_email and not record.csv_phone:
            continue

        short_num = record.unit_number
        owner_units = (
            db.query(OwnerUnit)
            .join(OwnerUnit.unit)
            .filter(Unit.unit_number.endswith(f"/{short_num}") | (Unit.unit_number == short_num))
            .filter(OwnerUnit.valid_to.is_(None))
            .all()
        )
        if not owner_units:
            continue

        for ou in owner_units:
            if ou.owner_id in seen_owners:
                continue
            seen_owners.add(ou.owner_id)

            owner = db.query(Owner).get(ou.owner_id)
            if not owner:
                continue

            will_email = bool(record.csv_email and not owner.email)
            will_phone = bool(record.csv_phone and not owner.phone)

            if not will_email and not will_phone:
                continue

            preview.append({
                "owner_id": owner.id,
                "owner_name": owner.display_name,
                "unit_number": record.unit_number,
                "current_email": owner.email or "",
                "csv_email": record.csv_email or "",
                "will_email": will_email,
                "current_phone": owner.phone or "",
                "csv_phone": record.csv_phone or "",
                "will_phone": will_phone,
            })

    preview.sort(key=lambda x: x["owner_name"])
    return preview


@router.get("/{session_id}/nahled-kontaktu")
async def contacts_preview(
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Preview contact transfers before applying."""
    session = db.query(SyncSession).get(session_id)
    if not session:
        return RedirectResponse("/synchronizace", status_code=302)

    preview = _build_contact_preview(session_id, db)
    back_url = f"/synchronizace/{session_id}"

    return templates.TemplateResponse("sync/contacts_preview.html", {
        "request": request,
        "active_nav": "kontroly",
        "session": session,
        "preview": preview,
        "total_email": sum(1 for p in preview if p["will_email"]),
        "total_phone": sum(1 for p in preview if p["will_phone"]),
        "back_url": back_url,
    })


@router.post("/{session_id}/aplikovat-kontakty")
async def apply_contacts(session_id: int, db: Session = Depends(get_db)):
    """Apply email and phone from CSV to matching owners."""
    records = (
        db.query(SyncRecord)
        .filter_by(session_id=session_id)
        .filter(SyncRecord.status == SyncStatus.MATCH)
        .all()
    )

    updated = 0
    seen_owners: set[int] = set()
    for record in records:
        if not record.unit_number:
            continue
        short_num = record.unit_number
        owner_units = (
            db.query(OwnerUnit)
            .join(OwnerUnit.unit)
            .filter(Unit.unit_number.endswith(f"/{short_num}") | (Unit.unit_number == short_num))
            .filter(OwnerUnit.valid_to.is_(None))
            .all()
        )

        for ou in owner_units:
            if ou.owner_id in seen_owners:
                continue
            seen_owners.add(ou.owner_id)

            owner = db.query(Owner).get(ou.owner_id)
            if not owner:
                continue

            changed = False
            if record.csv_email and not owner.email:
                owner.email = record.csv_email
                changed = True
            if record.csv_phone and not owner.phone:
                owner.phone = record.csv_phone
                changed = True

            if changed:
                owner.updated_at = datetime.utcnow()
                updated += 1

    db.commit()
    return RedirectResponse(f"/synchronizace/{session_id}", status_code=302)


@router.get("/{session_id}/vymena/{record_id}")
async def exchange_preview_single(
    session_id: int,
    record_id: int,
    filtr: str = "",
    request: Request = None,
    db: Session = Depends(get_db),
):
    """Preview owner exchange for a single unit."""
    session = db.query(SyncSession).get(session_id)
    if not session:
        return RedirectResponse("/synchronizace", status_code=302)

    previews = prepare_exchange_preview(db, [record_id])
    if not previews:
        return RedirectResponse(f"/synchronizace/{session_id}", status_code=302)

    stats = _exchange_stats(previews)
    back_url = f"/synchronizace/{session_id}"
    if filtr:
        back_url += f"?filtr={filtr}"
    back_url += f"#sync-{record_id}"

    return templates.TemplateResponse("sync/exchange_preview.html", {
        "request": request,
        "active_nav": "kontroly",
        "session": session,
        "previews": previews,
        "batch": False,
        "record_ids": [record_id],
        "stats": stats,
        "today": date.today().isoformat(),
        "back_url": back_url,
        "filtr": filtr,
    })


@router.post("/{session_id}/vymena/{record_id}/potvrdit")
async def exchange_confirm_single(
    session_id: int,
    record_id: int,
    exchange_date: str = Form(""),
    filtr: str = Form(""),
    db: Session = Depends(get_db),
):
    """Execute owner exchange for a single unit."""
    try:
        ed = date.fromisoformat(exchange_date) if exchange_date else date.today()
    except ValueError:
        ed = date.today()
    execute_exchange(db, [record_id], session_id, exchange_date=ed)
    url = f"/synchronizace/{session_id}"
    if filtr:
        url += f"?filtr={filtr}"
    url += f"#sync-{record_id}"
    return RedirectResponse(url, status_code=302)


@router.post("/{session_id}/vymena-hromadna")
async def exchange_preview_batch(
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Preview batch owner exchange for all DIFFERENCE records."""
    session = db.query(SyncSession).get(session_id)
    if not session:
        return RedirectResponse("/synchronizace", status_code=302)

    form = await request.form()
    filtr = form.get("filtr", "")

    # Get all DIFFERENCE + PENDING records for this session
    records = (
        db.query(SyncRecord)
        .filter_by(session_id=session_id, status=SyncStatus.DIFFERENCE, resolution=SyncResolution.PENDING)
        .order_by(cast(SyncRecord.unit_number, Integer).asc())
        .all()
    )
    record_ids = [r.id for r in records]
    if not record_ids:
        url = f"/synchronizace/{session_id}"
        if filtr:
            url += f"?filtr={filtr}"
        return RedirectResponse(url, status_code=302)

    previews = prepare_exchange_preview(db, record_ids)
    record_ids = [p["record"].id for p in previews]

    stats = _exchange_stats(previews)

    back_url = f"/synchronizace/{session_id}"
    if filtr:
        back_url += f"?filtr={filtr}"

    return templates.TemplateResponse("sync/exchange_preview.html", {
        "request": request,
        "active_nav": "kontroly",
        "session": session,
        "previews": previews,
        "batch": True,
        "record_ids": record_ids,
        "stats": stats,
        "today": date.today().isoformat(),
        "back_url": back_url,
        "filtr": filtr,
    })


@router.post("/{session_id}/vymena-hromadna/potvrdit")
async def exchange_confirm_batch(
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Execute batch owner exchange."""
    form = await request.form()
    raw_ids = form.get("record_ids", "")
    record_ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()]
    exchange_date_str = form.get("exchange_date", "")
    filtr = form.get("filtr", "")
    try:
        ed = date.fromisoformat(exchange_date_str) if exchange_date_str else date.today()
    except ValueError:
        ed = date.today()
    if record_ids:
        execute_exchange(db, record_ids, session_id, exchange_date=ed)
    url = f"/synchronizace/{session_id}"
    if filtr:
        url += f"?filtr={filtr}"
    return RedirectResponse(url, status_code=302)


def _exchange_stats(previews: list[dict]) -> dict:
    """Compute summary statistics for exchange preview."""
    total_units = len(previews)
    total_new = sum(
        1 for p in previews for o in p["new_owners"] if o["match_type"] == "new"
    )
    total_reused = sum(
        1 for p in previews for o in p["new_owners"] if o["match_type"] == "reuse"
    )
    total_possible = sum(
        1 for p in previews for o in p["new_owners"] if o["match_type"] == "possible"
    )
    return {
        "total_units": total_units,
        "total_new": total_new,
        "total_reused": total_reused,
        "total_possible": total_possible,
    }
