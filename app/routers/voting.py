import json
import logging
import shutil
from datetime import datetime, date
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, Query, Request, UploadFile
from fastapi.responses import JSONResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.config import settings
from app.database import get_db
from app.models import (
    Ballot, BallotStatus, BallotVote, Owner, OwnerUnit, SvjInfo, Voting,
    VotingItem, VotingStatus, VoteValue,
    ActivityAction, log_activity,
)
from app.services.word_parser import extract_voting_items, extract_voting_metadata
from app.services.voting_import import (
    read_excel_headers, preview_voting_import, execute_voting_import, validate_mapping,
)
from app.utils import build_list_url, is_htmx_partial, is_safe_path, setup_jinja_filters, strip_diacritics, validate_upload


logger = logging.getLogger(__name__)

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")
setup_jinja_filters(templates)


_VOTING_WIZARD_STEPS = [
    {"label": "Nastavení"},
    {"label": "Generování lístků"},
    {"label": "Zpracování"},
    {"label": "Výsledky"},
    {"label": "Uzavření"},
]


def _has_processed_ballots(voting) -> bool:
    """Check if voting has any processed ballots."""
    return any(b.status == BallotStatus.PROCESSED for b in voting.ballots)


def _voting_wizard(voting, current_step: int = None) -> dict:
    """Build wizard stepper context for voting workflow.
    current_step: 1-based step number for the current page.
                  If None, auto-computed from voting state (for list view).
    """
    status = voting.status
    has_processed = _has_processed_ballots(voting)

    # Auto-compute current_step if not provided (list view)
    if current_step is None:
        has_items = len(voting.items) > 0
        has_ballots = len(voting.ballots) > 0
        all_processed = has_ballots and all(
            b.status == BallotStatus.PROCESSED for b in voting.ballots
        )
        if status == VotingStatus.CLOSED:
            current_step = 5
        elif status == VotingStatus.ACTIVE and all_processed:
            current_step = 5
        elif status == VotingStatus.ACTIVE:
            current_step = 3
        elif status == VotingStatus.DRAFT and has_items:
            current_step = 2
        else:
            current_step = 1

    # Determine max completed step based on voting status
    if status == VotingStatus.CLOSED:
        max_done = 5
    elif status == VotingStatus.ACTIVE:
        max_done = 4 if has_processed else 2
    else:  # draft
        max_done = 1 if voting.items else 0

    steps = []
    for i, s in enumerate(_VOTING_WIZARD_STEPS, 1):
        if i < current_step and i <= max_done:
            step_status = "done"
        elif i == current_step:
            step_status = "done" if i <= max_done else "active"
        elif i <= max_done:
            step_status = "done"
        else:
            step_status = "pending"
        steps.append({"label": s["label"], "status": step_status})

    return {
        "wizard_steps": steps,
        "wizard_current": current_step,
        "wizard_total": len(_VOTING_WIZARD_STEPS),
        "wizard_label": _VOTING_WIZARD_STEPS[current_step - 1]["label"],
    }


def _get_declared_shares(db: Session) -> int:
    """Get total declared shares from SVJ administration settings."""
    svj_info = db.query(SvjInfo).first()
    return svj_info.total_shares if svj_info and svj_info.total_shares else 0


def _ballot_stats(voting, db: Session):
    """Compute ballot statistics for status bubbles."""
    total_ballots = len(voting.ballots)
    status_counts = {s.value: 0 for s in BallotStatus}
    for b in voting.ballots:
        status_counts[b.status.value] += 1
    total_processed_votes = sum(
        b.total_votes for b in voting.ballots
        if b.status == BallotStatus.PROCESSED and any(bv.vote is not None for bv in b.votes)
    )
    declared_shares = _get_declared_shares(db)
    quorum_reached = (
        total_processed_votes / declared_shares >= voting.quorum_threshold
        if declared_shares
        else False
    )
    total_generated_votes = sum(b.total_votes for b in voting.ballots)
    return {
        "total_ballots": total_ballots,
        "status_counts": status_counts,
        "total_processed_votes": total_processed_votes,
        "total_generated_votes": total_generated_votes,
        "declared_shares": declared_shares,
        "quorum_reached": quorum_reached,
    }


@router.get("/")
async def voting_list(
    request: Request,
    back: str = Query("", alias="back"),
    stav: str = Query("", alias="stav"),
    db: Session = Depends(get_db),
):
    q = db.query(Voting).options(
        joinedload(Voting.items),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    )
    if stav:
        q = q.filter(Voting.status == stav)
    votings = q.order_by(Voting.created_at.desc()).all()

    # Count per status (always from all votings, not filtered) — single GROUP BY query
    status_rows = db.query(Voting.status, func.count(Voting.id)).group_by(Voting.status).all()
    status_counts = {"all": sum(cnt for _, cnt in status_rows)}
    for st, cnt in status_rows:
        status_counts[st.value] = cnt

    # Compute stats per voting
    declared_shares = _get_declared_shares(db)
    voting_stats = {}
    for voting in votings:
        total = declared_shares or 1
        processed = [b for b in voting.ballots if b.status == BallotStatus.PROCESSED]
        voted = [b for b in processed if any(bv.vote is not None for bv in b.votes)]
        processed_votes = sum(b.total_votes for b in voted)

        # Per-item results
        item_results = []
        for item in voting.items:
            votes_for = 0
            votes_against = 0
            votes_abstain = 0
            for b in processed:
                for bv in b.votes:
                    if bv.voting_item_id == item.id:
                        if bv.vote == VoteValue.FOR:
                            votes_for += bv.votes_count
                        elif bv.vote == VoteValue.AGAINST:
                            votes_against += bv.votes_count
                        elif bv.vote == VoteValue.ABSTAIN:
                            votes_abstain += bv.votes_count
            item_results.append({
                "item": item,
                "votes_for": votes_for,
                "votes_against": votes_against,
                "votes_abstain": votes_abstain,
                "pct_for": round(votes_for / total * 100, 2) if total else 0,
                "pct_against": round(votes_against / total * 100, 2) if total else 0,
            })

        # Wizard steps (reuse shared helper)
        wizard = _voting_wizard(voting)

        voting_stats[voting.id] = {
            "processed_count": len(processed),
            "processed_votes": processed_votes,
            "quorum_pct": round(processed_votes / declared_shares * 100, 2) if declared_shares else 0,
            "quorum_reached": processed_votes / declared_shares >= voting.quorum_threshold if declared_shares else False,
            "item_results": item_results,
            "wizard_step": wizard["wizard_current"],
            **wizard,
        }

    list_url = build_list_url(request)

    return templates.TemplateResponse("voting/index.html", {
        "request": request,
        "active_nav": "voting",
        "votings": votings,
        "voting_stats": voting_stats,
        "status_counts": status_counts,
        "current_stav": stav,
        "back_url": back,
        "list_url": list_url,
    })


@router.get("/nova")
async def voting_create_page(
    request: Request,
    chyba: str = Query(""),
):
    wizard = {
        "wizard_steps": [
            {"label": s["label"], "status": "active" if i == 0 else "pending"}
            for i, s in enumerate(_VOTING_WIZARD_STEPS)
        ],
        "wizard_current": 1,
        "wizard_total": len(_VOTING_WIZARD_STEPS),
    }
    ctx = {
        "request": request,
        "active_nav": "voting",
        **wizard,
    }
    if chyba == "upload":
        ctx["flash_message"] = "Nahrání šablony selhalo. Ověřte, že soubor je platný .docx a není příliš velký."
        ctx["flash_type"] = "error"
    elif chyba:
        ctx["flash_message"] = chyba
        ctx["flash_type"] = "error"
    return templates.TemplateResponse("voting/create.html", ctx)


@router.post("/nova/nahled-metadat")
async def voting_preview_metadata(
    file: UploadFile = File(...),
):
    """AJAX endpoint: extract metadata from uploaded .docx and return JSON."""
    if not file.filename or not file.filename.endswith(".docx"):
        return JSONResponse({"error": "Nahrajte .docx soubor"}, status_code=400)

    err = await validate_upload(file, max_size_mb=10, allowed_extensions=[".docx"])
    if err:
        return JSONResponse({"error": err}, status_code=400)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    dest = settings.upload_dir / "word_templates" / f"{timestamp}_{file.filename}"
    dest.parent.mkdir(parents=True, exist_ok=True)
    try:
        with open(dest, "wb") as f:
            shutil.copyfileobj(file.file, f)
        meta = extract_voting_metadata(str(dest))
        items = extract_voting_items(str(dest))
        return JSONResponse({
            "meta": meta,
            "items_count": len(items),
        })
    except Exception as e:
        logger.exception("Chyba při extrakci metadat z DOCX šablony")
        return JSONResponse({"error": f"{type(e).__name__}: {e}"}, status_code=500)
    finally:
        # Clean up temp file — the real upload happens on form submit
        try:
            dest.unlink()
        except Exception:
            pass


@router.post("/nova")
async def voting_create(
    request: Request,
    title: str = Form(...),
    description: str = Form(""),
    start_date: str = Form(""),
    end_date: str = Form(""),
    quorum_threshold: float = Form(0.5),
    partial_owner_mode: str = Form("shared"),
    file: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    quorum_threshold = max(0.0, min(100.0, quorum_threshold))
    voting = Voting(
        title=title,
        description=description,
        quorum_threshold=quorum_threshold / 100,
        partial_owner_mode=partial_owner_mode,
    )
    if start_date:
        voting.start_date = date.fromisoformat(start_date)
    if end_date:
        voting.end_date = date.fromisoformat(end_date)

    # Handle Word template upload
    if file and file.filename:
        err = await validate_upload(file, max_size_mb=10, allowed_extensions=[".docx"])
        if err:
            return RedirectResponse("/hlasovani/nova?chyba=upload", status_code=302)

        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        dest = settings.upload_dir / "word_templates" / f"{timestamp}_{file.filename}"
        dest.parent.mkdir(parents=True, exist_ok=True)
        with open(dest, "wb") as f:
            shutil.copyfileobj(file.file, f)
        voting.template_path = str(dest)

        # Extract voting items from template
        extraction_warning = None
        try:
            items = extract_voting_items(str(dest))
            db.add(voting)
            db.flush()
            if items:
                for item_data in items:
                    item = VotingItem(
                        voting_id=voting.id,
                        order=item_data["order"],
                        title=item_data["title"],
                        description=item_data.get("description", ""),
                    )
                    db.add(item)
            else:
                extraction_warning = "sablona-prazdna"
        except Exception:
            logger.exception("Chyba při extrakci bodů z DOCX šablony")
            db.add(voting)
            db.flush()
            extraction_warning = "extrakce-selhala"

        # Extract metadata to pre-fill empty fields
        try:
            meta = extract_voting_metadata(str(dest))
            if not description.strip() and meta.get("description"):
                voting.description = meta["description"]
            if not start_date and meta.get("start_date"):
                voting.start_date = date.fromisoformat(meta["start_date"])
            if not end_date and meta.get("end_date"):
                voting.end_date = date.fromisoformat(meta["end_date"])
        except Exception:
            pass  # metadata extraction is best-effort
    else:
        db.add(voting)
        db.flush()

    # Calculate total votes (only current owner_units)
    total = db.query(func.sum(OwnerUnit.votes)).filter(OwnerUnit.valid_to.is_(None)).scalar() or 0
    voting.total_votes_possible = total

    log_activity(db, ActivityAction.CREATED, "voting", "hlasovani",
                 entity_id=voting.id, entity_name=voting.title)
    db.commit()
    redirect_url = f"/hlasovani/{voting.id}"
    if extraction_warning:
        redirect_url += f"?info={extraction_warning}"
    return RedirectResponse(redirect_url, status_code=302)


@router.get("/{voting_id}")
async def voting_detail(
    voting_id: int,
    request: Request,
    back: str = Query(""),
    q: str = Query(""),
    sort: str = Query("order"),
    order: str = Query("asc"),
    info: str = Query(""),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(
        joinedload(Voting.items),
        joinedload(Voting.ballots).joinedload(Ballot.owner).joinedload(Owner.units).joinedload(OwnerUnit.unit),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    # Snapshot warning: check if ballot total_votes differ from current owner data
    snapshot_warning = ""
    if voting.status == VotingStatus.ACTIVE:
        stale_count = 0
        for ballot in voting.ballots:
            current_votes = sum(ou.votes for ou in ballot.owner.current_units)
            if current_votes != ballot.total_votes:
                stale_count += 1
        if stale_count > 0:
            snapshot_warning = f"U {stale_count} {'lístku se změnil' if stale_count == 1 else 'lístků se změnil'} počet hlasů vlastníka od generování lístků (např. změna podílů nebo vlastníka jednotky)"

    # Calculate results per item
    declared = _get_declared_shares(db) or 1
    results = []
    for item in voting.items:
        votes_for = 0
        votes_against = 0
        votes_abstain = 0
        for ballot in voting.ballots:
            if ballot.status != BallotStatus.PROCESSED:
                continue
            for bv in ballot.votes:
                if bv.voting_item_id == item.id:
                    if bv.vote == VoteValue.FOR:
                        votes_for += bv.votes_count
                    elif bv.vote == VoteValue.AGAINST:
                        votes_against += bv.votes_count
                    elif bv.vote == VoteValue.ABSTAIN:
                        votes_abstain += bv.votes_count

        votes_missing = declared - votes_for - votes_against - votes_abstain
        results.append({
            "item": item,
            "votes_for": votes_for,
            "votes_against": votes_against,
            "votes_abstain": votes_abstain,
            "pct_for": round(votes_for / declared * 100, 2) if declared else 0,
            "pct_against": round(votes_against / declared * 100, 2) if declared else 0,
            "pct_abstain": round(votes_abstain / declared * 100, 2) if declared else 0,
            "votes_missing": votes_missing,
            "pct_missing": round(votes_missing / declared * 100, 2) if declared else 0,
        })

    # Search filter (diacritics-aware)
    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        results = [r for r in results if q_lower in r["item"].title.lower() or q_ascii in strip_diacritics(r["item"].title)]

    # Sort results
    sort_keys = {
        "order": lambda r: r["item"].order,
        "votes_for": lambda r: r["votes_for"],
        "pct_for": lambda r: r["pct_for"],
        "votes_against": lambda r: r["votes_against"],
        "pct_against": lambda r: r["pct_against"],
        "votes_abstain": lambda r: r["votes_abstain"],
        "votes_missing": lambda r: r["votes_missing"],
    }
    key_fn = sort_keys.get(sort, sort_keys["order"])
    results.sort(key=key_fn, reverse=(order == "desc"))

    back_url = back or "/hlasovani"
    back_label = "Zpět na přehled" if back == "/" else "Zpět na hlasování"

    has_processed = _has_processed_ballots(voting)
    if voting.status == VotingStatus.ACTIVE:
        detail_step = 5 if has_processed else 3
    elif voting.status == VotingStatus.CLOSED:
        detail_step = 5
    else:  # draft
        detail_step = 2 if voting.items else 1

    ctx = {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "results": results,
        "back_url": back_url,
        "back_label": back_label,
        "active_bubble": "",
        "show_close_voting": has_processed,
        "snapshot_warning": snapshot_warning,
        "q": q,
        "sort": sort,
        "order": order,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, detail_step),
    }

    # Flash messages from query params
    if info == "extrakce-selhala":
        ctx["flash_message"] = "Body hlasování nebyly extrahovány ze šablony (chyba formátu). Přidejte je ručně."
        ctx["flash_type"] = "warning"
    elif info == "sablona-prazdna":
        ctx["flash_message"] = "V šabloně nebyly nalezeny žádné body hlasování. Přidejte je ručně."
        ctx["flash_type"] = "warning"

    # HTMX partial: return only the results table
    if is_htmx_partial(request):
        return templates.TemplateResponse("voting/detail_results.html", ctx)

    return templates.TemplateResponse("voting/detail.html", ctx)


@router.post("/{voting_id}/generovat")
async def generate_ballots(
    voting_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(joinedload(Voting.items)).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    owners = db.query(Owner).filter_by(is_active=True).options(
        joinedload(Owner.units).joinedload(OwnerUnit.unit)
    ).all()

    created = 0
    processed_owner_ids = set()

    if voting.partial_owner_mode == "shared":
        # Group SJM co-owners: find owners sharing units with SJM ownership
        # Build unit→owners map for SJM units
        unit_sjm_owners = {}  # unit_id → [owner, ...]
        for owner in owners:
            for ou in owner.current_units:
                if ou.ownership_type and "SJM" in ou.ownership_type.upper():
                    unit_sjm_owners.setdefault(ou.unit_id, []).append(owner)

        # Build SJM groups — only pair owners on units with exactly 2 SJM owners
        # Units with >2 SJM owners (e.g. two couples co-owning) don't define pairs
        owner_to_group = {}
        group_id = 0
        for unit_id, group_owners in unit_sjm_owners.items():
            if len(group_owners) != 2:
                continue  # Skip multi-owner SJM units for pairing
            # Same connected components logic, but only for pair units
            existing_groups = {owner_to_group[o.id] for o in group_owners if o.id in owner_to_group}
            if existing_groups:
                target = min(existing_groups)
                for o in group_owners:
                    owner_to_group[o.id] = target
                for oid, gid in list(owner_to_group.items()):
                    if gid in existing_groups and gid != target:
                        owner_to_group[oid] = target
            else:
                for o in group_owners:
                    owner_to_group[o.id] = group_id
                group_id += 1

        # Handle SJM owners who only appear on multi-owner units
        # Group by identical SJM unit sets (likely married couples)
        unpaired_sjm = {}  # owner_id → owner
        for unit_id, group_owners in unit_sjm_owners.items():
            for owner in group_owners:
                if owner.id not in owner_to_group:
                    unpaired_sjm[owner.id] = owner

        # Build SJM unit set for each unpaired owner
        set_to_owners = {}  # frozenset(unit_ids) → [owner_id, ...]
        for oid, owner in unpaired_sjm.items():
            sjm_units = frozenset(
                ou.unit_id for ou in owner.current_units
                if ou.ownership_type and "SJM" in ou.ownership_type.upper()
            )
            set_to_owners.setdefault(sjm_units, []).append(oid)

        # Pair owners with identical SJM unit sets (exactly 2 = couple)
        for sjm_set, oids in set_to_owners.items():
            if len(oids) == 2:
                for oid in oids:
                    owner_to_group[oid] = group_id
                group_id += 1
            else:
                for oid in oids:
                    owner_to_group[oid] = group_id
                    group_id += 1

        # Invert: group_id → [owners]
        groups = {}
        for oid, gid in owner_to_group.items():
            groups.setdefault(gid, set()).add(oid)

        owner_by_id = {o.id: o for o in owners}

        # Create shared ballots for SJM groups
        for gid, member_ids in groups.items():
            members = sorted(
                [owner_by_id[oid] for oid in member_ids if oid in owner_by_id],
                key=lambda o: o.name_normalized or "",
            )
            if not members:
                continue
            primary = members[0]

            # Skip if ballot already exists for primary
            existing = db.query(Ballot).filter_by(
                voting_id=voting.id, owner_id=primary.id
            ).first()
            if existing:
                processed_owner_ids.update(member_ids)
                continue

            # Sum ALL members' votes (each owner has their own share)
            # Dedup only unit numbers for display
            seen_unit_ids = set()
            total_votes = 0
            unit_numbers = []
            for member in members:
                for ou in member.current_units:
                    total_votes += ou.votes
                    if ou.unit_id not in seen_unit_ids:
                        seen_unit_ids.add(ou.unit_id)
                        unit_numbers.append(str(ou.unit.unit_number))

            shared_names = ", ".join(m.display_name for m in members)

            ballot = Ballot(
                voting_id=voting.id,
                owner_id=primary.id,
                total_votes=total_votes,
                units_text=", ".join(unit_numbers),
                shared_owners_text=shared_names,
                status=BallotStatus.GENERATED,
            )
            db.add(ballot)
            db.flush()

            for item in voting.items:
                bv = BallotVote(
                    ballot_id=ballot.id,
                    voting_item_id=item.id,
                    votes_count=total_votes,
                )
                db.add(bv)

            processed_owner_ids.update(member_ids)
            created += 1

    # Create individual ballots for remaining owners (non-SJM or separate mode)
    for owner in owners:
        if owner.id in processed_owner_ids:
            continue

        if not owner.current_units:
            continue  # Skip owners without any units

        existing = db.query(Ballot).filter_by(
            voting_id=voting.id, owner_id=owner.id
        ).first()
        if existing:
            continue

        total_votes = sum(ou.votes for ou in owner.current_units)
        units_text = ", ".join(
            str(ou.unit.unit_number) for ou in owner.current_units
        )

        ballot = Ballot(
            voting_id=voting.id,
            owner_id=owner.id,
            total_votes=total_votes,
            units_text=units_text,
            status=BallotStatus.GENERATED,
        )

        db.add(ballot)
        db.flush()

        for item in voting.items:
            bv = BallotVote(
                ballot_id=ballot.id,
                voting_item_id=item.id,
                votes_count=total_votes,
            )
            db.add(bv)

        created += 1

    if voting.status == VotingStatus.DRAFT:
        voting.status = VotingStatus.ACTIVE

    log_activity(db, ActivityAction.STATUS_CHANGED, "voting", "hlasovani",
                 entity_id=voting.id, entity_name=voting.title,
                 description=f"Vygenerováno {created} lístků, stav → aktivní")
    db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)


@router.get("/{voting_id}/listky")
async def ballot_list(
    voting_id: int,
    request: Request,
    stav: str = Query(""),
    q: str = Query(""),
    sort: str = Query("owner"),
    order: str = Query("asc"),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(
        joinedload(Voting.items),
        joinedload(Voting.ballots).joinedload(Ballot.owner).joinedload(Owner.units).joinedload(OwnerUnit.unit),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    # Filter by status
    ballots = list(voting.ballots)
    if stav:
        try:
            stav_enum = BallotStatus(stav)
            ballots = [b for b in ballots if b.status == stav_enum]
        except ValueError:
            pass  # neplatný stav — ignorovat filtr

    # Search filter (diacritics-aware)
    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        ballots = [
            b for b in ballots
            if q_lower in (b.owner.display_name or "").lower()
            or q_ascii in strip_diacritics(b.owner.display_name or "")
            or q_lower in (b.units_text or "").lower()
        ]

    # Sort
    sort_keys = {
        "owner": lambda b: (b.owner.name_normalized or "").lower(),
        "units": lambda b: b.units_text or "",
        "votes": lambda b: b.total_votes,
        "status": lambda b: b.status.value if b.status else "",
        "proxy": lambda b: (b.proxy_holder_name or ""),
    }
    # Dynamic sort by voting item vote (e.g. sort=bod_3 for item id=3)
    if sort.startswith("bod_"):
        item_id = int(sort[4:])
        vote_order = {"for": 0, "against": 1, "abstain": 2}
        sort_keys[sort] = lambda b, _iid=item_id: next(
            (vote_order.get(bv.vote.value, 3) for bv in b.votes if bv.voting_item_id == _iid and bv.vote),
            4,
        )
    key_fn = sort_keys.get(sort, sort_keys["owner"])
    ballots = sorted(ballots, key=key_fn, reverse=(order == "desc"))

    list_url = build_list_url(request)

    has_processed = _has_processed_ballots(voting)
    ctx = {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "ballots": ballots,
        "current_stav": stav,
        "active_bubble": stav or "all",
        "show_close_voting": has_processed,
        "q": q,
        "sort": sort,
        "order": order,
        "list_url": list_url,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, 4 if has_processed else 3),
    }

    # HTMX partial: return only the table
    if is_htmx_partial(request):
        return templates.TemplateResponse("voting/ballots_table.html", ctx)

    return templates.TemplateResponse("voting/ballots.html", ctx)


@router.get("/{voting_id}/listek/{ballot_id}")
async def ballot_detail(
    voting_id: int,
    ballot_id: int,
    request: Request,
    back: str = Query(""),
    db: Session = Depends(get_db),
):
    ballot = db.query(Ballot).options(
        joinedload(Ballot.owner).joinedload(Owner.units).joinedload(OwnerUnit.unit),
        joinedload(Ballot.votes).joinedload(BallotVote.voting_item),
        joinedload(Ballot.voting).joinedload(Voting.items),
    ).filter_by(id=ballot_id, voting_id=voting_id).first()
    if not ballot:
        return RedirectResponse(f"/hlasovani/{voting_id}/listky", status_code=302)

    # Load full voting with ballots for stats/wizard
    voting = db.query(Voting).options(
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)

    back_url = back or f"/hlasovani/{voting_id}/listky"
    back_label = "Zpět na hlasovací lístky"
    has_processed = _has_processed_ballots(voting)

    ctx = {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "ballot": ballot,
        "back_url": back_url,
        "back_label": back_label,
        "active_bubble": "",
        "show_close_voting": has_processed,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, 3),
    }
    return templates.TemplateResponse("voting/ballot_detail.html", ctx)


@router.get("/{voting_id}/zpracovani")
async def process_page(
    voting_id: int,
    request: Request,
    q: str = Query(""),
    sort: str = Query("owner"),
    order: str = Query("asc"),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(
        joinedload(Voting.items),
        joinedload(Voting.ballots).joinedload(Ballot.owner).joinedload(Owner.units).joinedload(OwnerUnit.unit),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    # Direct DB query: ballot IDs that already have any vote recorded
    ballots_with_votes = set(
        row[0] for row in db.query(BallotVote.ballot_id)
        .filter(BallotVote.vote.isnot(None))
        .filter(BallotVote.ballot_id.in_([b.id for b in voting.ballots]))
        .all()
    )

    # Auto-fix: mark ballots with votes but wrong status as PROCESSED
    fixed = False
    for b in voting.ballots:
        if b.id in ballots_with_votes and b.status in (
            BallotStatus.GENERATED, BallotStatus.SENT, BallotStatus.RECEIVED,
        ):
            b.status = BallotStatus.PROCESSED
            b.processed_at = b.processed_at or datetime.utcnow()
            fixed = True
    if fixed:
        db.commit()

    # Get ballots that need processing (not processed, no votes)
    unprocessed = [
        b for b in voting.ballots
        if b.status in (BallotStatus.GENERATED, BallotStatus.SENT, BallotStatus.RECEIVED)
        and b.id not in ballots_with_votes
    ]

    # Hide ballots for inactive owners with no current units
    unprocessed = [
        b for b in unprocessed
        if not (not b.owner.is_active and not b.owner.current_units)
    ]

    # Search filter (diacritics-aware)
    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        unprocessed = [
            b for b in unprocessed
            if q_lower in (b.owner.display_name or "").lower()
            or q_ascii in strip_diacritics(b.owner.display_name or "")
            or q_lower in (b.units_text or "").lower()
        ]

    # Sort
    sort_keys = {
        "owner": lambda b: (b.owner.name_normalized or "").lower(),
        "units": lambda b: b.units_text or "",
        "votes": lambda b: b.total_votes,
    }
    key_fn = sort_keys.get(sort, sort_keys["owner"])
    unprocessed.sort(key=key_fn, reverse=(order == "desc"))

    has_processed = _has_processed_ballots(voting)
    ctx = {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "unprocessed": unprocessed,
        "active_bubble": "",
        "q": q,
        "sort": sort,
        "order": order,
        "show_close_voting": has_processed,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, 3),
    }

    if is_htmx_partial(request):
        return templates.TemplateResponse("voting/process_cards.html", ctx)

    return templates.TemplateResponse("voting/process.html", ctx)


@router.post("/{voting_id}/zpracovat/{ballot_id}")
async def process_ballot(
    voting_id: int,
    ballot_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    form_data = await request.form()
    ballot = db.query(Ballot).options(joinedload(Ballot.votes)).get(ballot_id)
    if not ballot or ballot.voting_id != voting_id:
        return RedirectResponse(f"/hlasovani/{voting_id}/zpracovani", status_code=302)

    # Collect votes and validate at least one is set
    has_any_vote = False
    for bv in ballot.votes:
        vote_key = f"vote_{bv.voting_item_id}"
        vote_value = form_data.get(vote_key)
        if vote_value:
            bv.vote = VoteValue(vote_value)
            bv.manually_verified = True
            has_any_vote = True

    if not has_any_vote:
        db.rollback()
        if request.headers.get("HX-Request"):
            voting = db.query(Voting).options(joinedload(Voting.items)).get(voting_id)
            return templates.TemplateResponse("partials/ballot_vote_error.html", {
                "request": request,
                "ballot": ballot,
                "voting": voting,
                "error": "Vyberte hlas alespoň u jednoho bodu.",
            })
        return RedirectResponse(f"/hlasovani/{voting_id}/zpracovani", status_code=302)

    ballot.status = BallotStatus.PROCESSED
    ballot.processed_at = datetime.utcnow()
    db.commit()

    if request.headers.get("HX-Request"):
        return templates.TemplateResponse("partials/ballot_processed.html", {
            "request": request,
            "ballot": ballot,
        })
    # If submitted from ballot detail, redirect there; otherwise to bulk processing
    referer = request.headers.get("referer", "")
    if f"/listek/{ballot_id}" in referer:
        return RedirectResponse(f"/hlasovani/{voting_id}/listek/{ballot_id}", status_code=302)
    return RedirectResponse(f"/hlasovani/{voting_id}/zpracovani", status_code=302)


@router.post("/{voting_id}/zpracovat-hromadne")
async def process_ballots_bulk(
    voting_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Batch process multiple ballots with the same votes."""
    form_data = await request.form()
    ballot_ids_raw = form_data.get("ballot_ids", "")
    ballot_ids = [int(x) for x in ballot_ids_raw.split(",") if x.strip()]

    if not ballot_ids:
        return RedirectResponse(f"/hlasovani/{voting_id}/zpracovani", status_code=302)

    voting = db.query(Voting).options(joinedload(Voting.items)).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    ballots = (
        db.query(Ballot).options(joinedload(Ballot.votes))
        .filter(Ballot.id.in_(ballot_ids), Ballot.voting_id == voting_id)
        .all()
    )

    # Check that at least one vote is selected in bulk form
    has_any_vote = any(form_data.get(f"vote_{item.id}") for item in voting.items)
    if not has_any_vote:
        return RedirectResponse(f"/hlasovani/{voting_id}/zpracovani", status_code=302)

    count = 0
    for ballot in ballots:
        for bv in ballot.votes:
            vote_key = f"vote_{bv.voting_item_id}"
            vote_value = form_data.get(vote_key)
            if vote_value:
                bv.vote = VoteValue(vote_value)
                bv.manually_verified = True
        ballot.status = BallotStatus.PROCESSED
        ballot.processed_at = datetime.utcnow()
        count += 1

    db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}/zpracovani", status_code=302)


@router.post("/{voting_id}/stav")
async def update_voting_status(
    voting_id: int,
    status: str = Form(...),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).get(voting_id)
    if voting:
        old = voting.status
        # Allowed transitions: active→closed, closed→active
        allowed = {
            (VotingStatus.ACTIVE, VotingStatus.CLOSED),
            (VotingStatus.CLOSED, VotingStatus.ACTIVE),
        }
        try:
            new_status = VotingStatus(status)
        except ValueError:
            return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)
        if (old, new_status) not in allowed:
            return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)
        voting.status = new_status
        voting.updated_at = datetime.utcnow()
        log_activity(db, ActivityAction.STATUS_CHANGED, "voting", "hlasovani",
                     entity_id=voting.id, entity_name=voting.title,
                     description=f"Stav: {old} → {status}")
        db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)


@router.get("/{voting_id}/exportovat")
async def voting_export(voting_id: int, db: Session = Depends(get_db)):
    """Export voting results to Excel."""
    from io import BytesIO
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    voting = db.query(Voting).options(
        joinedload(Voting.items),
        joinedload(Voting.ballots).joinedload(Ballot.owner),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    declared = _get_declared_shares(db) or 1
    items = sorted(voting.items, key=lambda i: i.order)
    processed = [b for b in voting.ballots if b.status == BallotStatus.PROCESSED]
    processed.sort(key=lambda b: (b.owner.name_normalized or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Výsledky hlasování"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    # Header row
    headers = ["Vlastník", "Jednotky", "Hlasy"]
    for item in items:
        headers.append(f"Bod {item.order}: {item.title}")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill

    # Data rows
    vote_labels = {"for": "PRO", "against": "PROTI", "abstain": "Zdržel se", "invalid": "Neplatný"}
    for row_idx, ballot in enumerate(processed, 2):
        ws.cell(row=row_idx, column=1, value=ballot.owner.display_name)
        ws.cell(row=row_idx, column=2, value=ballot.units_text or "")
        ws.cell(row=row_idx, column=3, value=ballot.total_votes)
        for item_idx, item in enumerate(items):
            bv = next((v for v in ballot.votes if v.voting_item_id == item.id), None)
            if bv and bv.vote:
                cell = ws.cell(row=row_idx, column=4 + item_idx, value=vote_labels.get(bv.vote.value, ""))
                if bv.vote == VoteValue.FOR:
                    cell.fill = green_fill
                elif bv.vote == VoteValue.AGAINST:
                    cell.fill = red_fill
            else:
                ws.cell(row=row_idx, column=4 + item_idx, value="—")

    # Summary row
    summary_row = len(processed) + 3
    ws.cell(row=summary_row, column=1, value="CELKEM").font = bold
    for item_idx, item in enumerate(items):
        votes_for = sum(
            bv.votes_count for b in processed for bv in b.votes
            if bv.voting_item_id == item.id and bv.vote == VoteValue.FOR
        )
        votes_against = sum(
            bv.votes_count for b in processed for bv in b.votes
            if bv.voting_item_id == item.id and bv.vote == VoteValue.AGAINST
        )
        pct_for = round(votes_for / declared * 100, 2) if declared else 0
        pct_against = round(votes_against / declared * 100, 2) if declared else 0
        cell = ws.cell(row=summary_row, column=4 + item_idx,
                       value=f"PRO: {votes_for} ({pct_for}%) | PROTI: {votes_against} ({pct_against}%)")
        cell.font = bold

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    buf = BytesIO()
    wb.save(buf)
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"hlasovani_{voting.id}_{timestamp}.xlsx"

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{voting_id}/smazat")
async def voting_delete(voting_id: int, db: Session = Depends(get_db)):
    """Delete a voting, its items, ballots, and associated files."""
    voting = db.query(Voting).options(joinedload(Voting.ballots)).get(voting_id)
    if voting:
        # Remove template file
        if voting.template_path:
            try:
                p = Path(voting.template_path)
                if p.exists():
                    p.unlink()
            except Exception:
                pass
        # Remove ballot PDF/scan files
        for ballot in voting.ballots:
            for attr in ("pdf_path", "scan_path"):
                fpath = getattr(ballot, attr, None)
                if fpath:
                    try:
                        p = Path(fpath)
                        if p.exists():
                            p.unlink()
                    except Exception:
                        pass
        # Cascade deletes VotingItem, Ballot, BallotVote
        log_activity(db, ActivityAction.DELETED, "voting", "hlasovani",
                     entity_id=voting.id, entity_name=voting.title)
        db.delete(voting)
        db.commit()
    return RedirectResponse("/hlasovani", status_code=302)


@router.post("/{voting_id}/smazat-bod/{item_id}")
async def delete_voting_item(
    voting_id: int,
    item_id: int,
    db: Session = Depends(get_db),
):
    item = db.query(VotingItem).filter_by(id=item_id, voting_id=voting_id).first()
    if item:
        db.delete(item)
        db.flush()
        # Renumber remaining items 1..N
        remaining = (
            db.query(VotingItem)
            .filter_by(voting_id=voting_id)
            .order_by(VotingItem.order)
            .all()
        )
        for idx, it in enumerate(remaining, 1):
            it.order = idx
        db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)


@router.post("/{voting_id}/pridat-bod")
async def add_voting_item(
    voting_id: int,
    title: str = Form(...),
    description: str = Form(""),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(joinedload(Voting.items)).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    max_order = max((i.order for i in voting.items), default=0)
    item = VotingItem(
        voting_id=voting.id,
        order=max_order + 1,
        title=title,
        description=description,
    )
    db.add(item)
    db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)


@router.post("/{voting_id}/bod/{item_id}/upravit")
async def edit_voting_item(
    voting_id: int,
    item_id: int,
    title: str = Form(...),
    description: str = Form(""),
    db: Session = Depends(get_db),
):
    """Edit title/description of a voting item (draft only)."""
    item = db.query(VotingItem).filter_by(id=item_id, voting_id=voting_id).first()
    if not item:
        return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)
    voting = db.query(Voting).get(voting_id)
    if not voting or voting.status != VotingStatus.DRAFT:
        return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)
    item.title = title.strip()
    item.description = description.strip()
    db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)


@router.post("/{voting_id}/bod/{item_id}/posunout")
async def move_voting_item(
    voting_id: int,
    item_id: int,
    direction: str = Form(...),
    db: Session = Depends(get_db),
):
    """Move a voting item up or down (draft only)."""
    voting = db.query(Voting).options(joinedload(Voting.items)).get(voting_id)
    if not voting or voting.status != VotingStatus.DRAFT:
        return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)

    items_sorted = sorted(voting.items, key=lambda i: i.order)
    target = next((i for i in items_sorted if i.id == item_id), None)
    if not target:
        return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)

    idx = items_sorted.index(target)
    if direction == "up" and idx > 0:
        neighbor = items_sorted[idx - 1]
        target.order, neighbor.order = neighbor.order, target.order
    elif direction == "down" and idx < len(items_sorted) - 1:
        neighbor = items_sorted[idx + 1]
        target.order, neighbor.order = neighbor.order, target.order

    db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)


@router.post("/{voting_id}/listek/{ballot_id}/opravit")
async def reset_ballot(
    voting_id: int,
    ballot_id: int,
    db: Session = Depends(get_db),
):
    """Reset a processed ballot back to GENERATED so it can be re-processed."""
    ballot = db.query(Ballot).options(
        joinedload(Ballot.votes), joinedload(Ballot.voting),
    ).filter_by(id=ballot_id, voting_id=voting_id).first()
    if not ballot:
        return RedirectResponse(f"/hlasovani/{voting_id}/listky", status_code=302)

    # Only allow reset if voting is active and ballot is processed
    if ballot.voting.status != VotingStatus.ACTIVE or ballot.status != BallotStatus.PROCESSED:
        return RedirectResponse(f"/hlasovani/{voting_id}/listek/{ballot_id}", status_code=302)

    ballot.status = BallotStatus.GENERATED
    ballot.processed_at = None
    for bv in ballot.votes:
        bv.vote = None
        bv.manually_verified = False
    db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}/listek/{ballot_id}", status_code=302)


@router.get("/{voting_id}/neodevzdane")
async def not_submitted(
    voting_id: int,
    request: Request,
    q: str = Query(""),
    sort: str = Query("owner"),
    order: str = Query("asc"),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(
        joinedload(Voting.ballots).joinedload(Ballot.owner).joinedload(Owner.units).joinedload(OwnerUnit.unit),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    missing = [
        b for b in voting.ballots
        if b.status not in (BallotStatus.PROCESSED,)
    ]

    # Search filter
    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        missing = [
            b for b in missing
            if q_lower in (b.owner.display_name or "").lower()
            or q_ascii in strip_diacritics(b.owner.display_name or "")
            or q_lower in (b.units_text or "").lower()
            or q_lower in (b.owner.email or "").lower()
        ]

    # Sorting
    SORT_KEYS = {
        "owner": lambda b: strip_diacritics(b.owner.display_name or ""),
        "units": lambda b: (b.units_text or "").lower(),
        "email": lambda b: (b.owner.email or "").lower(),
        "votes": lambda b: b.total_votes or 0,
        "status": lambda b: b.status.value,
    }
    sort_fn = SORT_KEYS.get(sort, SORT_KEYS["owner"])
    missing.sort(key=sort_fn, reverse=(order == "desc"))

    list_url = build_list_url(request)

    has_processed = _has_processed_ballots(voting)
    ctx = {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "missing": missing,
        "active_bubble": "neodevzdane",
        "show_close_voting": has_processed,
        "q": q,
        "sort": sort,
        "order": order,
        "list_url": list_url,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, 4 if has_processed else 3),
    }

    if is_htmx_partial(request):
        return templates.TemplateResponse("voting/not_submitted_table.html", ctx)

    return templates.TemplateResponse("voting/not_submitted.html", ctx)


# --- Import voting results from Excel ---


def _load_saved_mapping(voting: Voting, db: Session) -> Optional[dict]:
    """Load saved import mapping: per-voting first, then global fallback with item_id remapping."""
    # 1. Per-voting mapping (exact match)
    if voting.import_column_mapping:
        try:
            return json.loads(voting.import_column_mapping)
        except (json.JSONDecodeError, TypeError):
            pass

    # 2. Global fallback from SvjInfo
    svj = db.query(SvjInfo).first()
    if not svj or not svj.voting_import_mapping:
        return None

    try:
        mapping = json.loads(svj.voting_import_mapping)
    except (json.JSONDecodeError, TypeError):
        return None

    # Remap item_ids by position: global mapping has item_ids from previous voting
    items_sorted = sorted(voting.items, key=lambda i: i.order)
    old_items = mapping.get("item_mappings") or mapping.get("items", [])
    new_items = []
    for idx, im in enumerate(old_items):
        if idx < len(items_sorted):
            new_items.append({
                **im,
                "item_id": items_sorted[idx].id,
            })
    mapping["item_mappings"] = new_items
    if "items" in mapping:
        del mapping["items"]

    return mapping


def _save_mapping_global(mapping_json: str, db: Session):
    """Save mapping globally to SvjInfo for reuse across votings."""
    svj = db.query(SvjInfo).first()
    if svj:
        svj.voting_import_mapping = mapping_json
        db.commit()


@router.get("/{voting_id}/import")
async def import_upload_page(
    voting_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(
        joinedload(Voting.items),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    saved_mapping = _load_saved_mapping(voting, db)

    has_processed = _has_processed_ballots(voting)
    return templates.TemplateResponse("voting/import_upload.html", {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "saved_mapping": saved_mapping,
        "active_bubble": "",
        "show_close_voting": has_processed,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, 3),
    })


@router.post("/{voting_id}/import")
async def import_upload(
    voting_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(
        joinedload(Voting.items),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    has_processed = _has_processed_ballots(voting)
    err = await validate_upload(file, max_size_mb=50, allowed_extensions=[".xlsx", ".xls"]) if file.filename else "Nahrajte soubor ve formátu .xlsx"
    if err:
        return templates.TemplateResponse("voting/import_upload.html", {
            "request": request,
            "active_nav": "voting",
            "voting": voting,
            "saved_mapping": None,
            "active_bubble": "",
            "flash_message": err,
            "flash_type": "error",
            "show_close_voting": has_processed,
            **_ballot_stats(voting, db),
            **_voting_wizard(voting, 3),
        })

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    dest = settings.upload_dir / "excel" / f"{timestamp}_{file.filename}"
    dest.parent.mkdir(parents=True, exist_ok=True)
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)

    headers = read_excel_headers(str(dest))

    # Load saved mapping if available (per-voting or global fallback)
    saved_mapping = _load_saved_mapping(voting, db)

    return templates.TemplateResponse("voting/import_mapping.html", {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "headers": headers,
        "file_path": str(dest),
        "filename": file.filename,
        "saved_mapping": saved_mapping,
        "active_bubble": "",
        "show_close_voting": has_processed,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, 3),
    })


@router.post("/{voting_id}/import/nahled")
async def import_preview(
    voting_id: int,
    request: Request,
    file_path: str = Form(...),
    mapping_json: str = Form(...),
    save_mapping: str = Form(""),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(
        joinedload(Voting.items),
        joinedload(Voting.ballots).joinedload(Ballot.owner).joinedload(Owner.units).joinedload(OwnerUnit.unit),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    if not is_safe_path(Path(file_path), settings.upload_dir):
        return RedirectResponse(f"/hlasovani/{voting_id}/import", status_code=302)

    try:
        mapping = json.loads(mapping_json)
    except (json.JSONDecodeError, TypeError):
        return RedirectResponse(f"/hlasovani/{voting_id}/import", status_code=302)

    mapping_error = validate_mapping(mapping)
    if mapping_error:
        logger.warning("Invalid import mapping: %s", mapping_error)
        return RedirectResponse(f"/hlasovani/{voting_id}/import", status_code=302)

    # Save mapping for next time (only if user checked the box)
    if save_mapping:
        voting.import_column_mapping = mapping_json
        _save_mapping_global(mapping_json, db)

    preview = preview_voting_import(file_path, mapping, voting, db)

    # Build item lookup for template
    item_lookup = {item.id: item for item in voting.items}

    has_processed = _has_processed_ballots(voting)
    return templates.TemplateResponse("voting/import_preview.html", {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "preview": preview,
        "mapping": mapping,
        "mapping_json": mapping_json,
        "file_path": file_path,
        "item_lookup": item_lookup,
        "active_bubble": "",
        "show_close_voting": has_processed,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, 3),
    })


@router.post("/{voting_id}/import/potvrdit")
async def import_confirm(
    voting_id: int,
    request: Request,
    file_path: str = Form(...),
    mapping_json: str = Form(...),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(
        joinedload(Voting.items),
        joinedload(Voting.ballots).joinedload(Ballot.owner).joinedload(Owner.units).joinedload(OwnerUnit.unit),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    if not is_safe_path(Path(file_path), settings.upload_dir):
        return RedirectResponse(f"/hlasovani/{voting_id}/import", status_code=302)

    if not Path(file_path).exists():
        return RedirectResponse(f"/hlasovani/{voting_id}/import", status_code=302)

    try:
        mapping = json.loads(mapping_json)
    except (json.JSONDecodeError, TypeError):
        return RedirectResponse(f"/hlasovani/{voting_id}/import", status_code=302)

    mapping_error = validate_mapping(mapping)
    if mapping_error:
        logger.warning("Invalid import mapping: %s", mapping_error)
        return RedirectResponse(f"/hlasovani/{voting_id}/import", status_code=302)

    result = execute_voting_import(file_path, mapping, voting, db)

    # Save mapping globally for reuse in future votings
    _save_mapping_global(mapping_json, db)

    # Clean up uploaded Excel file after successful import
    try:
        Path(file_path).unlink(missing_ok=True)
    except Exception:
        pass

    log_activity(db, ActivityAction.IMPORTED, "voting", "hlasovani",
                 entity_id=voting.id, entity_name=voting.title,
                 description=f"Import výsledků: {result.get('processed', 0)} lístků")
    db.commit()

    has_processed = _has_processed_ballots(voting)
    return templates.TemplateResponse("voting/import_result.html", {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "result": result,
        "active_bubble": "",
        "show_close_voting": has_processed,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, 4),
    })
