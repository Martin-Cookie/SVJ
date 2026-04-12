from __future__ import annotations

import asyncio
import re
import threading
import time
from typing import Optional


from io import BytesIO

from fastapi import APIRouter, Depends, Form, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy import cast, Integer
from sqlalchemy.orm import Session, joinedload

from app.database import SessionLocal, get_db
from app.models import (
    EmailDeliveryStatus, MatchStatus, Owner, OwnerUnit, SendStatus,
    TaxDistribution, TaxDocument, TaxSession,
    ActivityAction, log_activity,
)
from app.models.smtp_profile import SmtpProfile
from app.services.email_service import create_smtp_connection, send_email
from app.utils import build_list_url, compute_eta, excel_auto_width, strip_diacritics, utcnow

from ._helpers import (
    logger, templates,
    _sending_progress, _sending_lock,
    _tax_wizard, _build_recipients, _find_coowners,
)

router = APIRouter()


def _auto_assign_unmatched_docs(db, session_id, owner_id, email, session, all_docs):
    """Auto-assign unmatched documents for units owned by this owner. Returns True if any assigned.

    all_docs musí mít eager-loaded distributions (joinedload), jinak se použije
    batch query pro načtení distribucí jedním dotazem.
    """
    changed = False
    owner_unit_numbers = {
        str(ou.unit.unit_number)
        for ou in db.query(OwnerUnit)
        .filter_by(owner_id=owner_id)
        .options(joinedload(OwnerUnit.unit))
        .all()
    }

    # Batch query: načíst všechny distribuce pro dokumenty v session jedním dotazem
    doc_ids = [doc.id for doc in all_docs]
    all_dists = (
        db.query(TaxDistribution)
        .filter(TaxDistribution.document_id.in_(doc_ids))
        .all()
    ) if doc_ids else []
    # Indexovat distribuce podle document_id
    dists_by_doc = {}
    for d in all_dists:
        dists_by_doc.setdefault(d.document_id, []).append(d)

    for doc in all_docs:
        if str(doc.unit_number) not in owner_unit_numbers:
            continue
        doc_dists = dists_by_doc.get(doc.id, [])
        if any(d.owner_id == owner_id for d in doc_dists):
            continue
        if not all(d.match_status == MatchStatus.UNMATCHED for d in doc_dists):
            continue
        for d in doc_dists:
            db.delete(d)
        co_owner_ids = _find_coowners(
            owner_id, str(doc.unit_number),
            session.year if session else None, db,
        )
        for oid in co_owner_ids:
            if oid == owner_id:
                dist_email = email or None
            else:
                o = db.query(Owner).get(oid)
                dist_email = o.email if o else None
            db.add(TaxDistribution(
                document_id=doc.id, owner_id=oid,
                match_status=MatchStatus.MANUAL, email_address_used=dist_email,
            ))
            changed = True
    return changed


def _build_single_recipient(db, session_id, dist):
    """Build a single recipient dict for HTMX row rebuild. Returns None if owner not found."""
    if dist.owner_id:
        owner = db.query(Owner).get(dist.owner_id)
        if not owner:
            return None
        relevant_dists = (
            db.query(TaxDistribution)
            .filter(
                TaxDistribution.document_id.in_(
                    db.query(TaxDocument.id).filter_by(session_id=session_id)
                ),
                TaxDistribution.owner_id == dist.owner_id,
            )
            .options(joinedload(TaxDistribution.document))
            .all()
        )
        docs_list, dist_ids, used_email, email_status = [], [], None, "pending"
        for rd in relevant_dists:
            docs_list.append({
                "id": rd.document_id,
                "filename": rd.document.filename or "",
                "file_path": rd.document.file_path,
            })
            dist_ids.append(rd.id)
            if rd.email_address_used:
                used_email = rd.email_address_used
            if rd.email_status:
                email_status = rd.email_status.value if hasattr(rd.email_status, 'value') else rd.email_status
        primary = owner.email or ""
        secondary = owner.email_secondary or ""
        final = used_email or primary or secondary or ""
        return {
            "key": f"owner_{dist.owner_id}",
            "name": owner.display_name,
            "email": final,
            "primary_email": primary,
            "secondary_email": secondary,
            "selected_emails": [e.strip() for e in final.split(",") if e.strip()] if final else [],
            "has_dual_email": bool(primary and secondary and primary != secondary),
            "docs": docs_list, "dist_ids": dist_ids,
            "owner_id": owner.id, "is_external": False, "email_status": email_status,
        }
    else:
        doc = db.query(TaxDocument).get(dist.document_id) if dist.document_id else None
        return {
            "key": f"ext_{dist.id}",
            "name": dist.ad_hoc_name or "Externí příjemce",
            "email": dist.ad_hoc_email or "",
            "primary_email": dist.ad_hoc_email or "",
            "secondary_email": "",
            "selected_emails": [dist.ad_hoc_email] if dist.ad_hoc_email else [],
            "has_dual_email": False,
            "docs": [{"id": doc.id, "filename": doc.filename or "", "file_path": doc.file_path}] if doc else [],
            "dist_ids": [dist.id], "owner_id": None, "is_external": True,
            "email_status": dist.email_status.value if dist.email_status and hasattr(dist.email_status, 'value') else "pending",
        }


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

@router.get("/{session_id}/rozeslat/export/{fmt}")
async def tax_send_export(
    session_id: int,
    fmt: str,
    q: str = Query(""),
    filtr: str = Query(""),
    sort: str = Query("name"),
    order: str = Query("asc"),
    db: Session = Depends(get_db),
):
    """Export filtrovaného seznamu příjemců do Excelu nebo CSV."""
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if fmt not in ("xlsx", "csv"):
        return RedirectResponse(f"/dane/{session_id}/rozeslat", status_code=302)

    session = db.query(TaxSession).get(session_id)
    if not session:
        return RedirectResponse("/dane", status_code=302)

    documents = (
        db.query(TaxDocument)
        .filter_by(session_id=session_id)
        .options(
            joinedload(TaxDocument.distributions)
            .joinedload(TaxDistribution.owner),
        )
        .all()
    )
    all_recipients = _build_recipients(documents)

    # Filter — same logic as send preview
    if filtr == "with_email":
        recipients = [r for r in all_recipients if r["email"]]
    elif filtr == "no_email":
        recipients = [r for r in all_recipients if not r["email"]]
    elif filtr == "pending":
        recipients = [r for r in all_recipients if r["email_status"] in ("pending", "queued")]
    elif filtr == "sent":
        recipients = [r for r in all_recipients if r["email_status"] == "sent"]
    elif filtr == "failed":
        recipients = [r for r in all_recipients if r["email_status"] == "failed"]
    else:
        recipients = all_recipients

    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        recipients = [
            r for r in recipients
            if q_lower in r["name"].lower()
            or q_ascii in strip_diacritics(r["name"])
            or q_lower in (r["email"] or "").lower()
            or any(q_lower in d["filename"].lower() for d in r["docs"])
        ]

    SEND_SORT_KEYS = {
        "name": lambda r: strip_diacritics(r["name"]),
        "email": lambda r: (r["email"] or "").lower(),
        "docs": lambda r: len(r["docs"]),
        "status": lambda r: r["email_status"],
    }
    sort_fn = SEND_SORT_KEYS.get(sort, SEND_SORT_KEYS["name"])
    recipients.sort(key=sort_fn, reverse=(order == "desc"))

    # Build rows
    headers = ["Prijemce", "Email", "Dokumenty", "Stav"]
    status_labels = {"pending": "Ceka", "queued": "Ve fronte", "sent": "Odeslano", "failed": "Chyba"}
    rows = []
    for r in recipients:
        rows.append([
            r["name"],
            r["email"] or "",
            ", ".join(d["filename"] for d in r["docs"]),
            status_labels.get(r["email_status"], r["email_status"]),
        ])

    # Filename with filter suffix
    timestamp = datetime.now().strftime("%Y%m%d")
    suffix_map = {
        "with_email": "s_emailem", "no_email": "bez_emailu",
        "pending": "cekajici", "sent": "odeslano", "failed": "chyba",
    }
    suffix = f"_{suffix_map[filtr]}" if filtr in suffix_map else "_vsichni"
    filename = f"rozeslani_{session_id}{suffix}_{timestamp}.{fmt}"

    if fmt == "csv":
        import csv
        import io
        buf = io.StringIO()
        buf.write("\ufeff")  # BOM for Excel
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(headers)
        writer.writerows(rows)
        from fastapi.responses import Response
        return Response(
            content=buf.getvalue().encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Rozeslani"
    bold = Font(bold=True)

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = bold

    for row_idx, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=val)

    excel_auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Send preview endpoints
# ---------------------------------------------------------------------------

@router.get("/{session_id}/rozeslat")
async def tax_send_preview(
    session_id: int,
    request: Request,
    q: str = Query(""),
    filtr: str = Query(""),
    sort: str = Query("name"),
    order: str = Query("asc"),
    back: str = Query("", alias="back"),
    varovani: str = Query(""),
    db: Session = Depends(get_db),
):
    """Náhled rozesílání s příjemci, emaily a nastavením."""
    session = db.query(TaxSession).get(session_id)
    if not session:
        return RedirectResponse("/dane", status_code=302)

    # If sending is in progress, redirect to progress page
    with _sending_lock:
        progress = _sending_progress.get(session_id)
        sending_in_progress = progress and not progress.get("done")
    if sending_in_progress:
        return RedirectResponse(f"/dane/{session_id}/rozeslat/prubeh", status_code=302)

    # Edge case: DB says SENDING but no progress dict (server restart)
    if session.send_status == SendStatus.SENDING and not progress:
        session.send_status = SendStatus.PAUSED
        db.commit()

    # Sdílená nastavení z SvjInfo jako fallback
    from app.models import SvjInfo
    svj = db.query(SvjInfo).first()

    documents = (
        db.query(TaxDocument)
        .filter_by(session_id=session_id)
        .options(
            joinedload(TaxDocument.distributions)
            .joinedload(TaxDistribution.owner),
        )
        .order_by(cast(TaxDocument.unit_number, Integer), TaxDocument.unit_letter)
        .all()
    )

    # Count skipped AUTO_MATCHED distributions
    skipped_auto_count = sum(
        1 for doc in documents
        for dist in doc.distributions
        if dist.match_status == MatchStatus.AUTO_MATCHED
    )

    all_recipients = _build_recipients(documents)
    total_recipients = len(all_recipients)
    with_email = sum(1 for r in all_recipients if r["email"])
    without_email = total_recipients - with_email

    # Count by status for filter pills
    count_pending = sum(1 for r in all_recipients if r["email_status"] in ("pending", "queued"))
    count_sent = sum(1 for r in all_recipients if r["email_status"] == "sent")
    count_failed = sum(1 for r in all_recipients if r["email_status"] == "failed")

    # Auto-fix: all recipients sent but status not COMPLETED
    if total_recipients > 0 and count_sent == total_recipients and session.send_status != SendStatus.COMPLETED:
        session.send_status = SendStatus.COMPLETED
        db.commit()

    # Filter by status
    if filtr == "with_email":
        recipients = [r for r in all_recipients if r["email"]]
    elif filtr == "no_email":
        recipients = [r for r in all_recipients if not r["email"]]
    elif filtr == "pending":
        recipients = [r for r in all_recipients if r["email_status"] in ("pending", "queued")]
    elif filtr == "sent":
        recipients = [r for r in all_recipients if r["email_status"] == "sent"]
    elif filtr == "failed":
        recipients = [r for r in all_recipients if r["email_status"] == "failed"]
    else:
        recipients = all_recipients

    # Search filtering
    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        recipients = [
            r for r in recipients
            if q_lower in r["name"].lower()
            or q_ascii in strip_diacritics(r["name"])
            or q_lower in (r["email"] or "").lower()
            or any(q_lower in d["filename"].lower() for d in r["docs"])
        ]

    # Sorting
    SEND_SORT_KEYS = {
        "name": lambda r: strip_diacritics(r["name"]),
        "email": lambda r: (r["email"] or "").lower(),
        "docs": lambda r: len(r["docs"]),
        "status": lambda r: r["email_status"],
    }
    sort_fn = SEND_SORT_KEYS.get(sort, SEND_SORT_KEYS["name"])
    recipients.sort(key=sort_fn, reverse=(order == "desc"))

    back_url = back or f"/dane/{session_id}"

    list_url = build_list_url(request)

    smtp_profiles = db.query(SmtpProfile).order_by(SmtpProfile.is_default.desc(), SmtpProfile.id).all()

    ctx = {
        "request": request,
        "active_nav": "tax",
        "session": session,
        "smtp_profiles": smtp_profiles,
        "recipients": recipients,
        "total_recipients": total_recipients,
        "with_email": with_email,
        "without_email": without_email,
        "count_pending": count_pending,
        "count_sent": count_sent,
        "count_failed": count_failed,
        "back_url": back_url,
        "list_url": list_url,
        "q": q,
        "filtr": filtr,
        "sort": sort,
        "order": order,
        "test_email_value": session.test_email_address or (svj.send_test_email_address if svj else "") or "",
        "all_documents": documents,
        "skipped_auto_count": skipped_auto_count,
        **_tax_wizard(session, 3),
    }

    if varovani == "test-zneplatnen":
        ctx["flash_message"] = "Obsah emailu byl změněn — je nutné odeslat nový testovací email."
        ctx["flash_type"] = "warning"

    if request.headers.get("HX-Request") and not request.headers.get("HX-Boosted"):
        return templates.TemplateResponse(request, "partials/tax_send_body.html", ctx)

    return templates.TemplateResponse(request, "tax/send.html", ctx)


@router.post("/{session_id}/rozeslat/email/{dist_id}")
async def update_recipient_email(
    session_id: int,
    dist_id: int,
    request: Request,
    email: str = Form(""),
    db: Session = Depends(get_db),
):
    """Změna emailové adresy příjemce pro rozesílání."""
    dist = db.query(TaxDistribution).get(dist_id)
    if not dist:
        return RedirectResponse(f"/dane/{session_id}/rozeslat", status_code=302)

    email = email.strip()

    # Basic email format validation (#26)
    if email and not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        # Return current row unchanged with no update
        documents = (
            db.query(TaxDocument)
            .filter_by(session_id=session_id)
            .options(joinedload(TaxDocument.distributions).joinedload(TaxDistribution.owner))
            .all()
        )
        recipients = _build_recipients(documents)
        key = f"owner_{dist.owner_id}" if dist.owner_id else f"ext_{dist.id}"
        recipient = next((r for r in recipients if r["key"] == key), None)
        if recipient:
            return templates.TemplateResponse(request, "partials/tax_recipient_row.html", {
                "r": recipient,
                "session": db.query(TaxSession).get(session_id),
                "list_url": build_list_url(request),
            })
        return RedirectResponse(f"/dane/{session_id}/rozeslat", status_code=302)

    assignments_changed = False

    if dist.owner_id:
        session = db.query(TaxSession).get(session_id)
        all_docs = (
            db.query(TaxDocument)
            .filter_by(session_id=session_id)
            .options(joinedload(TaxDocument.distributions))
            .all()
        )
        doc_ids = [d.id for d in all_docs]

        # Propagate to all distributions of this owner in this session
        # Skip SENT distributions to preserve historical record (#12)
        sibling_dists = (
            db.query(TaxDistribution)
            .filter(
                TaxDistribution.document_id.in_(doc_ids),
                TaxDistribution.owner_id == dist.owner_id,
            )
            .all()
        )
        for d in sibling_dists:
            if d.email_status != EmailDeliveryStatus.SENT:
                d.email_address_used = email or None

        # If owner has no email, save it to their profile too
        owner = db.query(Owner).get(dist.owner_id)
        if owner and email and not owner.email:
            owner.email = email

        # Auto-assign unmatched documents for units owned by this owner
        assignments_changed = _auto_assign_unmatched_docs(
            db, session_id, dist.owner_id, email, session, all_docs,
        )
    else:
        # Ad-hoc recipient
        dist.ad_hoc_email = email or None

    db.commit()

    # If new assignments were created, full page refresh is needed
    if assignments_changed:
        redirect_url = f"/dane/{session_id}/rozeslat"
        if request.headers.get("HX-Request"):
            return HTMLResponse(
                status_code=200,
                headers={"HX-Redirect": redirect_url},
            )
        return RedirectResponse(redirect_url, status_code=302)

    # Rebuild only the relevant recipient row (optimized — no full rebuild)
    recipient = _build_single_recipient(db, session_id, dist)
    if not recipient:
        return RedirectResponse(f"/dane/{session_id}/rozeslat", status_code=302)

    return templates.TemplateResponse(request, "partials/tax_recipient_row.html", {
        "r": recipient,
        "session": db.query(TaxSession).get(session_id),
        "list_url": build_list_url(request),
    })


@router.post("/{session_id}/rozeslat/email-vyber/{dist_id}")
async def toggle_recipient_email(
    session_id: int,
    dist_id: int,
    request: Request,
    email: str = Form(""),
    checked: str = Form("true"),
    db: Session = Depends(get_db),
):
    """Toggle an individual email address on/off for a recipient."""
    dist = db.query(TaxDistribution).get(dist_id)
    if not dist:
        return RedirectResponse(f"/dane/{session_id}/rozeslat", status_code=302)

    email = email.strip()
    is_checked = checked == "true"

    # Parse current selected emails
    current = set()
    if dist.email_address_used:
        current = {e.strip() for e in dist.email_address_used.split(",") if e.strip()}

    # Add or remove email
    if is_checked and email:
        current.add(email)
    elif not is_checked and email:
        current.discard(email)

    new_value = ",".join(sorted(current)) if current else None

    # Propagate to all sibling distributions of same owner in this session
    # Skip distributions that are already SENT to preserve historical record (#6)
    if dist.owner_id:
        all_docs = (
            db.query(TaxDocument)
            .filter_by(session_id=session_id)
            .options(joinedload(TaxDocument.distributions))
            .all()
        )
        doc_ids = [d.id for d in all_docs]
        sibling_dists = (
            db.query(TaxDistribution)
            .filter(
                TaxDistribution.document_id.in_(doc_ids),
                TaxDistribution.owner_id == dist.owner_id,
            )
            .all()
        )
        for d in sibling_dists:
            if d.email_status != EmailDeliveryStatus.SENT:
                d.email_address_used = new_value
    else:
        if dist.email_status != EmailDeliveryStatus.SENT:
            dist.email_address_used = new_value

    db.commit()

    # Rebuild and return updated row
    documents = (
        db.query(TaxDocument)
        .filter_by(session_id=session_id)
        .options(
            joinedload(TaxDocument.distributions)
            .joinedload(TaxDistribution.owner),
        )
        .order_by(cast(TaxDocument.unit_number, Integer), TaxDocument.unit_letter)
        .all()
    )
    recipients = _build_recipients(documents)

    key = f"owner_{dist.owner_id}" if dist.owner_id else f"ext_{dist.id}"
    recipient = next((r for r in recipients if r["key"] == key), None)
    if not recipient:
        return RedirectResponse(f"/dane/{session_id}/rozeslat", status_code=302)

    list_url = build_list_url(request)

    return templates.TemplateResponse(request, "partials/tax_recipient_row.html", {
        "r": recipient,
        "session": db.query(TaxSession).get(session_id),
        "list_url": list_url,
    })


@router.post("/{session_id}/rozeslat/test")
async def send_test_email(
    session_id: int,
    request: Request,
    test_email: str = Form(...),
    test_doc_id: int = Form(0),
    email_subject: str = Form(""),
    email_body: str = Form(""),
    db: Session = Depends(get_db),
):
    """Odeslání testovacího emailu s vybraným dokumentem."""
    session = db.query(TaxSession).get(session_id)
    if not session:
        return RedirectResponse("/dane", status_code=302)

    # Save subject and body from the form (user may not have clicked Uložit)
    if email_subject:
        session.email_subject = email_subject
    if email_body:
        session.email_body = email_body
    if email_subject or email_body:
        db.commit()

    # Find specific or first document for test attachment
    if test_doc_id:
        test_doc = db.query(TaxDocument).filter_by(
            id=test_doc_id, session_id=session_id
        ).first()
    else:
        test_doc = None
    if not test_doc:
        test_doc = (
            db.query(TaxDocument)
            .filter_by(session_id=session_id)
            .order_by(TaxDocument.id)
            .first()
        )

    attachments = [test_doc.file_path] if test_doc else []

    result = await asyncio.to_thread(
        send_email,
        to_email=test_email.strip(),
        to_name="Test",
        subject=f"[TEST] {session.email_subject or 'Test email'}",
        body_html=session.email_body or "",
        attachments=attachments,
        module="tax",
        reference_id=session.id,
        db=db,
        smtp_profile_id=session.smtp_profile_id,
    )

    if result["success"]:
        flash_message = f"Testovací email odeslán na {test_email}"
        flash_type = "success"
        session.test_email_passed = True
        session.test_email_address = test_email.strip()
        db.commit()
    else:
        flash_message = f"Chyba při odesílání: {result['error']}"
        flash_type = "error"

    # Redirect back with flash
    documents = (
        db.query(TaxDocument)
        .filter_by(session_id=session_id)
        .options(
            joinedload(TaxDocument.distributions)
            .joinedload(TaxDistribution.owner),
        )
        .all()
    )

    skipped_auto_count = sum(
        1 for doc in documents
        for dist in doc.distributions
        if dist.match_status == MatchStatus.AUTO_MATCHED
    )

    recipients = _build_recipients(documents)
    total_recipients = len(recipients)
    with_email = sum(1 for r in recipients if r["email"])

    all_recipients = recipients
    count_pending = sum(1 for r in all_recipients if r["email_status"] in ("pending", "queued"))
    count_sent = sum(1 for r in all_recipients if r["email_status"] == "sent")
    count_failed = sum(1 for r in all_recipients if r["email_status"] == "failed")

    back_url = f"/dane/{session_id}/rozeslat"
    smtp_profiles = db.query(SmtpProfile).order_by(SmtpProfile.is_default.desc(), SmtpProfile.id).all()
    return templates.TemplateResponse(request, "tax/send.html", {
        "active_nav": "tax",
        "session": session,
        "smtp_profiles": smtp_profiles,
        "recipients": recipients,
        "total_recipients": total_recipients,
        "with_email": with_email,
        "without_email": total_recipients - with_email,
        "count_pending": count_pending,
        "count_sent": count_sent,
        "count_failed": count_failed,
        "back_url": f"/dane/{session_id}",
        "list_url": back_url,
        "flash_message": flash_message,
        "flash_type": flash_type,
        "test_email_value": test_email.strip(),
        "all_documents": documents,
        "skipped_auto_count": skipped_auto_count,
        **_tax_wizard(session, 3),
    })


@router.post("/{session_id}/rozeslat/nastaveni")
async def save_send_settings(
    session_id: int,
    request: Request,
    email_subject: str = Form(""),
    email_body: str = Form(""),
    send_batch_size: int = Form(10),
    send_batch_interval: int = Form(5),
    send_confirm_each_batch: bool = Form(False),
    test_email_inline: str = Form(""),
    smtp_profile_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
):
    """Uložení nastavení rozesílání (předmět, tělo, dávkování)."""
    session = db.query(TaxSession).get(session_id)
    if not session:
        return RedirectResponse("/dane", status_code=302)

    # Invalidate test if email content changed (#8)
    test_invalidated = False
    if (email_subject != (session.email_subject or "")
            or email_body != (session.email_body or "")):
        if session.test_email_passed:
            test_invalidated = True
        session.test_email_passed = False

    session.email_subject = email_subject
    session.email_body = email_body
    session.send_batch_size = send_batch_size
    session.send_batch_interval = send_batch_interval
    session.send_confirm_each_batch = send_confirm_each_batch
    if smtp_profile_id:
        session.smtp_profile_id = smtp_profile_id
    if test_email_inline.strip():
        session.test_email_address = test_email_inline.strip()
    # Only set READY if session is in appropriate state (not COMPLETED/SENDING/PAUSED)
    if session.send_status not in (SendStatus.DRAFT, SendStatus.COMPLETED, SendStatus.SENDING, SendStatus.PAUSED):
        session.send_status = SendStatus.READY
    db.commit()

    redirect_url = f"/dane/{session_id}/rozeslat?config=open"
    if test_invalidated:
        redirect_url += "&varovani=test-zneplatnen"
    return RedirectResponse(redirect_url, status_code=302)


# ---------------------------------------------------------------------------
# Batch email sending
# ---------------------------------------------------------------------------


def _sending_eta(progress: dict) -> dict:
    """Compute ETA fields from sending progress dict."""

    sent = progress["sent"]
    failed = progress["failed"]
    current = sent + failed
    eta = compute_eta(current, progress["total"], progress["started_at"])

    return {
        "total": progress["total"],
        "sent": sent,
        "failed": failed,
        "current": current,
        **eta,
        "current_recipient": progress.get("current_recipient", ""),
        "done": progress.get("done", False),
        "error": progress.get("error"),
        "paused": progress.get("paused", False),
        "waiting_batch_confirm": progress.get("waiting_batch_confirm", False),
        "batch_number": progress.get("batch_number", 0),
        "total_batches": progress.get("total_batches", 0),
    }


def _send_emails_batch(session_id: int, recipient_data: list, email_subject: str,
                        email_body: str, batch_size: int, batch_interval: int,
                        confirm_each_batch: bool, smtp_profile_id: Optional[int] = None):
    """Background thread: send emails in batches."""
    db = SessionLocal()
    try:
        # Split into batches
        batches = []
        for i in range(0, len(recipient_data), batch_size):
            batches.append(recipient_data[i:i + batch_size])

        with _sending_lock:
            _sending_progress[session_id]["total_batches"] = len(batches)

        for batch_idx, batch in enumerate(batches):
            with _sending_lock:
                _sending_progress[session_id]["batch_number"] = batch_idx + 1

            # Create shared SMTP connection per batch (#25)
            smtp_conn = None
            try:
                smtp_conn = create_smtp_connection(profile_id=smtp_profile_id)
            except Exception:
                logger.warning("Failed to create shared SMTP connection, falling back to per-email")

            for rcpt in batch:
                # Check paused
                while True:
                    with _sending_lock:
                        paused = _sending_progress[session_id].get("paused")
                        done = _sending_progress[session_id].get("done")
                    if not paused:
                        break
                    if done:
                        if smtp_conn:
                            try:
                                smtp_conn.quit()
                            except Exception:
                                logger.debug("SMTP quit failed during pause/cancel cleanup", exc_info=True)
                        return
                    time.sleep(0.5)

                with _sending_lock:
                    _sending_progress[session_id]["current_recipient"] = rcpt["name"]

                # Gather only unsent attachment file paths
                unsent_docs = [d for d in rcpt["docs"] if not d.get("sent")]
                attachments = [d["file_path"] for d in unsent_docs] if unsent_docs else [d["file_path"] for d in rcpt["docs"]]
                unsent_dist_ids = [d["dist_id"] for d in unsent_docs] if unsent_docs else rcpt["dist_ids"]

                # Send email — wrapped in try/except so one failure
                # never kills the entire batch
                try:
                    result = send_email(
                        to_email=rcpt["email"],
                        to_name=rcpt["name"],
                        subject=email_subject,
                        body_html=email_body,
                        attachments=attachments,
                        module="tax",
                        reference_id=session_id,
                        db=db,
                        smtp_server=smtp_conn,
                        smtp_profile_id=smtp_profile_id,
                    )
                except Exception as exc:
                    logger.exception("Neočekávaná chyba při odesílání pro %s (%s)",
                                     rcpt["name"], rcpt["email"])
                    result = {"success": False, "error": str(exc)}
                    # Shared SMTP connection is likely dead — recreate
                    smtp_conn = None
                    try:
                        smtp_conn = create_smtp_connection(profile_id=smtp_profile_id)
                    except Exception:
                        logger.warning("Nepodařilo se obnovit SMTP spojení")

                # Batch-update distribution statuses in DB (avoid N+1)
                dists = (
                    db.query(TaxDistribution)
                    .filter(TaxDistribution.id.in_(unsent_dist_ids))
                    .all()
                )
                for dist in dists:
                    if result["success"]:
                        dist.email_status = EmailDeliveryStatus.SENT
                        dist.email_sent = True
                        dist.email_sent_at = utcnow()
                        dist.email_address_used = rcpt["email"]
                        dist.email_error = None
                    else:
                        dist.email_status = EmailDeliveryStatus.FAILED
                        dist.email_error = result.get("error", "Unknown error")
                        dist.email_address_used = rcpt["email"]

                db.commit()

                with _sending_lock:
                    if result["success"]:
                        _sending_progress[session_id]["sent"] += 1
                    else:
                        _sending_progress[session_id]["failed"] += 1

            # Close shared SMTP connection after batch
            if smtp_conn:
                try:
                    smtp_conn.quit()
                except Exception:
                    logger.debug("SMTP quit failed after batch cleanup", exc_info=True)
                smtp_conn = None

            # After batch: wait for confirmation or interval
            if batch_idx < len(batches) - 1:  # not last batch
                if confirm_each_batch:
                    with _sending_lock:
                        _sending_progress[session_id]["waiting_batch_confirm"] = True
                    while True:
                        with _sending_lock:
                            waiting = _sending_progress[session_id].get("waiting_batch_confirm")
                            done = _sending_progress[session_id].get("done")
                        if not waiting:
                            break
                        if done:
                            return
                        time.sleep(0.5)
                else:
                    # Wait batch_interval seconds (but check for pause)
                    for _ in range(batch_interval * 2):
                        with _sending_lock:
                            done = _sending_progress[session_id].get("done")
                        if done:
                            return
                        time.sleep(0.5)

        # Complete
        session = db.query(TaxSession).get(session_id)
        if session:
            session.send_status = SendStatus.COMPLETED
            db.commit()

    except Exception as e:
        logger.exception("Error in batch email sending for session %s", session_id)
        with _sending_lock:
            _sending_progress[session_id]["error"] = str(e)
    finally:
        with _sending_lock:
            _sending_progress[session_id]["done"] = True
            _sending_progress[session_id]["current_recipient"] = ""
            _sending_progress[session_id]["finished_at"] = time.monotonic()
        db.close()


@router.post("/{session_id}/rozeslat/odeslat")
async def start_batch_send(
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Start batch email sending for selected recipients."""
    logger.info("start_batch_send called for session %s", session_id)
    session = db.query(TaxSession).get(session_id)
    if not session:
        logger.warning("Session %s not found", session_id)
        return RedirectResponse("/dane", status_code=302)

    if not session.test_email_passed:
        logger.warning("Session %s: test email not passed", session_id)
        return RedirectResponse(f"/dane/{session_id}/rozeslat", status_code=302)

    # Check no concurrent sending
    with _sending_lock:
        progress = _sending_progress.get(session_id)
        if progress and not progress.get("done"):
            logger.info("Session %s: concurrent sending in progress", session_id)
            return RedirectResponse(f"/dane/{session_id}/rozeslat/prubeh", status_code=302)

    # Get selected keys from form
    form = await request.form()
    selected_keys = form.getlist("selected_keys")
    logger.info("Session %s: received %d selected keys", session_id, len(selected_keys))
    if not selected_keys:
        logger.warning("Session %s: no selected keys in form", session_id)
        return RedirectResponse(f"/dane/{session_id}/rozeslat", status_code=302)

    # Build recipients
    documents = (
        db.query(TaxDocument)
        .filter_by(session_id=session_id)
        .options(
            joinedload(TaxDocument.distributions)
            .joinedload(TaxDistribution.owner),
        )
        .all()
    )
    all_recipients = _build_recipients(documents)

    # Filter to selected
    selected_set = set(selected_keys)
    recipients_to_send = [r for r in all_recipients if r["key"] in selected_set and r["email"]]
    logger.info("Session %s: %d recipients to send (from %d all, %d selected)",
                session_id, len(recipients_to_send), len(all_recipients), len(selected_set))

    if not recipients_to_send:
        logger.warning("Session %s: no recipients with email in selection", session_id)
        return RedirectResponse(f"/dane/{session_id}/rozeslat", status_code=302)

    # Mark only unsent distributions as QUEUED
    for rcpt in recipients_to_send:
        unsent_dist_ids = [d["dist_id"] for d in rcpt["docs"] if not d.get("sent")]
        target_ids = unsent_dist_ids if unsent_dist_ids else rcpt["dist_ids"]
        for dist_id in target_ids:
            dist = db.query(TaxDistribution).get(dist_id)
            if dist and dist.email_status != EmailDeliveryStatus.SENT:
                dist.email_status = EmailDeliveryStatus.QUEUED

    session.send_status = SendStatus.SENDING
    log_activity(db, ActivityAction.STATUS_CHANGED, "tax_session", "dane",
                 entity_id=session.id, entity_name=session.title,
                 description=f"Rozesílka zahájena: {len(recipients_to_send)} příjemců")
    db.commit()

    # Initialize progress
    with _sending_lock:
        _sending_progress[session_id] = {
            "total": len(recipients_to_send),
            "sent": 0,
            "failed": 0,
            "current_recipient": "",
            "done": False,
            "error": None,
            "started_at": time.monotonic(),
            "paused": False,
            "waiting_batch_confirm": False,
            "batch_number": 0,
            "total_batches": 0,
        }

    # Start background thread
    thread = threading.Thread(
        target=_send_emails_batch,
        args=(
            session_id,
            recipients_to_send,
            session.email_subject or "",
            session.email_body or "",
            session.send_batch_size or 10,
            session.send_batch_interval or 5,
            session.send_confirm_each_batch or False,
            session.smtp_profile_id,
        ),
        daemon=True,
    )
    thread.start()

    return RedirectResponse(f"/dane/{session_id}/rozeslat/prubeh", status_code=302)


@router.get("/{session_id}/rozeslat/prubeh")
async def sending_progress_page(
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Show progress page while emails are being sent."""
    session = db.query(TaxSession).get(session_id)
    if not session:
        return RedirectResponse("/dane", status_code=302)

    with _sending_lock:
        progress = _sending_progress.get(session_id)
        if not progress:
            return RedirectResponse(f"/dane/{session_id}/rozeslat", status_code=302)
        progress = dict(progress)

    return templates.TemplateResponse(request, "tax/sending.html", {
        "active_nav": "tax",
        "session": session,
        **_sending_eta(progress),
        **_tax_wizard(session, 3),
    })


@router.get("/{session_id}/rozeslat/prubeh-stav")
async def sending_progress_status(
    session_id: int,
    request: Request,
):
    """HTMX polling endpoint — returns progress partial or redirect when done."""
    with _sending_lock:
        progress = _sending_progress.get(session_id)
        if not progress:
            response = HTMLResponse("")
            response.headers["HX-Redirect"] = f"/dane/{session_id}/rozeslat"
            return response
        # Po dokončení počkat 3 sekundy, aby uživatel viděl výsledek
        if progress.get("done"):
            finished_at = progress.get("finished_at", 0)
            if time.monotonic() - finished_at >= 3:
                _sending_progress.pop(session_id, None)
                response = HTMLResponse("")
                response.headers["HX-Redirect"] = f"/dane/{session_id}/rozeslat"
                return response
        progress = dict(progress)  # snapshot under lock

    return templates.TemplateResponse(request, "partials/_send_progress_inner.html", {
        "session_id": session_id,
        **_sending_eta(progress),
    })


@router.post("/{session_id}/rozeslat/pozastavit")
async def pause_sending(
    session_id: int,
    db: Session = Depends(get_db),
):
    """Pause the sending process."""
    with _sending_lock:
        progress = _sending_progress.get(session_id)
        if progress and not progress.get("done"):
            progress["paused"] = True

    session = db.query(TaxSession).get(session_id)
    if session:
        session.send_status = SendStatus.PAUSED
        db.commit()

    return RedirectResponse(f"/dane/{session_id}/rozeslat/prubeh", status_code=302)


@router.post("/{session_id}/rozeslat/pokracovat")
async def resume_sending(
    session_id: int,
    db: Session = Depends(get_db),
):
    """Resume the sending process (also confirms batch)."""
    with _sending_lock:
        progress = _sending_progress.get(session_id)
        if progress and not progress.get("done"):
            progress["paused"] = False
            progress["waiting_batch_confirm"] = False

    session = db.query(TaxSession).get(session_id)
    if session:
        session.send_status = SendStatus.SENDING
        db.commit()

    return RedirectResponse(f"/dane/{session_id}/rozeslat/prubeh", status_code=302)


@router.post("/{session_id}/rozeslat/zrusit")
async def cancel_sending(
    session_id: int,
    db: Session = Depends(get_db),
):
    """Cancel the sending process — stop thread, reset QUEUED distributions to PENDING."""
    with _sending_lock:
        progress = _sending_progress.get(session_id)
        if progress and not progress.get("done"):
            progress["done"] = True  # signal thread to stop
            progress["finished_at"] = time.monotonic()

    # Reset QUEUED distributions back to PENDING
    queued_dists = (
        db.query(TaxDistribution)
        .join(TaxDistribution.document)
        .filter(
            TaxDocument.session_id == session_id,
            TaxDistribution.email_status == EmailDeliveryStatus.QUEUED,
        )
        .all()
    )
    for dist in queued_dists:
        dist.email_status = EmailDeliveryStatus.PENDING

    session = db.query(TaxSession).get(session_id)
    if session:
        session.send_status = SendStatus.PAUSED
        db.commit()

    return RedirectResponse(f"/dane/{session_id}/rozeslat", status_code=302)


@router.post("/{session_id}/rozeslat/retry")
async def retry_failed(
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Retry all failed recipients."""
    session = db.query(TaxSession).get(session_id)
    if not session:
        return RedirectResponse("/dane", status_code=302)

    # Check no concurrent sending
    with _sending_lock:
        progress = _sending_progress.get(session_id)
        if progress and not progress.get("done"):
            return RedirectResponse(f"/dane/{session_id}/rozeslat/prubeh", status_code=302)

    # Build recipients and filter to failed only
    documents = (
        db.query(TaxDocument)
        .filter_by(session_id=session_id)
        .options(
            joinedload(TaxDocument.distributions)
            .joinedload(TaxDistribution.owner),
        )
        .all()
    )
    all_recipients = _build_recipients(documents)
    failed_recipients = [r for r in all_recipients if r["email_status"] == "failed" and r["email"]]

    if not failed_recipients:
        return RedirectResponse(f"/dane/{session_id}/rozeslat", status_code=302)

    # Reset failed distributions to QUEUED
    for rcpt in failed_recipients:
        for dist_id in rcpt["dist_ids"]:
            dist = db.query(TaxDistribution).get(dist_id)
            if dist and dist.email_status == EmailDeliveryStatus.FAILED:
                dist.email_status = EmailDeliveryStatus.QUEUED
                dist.email_error = None

    session.send_status = SendStatus.SENDING
    db.commit()

    # Initialize progress
    with _sending_lock:
        _sending_progress[session_id] = {
            "total": len(failed_recipients),
            "sent": 0,
            "failed": 0,
            "current_recipient": "",
            "done": False,
            "error": None,
            "started_at": time.monotonic(),
            "paused": False,
            "waiting_batch_confirm": False,
            "batch_number": 0,
            "total_batches": 0,
        }

    # Start background thread
    thread = threading.Thread(
        target=_send_emails_batch,
        args=(
            session_id,
            failed_recipients,
            session.email_subject or "",
            session.email_body or "",
            session.send_batch_size or 10,
            session.send_batch_interval or 5,
            session.send_confirm_each_batch or False,
            session.smtp_profile_id,
        ),
        daemon=True,
    )
    thread.start()

    return RedirectResponse(f"/dane/{session_id}/rozeslat/prubeh", status_code=302)
