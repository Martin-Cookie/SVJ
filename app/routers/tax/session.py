from __future__ import annotations

import shutil
import threading
import time as _time
from datetime import datetime
from html import escape
from io import BytesIO
from pathlib import Path
from typing import List
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, Form, Query, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import cast, Integer
from sqlalchemy.orm import Session, joinedload

from app.config import settings
from app.database import get_db
from app.models import (
    EmailTemplate, MatchStatus, Owner, OwnerUnit, SendStatus,
    TaxDistribution, TaxDocument, TaxSession,
    ActivityAction, log_activity,
)
from app.utils import UPLOAD_LIMITS, build_list_url, excel_auto_width, is_htmx_partial, is_safe_path, strip_diacritics, validate_uploads

from ._helpers import (
    logger, templates,
    _processing_progress, _processing_lock,
    _TAX_WIZARD_STEPS, _tax_wizard, _session_stats, _unit_by_number,
)
from .processing import _process_tax_files

router = APIRouter()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/")
async def tax_list(request: Request, back: str = Query("", alias="back"), stav: str = Query("", alias="stav"), db: Session = Depends(get_db)):
    """Seznam daňových sessions s filtrováním podle stavu."""
    sessions = (
        db.query(TaxSession)
        .options(
            joinedload(TaxSession.documents)
            .joinedload(TaxDocument.distributions),
        )
        .order_by(TaxSession.created_at.desc())
        .all()
    )
    list_url = build_list_url(request)

    # Compute per-session stats for progress display
    session_stats = {}
    for s in sessions:
        total = len(s.documents)
        confirmed = 0
        for doc in s.documents:
            if doc.distributions and all(
                d.match_status in (MatchStatus.CONFIRMED, MatchStatus.MANUAL)
                for d in doc.distributions
            ):
                confirmed += 1
        # Determine wizard step for list view
        send_status = s.send_status.value if s.send_status else "draft"
        if send_status == "completed":
            wizard_step, wizard_label = 4, "Dokončeno"
        elif send_status in ("sending", "paused", "ready"):
            wizard_step, wizard_label = 3, "Rozesílka"
        elif total > 0 and confirmed < total:
            wizard_step, wizard_label = 2, "Přiřazení"
        elif total == 0:
            wizard_step, wizard_label = 1, "Nahrání PDF"
        else:
            wizard_step, wizard_label = 3, "Rozesílka"

        # Determine max_done for list wizard
        if send_status == "completed":
            list_max_done = 4
        elif send_status in ("sending", "paused", "ready"):
            list_max_done = 2
        else:
            list_max_done = 0
        # If documents exist, step 1 (Nahrání PDF) is always done
        if total > 0 and list_max_done < 1:
            list_max_done = 1

        # Build wizard steps for compact stepper
        wiz_steps = []
        for i, ws in enumerate(_TAX_WIZARD_STEPS, 1):
            if i < wizard_step:
                wiz_steps.append({"label": ws["label"], "status": "done"})
            elif i == wizard_step:
                wiz_steps.append({"label": ws["label"], "status": "done" if i <= list_max_done else "active"})
            else:
                wiz_steps.append({"label": ws["label"], "status": "pending"})

        session_stats[s.id] = {
            "total": total,
            "confirmed": confirmed,
            "pct": int(confirmed / total * 100) if total > 0 else 0,
            "wizard_step": wizard_step,
            "wizard_label": wizard_label,
            "wizard_steps": wiz_steps,
            "wizard_current": wizard_step,
            "wizard_total": len(_TAX_WIZARD_STEPS),
        }

    # Compute status counts for filter bubbles
    status_counts = {"all": len(sessions), "draft": 0, "ready": 0, "sending": 0, "completed": 0}
    session_status_map = {}
    for s in sessions:
        ss = s.send_status.value if s.send_status else "draft"
        if ss in ("sending", "paused"):
            cat = "sending"
        else:
            cat = ss
        status_counts[cat] = status_counts.get(cat, 0) + 1
        session_status_map[s.id] = cat

    # Filter sessions by status
    if stav and stav in ("draft", "ready", "sending", "completed"):
        sessions = [s for s in sessions if session_status_map[s.id] == stav]

    back_label = ""
    if back:
        back_label = (
            "Zpět na přehled" if back == "/" or back.startswith("/?")
            else "Zpět na nastavení" if back.startswith("/nastaveni")
            else "Zpět"
        )

    return templates.TemplateResponse("tax/index.html", {
        "request": request,
        "active_nav": "tax",
        "sessions": sessions,
        "back_url": back,
        "back_label": back_label,
        "list_url": list_url,
        "session_stats": session_stats,
        "status_counts": status_counts,
        "current_stav": stav,
    })


@router.get("/nova")
async def tax_create_page(
    request: Request,
    chyba: str = Query("", alias="chyba"),
    db: Session = Depends(get_db),
):
    """Formulář pro vytvoření nové daňové session."""
    # Wizard step 1 for new session (no session object yet, build manually)
    steps = [{"label": s["label"], "status": "active" if i == 0 else "pending"} for i, s in enumerate(_TAX_WIZARD_STEPS)]
    email_templates = (
        db.query(EmailTemplate)
        .order_by(EmailTemplate.order, EmailTemplate.name)
        .all()
    )
    ctx = {
        "request": request,
        "active_nav": "tax",
        "wizard_steps": steps,
        "wizard_current": 1,
        "wizard_total": len(_TAX_WIZARD_STEPS),
        "email_templates": email_templates,
    }
    if chyba:
        ctx["flash_message"] = chyba
        ctx["flash_type"] = "error"
    return templates.TemplateResponse("tax/upload.html", ctx)


@router.post("/nova")
async def tax_create(
    request: Request,
    title: str = Form(...),
    email_body: str = Form(""),
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    """Vytvoření nové daňové session s nahráním PDF souborů."""
    # Filter to PDF files only (webkitdirectory sends all files including .DS_Store)
    pdf_files = [f for f in files if f.filename and f.filename.lower().endswith(".pdf")]
    if not pdf_files:
        return RedirectResponse(
            f"/dane/nova?chyba={quote('Nebyly nalezeny žádné PDF soubory. Zkontrolujte, zda vybraný adresář obsahuje soubory s příponou .pdf.')}",
            status_code=302,
        )

    err = await validate_uploads(pdf_files, **UPLOAD_LIMITS["pdf"])
    if err:
        return RedirectResponse(f"/dane/nova?chyba={quote(err)}", status_code=302)

    year = datetime.now().year
    session = TaxSession(
        title=title,
        year=year,
        email_subject=title,
        email_body=email_body,
    )
    db.add(session)
    db.flush()

    # Save PDF files to disk (fast I/O only — no extraction yet)
    upload_dir = settings.upload_dir / "tax_pdfs" / f"session_{session.id}"
    upload_dir.mkdir(parents=True, exist_ok=True)

    saved_files = []
    for file in pdf_files:
        basename = Path(file.filename).name
        dest = upload_dir / basename
        with open(dest, "wb") as f:
            shutil.copyfileobj(file.file, f)
        saved_files.append(str(dest))

    if not saved_files:
        db.rollback()
        return RedirectResponse("/dane", status_code=302)

    log_activity(db, ActivityAction.CREATED, "tax_session", "dane",
                 entity_id=session.id, entity_name=session.title,
                 description=f"Nahráno {len(saved_files)} PDF souborů")
    db.commit()

    # Initialize progress tracker
    with _processing_lock:
        _processing_progress[session.id] = {
            "total": len(saved_files),
            "current": 0,
            "current_file": "",
            "done": False,
            "error": None,
            "started_at": _time.monotonic(),
        }

    # Start background processing thread
    thread = threading.Thread(
        target=_process_tax_files,
        args=(session.id, saved_files, year),
        daemon=True,
    )
    thread.start()

    return RedirectResponse(f"/dane/{session.id}/zpracovani", status_code=302)


@router.get("/{session_id}/upload")
async def tax_upload_page(
    session_id: int,
    request: Request,
    back: str = Query("", alias="back"),
    chyba: str = Query("", alias="chyba"),
    db: Session = Depends(get_db),
):
    """Upload additional PDFs to an existing session."""
    session = db.query(TaxSession).options(
        joinedload(TaxSession.documents),
    ).get(session_id)
    if not session:
        return RedirectResponse("/dane", status_code=302)

    has_documents = len(session.documents) > 0
    ctx = {
        "request": request,
        "active_nav": "tax",
        "session": session,
        "has_documents": has_documents,
        "back_url": back or f"/dane/{session_id}",
        **_tax_wizard(session, 1, has_documents=has_documents),
    }
    if chyba:
        ctx["flash_message"] = chyba
        ctx["flash_type"] = "error"
    return templates.TemplateResponse("tax/upload_additional.html", ctx)


@router.post("/{session_id}/upload")
async def tax_upload_additional(
    session_id: int,
    request: Request,
    import_mode: str = Form("append"),
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    """Upload additional PDFs to an existing session with append/overwrite mode."""
    # Filter to PDF files only (webkitdirectory sends all files including .DS_Store)
    pdf_files = [f for f in files if f.filename and f.filename.lower().endswith(".pdf")]
    if not pdf_files:
        return RedirectResponse(
            f"/dane/{session_id}/upload?chyba={quote('Nebyly nalezeny žádné PDF soubory.')}",
            status_code=302,
        )

    err = await validate_uploads(pdf_files, **UPLOAD_LIMITS["pdf"])
    if err:
        return RedirectResponse(f"/dane/{session_id}/upload?chyba={quote(err)}", status_code=302)

    session = db.query(TaxSession).options(
        joinedload(TaxSession.documents).joinedload(TaxDocument.distributions),
    ).get(session_id)
    if not session:
        return RedirectResponse("/dane", status_code=302)

    # Save new PDF files to disk FIRST (before deleting old ones) (#21)
    upload_dir = settings.upload_dir / "tax_pdfs" / f"session_{session_id}"
    upload_dir.mkdir(parents=True, exist_ok=True)

    saved_files = []
    for file in pdf_files:
        basename = Path(file.filename).name
        dest = upload_dir / basename
        with open(dest, "wb") as f:
            shutil.copyfileobj(file.file, f)
        saved_files.append(str(dest))

    # If overwrite mode, delete existing documents and files AFTER new files are saved
    if import_mode == "overwrite":
        old_file_paths = []
        for doc in session.documents:
            if doc.file_path and str(doc.file_path) not in saved_files:
                old_file_paths.append(doc.file_path)
            for dist in doc.distributions:
                db.delete(dist)
            db.delete(doc)
        db.flush()
        # Remove old files from disk (only those not overwritten by new uploads)
        for fp in old_file_paths:
            try:
                Path(fp).unlink()
            except Exception:
                logger.debug("Failed to clean up old file: %s", fp)

    if not saved_files:
        return RedirectResponse(f"/dane/{session_id}", status_code=302)

    db.commit()

    # Initialize progress tracker
    with _processing_lock:
        _processing_progress[session_id] = {
            "total": len(saved_files),
            "current": 0,
            "current_file": "",
            "done": False,
            "error": None,
            "started_at": _time.monotonic(),
        }

    # Start background processing thread
    thread = threading.Thread(
        target=_process_tax_files,
        args=(session_id, saved_files, session.year),
        daemon=True,
    )
    thread.start()

    return RedirectResponse(f"/dane/{session_id}/zpracovani", status_code=302)


@router.get("/{session_id}")
async def tax_detail(
    session_id: int,
    request: Request,
    back: str = Query("", alias="back"),
    filtr: str = Query("", alias="filtr"),
    q: str = Query("", alias="q"),
    sort: str = Query("unit_number", alias="sort"),
    order: str = Query("asc", alias="order"),
    stranka: int = Query(1, alias="stranka"),
    db: Session = Depends(get_db),
):
    """Detail daňové session s dokumenty, párováním a filtrováním."""
    session = db.query(TaxSession).get(session_id)
    if not session:
        return RedirectResponse("/dane", status_code=302)

    is_partial = is_htmx_partial(request)

    # Light join: no Owner.units chain (unit numbers come from separate owners query)
    all_documents = (
        db.query(TaxDocument)
        .filter_by(session_id=session_id)
        .options(
            joinedload(TaxDocument.distributions)
            .joinedload(TaxDistribution.owner),
        )
        .order_by(cast(TaxDocument.unit_number, Integer), TaxDocument.unit_letter)
        .all()
    )

    # Apply filter
    if filtr == "confirmed":
        documents = [
            d for d in all_documents
            if d.distributions and all(
                x.match_status in (MatchStatus.CONFIRMED, MatchStatus.MANUAL)
                for x in d.distributions
            )
        ]
    elif filtr == "auto":
        documents = [
            d for d in all_documents
            if d.distributions and any(
                x.match_status == MatchStatus.AUTO_MATCHED for x in d.distributions
            ) and not all(
                x.match_status in (MatchStatus.CONFIRMED, MatchStatus.MANUAL)
                for x in d.distributions
            )
        ]
    elif filtr == "unmatched":
        documents = [
            d for d in all_documents
            if not d.distributions or any(
                x.match_status == MatchStatus.UNMATCHED for x in d.distributions
            )
        ]
    elif filtr == "missing":
        documents = []
    else:
        documents = all_documents

    # Search filtering
    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        documents = [
            d for d in documents
            if q_lower in (d.filename or "").lower()
            or q_lower in (d.extracted_owner_name or "").lower()
            or q_ascii in strip_diacritics(d.extracted_owner_name or "")
            or q_lower in str(d.unit_number or "")
            or any(
                q_ascii in strip_diacritics(dist.owner.display_name)
                for dist in d.distributions if dist.owner
            )
        ]

    # Sorting
    SORT_KEYS = {
        "filename": lambda d: (d.filename or "").lower(),
        "unit_number": lambda d: (int(d.unit_number) if d.unit_number and d.unit_number.isdigit() else 0, d.unit_letter or ""),
        "extracted": lambda d: (d.extracted_owner_name or "").lower(),
        "owner": lambda d: next(
            (strip_diacritics(dist.owner.display_name) for dist in d.distributions if dist.owner),
            "zzz"
        ),
        "confidence": lambda d: next(
            (dist.match_confidence or 0 for dist in d.distributions if dist.match_confidence),
            0
        ),
    }
    sort_fn = SORT_KEYS.get(sort, SORT_KEYS["unit_number"])
    documents.sort(key=sort_fn, reverse=(order == "desc"))

    # Pagination
    per_page = 100
    total_filtered = len(documents)
    total_pages = max(1, (total_filtered + per_page - 1) // per_page)
    stranka = max(1, min(stranka, total_pages))
    start = (stranka - 1) * per_page
    documents = documents[start:start + per_page]

    is_locked = session.send_status in (SendStatus.READY, SendStatus.SENDING, SendStatus.PAUSED, SendStatus.COMPLETED) if session.send_status else False
    list_url = build_list_url(request)

    pagination = {
        "current_page": stranka,
        "total_pages": total_pages,
        "total_filtered": total_filtered,
        "per_page": per_page,
    }

    # HTMX partial response — skip stats, missing units, owners (not in tbody)
    if is_partial:
        return templates.TemplateResponse("partials/tax_table_body.html", {
            "request": request,
            "documents": documents,
            "is_locked": is_locked,
            "list_url": list_url,
            "unit_by_number": _unit_by_number(db),
            **pagination,
        })

    # --- Full page: stats + missing units + owners ---
    stats = _session_stats(all_documents)

    doc_unit_numbers = {d.unit_number for d in all_documents if d.unit_number}
    current_ous = (
        db.query(OwnerUnit)
        .filter(OwnerUnit.valid_to.is_(None))
        .options(joinedload(OwnerUnit.unit), joinedload(OwnerUnit.owner))
        .all()
    )
    missing_units = {}
    for ou in current_ous:
        unum = str(ou.unit.unit_number)
        if unum not in doc_unit_numbers:
            if unum not in missing_units:
                missing_units[unum] = {"unit": ou.unit, "owners": []}
            missing_units[unum]["owners"].append(ou.owner)
    missing_list = sorted(missing_units.values(), key=lambda m: m["unit"].unit_number)
    stats["stat_missing"] = len(missing_list)

    owners = (
        db.query(Owner)
        .filter_by(is_active=True)
        .options(joinedload(Owner.units).joinedload(OwnerUnit.unit))
        .order_by(Owner.name_normalized)
        .all()
    )

    back_url = back or "/dane"
    back_label = (
        "Zpět na přehled" if back == "/"
        else "Zpět na seznam rozesílek" if back.startswith("/dane") and "/" not in back.lstrip("/dane")
        else "Zpět na rozesílání"
    )

    return templates.TemplateResponse("tax/matching.html", {
        "request": request,
        "active_nav": "tax",
        "session": session,
        "documents": documents,
        "owners": owners,
        "back_url": back_url,
        "back_label": back_label,
        "list_url": list_url,
        "filtr": filtr,
        "q": q,
        "sort": sort,
        "order": order,
        "is_locked": is_locked,
        "unit_by_number": _unit_by_number(db),
        "missing_list": missing_list,
        **stats,
        **pagination,
        **_tax_wizard(session, 2, has_documents=len(documents) > 0),
    })


@router.post("/{session_id}/prejmenovat")
async def rename_session(
    session_id: int,
    request: Request,
    title: str = Form(...),
    db: Session = Depends(get_db),
):
    """Přejmenování daňové session."""
    session = db.query(TaxSession).get(session_id)
    if not session:
        return RedirectResponse("/dane", status_code=302)

    session.title = title.strip()
    db.commit()

    t = escape(session.title)
    return HTMLResponse(
        f'<div id="session-title-area">'
        f'<div class="flex items-center gap-2">'
        f'<h1 class="text-2xl font-bold text-gray-800">{t}</h1>'
        f'<button type="button" onclick="showTitleEdit()" class="text-gray-400 hover:text-gray-600" title="Přejmenovat">'
        f'<svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24">'
        f'<path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" '
        f'd="M15.232 5.232l3.536 3.536m-2.036-5.036a2.5 2.5 0 113.536 3.536L6.5 21.036H3v-3.572L16.732 3.732z"/>'
        f'</svg></button></div>'
        f'<p class="text-sm text-green-600 mt-1">Uloženo</p>'
        f'<script>hideTitleEdit()</script>'
        f'</div>'
    )


@router.post("/{session_id}/dokoncit")
async def finalize_session(
    session_id: int,
    db: Session = Depends(get_db),
):
    """Lock the session — no more editing, can proceed to sending."""
    session = db.query(TaxSession).get(session_id)
    if session:
        session.send_status = SendStatus.READY
        log_activity(db, ActivityAction.STATUS_CHANGED, "tax_session", "dane",
                     entity_id=session.id, entity_name=session.title,
                     description="Stav: koncept → připraveno k odeslání")
        db.commit()
    return RedirectResponse(f"/dane/{session_id}", status_code=302)


@router.post("/{session_id}/znovu-otevrit")
async def reopen_session(
    session_id: int,
    db: Session = Depends(get_db),
):
    """Unlock the session for further editing."""
    session = db.query(TaxSession).get(session_id)
    if session:
        session.send_status = SendStatus.DRAFT
        log_activity(db, ActivityAction.STATUS_CHANGED, "tax_session", "dane",
                     entity_id=session.id, entity_name=session.title,
                     description="Stav: znovu otevřeno pro úpravy")
        db.commit()
    return RedirectResponse(f"/dane/{session_id}", status_code=302)


@router.get("/{session_id}/exportovat")
async def tax_export(session_id: int, db: Session = Depends(get_db)):
    """Export distribution overview to Excel."""
    session = db.query(TaxSession).options(
        joinedload(TaxSession.documents)
        .joinedload(TaxDocument.distributions)
        .joinedload(TaxDistribution.owner),
    ).get(session_id)
    if not session:
        return RedirectResponse("/dane", status_code=302)

    wb = Workbook()
    ws = wb.active
    ws.title = "Přehled rozesílání"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    sent_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    failed_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    headers = ["Dokument", "Jednotka", "Vlastník", "Email", "Status", "Odesláno"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill

    email_status_labels = {
        "pending": "Čeká",
        "queued": "Ve frontě",
        "sent": "Odesláno",
        "failed": "Chyba",
        "skipped": "Přeskočeno",
    }

    row_idx = 2
    for doc in sorted(session.documents, key=lambda d: d.filename or ""):
        for dist in doc.distributions:
            ws.cell(row=row_idx, column=1, value=doc.filename or "")
            ws.cell(row=row_idx, column=2, value=doc.unit_number or "")
            if dist.owner:
                ws.cell(row=row_idx, column=3, value=dist.owner.display_name)
            elif dist.ad_hoc_name:
                ws.cell(row=row_idx, column=3, value=dist.ad_hoc_name)
            ws.cell(row=row_idx, column=4, value=dist.email_address_used or (dist.owner.email if dist.owner else dist.ad_hoc_email) or "")
            status_val = dist.email_status.value if dist.email_status else "pending"
            status_cell = ws.cell(row=row_idx, column=5, value=email_status_labels.get(status_val, status_val))
            if status_val == "sent":
                status_cell.fill = sent_fill
            elif status_val == "failed":
                status_cell.fill = failed_fill
            ws.cell(row=row_idx, column=6, value=dist.email_sent_at.strftime("%d.%m.%Y %H:%M") if dist.email_sent_at else "")
            row_idx += 1

    excel_auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"rozeslani_{session_id}_{timestamp}.xlsx"

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{session_id}/smazat")
async def delete_session(
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Delete session, its documents, distributions, and files from disk."""
    session = db.query(TaxSession).get(session_id)
    if not session:
        return RedirectResponse("/dane", status_code=302)

    # Delete files from disk
    upload_dir = settings.upload_dir / "tax_pdfs" / f"session_{session_id}"
    try:
        if upload_dir.exists():
            shutil.rmtree(upload_dir)
    except Exception:
        logger.debug("Failed to clean up upload dir: %s", upload_dir)

    log_activity(db, ActivityAction.DELETED, "tax_session", "dane",
                 entity_id=session.id, entity_name=session.title)
    db.delete(session)
    db.commit()

    return RedirectResponse("/dane", status_code=302)


# ---------------------------------------------------------------------------
# PDF serving
# ---------------------------------------------------------------------------

@router.get("/{session_id}/dokument/{doc_id}")
async def serve_document(
    session_id: int,
    doc_id: int,
    db: Session = Depends(get_db),
):
    """Serve a tax PDF file for in-browser viewing."""
    doc = db.query(TaxDocument).filter_by(id=doc_id, session_id=session_id).first()
    if not doc or not doc.file_path:
        return RedirectResponse(f"/dane/{session_id}", status_code=302)

    path = Path(doc.file_path)
    if not path.exists() or not is_safe_path(path, settings.upload_dir):
        return RedirectResponse(f"/dane/{session_id}", status_code=302)

    return FileResponse(path, media_type="application/pdf")
