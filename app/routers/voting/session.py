import shutil
from datetime import datetime, date
from io import BytesIO
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, Query, Request, UploadFile
from fastapi.responses import JSONResponse, RedirectResponse, Response
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.config import settings
from app.database import get_db
from app.models import (
    Ballot, BallotStatus, BallotVote, Owner, OwnerUnit, Voting,
    VotingItem, VotingStatus, VoteValue,
    ActivityAction, log_activity,
)
from app.services.word_parser import extract_voting_items, extract_voting_metadata
from app.utils import build_list_url, excel_auto_width, is_htmx_partial, strip_diacritics, validate_upload

from ._helpers import (
    _VOTING_WIZARD_STEPS,
    _ballot_stats,
    _get_declared_shares,
    _has_processed_ballots,
    _voting_wizard,
    logger,
    templates,
)


router = APIRouter()


@router.get("/")
async def voting_list(
    request: Request,
    back: str = Query("", alias="back"),
    stav: str = Query("", alias="stav"),
    db: Session = Depends(get_db),
):
    q = db.query(Voting).options(
        joinedload(Voting.items),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    )
    if stav:
        q = q.filter(Voting.status == stav)
    votings = q.order_by(Voting.created_at.desc()).all()

    # Count per status (always from all votings, not filtered) — single GROUP BY query
    status_rows = db.query(Voting.status, func.count(Voting.id)).group_by(Voting.status).all()
    status_counts = {"all": sum(cnt for _, cnt in status_rows)}
    for st, cnt in status_rows:
        status_counts[st.value] = cnt

    # Compute stats per voting
    declared_shares = _get_declared_shares(db)
    voting_stats = {}
    for voting in votings:
        total = declared_shares or 1
        processed = [b for b in voting.ballots if b.status == BallotStatus.PROCESSED]
        voted = [b for b in processed if any(bv.vote is not None for bv in b.votes)]
        processed_votes = sum(b.total_votes for b in voted)

        # Per-item results
        item_results = []
        for item in voting.items:
            votes_for = 0
            votes_against = 0
            votes_abstain = 0
            for b in processed:
                for bv in b.votes:
                    if bv.voting_item_id == item.id:
                        if bv.vote == VoteValue.FOR:
                            votes_for += bv.votes_count
                        elif bv.vote == VoteValue.AGAINST:
                            votes_against += bv.votes_count
                        elif bv.vote == VoteValue.ABSTAIN:
                            votes_abstain += bv.votes_count
            item_results.append({
                "item": item,
                "votes_for": votes_for,
                "votes_against": votes_against,
                "votes_abstain": votes_abstain,
                "pct_for": round(votes_for / total * 100, 2) if total else 0,
                "pct_against": round(votes_against / total * 100, 2) if total else 0,
            })

        # Wizard steps (reuse shared helper)
        wizard = _voting_wizard(voting)

        voting_stats[voting.id] = {
            "processed_count": len(processed),
            "processed_votes": processed_votes,
            "quorum_pct": round(processed_votes / declared_shares * 100, 2) if declared_shares else 0,
            "quorum_reached": processed_votes / declared_shares >= voting.quorum_threshold if declared_shares else False,
            "item_results": item_results,
            "wizard_step": wizard["wizard_current"],
            **wizard,
        }

    list_url = build_list_url(request)

    return templates.TemplateResponse("voting/index.html", {
        "request": request,
        "active_nav": "voting",
        "votings": votings,
        "voting_stats": voting_stats,
        "status_counts": status_counts,
        "current_stav": stav,
        "back_url": back,
        "list_url": list_url,
    })


@router.get("/nova")
async def voting_create_page(
    request: Request,
    chyba: str = Query(""),
    db: Session = Depends(get_db),
):
    wizard = {
        "wizard_steps": [
            {"label": s["label"], "status": "active" if i == 0 else "pending"}
            for i, s in enumerate(_VOTING_WIZARD_STEPS)
        ],
        "wizard_current": 1,
        "wizard_total": len(_VOTING_WIZARD_STEPS),
    }
    ctx = {
        "request": request,
        "active_nav": "voting",
        "declared_shares": _get_declared_shares(db),
        **wizard,
    }
    if chyba == "upload":
        ctx["flash_message"] = "Nahrání šablony selhalo. Ověřte, že soubor je platný .docx a není příliš velký."
        ctx["flash_type"] = "error"
    elif chyba:
        ctx["flash_message"] = chyba
        ctx["flash_type"] = "error"
    return templates.TemplateResponse("voting/create.html", ctx)


@router.post("/nova/nahled-metadat")
async def voting_preview_metadata(
    file: UploadFile = File(...),
):
    """AJAX endpoint: extract metadata from uploaded .docx and return JSON."""
    if not file.filename or not file.filename.endswith(".docx"):
        return JSONResponse({"error": "Nahrajte .docx soubor"}, status_code=400)

    err = await validate_upload(file, max_size_mb=10, allowed_extensions=[".docx"])
    if err:
        return JSONResponse({"error": err}, status_code=400)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    dest = settings.upload_dir / "word_templates" / f"{timestamp}_{file.filename}"
    dest.parent.mkdir(parents=True, exist_ok=True)
    try:
        with open(dest, "wb") as f:
            shutil.copyfileobj(file.file, f)
        meta = extract_voting_metadata(str(dest))
        items = extract_voting_items(str(dest))
        return JSONResponse({
            "meta": meta,
            "items_count": len(items),
        })
    except Exception as e:
        logger.exception("Chyba při extrakci metadat z DOCX šablony")
        return JSONResponse({"error": f"{type(e).__name__}: {e}"}, status_code=500)
    finally:
        # Clean up temp file — the real upload happens on form submit
        try:
            dest.unlink()
        except Exception:
            logger.debug("Failed to clean up file: %s", dest)


@router.post("/nova")
async def voting_create(
    request: Request,
    title: str = Form(...),
    description: str = Form(""),
    start_date: str = Form(""),
    end_date: str = Form(""),
    quorum_threshold: float = Form(0.5),
    partial_owner_mode: str = Form("shared"),
    file: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    quorum_threshold = max(0.0, min(100.0, quorum_threshold))
    voting = Voting(
        title=title,
        description=description,
        quorum_threshold=quorum_threshold / 100,
        partial_owner_mode=partial_owner_mode,
    )
    try:
        if start_date:
            voting.start_date = date.fromisoformat(start_date)
        if end_date:
            voting.end_date = date.fromisoformat(end_date)
    except ValueError:
        pass  # invalid date from form — ignore, keep None

    # Handle Word template upload
    if file and file.filename:
        err = await validate_upload(file, max_size_mb=10, allowed_extensions=[".docx"])
        if err:
            return RedirectResponse("/hlasovani/nova?chyba=upload", status_code=302)

        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        dest = settings.upload_dir / "word_templates" / f"{timestamp}_{file.filename}"
        dest.parent.mkdir(parents=True, exist_ok=True)
        with open(dest, "wb") as f:
            shutil.copyfileobj(file.file, f)
        voting.template_path = str(dest)

        # Extract voting items from template
        extraction_warning = None
        try:
            items = extract_voting_items(str(dest))
            db.add(voting)
            db.flush()
            if items:
                for item_data in items:
                    item = VotingItem(
                        voting_id=voting.id,
                        order=item_data["order"],
                        title=item_data["title"],
                        description=item_data.get("description", ""),
                    )
                    db.add(item)
            else:
                extraction_warning = "sablona-prazdna"
        except Exception:
            logger.exception("Chyba při extrakci bodů z DOCX šablony")
            db.add(voting)
            db.flush()
            extraction_warning = "extrakce-selhala"

        # Extract metadata to pre-fill empty fields
        try:
            meta = extract_voting_metadata(str(dest))
            if not description.strip() and meta.get("description"):
                voting.description = meta["description"]
            if not start_date and meta.get("start_date"):
                voting.start_date = date.fromisoformat(meta["start_date"])
            if not end_date and meta.get("end_date"):
                voting.end_date = date.fromisoformat(meta["end_date"])
        except Exception:
            pass  # metadata extraction is best-effort
    else:
        db.add(voting)
        db.flush()

    # Calculate total votes (only current owner_units)
    total = db.query(func.sum(OwnerUnit.votes)).filter(OwnerUnit.valid_to.is_(None)).scalar() or 0
    voting.total_votes_possible = total

    log_activity(db, ActivityAction.CREATED, "voting", "hlasovani",
                 entity_id=voting.id, entity_name=voting.title)
    db.commit()
    redirect_url = f"/hlasovani/{voting.id}"
    if extraction_warning:
        redirect_url += f"?info={extraction_warning}"
    return RedirectResponse(redirect_url, status_code=302)


@router.get("/{voting_id}")
async def voting_detail(
    voting_id: int,
    request: Request,
    back: str = Query(""),
    q: str = Query(""),
    sort: str = Query("order"),
    order: str = Query("asc"),
    info: str = Query(""),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(
        joinedload(Voting.items),
        joinedload(Voting.ballots).joinedload(Ballot.owner).joinedload(Owner.units).joinedload(OwnerUnit.unit),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    # Snapshot warning: check if ballot total_votes differ from current owner data
    snapshot_warning = ""
    if voting.status == VotingStatus.ACTIVE:
        stale_count = 0
        for ballot in voting.ballots:
            current_votes = sum(ou.votes for ou in ballot.owner.current_units)
            if current_votes != ballot.total_votes:
                stale_count += 1
        if stale_count > 0:
            snapshot_warning = f"U {stale_count} {'lístku se změnil' if stale_count == 1 else 'lístků se změnil'} počet hlasů vlastníka od generování lístků (např. změna podílů nebo vlastníka jednotky)"

    # Calculate results per item
    declared = _get_declared_shares(db) or 1
    results = []
    for item in voting.items:
        votes_for = 0
        votes_against = 0
        votes_abstain = 0
        for ballot in voting.ballots:
            if ballot.status != BallotStatus.PROCESSED:
                continue
            for bv in ballot.votes:
                if bv.voting_item_id == item.id:
                    if bv.vote == VoteValue.FOR:
                        votes_for += bv.votes_count
                    elif bv.vote == VoteValue.AGAINST:
                        votes_against += bv.votes_count
                    elif bv.vote == VoteValue.ABSTAIN:
                        votes_abstain += bv.votes_count

        votes_missing = declared - votes_for - votes_against - votes_abstain
        results.append({
            "item": item,
            "votes_for": votes_for,
            "votes_against": votes_against,
            "votes_abstain": votes_abstain,
            "pct_for": round(votes_for / declared * 100, 2) if declared else 0,
            "pct_against": round(votes_against / declared * 100, 2) if declared else 0,
            "pct_abstain": round(votes_abstain / declared * 100, 2) if declared else 0,
            "votes_missing": votes_missing,
            "pct_missing": round(votes_missing / declared * 100, 2) if declared else 0,
        })

    # Search filter (diacritics-aware)
    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        results = [r for r in results if q_lower in r["item"].title.lower() or q_ascii in strip_diacritics(r["item"].title)]

    # Sort results
    sort_keys = {
        "order": lambda r: r["item"].order,
        "votes_for": lambda r: r["votes_for"],
        "pct_for": lambda r: r["pct_for"],
        "votes_against": lambda r: r["votes_against"],
        "pct_against": lambda r: r["pct_against"],
        "votes_abstain": lambda r: r["votes_abstain"],
        "votes_missing": lambda r: r["votes_missing"],
    }
    key_fn = sort_keys.get(sort, sort_keys["order"])
    results.sort(key=key_fn, reverse=(order == "desc"))

    back_url = back or "/hlasovani"
    back_label = "Zpět na přehled" if back == "/" else "Zpět na hlasování"

    has_processed = _has_processed_ballots(voting)
    if voting.status == VotingStatus.ACTIVE:
        detail_step = 5 if has_processed else 3
    elif voting.status == VotingStatus.CLOSED:
        detail_step = 5
    else:  # draft
        detail_step = 2 if voting.items else 1

    ctx = {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "results": results,
        "back_url": back_url,
        "back_label": back_label,
        "active_bubble": "",
        "show_close_voting": has_processed,
        "snapshot_warning": snapshot_warning,
        "q": q,
        "sort": sort,
        "order": order,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, detail_step),
    }

    # Flash messages from query params
    if info == "extrakce-selhala":
        ctx["flash_message"] = "Body hlasování nebyly extrahovány ze šablony (chyba formátu). Přidejte je ručně."
        ctx["flash_type"] = "warning"
    elif info == "sablona-prazdna":
        ctx["flash_message"] = "V šabloně nebyly nalezeny žádné body hlasování. Přidejte je ručně."
        ctx["flash_type"] = "warning"
    elif info == "zadne-body":
        ctx["flash_message"] = "Nejdříve přidejte alespoň jeden bod hlasování."
        ctx["flash_type"] = "error"

    # HTMX partial: return only the results table
    if is_htmx_partial(request):
        return templates.TemplateResponse("voting/detail_results.html", ctx)

    return templates.TemplateResponse("voting/detail.html", ctx)


@router.post("/{voting_id}/generovat")
async def generate_ballots(
    voting_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(joinedload(Voting.items)).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    if not voting.items:
        return RedirectResponse(
            f"/hlasovani/{voting_id}?info=zadne-body", status_code=302
        )

    owners = db.query(Owner).filter_by(is_active=True).options(
        joinedload(Owner.units).joinedload(OwnerUnit.unit)
    ).all()

    created = 0
    processed_owner_ids = set()

    if voting.partial_owner_mode == "shared":
        # Group SJM co-owners: find owners sharing units with SJM ownership
        # Build unit→owners map for SJM units
        unit_sjm_owners = {}  # unit_id → [owner, ...]
        for owner in owners:
            for ou in owner.current_units:
                if ou.ownership_type and "SJM" in ou.ownership_type.upper():
                    unit_sjm_owners.setdefault(ou.unit_id, []).append(owner)

        # Build SJM groups — only pair owners on units with exactly 2 SJM owners
        # Units with >2 SJM owners (e.g. two couples co-owning) don't define pairs
        owner_to_group = {}
        group_id = 0
        for unit_id, group_owners in unit_sjm_owners.items():
            if len(group_owners) != 2:
                continue  # Skip multi-owner SJM units for pairing
            # Same connected components logic, but only for pair units
            existing_groups = {owner_to_group[o.id] for o in group_owners if o.id in owner_to_group}
            if existing_groups:
                target = min(existing_groups)
                for o in group_owners:
                    owner_to_group[o.id] = target
                for oid, gid in list(owner_to_group.items()):
                    if gid in existing_groups and gid != target:
                        owner_to_group[oid] = target
            else:
                for o in group_owners:
                    owner_to_group[o.id] = group_id
                group_id += 1

        # Handle SJM owners who only appear on multi-owner units
        # Group by identical SJM unit sets (likely married couples)
        unpaired_sjm = {}  # owner_id → owner
        for unit_id, group_owners in unit_sjm_owners.items():
            for owner in group_owners:
                if owner.id not in owner_to_group:
                    unpaired_sjm[owner.id] = owner

        # Build SJM unit set for each unpaired owner
        set_to_owners = {}  # frozenset(unit_ids) → [owner_id, ...]
        for oid, owner in unpaired_sjm.items():
            sjm_units = frozenset(
                ou.unit_id for ou in owner.current_units
                if ou.ownership_type and "SJM" in ou.ownership_type.upper()
            )
            set_to_owners.setdefault(sjm_units, []).append(oid)

        # Pair owners with identical SJM unit sets (exactly 2 = couple)
        for sjm_set, oids in set_to_owners.items():
            if len(oids) == 2:
                for oid in oids:
                    owner_to_group[oid] = group_id
                group_id += 1
            else:
                for oid in oids:
                    owner_to_group[oid] = group_id
                    group_id += 1

        # Invert: group_id → [owners]
        groups = {}
        for oid, gid in owner_to_group.items():
            groups.setdefault(gid, set()).add(oid)

        owner_by_id = {o.id: o for o in owners}

        # Create shared ballots for SJM groups
        for gid, member_ids in groups.items():
            members = sorted(
                [owner_by_id[oid] for oid in member_ids if oid in owner_by_id],
                key=lambda o: o.name_normalized or "",
            )
            if not members:
                continue
            primary = members[0]

            # Skip if ballot already exists for primary
            existing = db.query(Ballot).filter_by(
                voting_id=voting.id, owner_id=primary.id
            ).first()
            if existing:
                processed_owner_ids.update(member_ids)
                continue

            # Sum ALL members' votes (each owner has their own share)
            # Dedup only unit numbers for display
            seen_unit_ids = set()
            total_votes = 0
            unit_numbers = []
            for member in members:
                for ou in member.current_units:
                    total_votes += ou.votes
                    if ou.unit_id not in seen_unit_ids:
                        seen_unit_ids.add(ou.unit_id)
                        unit_numbers.append(str(ou.unit.unit_number))

            shared_names = ", ".join(m.display_name for m in members)

            ballot = Ballot(
                voting_id=voting.id,
                owner_id=primary.id,
                total_votes=total_votes,
                units_text=", ".join(unit_numbers),
                shared_owners_text=shared_names,
                status=BallotStatus.GENERATED,
            )
            db.add(ballot)
            db.flush()

            for item in voting.items:
                bv = BallotVote(
                    ballot_id=ballot.id,
                    voting_item_id=item.id,
                    votes_count=total_votes,
                )
                db.add(bv)

            processed_owner_ids.update(member_ids)
            created += 1

    # Create individual ballots for remaining owners (non-SJM or separate mode)
    for owner in owners:
        if owner.id in processed_owner_ids:
            continue

        if not owner.current_units:
            continue  # Skip owners without any units

        existing = db.query(Ballot).filter_by(
            voting_id=voting.id, owner_id=owner.id
        ).first()
        if existing:
            continue

        total_votes = sum(ou.votes for ou in owner.current_units)
        units_text = ", ".join(
            str(ou.unit.unit_number) for ou in owner.current_units
        )

        ballot = Ballot(
            voting_id=voting.id,
            owner_id=owner.id,
            total_votes=total_votes,
            units_text=units_text,
            status=BallotStatus.GENERATED,
        )

        db.add(ballot)
        db.flush()

        for item in voting.items:
            bv = BallotVote(
                ballot_id=ballot.id,
                voting_item_id=item.id,
                votes_count=total_votes,
            )
            db.add(bv)

        created += 1

    if voting.status == VotingStatus.DRAFT:
        voting.status = VotingStatus.ACTIVE

    log_activity(db, ActivityAction.STATUS_CHANGED, "voting", "hlasovani",
                 entity_id=voting.id, entity_name=voting.title,
                 description=f"Vygenerováno {created} lístků, stav → aktivní")
    db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)


@router.post("/{voting_id}/stav")
async def update_voting_status(
    voting_id: int,
    status: str = Form(...),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).get(voting_id)
    if voting:
        old = voting.status
        # Allowed transitions: active→closed, closed→active
        allowed = {
            (VotingStatus.ACTIVE, VotingStatus.CLOSED),
            (VotingStatus.CLOSED, VotingStatus.ACTIVE),
        }
        try:
            new_status = VotingStatus(status)
        except ValueError:
            return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)
        if (old, new_status) not in allowed:
            return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)
        voting.status = new_status
        voting.updated_at = datetime.utcnow()
        log_activity(db, ActivityAction.STATUS_CHANGED, "voting", "hlasovani",
                     entity_id=voting.id, entity_name=voting.title,
                     description=f"Stav: {old} → {status}")
        db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)


@router.get("/{voting_id}/exportovat")
async def voting_export(voting_id: int, stav: str = "", db: Session = Depends(get_db)):
    """Export voting results to Excel. Optional stav filter: generated, sent, processed (default all processed)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    voting = db.query(Voting).options(
        joinedload(Voting.items),
        joinedload(Voting.ballots).joinedload(Ballot.owner),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    declared = _get_declared_shares(db) or 1
    items = sorted(voting.items, key=lambda i: i.order)

    # Filter ballots by status
    status_filter = {
        "generated": BallotStatus.GENERATED,
        "sent": BallotStatus.SENT,
        "processed": BallotStatus.PROCESSED,
    }
    if stav and stav in status_filter:
        processed = [b for b in voting.ballots if b.status == status_filter[stav]]
    else:
        processed = list(voting.ballots)
    processed.sort(key=lambda b: (b.owner.name_normalized or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Výsledky hlasování"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    # Disclaimer for active votings
    start_row = 1
    if voting.status != VotingStatus.CLOSED:
        from openpyxl.styles import Alignment
        warn_cell = ws.cell(row=1, column=1,
                            value=f"Průběžné výsledky ke dni {datetime.utcnow().strftime('%d.%m.%Y')} — hlasování stále probíhá")
        warn_cell.font = Font(bold=True, color="FF6600")
        warn_cell.alignment = Alignment(wrap_text=False)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        start_row = 3

    # Header row
    headers = ["Vlastník", "Jednotky", "Hlasy"]
    for item in items:
        headers.append(f"Bod {item.order}: {item.title}")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill

    # Data rows
    vote_labels = {"for": "PRO", "against": "PROTI", "abstain": "Zdržel se", "invalid": "Neplatný"}
    for row_idx, ballot in enumerate(processed, start_row + 1):
        ws.cell(row=row_idx, column=1, value=ballot.shared_owners_text or ballot.owner.display_name)
        ws.cell(row=row_idx, column=2, value=ballot.units_text or "")
        ws.cell(row=row_idx, column=3, value=ballot.total_votes)
        for item_idx, item in enumerate(items):
            bv = next((v for v in ballot.votes if v.voting_item_id == item.id), None)
            if bv and bv.vote:
                cell = ws.cell(row=row_idx, column=4 + item_idx, value=vote_labels.get(bv.vote.value, ""))
                if bv.vote == VoteValue.FOR:
                    cell.fill = green_fill
                elif bv.vote == VoteValue.AGAINST:
                    cell.fill = red_fill
            else:
                ws.cell(row=row_idx, column=4 + item_idx, value="—")

    # Summary row
    summary_row = start_row + len(processed) + 2
    ws.cell(row=summary_row, column=1, value="CELKEM").font = bold
    for item_idx, item in enumerate(items):
        votes_for = sum(
            bv.votes_count for b in processed for bv in b.votes
            if bv.voting_item_id == item.id and bv.vote == VoteValue.FOR
        )
        votes_against = sum(
            bv.votes_count for b in processed for bv in b.votes
            if bv.voting_item_id == item.id and bv.vote == VoteValue.AGAINST
        )
        pct_for = round(votes_for / declared * 100, 2) if declared else 0
        pct_against = round(votes_against / declared * 100, 2) if declared else 0
        cell = ws.cell(row=summary_row, column=4 + item_idx,
                       value=f"PRO: {votes_for} ({pct_for}%) | PROTI: {votes_against} ({pct_against}%)")
        cell.font = bold

    excel_auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    timestamp = datetime.utcnow().strftime("%Y%m%d")
    stav_labels = {"generated": "nezpracovane", "sent": "odeslane", "processed": "zpracovane"}
    stav_suffix = f"_{stav_labels[stav]}" if stav and stav in stav_labels else "_vsechny"
    filename = f"hlasovani_{voting.id}{stav_suffix}_{timestamp}.xlsx"

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{voting_id}/smazat")
async def voting_delete(voting_id: int, db: Session = Depends(get_db)):
    """Delete a voting, its items, ballots, and associated files."""
    voting = db.query(Voting).options(joinedload(Voting.ballots)).get(voting_id)
    if voting:
        # Remove template file
        if voting.template_path:
            try:
                p = Path(voting.template_path)
                if p.exists():
                    p.unlink()
            except Exception:
                logger.debug("Failed to clean up template: %s", voting.template_path)
        # Remove ballot PDF/scan files
        for ballot in voting.ballots:
            for attr in ("pdf_path", "scan_path"):
                fpath = getattr(ballot, attr, None)
                if fpath:
                    try:
                        p = Path(fpath)
                        if p.exists():
                            p.unlink()
                    except Exception:
                        logger.debug("Failed to clean up ballot file: %s", fpath)
        # Cascade deletes VotingItem, Ballot, BallotVote
        log_activity(db, ActivityAction.DELETED, "voting", "hlasovani",
                     entity_id=voting.id, entity_name=voting.title)
        db.delete(voting)
        db.commit()
    return RedirectResponse("/hlasovani", status_code=302)


@router.post("/{voting_id}/smazat-bod/{item_id}")
async def delete_voting_item(
    voting_id: int,
    item_id: int,
    db: Session = Depends(get_db),
):
    item = db.query(VotingItem).filter_by(id=item_id, voting_id=voting_id).first()
    if item:
        db.delete(item)
        db.flush()
        # Renumber remaining items 1..N
        remaining = (
            db.query(VotingItem)
            .filter_by(voting_id=voting_id)
            .order_by(VotingItem.order)
            .all()
        )
        for idx, it in enumerate(remaining, 1):
            it.order = idx
        db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)


@router.post("/{voting_id}/pridat-bod")
async def add_voting_item(
    voting_id: int,
    title: str = Form(...),
    description: str = Form(""),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(joinedload(Voting.items)).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    max_order = max((i.order for i in voting.items), default=0)
    item = VotingItem(
        voting_id=voting.id,
        order=max_order + 1,
        title=title,
        description=description,
    )
    db.add(item)
    db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)


@router.post("/{voting_id}/bod/{item_id}/upravit")
async def edit_voting_item(
    voting_id: int,
    item_id: int,
    title: str = Form(...),
    description: str = Form(""),
    db: Session = Depends(get_db),
):
    """Edit title/description of a voting item (draft only)."""
    item = db.query(VotingItem).filter_by(id=item_id, voting_id=voting_id).first()
    if not item:
        return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)
    voting = db.query(Voting).get(voting_id)
    if not voting or voting.status != VotingStatus.DRAFT:
        return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)
    item.title = title.strip()
    item.description = description.strip()
    db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)


@router.post("/{voting_id}/bod/{item_id}/posunout")
async def move_voting_item(
    voting_id: int,
    item_id: int,
    direction: str = Form(...),
    db: Session = Depends(get_db),
):
    """Move a voting item up or down (draft only)."""
    voting = db.query(Voting).options(joinedload(Voting.items)).get(voting_id)
    if not voting or voting.status != VotingStatus.DRAFT:
        return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)

    items_sorted = sorted(voting.items, key=lambda i: i.order)
    target = next((i for i in items_sorted if i.id == item_id), None)
    if not target:
        return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)

    idx = items_sorted.index(target)
    if direction == "up" and idx > 0:
        neighbor = items_sorted[idx - 1]
        target.order, neighbor.order = neighbor.order, target.order
    elif direction == "down" and idx < len(items_sorted) - 1:
        neighbor = items_sorted[idx + 1]
        target.order, neighbor.order = neighbor.order, target.order

    db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)
