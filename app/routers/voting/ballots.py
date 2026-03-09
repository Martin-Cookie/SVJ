from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import FileResponse, RedirectResponse
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import (
    Ballot, BallotStatus, BallotVote, Owner, OwnerUnit, Voting,
    VotingStatus, VoteValue,
)
from app.utils import build_list_url, is_htmx_partial, strip_diacritics

from ._helpers import (
    _ballot_stats,
    _has_processed_ballots,
    _voting_wizard,
    templates,
)


router = APIRouter()


@router.get("/{voting_id}/listky")
async def ballot_list(
    voting_id: int,
    request: Request,
    stav: str = Query(""),
    q: str = Query(""),
    sort: str = Query("owner"),
    order: str = Query("asc"),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(
        joinedload(Voting.items),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    # SQL-based ballot query with filters
    needs_owner_join = False
    ballot_query = db.query(Ballot).options(
        joinedload(Ballot.owner).joinedload(Owner.units).joinedload(OwnerUnit.unit),
        joinedload(Ballot.votes),
    ).filter(Ballot.voting_id == voting_id)

    # Filter by status (SQL)
    if stav:
        try:
            stav_enum = BallotStatus(stav)
            ballot_query = ballot_query.filter(Ballot.status == stav_enum)
        except ValueError:
            pass

    # Search filter (SQL via name_normalized)
    if q:
        q_ascii = strip_diacritics(q)
        q_pattern = f"%{q_ascii}%"
        ballot_query = ballot_query.join(Ballot.owner).filter(
            Owner.name_normalized.like(q_pattern)
        )
        needs_owner_join = True

    # SQL sort for simple columns
    BALLOT_SORT_SQL = {"owner", "votes", "status"}
    if sort in BALLOT_SORT_SQL:
        if sort == "owner":
            if not needs_owner_join:
                ballot_query = ballot_query.join(Ballot.owner)
            col = Owner.name_normalized
        elif sort == "votes":
            col = Ballot.total_votes
        else:
            col = Ballot.status
        ballot_query = ballot_query.order_by(
            col.desc().nulls_last() if order == "desc" else col.asc().nulls_last()
        )

    ballots = ballot_query.all()

    # Python-side sort for complex keys (units, proxy, vote columns)
    if sort not in BALLOT_SORT_SQL:
        sort_keys = {
            "units": lambda b: b.units_text or "",
            "proxy": lambda b: (b.proxy_holder_name or ""),
        }
        if sort.startswith("bod_"):
            try:
                item_id = int(sort[4:])
                vote_order = {"for": 0, "against": 1, "abstain": 2}
                sort_keys[sort] = lambda b, _iid=item_id: next(
                    (vote_order.get(bv.vote.value, 3) for bv in b.votes if bv.voting_item_id == _iid and bv.vote),
                    4,
                )
            except ValueError:
                pass
        key_fn = sort_keys.get(sort, lambda b: (b.owner.name_normalized or ""))
        ballots = sorted(ballots, key=key_fn, reverse=(order == "desc"))

    list_url = build_list_url(request)

    has_processed = _has_processed_ballots(voting)
    ctx = {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "ballots": ballots,
        "current_stav": stav,
        "active_bubble": stav or "all",
        "show_close_voting": has_processed,
        "q": q,
        "sort": sort,
        "order": order,
        "list_url": list_url,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, 4 if has_processed else 3),
    }

    # HTMX partial: return only the table
    if is_htmx_partial(request):
        return templates.TemplateResponse("voting/ballots_table.html", ctx)

    return templates.TemplateResponse("voting/ballots.html", ctx)


@router.get("/{voting_id}/listek/{ballot_id}")
async def ballot_detail(
    voting_id: int,
    ballot_id: int,
    request: Request,
    back: str = Query(""),
    db: Session = Depends(get_db),
):
    ballot = db.query(Ballot).options(
        joinedload(Ballot.owner).joinedload(Owner.units).joinedload(OwnerUnit.unit),
        joinedload(Ballot.votes).joinedload(BallotVote.voting_item),
        joinedload(Ballot.voting).joinedload(Voting.items),
    ).filter_by(id=ballot_id, voting_id=voting_id).first()
    if not ballot:
        return RedirectResponse(f"/hlasovani/{voting_id}/listky", status_code=302)

    # Load full voting with ballots for stats/wizard
    voting = db.query(Voting).options(
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)

    back_url = back or f"/hlasovani/{voting_id}/listky"
    back_label = "Zpět na hlasovací lístky"
    has_processed = _has_processed_ballots(voting)

    ctx = {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "ballot": ballot,
        "back_url": back_url,
        "back_label": back_label,
        "active_bubble": "",
        "show_close_voting": has_processed,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, 3),
    }
    return templates.TemplateResponse("voting/ballot_detail.html", ctx)


@router.get("/{voting_id}/zpracovani")
async def process_page(
    voting_id: int,
    request: Request,
    q: str = Query(""),
    sort: str = Query("owner"),
    order: str = Query("asc"),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(
        joinedload(Voting.items),
        joinedload(Voting.ballots).joinedload(Ballot.owner).joinedload(Owner.units).joinedload(OwnerUnit.unit),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    # Direct DB query: ballot IDs that already have any vote recorded
    ballots_with_votes = set(
        row[0] for row in db.query(BallotVote.ballot_id)
        .filter(BallotVote.vote.isnot(None))
        .filter(BallotVote.ballot_id.in_([b.id for b in voting.ballots]))
        .all()
    )

    # Auto-fix: mark ballots with votes but wrong status as PROCESSED
    fixed = False
    for b in voting.ballots:
        if b.id in ballots_with_votes and b.status in (
            BallotStatus.GENERATED, BallotStatus.SENT, BallotStatus.RECEIVED,
        ):
            b.status = BallotStatus.PROCESSED
            b.processed_at = b.processed_at or datetime.utcnow()
            fixed = True
    if fixed:
        db.commit()

    # Get ballots that need processing (not processed, no votes)
    unprocessed = [
        b for b in voting.ballots
        if b.status in (BallotStatus.GENERATED, BallotStatus.SENT, BallotStatus.RECEIVED)
        and b.id not in ballots_with_votes
    ]

    # Hide ballots for inactive owners with no current units
    unprocessed = [
        b for b in unprocessed
        if not (not b.owner.is_active and not b.owner.current_units)
    ]

    # Search filter (diacritics-aware)
    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        unprocessed = [
            b for b in unprocessed
            if q_lower in (b.owner.display_name or "").lower()
            or q_ascii in strip_diacritics(b.owner.display_name or "")
            or q_lower in (b.units_text or "").lower()
        ]

    # Sort
    sort_keys = {
        "owner": lambda b: (b.owner.name_normalized or "").lower(),
        "units": lambda b: b.units_text or "",
        "votes": lambda b: b.total_votes,
    }
    key_fn = sort_keys.get(sort, sort_keys["owner"])
    unprocessed.sort(key=key_fn, reverse=(order == "desc"))

    has_processed = _has_processed_ballots(voting)
    ctx = {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "unprocessed": unprocessed,
        "active_bubble": "",
        "q": q,
        "sort": sort,
        "order": order,
        "show_close_voting": has_processed,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, 3),
    }

    if is_htmx_partial(request):
        return templates.TemplateResponse("voting/process_cards.html", ctx)

    return templates.TemplateResponse("voting/process.html", ctx)


@router.post("/{voting_id}/zpracovat/{ballot_id}")
async def process_ballot(
    voting_id: int,
    ballot_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    form_data = await request.form()
    ballot = db.query(Ballot).options(joinedload(Ballot.votes)).get(ballot_id)
    if not ballot or ballot.voting_id != voting_id:
        return RedirectResponse(f"/hlasovani/{voting_id}/zpracovani", status_code=302)

    # Check voting is active
    voting = db.query(Voting).get(voting_id)
    if not voting or voting.status != VotingStatus.ACTIVE:
        return RedirectResponse(f"/hlasovani/{voting_id}", status_code=302)

    # Collect votes and validate at least one is set
    has_any_vote = False
    for bv in ballot.votes:
        vote_key = f"vote_{bv.voting_item_id}"
        vote_value = form_data.get(vote_key)
        if vote_value:
            bv.vote = VoteValue(vote_value)
            bv.manually_verified = True
            has_any_vote = True

    if not has_any_vote:
        db.rollback()
        if request.headers.get("HX-Request"):
            voting = db.query(Voting).options(joinedload(Voting.items)).get(voting_id)
            return templates.TemplateResponse("partials/ballot_vote_error.html", {
                "request": request,
                "ballot": ballot,
                "voting": voting,
                "error": "Vyberte hlas alespoň u jednoho bodu.",
            })
        return RedirectResponse(f"/hlasovani/{voting_id}/zpracovani", status_code=302)

    ballot.status = BallotStatus.PROCESSED
    ballot.processed_at = datetime.utcnow()
    db.commit()

    if request.headers.get("HX-Request"):
        return templates.TemplateResponse("partials/ballot_processed.html", {
            "request": request,
            "ballot": ballot,
        })
    # If submitted from ballot detail, redirect there; otherwise to bulk processing
    referer = request.headers.get("referer", "")
    if f"/listek/{ballot_id}" in referer:
        return RedirectResponse(f"/hlasovani/{voting_id}/listek/{ballot_id}", status_code=302)
    return RedirectResponse(f"/hlasovani/{voting_id}/zpracovani", status_code=302)


@router.post("/{voting_id}/zpracovat-hromadne")
async def process_ballots_bulk(
    voting_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Batch process multiple ballots with the same votes."""
    form_data = await request.form()
    ballot_ids_raw = form_data.get("ballot_ids", "")
    ballot_ids = [int(x) for x in ballot_ids_raw.split(",") if x.strip()]

    if not ballot_ids:
        return RedirectResponse(f"/hlasovani/{voting_id}/zpracovani", status_code=302)

    voting = db.query(Voting).options(joinedload(Voting.items)).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    ballots = (
        db.query(Ballot).options(joinedload(Ballot.votes))
        .filter(Ballot.id.in_(ballot_ids), Ballot.voting_id == voting_id)
        .all()
    )

    # Check that at least one vote is selected in bulk form
    has_any_vote = any(form_data.get(f"vote_{item.id}") for item in voting.items)
    if not has_any_vote:
        return RedirectResponse(f"/hlasovani/{voting_id}/zpracovani", status_code=302)

    count = 0
    for ballot in ballots:
        for bv in ballot.votes:
            vote_key = f"vote_{bv.voting_item_id}"
            vote_value = form_data.get(vote_key)
            if vote_value:
                bv.vote = VoteValue(vote_value)
                bv.manually_verified = True
        ballot.status = BallotStatus.PROCESSED
        ballot.processed_at = datetime.utcnow()
        count += 1

    db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}/zpracovani", status_code=302)


@router.post("/{voting_id}/listek/{ballot_id}/opravit")
async def reset_ballot(
    voting_id: int,
    ballot_id: int,
    db: Session = Depends(get_db),
):
    """Reset a processed ballot back to GENERATED so it can be re-processed."""
    ballot = db.query(Ballot).options(
        joinedload(Ballot.votes), joinedload(Ballot.voting),
    ).filter_by(id=ballot_id, voting_id=voting_id).first()
    if not ballot:
        return RedirectResponse(f"/hlasovani/{voting_id}/listky", status_code=302)

    # Only allow reset if voting is active and ballot is processed
    if ballot.voting.status != VotingStatus.ACTIVE or ballot.status != BallotStatus.PROCESSED:
        return RedirectResponse(f"/hlasovani/{voting_id}/listek/{ballot_id}", status_code=302)

    ballot.status = BallotStatus.GENERATED
    ballot.processed_at = None
    for bv in ballot.votes:
        bv.vote = None
        bv.manually_verified = False
    db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}/listek/{ballot_id}", status_code=302)


@router.post("/{voting_id}/listky/hromadny-reset")
async def bulk_reset_ballots(
    voting_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Bulk reset selected processed ballots back to GENERATED."""
    from fastapi import Form as FastForm
    form_data = await request.form()
    ballot_ids_raw = form_data.get("ballot_ids", "")
    ballot_ids = [int(x) for x in ballot_ids_raw.split(",") if x.strip()]

    if not ballot_ids:
        return RedirectResponse(f"/hlasovani/{voting_id}/listky", status_code=302)

    voting = db.query(Voting).get(voting_id)
    if not voting or voting.status != VotingStatus.ACTIVE:
        return RedirectResponse(f"/hlasovani/{voting_id}/listky", status_code=302)

    ballots = (
        db.query(Ballot).options(joinedload(Ballot.votes))
        .filter(Ballot.id.in_(ballot_ids), Ballot.voting_id == voting_id, Ballot.status == BallotStatus.PROCESSED)
        .all()
    )

    for ballot in ballots:
        ballot.status = BallotStatus.GENERATED
        ballot.processed_at = None
        for bv in ballot.votes:
            bv.vote = None
            bv.manually_verified = False

    db.commit()
    return RedirectResponse(f"/hlasovani/{voting_id}/listky", status_code=302)


@router.get("/{voting_id}/listek/{ballot_id}/pdf")
async def ballot_pdf_download(
    voting_id: int,
    ballot_id: int,
    db: Session = Depends(get_db),
):
    """Download the generated PDF for a ballot."""
    ballot = db.query(Ballot).filter_by(id=ballot_id, voting_id=voting_id).first()
    if not ballot or not ballot.pdf_path:
        return RedirectResponse(f"/hlasovani/{voting_id}/listek/{ballot_id}", status_code=302)
    pdf = Path(ballot.pdf_path)
    if not pdf.exists():
        return RedirectResponse(f"/hlasovani/{voting_id}/listek/{ballot_id}", status_code=302)
    return FileResponse(
        str(pdf),
        media_type="application/pdf",
        filename=pdf.name,
    )


@router.get("/{voting_id}/neodevzdane")
async def not_submitted(
    voting_id: int,
    request: Request,
    q: str = Query(""),
    sort: str = Query("owner"),
    order: str = Query("asc"),
    db: Session = Depends(get_db),
):
    voting = db.query(Voting).options(
        joinedload(Voting.ballots).joinedload(Ballot.owner).joinedload(Owner.units).joinedload(OwnerUnit.unit),
        joinedload(Voting.ballots).joinedload(Ballot.votes),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    missing = [
        b for b in voting.ballots
        if b.status not in (BallotStatus.PROCESSED,)
    ]

    # Search filter
    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        missing = [
            b for b in missing
            if q_lower in (b.owner.display_name or "").lower()
            or q_ascii in strip_diacritics(b.owner.display_name or "")
            or q_lower in (b.units_text or "").lower()
            or q_lower in (b.owner.email or "").lower()
        ]

    # Sorting
    SORT_KEYS = {
        "owner": lambda b: strip_diacritics(b.owner.display_name or ""),
        "units": lambda b: (b.units_text or "").lower(),
        "email": lambda b: (b.owner.email or "").lower(),
        "votes": lambda b: b.total_votes or 0,
        "status": lambda b: b.status.value,
    }
    sort_fn = SORT_KEYS.get(sort, SORT_KEYS["owner"])
    missing.sort(key=sort_fn, reverse=(order == "desc"))

    list_url = build_list_url(request)

    has_processed = _has_processed_ballots(voting)
    ctx = {
        "request": request,
        "active_nav": "voting",
        "voting": voting,
        "missing": missing,
        "active_bubble": "neodevzdane",
        "show_close_voting": has_processed,
        "q": q,
        "sort": sort,
        "order": order,
        "list_url": list_url,
        **_ballot_stats(voting, db),
        **_voting_wizard(voting, 4 if has_processed else 3),
    }

    if is_htmx_partial(request):
        return templates.TemplateResponse("voting/not_submitted_table.html", ctx)

    return templates.TemplateResponse("voting/not_submitted.html", ctx)


@router.get("/{voting_id}/neodevzdane/exportovat")
async def export_not_submitted(voting_id: int, db: Session = Depends(get_db)):
    """Export not-submitted ballots to Excel."""
    from io import BytesIO
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font

    voting = db.query(Voting).options(
        joinedload(Voting.ballots).joinedload(Ballot.owner).joinedload(Owner.units).joinedload(OwnerUnit.unit),
    ).get(voting_id)
    if not voting:
        return RedirectResponse("/hlasovani", status_code=302)

    missing = [b for b in voting.ballots if b.status not in (BallotStatus.PROCESSED,)]
    missing.sort(key=lambda b: strip_diacritics(b.owner.display_name or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Neodevzdané lístky"
    bold = Font(bold=True)

    headers = ["Vlastník", "Jednotky", "Email", "Telefon", "Hlasy", "Stav"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold

    for row_idx, ballot in enumerate(missing, 2):
        owner = ballot.owner
        ws.cell(row=row_idx, column=1, value=owner.display_name)
        ws.cell(row=row_idx, column=2, value=ballot.units_text or "")
        ws.cell(row=row_idx, column=3, value=owner.email or "")
        ws.cell(row=row_idx, column=4, value=owner.phone or "")
        ws.cell(row=row_idx, column=5, value=ballot.total_votes or 0)
        ws.cell(row=row_idx, column=6, value=ballot.status.value)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"neodevzdane_{voting.title[:30]}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
