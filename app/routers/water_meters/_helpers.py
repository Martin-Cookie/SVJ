from __future__ import annotations

import logging
import re
from collections import defaultdict
from datetime import date, datetime

from app.models import MeterType
from app.utils import templates

logger = logging.getLogger(__name__)


def compute_consumption(meter) -> float | None:
    """Compute consumption = last reading - second to last reading.

    Returns None if fewer than 2 readings exist.
    """
    if not meter.readings or len(meter.readings) < 2:
        return None
    sorted_r = sorted(meter.readings, key=lambda r: r.reading_date)
    last = sorted_r[-1].value
    prev = sorted_r[-2].value
    if last is None or prev is None:
        return None
    return round(last - prev, 3)


def compute_deviations(meters: list) -> dict:
    """Compute per-meter deviation from average consumption per type.

    Returns dict: {meter_id: {"consumption": float|None, "deviation_pct": float|None}}
    """
    # Compute consumption for each meter
    consumption_map = {}
    for m in meters:
        consumption_map[m.id] = compute_consumption(m)

    # Average per type (SV/TV separately)
    type_sums = {}   # {MeterType: [sum, count]}
    for m in meters:
        c = consumption_map[m.id]
        if c is not None and c >= 0:
            key = m.meter_type
            if key not in type_sums:
                type_sums[key] = [0.0, 0]
            type_sums[key][0] += c
            type_sums[key][1] += 1

    type_avg = {}
    for key, (total, count) in type_sums.items():
        type_avg[key] = total / count if count > 0 else 0

    # Compute deviation for each meter
    result = {}
    for m in meters:
        c = consumption_map[m.id]
        avg = type_avg.get(m.meter_type, 0)
        deviation = None
        if c is not None and avg > 0:
            deviation = round((c - avg) / avg * 100, 1)
        result[m.id] = {
            "consumption": c,
            "deviation_pct": deviation,
        }

    return result


def parse_unit_label(label: str) -> tuple[int | None, str, str]:
    """Parse Techem unit label to (unit_number, section_letter, suffix).

    Formats:
        'A 111'    → (111, 'A', '')
        'B 212 A'  → (212, 'B', 'A')
        'C 143B'   → (143, 'C', 'B')
        'AK 11'    → (11, 'AK', '')
        '0'        → (None, '', '')
        ''         → (None, '', '')
    """
    if not label or label.strip() == "0":
        return None, "", ""
    label = label.strip()
    # Format: "A 111", "B 212 A", "C 143B", "AK 11"
    m = re.match(r"([A-Za-z]+)\s*(\d+)\s*([A-Za-z]*)", label)
    if m:
        return int(m.group(2)), m.group(1).upper(), m.group(3).upper()
    # Fallback: just number "111"
    m = re.match(r"(\d+)", label)
    if m:
        return int(m.group(1)), "", ""
    return None, "", ""


def _parse_header_date(header: str) -> date | None:
    """Parse date from column header like '31.1.25' or '28.2.25'.

    Returns date object or None.
    """
    header = header.strip()
    for fmt in ("%d.%m.%y", "%d.%m.%Y", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(header, fmt).date()
        except ValueError:
            continue
    return None


def parse_techem_xls(file_path: str, mapping: dict | None = None,
                     sheet_name: str | None = None, header_row: int = 1) -> list[dict]:
    """Parse Techem XLS file with monthly reading columns.

    If mapping is provided, uses mapping["fields"] for column indices.
    Otherwise falls back to auto-detection.

    Techem format has:
      - Fixed columns: unit label, user name, meter type, serial, etc.
      - Date columns (13+): monthly readings (31.1.25, 28.2.25, ...)

    Returns list of dicts, one per meter, each with keys:
        unit_label, unit_number, unit_letter, meter_serial,
        meter_type, location, user_name, readings: [{date, value}, ...]
    """
    import xlrd

    book = xlrd.open_workbook(file_path)
    if sheet_name and sheet_name in book.sheet_names():
        sheet = book.sheet_by_name(sheet_name)
    else:
        sheet = book.sheet_by_index(0)

    header_idx = header_row - 1  # xlrd is 0-based
    data_start = header_row      # data starts on next row (0-based = header_row)

    if sheet.nrows <= data_start:
        return []

    headers = [str(sheet.cell_value(header_idx, c)).strip() for c in range(sheet.ncols)]

    # Resolve column indices from mapping or auto-detect
    col = {}
    if mapping and "fields" in mapping:
        mf = mapping["fields"]
        for key in ("unit_label", "meter_serial", "meter_type", "location"):
            if key in mf:
                col[key] = int(mf[key])
    else:
        # Auto-detect fixed columns by header name
        for i, h in enumerate(headers):
            hl = h.lower()
            if "zákazník" in hl and "číslo" in hl.lower().replace("zákazník", "").replace("jednotky", ""):
                col["unit_label"] = i
            elif "jméno" in hl and "2" not in h:
                col.setdefault("user_name", i)
            elif "typ přístroje" in hl or "typ měřidla" in hl:
                col["meter_type"] = i
            elif "číslo přístroje" in hl or "evidenční" in hl:
                col["meter_serial"] = i
            elif "poloha" in hl:
                col["location"] = i
            elif "místnost" in hl and "číslo" not in hl and "zkratka" not in hl:
                col.setdefault("room", i)

        # Fallback: exact name matching for critical columns
        exact_map = {
            "Číslo jednotky Zákazník": "unit_label",
            "Jméno uživatele": "user_name",
            "Typ přístroje": "meter_type",
            "Číslo přístroje": "meter_serial",
            "Poloha": "location",
        }
        for i, h in enumerate(headers):
            for name, key in exact_map.items():
                if h == name and key not in col:
                    col[key] = i

    # Detect date columns (monthly readings) — always auto-detected from headers
    date_columns = []  # [(col_index, date_object), ...]
    for i, h in enumerate(headers):
        d = _parse_header_date(h)
        if d:
            date_columns.append((i, d))

    # Sort date columns chronologically
    date_columns.sort(key=lambda x: x[1])

    rows = []
    for r in range(data_start, sheet.nrows):
        # Unit label
        unit_idx = col.get("unit_label", 7)
        raw_label = str(sheet.cell_value(r, unit_idx)).strip()
        unit_number, unit_letter, unit_suffix = parse_unit_label(raw_label)

        # User name
        name_idx = col.get("user_name", 1)
        user_name = str(sheet.cell_value(r, name_idx)).strip()

        # Meter serial
        serial_idx = col.get("meter_serial", 11)
        meter_serial = str(sheet.cell_value(r, serial_idx)).strip()
        # Remove trailing .0 from numbers read as float
        if meter_serial.endswith(".0"):
            meter_serial = meter_serial[:-2]

        # Meter type
        type_idx = col.get("meter_type", 10)
        raw_type = str(sheet.cell_value(r, type_idx)).strip().upper()
        if "SV" in raw_type or "STUD" in raw_type:
            meter_type = "cold"
        elif "TV" in raw_type or "TEPL" in raw_type:
            meter_type = "hot"
        else:
            meter_type = "cold"

        # Location
        loc_idx = col.get("location")
        location = str(sheet.cell_value(r, loc_idx)).strip() if loc_idx is not None else ""

        # Skip rows without meter serial
        if not meter_serial:
            continue

        # Parse monthly readings
        readings = []
        for col_idx, col_date in date_columns:
            raw_val = sheet.cell_value(r, col_idx)
            if isinstance(raw_val, (int, float)) and raw_val != 0:
                readings.append({"date": col_date, "value": float(raw_val)})
            elif isinstance(raw_val, str):
                cleaned = raw_val.replace(",", ".").replace(" ", "").strip()
                if cleaned:
                    try:
                        readings.append({"date": col_date, "value": float(cleaned)})
                    except ValueError:
                        pass

        rows.append({
            "unit_label": raw_label,
            "unit_number": unit_number,
            "unit_letter": unit_letter,
            "meter_serial": meter_serial,
            "meter_type": meter_type,
            "location": location,
            "user_name": user_name,
            "readings": readings,
        })

    return rows


def parse_water_readings_row_format(
    file_path: str, mapping: dict,
    sheet_name: str | None = None, header_row: int = 1,
) -> list[dict]:
    """Parse row-based water meter XLSX (each row = one reading).

    Columns: unit_label, user_name, meter_serial, meter_type, location,
             reading_date, reading_value.
    Rows are grouped by meter_serial into one dict per meter.

    Returns same format as parse_techem_xls:
        [{unit_label, unit_number, unit_letter, meter_serial,
          meter_type, location, user_name, readings: [{date, value}]}, ...]
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    mf = mapping.get("fields", {})
    col = {}
    for key in ("unit_label", "meter_serial", "meter_type",
                "location", "reading_date", "reading_value"):
        if key in mf:
            col[key] = int(mf[key])

    # Read all rows into memory (read_only mode needs iteration)
    all_rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    wb.close()

    # Group by meter serial
    meters: dict[str, dict] = {}  # serial → meter dict
    for row in all_rows:
        if not row or len(row) <= max(col.values()):
            continue

        # Meter serial
        serial_raw = row[col["meter_serial"]]
        if serial_raw is None:
            continue
        serial = str(serial_raw).strip()
        if serial.endswith(".0"):
            serial = serial[:-2]
        if not serial or serial == "None":
            continue

        # Parse reading
        rd_raw = row[col.get("reading_date", 0)]
        rv_raw = row[col.get("reading_value", 0)]

        reading_date = None
        if isinstance(rd_raw, datetime):
            reading_date = rd_raw.date()
        elif isinstance(rd_raw, date):
            reading_date = rd_raw
        elif isinstance(rd_raw, str):
            reading_date = _parse_header_date(rd_raw)

        reading_value = None
        if isinstance(rv_raw, (int, float)):
            reading_value = float(rv_raw)
        elif isinstance(rv_raw, str):
            cleaned = rv_raw.replace(",", ".").replace(" ", "").strip()
            if cleaned:
                try:
                    reading_value = float(cleaned)
                except ValueError:
                    pass

        if serial not in meters:
            # Parse unit label
            ul_raw = row[col.get("unit_label", 0)]
            raw_label = str(ul_raw).strip() if ul_raw is not None else ""
            unit_number, unit_letter, unit_suffix = parse_unit_label(raw_label)

            # Meter type
            mt_raw = row[col.get("meter_type", 0)]
            raw_type = str(mt_raw).strip().upper() if mt_raw is not None else ""
            if "SV" in raw_type or "STUD" in raw_type:
                meter_type = "cold"
            elif "TV" in raw_type or "TEPL" in raw_type:
                meter_type = "hot"
            else:
                meter_type = "cold"

            # Location
            loc_raw = row[col["location"]] if "location" in col else None
            location = str(loc_raw).strip() if loc_raw is not None else ""

            meters[serial] = {
                "unit_label": raw_label,
                "unit_number": unit_number,
                "unit_letter": unit_letter,
                "meter_serial": serial,
                "meter_type": meter_type,
                "location": location,
                "readings": [],
            }

        # Add reading
        if reading_date is not None and reading_value is not None:
            meters[serial]["readings"].append({
                "date": reading_date,
                "value": reading_value,
            })

    # Sort readings chronologically per meter
    for m in meters.values():
        m["readings"].sort(key=lambda r: r["date"])

    return list(meters.values())
