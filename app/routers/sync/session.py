import logging
import shutil
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, File, Query, Request, UploadFile
from fastapi.responses import FileResponse, RedirectResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import and_, cast, func, Integer, or_
from sqlalchemy.orm import Session, joinedload

from app.config import settings
from app.database import get_db
from app.models import (
    Owner, OwnerUnit, ShareCheckSession, SyncRecord, SyncResolution,
    SyncSession, SyncStatus, Unit,
)
from app.services.csv_comparator import compare_owners, parse_sousede_csv
from app.utils import (
    UPLOAD_LIMITS, build_list_url, excel_auto_width, is_htmx_partial,
    strip_diacritics, templates, validate_upload,
)
from ._helpers import SYNC_SORT_COLUMNS

logger = logging.getLogger(__name__)

router = APIRouter()


@router.get("/")
async def sync_list(
    request: Request,
    sync_q: str = Query("", alias="sync_q"),
    sync_sort: str = Query("date", alias="sync_sort"),
    sync_order: str = Query("desc", alias="sync_order"),
    sc_q: str = Query("", alias="sc_q"),
    sc_sort: str = Query("date", alias="sc_sort"),
    sc_order: str = Query("desc", alias="sc_order"),
    # Legacy params (HTMX search uses these)
    q: str = Query(""),
    sort: str = Query(""),
    order: str = Query(""),
    back: str = Query("", alias="back"),
    chyba: str = Query("", alias="chyba"),
    db: Session = Depends(get_db),
):
    """Seznam synchronizačních a kontrolních sessions."""
    # Legacy fallback: if old q/sort/order params are used (HTMX partial)
    _sync_q = sync_q or q
    _sync_sort = sync_sort if sync_sort != "date" or not sort else sort
    _sync_order = sync_order if sync_order != "desc" or not order else order

    # --- Sync sessions ---
    sync_sessions = db.query(SyncSession).order_by(SyncSession.created_at.desc()).all()
    if _sync_q:
        q_lower = _sync_q.lower()
        q_ascii = strip_diacritics(_sync_q)
        sync_sessions = [
            s for s in sync_sessions
            if q_lower in (s.csv_filename or "").lower()
            or q_ascii in strip_diacritics(s.csv_filename or "")
            or q_lower in s.created_at.strftime("%d.%m.%Y %H:%M")
        ]

    SYNC_SORT_KEYS = {
        "filename": lambda s: (s.csv_filename or "").lower(),
        "date": lambda s: s.created_at,
        "matches": lambda s: s.total_matches or 0,
        "differences": lambda s: (s.total_differences or 0) + (s.total_missing or 0),
    }
    sort_fn = SYNC_SORT_KEYS.get(_sync_sort, SYNC_SORT_KEYS["date"])
    sync_sessions.sort(key=sort_fn, reverse=(_sync_order == "desc"))

    list_url = build_list_url(request)

    # HTMX partial for sync search
    if is_htmx_partial(request):
        ctx = {
            "request": request,
            "sessions": sync_sessions,
            "list_url": list_url,
            "q": _sync_q,
        }
        return templates.TemplateResponse("partials/sync_list_body.html", ctx)

    # --- Share check sessions (for full page only) ---
    sc_sessions = db.query(ShareCheckSession).order_by(ShareCheckSession.created_at.desc()).all()
    if sc_q:
        q_lower = sc_q.lower()
        q_ascii = strip_diacritics(sc_q)
        sc_sessions = [
            s for s in sc_sessions
            if q_lower in (s.filename or "").lower()
            or q_ascii in strip_diacritics(s.filename or "")
            or q_lower in s.created_at.strftime("%d.%m.%Y %H:%M")
        ]

    SC_SORT_KEYS = {
        "filename": lambda s: (s.filename or "").lower(),
        "date": lambda s: s.created_at,
        "matches": lambda s: s.total_matches or 0,
        "differences": lambda s: (s.total_differences or 0) + (s.total_missing_db or 0) + (s.total_missing_file or 0),
    }
    sc_sort_fn = SC_SORT_KEYS.get(sc_sort, SC_SORT_KEYS["date"])
    sc_sessions.sort(key=sc_sort_fn, reverse=(sc_order == "desc"))

    # Flash messages from query params
    _CHYBA_MSG = {
        "upload": "Nepodporovaný formát nebo příliš velký soubor.",
        "soubor": "Nebyl vybrán žádný soubor.",
        "cesta": "Neplatná cesta k souboru.",
        "hlavicky": "Nelze přečíst hlavičky souboru.",
        "prazdny": "Soubor neobsahuje žádná data k porovnání.",
    }
    flash_message = _CHYBA_MSG.get(chyba, "")
    flash_type = "error" if flash_message else ""

    ctx = {
        "request": request,
        "active_nav": "kontroly",
        "sync_sessions": sync_sessions,
        "sc_sessions": sc_sessions,
        "back_url": back,
        "list_url": list_url,
        "sync_q": _sync_q,
        "sync_sort": _sync_sort,
        "sync_order": _sync_order,
        "sc_q": sc_q,
        "sc_sort": sc_sort,
        "sc_order": sc_order,
        # Legacy: partials use "sessions" and "q"
        "sessions": sync_sessions,
        "q": _sync_q,
        "flash_message": flash_message,
        "flash_type": flash_type,
    }

    return templates.TemplateResponse("sync/index.html", ctx)


@router.post("/{session_id}/smazat")
async def sync_delete(session_id: int, db: Session = Depends(get_db)):
    """Delete a sync session, its records, and the uploaded CSV file."""
    session = db.query(SyncSession).get(session_id)
    if session:
        # Remove CSV file
        try:
            p = Path(session.csv_path)
            if p.exists():
                p.unlink()
        except Exception:
            logger.debug("Failed to clean up CSV: %s", session.csv_path)
        # Cascade deletes SyncRecord entries
        db.delete(session)
        db.commit()
    return RedirectResponse("/synchronizace", status_code=302)


@router.get("/nova")
async def sync_create_page():
    """Přesměrování na stránku synchronizace."""
    return RedirectResponse("/synchronizace", status_code=302)


@router.post("/nova")
async def sync_create(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Nahrání CSV souboru a vytvoření nové synchronizační session."""
    if not file.filename:
        return RedirectResponse("/synchronizace/nova", status_code=302)

    err = await validate_upload(file, **UPLOAD_LIMITS["csv"])
    if err:
        return RedirectResponse("/synchronizace?chyba=upload", status_code=302)

    # Save CSV
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = settings.upload_dir / "csv" / f"{timestamp}_{file.filename}"
    dest.parent.mkdir(parents=True, exist_ok=True)
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)

    # Read CSV content - try multiple encodings
    csv_content = None
    for encoding in ["utf-8", "cp1250", "latin-1"]:
        try:
            with open(dest, "r", encoding=encoding) as f:
                csv_content = f.read()
            break
        except (UnicodeDecodeError, OSError):
            continue

    if not csv_content:
        return RedirectResponse("/synchronizace", status_code=302)

    csv_records = parse_sousede_csv(csv_content)

    # Get current Excel data
    owners = db.query(Owner).filter_by(is_active=True).options(
        joinedload(Owner.units)
    ).all()

    excel_data = []
    for owner in owners:
        for ou in owner.current_units:
            unit_num = str(ou.unit.unit_number)
            excel_data.append({
                "unit_number": unit_num,
                "owner_name": owner.display_name,
                "first_name": owner.first_name,
                "last_name": owner.last_name or "",
                "name_normalized": owner.name_normalized,
                "owner_type": owner.owner_type.value,
                "space_type": ou.unit.space_type or "",
                "podil_scd": ou.unit.podil_scd or 0,
                "ownership_type": ou.ownership_type or "",
            })

    # Compare
    comparison = compare_owners(csv_records, excel_data)

    # Create session and records
    session = SyncSession(
        csv_filename=file.filename,
        csv_path=str(dest),
        total_records=len(comparison),
        total_matches=sum(1 for c in comparison if c["status"] == SyncStatus.MATCH),
        total_name_order=sum(1 for c in comparison if c["status"] == SyncStatus.NAME_ORDER),
        total_differences=sum(1 for c in comparison if c["status"] == SyncStatus.DIFFERENCE),
        total_missing=sum(
            1 for c in comparison
            if c["status"] in (SyncStatus.MISSING_CSV, SyncStatus.MISSING_EXCEL)
        ),
    )
    db.add(session)
    db.flush()

    for comp in comparison:
        record = SyncRecord(
            session_id=session.id,
            unit_number=comp["unit_number"],
            csv_owner_name=comp.get("csv_owner_name"),
            excel_owner_name=comp.get("excel_owner_name"),
            csv_ownership_type=comp.get("csv_ownership_type"),
            excel_ownership_type=comp.get("excel_ownership_type"),
            csv_space_type=comp.get("csv_space_type"),
            excel_space_type=comp.get("excel_space_type"),
            excel_podil_scd=comp.get("excel_podil_scd"),
            csv_share=comp.get("csv_share"),
            csv_email=comp.get("csv_email", ""),
            csv_phone=comp.get("csv_phone", ""),
            status=comp["status"],
            match_details=comp.get("match_details"),
            resolution=(
                SyncResolution.ACCEPTED
                if comp["status"] in (SyncStatus.MATCH, SyncStatus.NAME_ORDER)
                else SyncResolution.PENDING
            ),
        )
        db.add(record)

    db.commit()
    return RedirectResponse(f"/synchronizace/{session.id}", status_code=302)


@router.get("/{session_id}")
async def sync_detail(
    session_id: int,
    request: Request,
    q: str = Query(""),
    filtr: str = Query("", alias="filtr"),
    sort: str = Query("unit", alias="sort"),
    order: str = Query("asc", alias="order"),
    back: str = Query("", alias="back"),
    db: Session = Depends(get_db),
):
    """Detail synchronizační session s porovnáním záznamů."""
    session = db.query(SyncSession).get(session_id)
    if not session:
        return RedirectResponse("/synchronizace", status_code=302)

    # Field difference condition (name, type, ownership, or share differs)
    # Use coalesce so that NULL vs "value" counts as a difference
    field_diff = or_(
        func.coalesce(SyncRecord.csv_owner_name, '') != func.coalesce(SyncRecord.excel_owner_name, ''),
        func.coalesce(SyncRecord.csv_space_type, '') != func.coalesce(SyncRecord.excel_space_type, ''),
        func.coalesce(SyncRecord.csv_ownership_type, '') != func.coalesce(SyncRecord.excel_ownership_type, ''),
        func.coalesce(SyncRecord.excel_podil_scd, 0) != func.coalesce(SyncRecord.csv_share, 0),
    )

    # Podíl-only difference condition
    podil_diff = (
        func.coalesce(SyncRecord.excel_podil_scd, 0) != func.coalesce(SyncRecord.csv_share, 0)
    )

    base = db.query(SyncRecord).filter_by(session_id=session_id)
    # Partial match = owner matched (MATCH status) but some field differs
    total_partial = base.filter(SyncRecord.status == SyncStatus.MATCH, field_diff).count()
    # Full match = MATCH status and no field differences
    total_full_match = session.total_matches - total_partial
    # Records where podíl specifically differs
    total_podil_diff = base.filter(podil_diff).count()
    # Sum of podíl values for differing records only
    podil_diff_sums = db.query(
        func.sum(SyncRecord.excel_podil_scd),
        func.sum(SyncRecord.csv_share),
    ).filter_by(session_id=session_id).filter(podil_diff).one()
    podil_diff_excel = podil_diff_sums[0] or 0
    podil_diff_csv = podil_diff_sums[1] or 0
    # Sum of podíl values (total)
    sums = db.query(
        func.sum(SyncRecord.excel_podil_scd),
        func.sum(SyncRecord.csv_share),
    ).filter_by(session_id=session_id).one()
    total_excel_podil = sums[0] or 0
    total_csv_podil = sums[1] or 0

    # Sum of podíl per filter category
    def _podil_sums(flt):
        r = db.query(
            func.sum(SyncRecord.excel_podil_scd),
            func.sum(SyncRecord.csv_share),
        ).filter_by(session_id=session_id).filter(flt).one()
        return r[0] or 0, r[1] or 0

    match_excel, match_csv = _podil_sums(and_(SyncRecord.status == SyncStatus.MATCH, ~field_diff))
    partial_excel, partial_csv = _podil_sums(and_(SyncRecord.status == SyncStatus.MATCH, field_diff))
    name_order_excel, name_order_csv = _podil_sums(SyncRecord.status == SyncStatus.NAME_ORDER)
    diff_excel, diff_csv = _podil_sums(SyncRecord.status == SyncStatus.DIFFERENCE)
    missing_excel, missing_csv = _podil_sums(SyncRecord.status.in_([SyncStatus.MISSING_CSV, SyncStatus.MISSING_EXCEL]))

    query = base
    if filtr == "match":
        query = query.filter(SyncRecord.status == SyncStatus.MATCH, ~field_diff)
    elif filtr == "partial":
        query = query.filter(SyncRecord.status == SyncStatus.MATCH, field_diff)
    elif filtr == "name_order":
        query = query.filter(SyncRecord.status == SyncStatus.NAME_ORDER)
    elif filtr == "difference":
        query = query.filter(SyncRecord.status == SyncStatus.DIFFERENCE)
    elif filtr == "podil_diff":
        query = query.filter(podil_diff)
    elif filtr == "missing":
        query = query.filter(SyncRecord.status.in_([SyncStatus.MISSING_CSV, SyncStatus.MISSING_EXCEL]))

    # Sorting
    sort_col = SYNC_SORT_COLUMNS.get(sort)
    if sort_col is not None:
        if order == "desc":
            query = query.order_by(sort_col.desc().nulls_last())
        else:
            query = query.order_by(sort_col.asc().nulls_last())
    else:
        query = query.order_by(cast(SyncRecord.unit_number, Integer).asc())
    records = query.all()

    # Search filtering
    if q:
        q_lower = q.lower()
        q_ascii = strip_diacritics(q)
        records = [
            r for r in records
            if q_lower in str(r.unit_number or "")
            or q_ascii in strip_diacritics(r.excel_owner_name or "")
            or q_ascii in strip_diacritics(r.csv_owner_name or "")
            or q_lower in (r.excel_space_type or "").lower()
            or q_lower in (r.excel_ownership_type or "").lower()
        ]

    # Build unit_number → [(owner_id, owner_name), ...] mapping for clickable owner links
    # and unit_number → unit_id mapping for clickable unit links
    owner_map = {}
    unit_map = {}
    owner_units = (
        db.query(OwnerUnit.owner_id, Unit.unit_number, Owner.name_with_titles, Unit.id)
        .join(OwnerUnit.unit)
        .join(Owner, OwnerUnit.owner_id == Owner.id)
        .filter(OwnerUnit.valid_to.is_(None))
        .all()
    )
    for oid, unit_num, oname, unit_id in owner_units:
        short = str(unit_num)
        owner_map.setdefault(short, []).append((oid, oname))
        unit_map[short] = unit_id

    back_url = back or "/synchronizace"
    back_label = "Zpět na přehled" if back == "/" else "Zpět na kontroly" if "/synchronizace" in (back or "") else "Zpět"

    return templates.TemplateResponse("sync/compare.html", {
        "request": request,
        "active_nav": "kontroly",
        "session": session,
        "records": records,
        "filtr": filtr,
        "sort": sort,
        "order": order,
        "q": q,
        "back_url": back_url,
        "back_label": back_label,
        "total_full_match": total_full_match,
        "total_partial": total_partial,
        "total_podil_diff": total_podil_diff,
        "podil_diff_excel": podil_diff_excel,
        "podil_diff_csv": podil_diff_csv,
        "total_excel_podil": total_excel_podil,
        "total_csv_podil": total_csv_podil,
        "match_excel": match_excel,
        "match_csv": match_csv,
        "partial_excel": partial_excel,
        "partial_csv": partial_csv,
        "name_order_excel": name_order_excel,
        "name_order_csv": name_order_csv,
        "diff_excel": diff_excel,
        "diff_csv": diff_csv,
        "missing_excel": missing_excel,
        "missing_csv": missing_csv,
        "owner_map": owner_map,
        "unit_map": unit_map,
    })


@router.post("/{session_id}/exportovat")
async def export_excel(
    session_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Export porovnání synchronizační session do Excelu."""
    form_data = await request.form()
    filtr = form_data.get("filtr", "")

    session = db.query(SyncSession).get(session_id)
    if not session:
        return RedirectResponse("/synchronizace", status_code=302)

    # Apply same filter logic as sync_detail
    field_diff = or_(
        func.coalesce(SyncRecord.csv_owner_name, '') != func.coalesce(SyncRecord.excel_owner_name, ''),
        func.coalesce(SyncRecord.csv_space_type, '') != func.coalesce(SyncRecord.excel_space_type, ''),
        func.coalesce(SyncRecord.csv_ownership_type, '') != func.coalesce(SyncRecord.excel_ownership_type, ''),
        func.coalesce(SyncRecord.excel_podil_scd, 0) != func.coalesce(SyncRecord.csv_share, 0),
    )
    podil_diff = (
        func.coalesce(SyncRecord.excel_podil_scd, 0) != func.coalesce(SyncRecord.csv_share, 0)
    )

    query = db.query(SyncRecord).filter_by(session_id=session_id)
    if filtr == "match":
        query = query.filter(SyncRecord.status == SyncStatus.MATCH, ~field_diff)
    elif filtr == "partial":
        query = query.filter(SyncRecord.status == SyncStatus.MATCH, field_diff)
    elif filtr == "name_order":
        query = query.filter(SyncRecord.status == SyncStatus.NAME_ORDER)
    elif filtr == "difference":
        query = query.filter(SyncRecord.status == SyncStatus.DIFFERENCE)
    elif filtr == "podil_diff":
        query = query.filter(podil_diff)
    elif filtr == "missing":
        query = query.filter(SyncRecord.status.in_([SyncStatus.MISSING_CSV, SyncStatus.MISSING_EXCEL]))

    records = query.order_by(cast(SyncRecord.unit_number, Integer).asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Porovnání"

    headers = [
        "Jednotka", "Vlastník (Evidence)", "Vlastník (CSV)",
        "Typ prostoru (Ev)", "Typ prostoru (CSV)",
        "Vlastnictví (Ev)", "Vlastnictví (CSV)",
        "Podíl (Ev)", "Podíl (CSV)",
        "Shoda", "Stav", "Akce",
    ]
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    _STATUS_LABELS = {
        "match": "Shoda",
        "name_order": "Přeházená jména",
        "difference": "Rozdíl",
        "missing_csv": "Chybí v CSV",
        "missing_excel": "Chybí v evidenci",
    }
    _RESOLUTION_LABELS = {
        "pending": "Čeká",
        "accepted": "OK",
        "rejected": "Zamítnuto",
        "manual_edit": "Upraveno",
        "exchanged": "Výměna",
    }

    diff_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    for r in records:
        match_pct = (r.match_details.split("|")[0].strip() if r.match_details else "")
        row = [
            r.unit_number,
            r.excel_owner_name or "",
            r.csv_owner_name or "",
            r.excel_space_type or "",
            r.csv_space_type or "",
            r.excel_ownership_type or "",
            r.csv_ownership_type or "",
            r.excel_podil_scd,
            r.csv_share,
            match_pct,
            _STATUS_LABELS.get(r.status.value, r.status.value),
            _RESOLUTION_LABELS.get(r.resolution.value, r.resolution.value),
        ]
        ws.append(row)

        # Highlight differing cells
        row_num = ws.max_row
        if (r.excel_owner_name or "") != (r.csv_owner_name or ""):
            ws.cell(row=row_num, column=2).fill = diff_fill
            ws.cell(row=row_num, column=3).fill = diff_fill
        if (r.excel_space_type or "") != (r.csv_space_type or ""):
            ws.cell(row=row_num, column=4).fill = diff_fill
            ws.cell(row=row_num, column=5).fill = diff_fill
        if (r.excel_ownership_type or "") != (r.csv_ownership_type or ""):
            ws.cell(row=row_num, column=6).fill = diff_fill
            ws.cell(row=row_num, column=7).fill = diff_fill
        if (r.excel_podil_scd or 0) != (r.csv_share or 0):
            ws.cell(row=row_num, column=8).fill = diff_fill
            ws.cell(row=row_num, column=9).fill = diff_fill

    excel_auto_width(ws)

    _FILTR_LABELS = {
        "": "vse", "match": "shoda", "partial": "castecna",
        "name_order": "jmena", "difference": "rozdily",
        "podil_diff": "podily", "missing": "chybi",
    }
    suffix = _FILTR_LABELS.get(filtr, filtr or "vse")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"porovnani_{suffix}_{timestamp}.xlsx"
    output_path = settings.generated_dir / "exports" / filename
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return FileResponse(
        str(output_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
