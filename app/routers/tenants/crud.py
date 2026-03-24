import csv
import io
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, Form, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import Owner, OwnerType, Space, SpaceTenant, Tenant
from app.utils import (
    build_list_url, build_name_with_titles, excel_auto_width,
    is_htmx_partial, is_valid_email, strip_diacritics, templates, utcnow,
)

from ._helpers import _filter_tenants, _tenant_stats, logger

router = APIRouter()


# ── Create ────────────────────────────────────────────────────────────


@router.get("/novy-formular")
async def tenant_create_form(request: Request):
    """Formulář pro vytvoření nového nájemce."""
    return templates.TemplateResponse("tenants/partials/_create_form.html", {
        "request": request,
    })


@router.post("/novy")
async def tenant_create(
    request: Request,
    first_name: str = Form(""),
    last_name: str = Form(""),
    title: str = Form(""),
    tenant_type: str = Form("physical"),
    birth_number: str = Form(""),
    company_id: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    """Vytvoření nového nájemce."""
    first_name = first_name.strip()
    last_name = last_name.strip()

    if not first_name and not last_name:
        return templates.TemplateResponse("tenants/partials/_create_form.html", {
            "request": request,
            "error": "Jméno nebo příjmení je povinné.",
        })

    if email.strip() and not is_valid_email(email.strip()):
        return templates.TemplateResponse("tenants/partials/_create_form.html", {
            "request": request,
            "error": "Neplatný formát emailu.",
        })

    name_with_titles = build_name_with_titles(title.strip(), first_name, last_name)
    name_normalized = strip_diacritics(f"{last_name} {first_name}".strip())

    t_type = OwnerType.LEGAL_ENTITY if tenant_type == "legal" else OwnerType.PHYSICAL

    tenant = Tenant(
        first_name=first_name or None,
        last_name=last_name or None,
        title=title.strip() or None,
        name_with_titles=name_with_titles,
        name_normalized=name_normalized,
        tenant_type=t_type,
        birth_number=birth_number.strip() or None,
        company_id=company_id.strip() or None,
        phone=phone.strip() or None,
        email=email.strip() or None,
        note=note.strip() or None,
        data_source="manual",
        is_active=True,
        created_at=utcnow(),
    )
    db.add(tenant)
    db.commit()

    if request.headers.get("HX-Request"):
        return HTMLResponse(
            content=f'<p class="text-sm text-green-600 p-4">Nájemce {tenant.display_name} vytvořen. '
                    f'<a href="/najemci/{tenant.id}" class="text-blue-600 hover:underline">Zobrazit</a></p>',
        )
    return RedirectResponse(f"/najemci/{tenant.id}", status_code=302)


# ── Per-section inline edit — Identity ───────────────────────────────


@router.get("/{tenant_id}/identita-formular")
async def tenant_identity_form(
    tenant_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Formulář editace identifikace nájemce."""
    tenant = db.query(Tenant).options(joinedload(Tenant.owner)).get(tenant_id)
    if not tenant:
        return RedirectResponse("/najemci", status_code=302)
    return templates.TemplateResponse("tenants/partials/_tenant_identity_form.html", {
        "request": request, "tenant": tenant,
    })


@router.get("/{tenant_id}/identita-info")
async def tenant_identity_info(
    tenant_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Display identifikace nájemce."""
    tenant = db.query(Tenant).options(joinedload(Tenant.owner)).get(tenant_id)
    if not tenant:
        return RedirectResponse("/najemci", status_code=302)
    return templates.TemplateResponse("tenants/partials/_tenant_identity_info.html", {
        "request": request, "tenant": tenant,
    })


@router.post("/{tenant_id}/identita-upravit")
async def tenant_identity_update(
    tenant_id: int,
    request: Request,
    tenant_type: str = Form("physical"),
    title: str = Form(""),
    last_name: str = Form(""),
    first_name_physical: str = Form(""),
    first_name_legal: str = Form(""),
    birth_number: str = Form(""),
    company_id: str = Form(""),
    db: Session = Depends(get_db),
):
    """Uložení identifikace nájemce."""
    tenant = db.query(Tenant).options(joinedload(Tenant.owner)).get(tenant_id)
    if not tenant:
        return RedirectResponse("/najemci", status_code=302)
    if tenant.is_linked:
        return templates.TemplateResponse("tenants/partials/_tenant_identity_info.html", {
            "request": request, "tenant": tenant,
            "error": "Propojený nájemce — upravte údaje v detailu vlastníka.",
        })

    is_legal = tenant_type == "legal"
    first_name = first_name_legal.strip() if is_legal else first_name_physical.strip()
    last_name = last_name.strip()

    if not first_name and not last_name:
        return templates.TemplateResponse("tenants/partials/_tenant_identity_form.html", {
            "request": request, "tenant": tenant,
            "error": "Jméno nebo příjmení je povinné.",
        })

    tenant.tenant_type = OwnerType.LEGAL_ENTITY if is_legal else OwnerType.PHYSICAL
    tenant.first_name = first_name or None
    tenant.last_name = last_name or None
    tenant.title = title.strip() or None
    tenant.name_with_titles = build_name_with_titles(title.strip(), first_name, last_name)
    tenant.name_normalized = strip_diacritics(f"{last_name} {first_name}".strip())
    tenant.birth_number = birth_number.strip() or None
    tenant.company_id = company_id.strip() or None
    tenant.updated_at = utcnow()
    db.commit()

    return templates.TemplateResponse("tenants/partials/_tenant_identity_info.html", {
        "request": request, "tenant": tenant, "saved": True,
    })


# ── Per-section inline edit — Contacts ───────────────────────────────


@router.get("/{tenant_id}/kontakt-formular")
async def tenant_contact_form(
    tenant_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Formulář editace kontaktů nájemce."""
    tenant = db.query(Tenant).options(joinedload(Tenant.owner)).get(tenant_id)
    if not tenant:
        return RedirectResponse("/najemci", status_code=302)
    return templates.TemplateResponse("tenants/partials/_tenant_contact_form.html", {
        "request": request, "tenant": tenant,
    })


@router.get("/{tenant_id}/kontakt-info")
async def tenant_contact_info(
    tenant_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Display kontaktů nájemce."""
    tenant = db.query(Tenant).options(joinedload(Tenant.owner)).get(tenant_id)
    if not tenant:
        return RedirectResponse("/najemci", status_code=302)
    return templates.TemplateResponse("tenants/partials/_tenant_contact_info.html", {
        "request": request, "tenant": tenant,
    })


@router.post("/{tenant_id}/kontakt-upravit")
async def tenant_contact_update(
    tenant_id: int,
    request: Request,
    email: str = Form(""),
    email_secondary: str = Form(""),
    phone: str = Form(""),
    phone_secondary: str = Form(""),
    phone_landline: str = Form(""),
    db: Session = Depends(get_db),
):
    """Uložení kontaktů nájemce."""
    tenant = db.query(Tenant).options(joinedload(Tenant.owner)).get(tenant_id)
    if not tenant:
        return RedirectResponse("/najemci", status_code=302)
    if tenant.is_linked:
        return templates.TemplateResponse("tenants/partials/_tenant_contact_info.html", {
            "request": request, "tenant": tenant,
            "error": "Propojený nájemce — upravte údaje v detailu vlastníka.",
        })

    if email.strip() and not is_valid_email(email.strip()):
        return templates.TemplateResponse("tenants/partials/_tenant_contact_form.html", {
            "request": request, "tenant": tenant,
            "error": "Neplatný formát emailu.",
        })
    if email_secondary.strip() and not is_valid_email(email_secondary.strip()):
        return templates.TemplateResponse("tenants/partials/_tenant_contact_form.html", {
            "request": request, "tenant": tenant,
            "error": "Neplatný formát sekundárního emailu.",
        })

    tenant.email = email.strip() or None
    tenant.email_secondary = email_secondary.strip() or None
    tenant.phone = phone.strip() or None
    tenant.phone_secondary = phone_secondary.strip() or None
    tenant.phone_landline = phone_landline.strip() or None
    tenant.updated_at = utcnow()
    db.commit()

    return templates.TemplateResponse("tenants/partials/_tenant_contact_info.html", {
        "request": request, "tenant": tenant, "saved": True,
    })


# ── Per-section inline edit — Address ────────────────────────────────


def _address_ctx(tenant, prefix: str, **extra) -> dict:
    """Build template context for address partials."""
    if prefix == "perm":
        return {
            "tenant": tenant, "prefix": "perm", "address_label": "Trvalá adresa",
            "street": tenant.perm_street, "district": tenant.perm_district,
            "city": tenant.perm_city, "zip": tenant.perm_zip,
            "country": tenant.perm_country, **extra,
        }
    return {
        "tenant": tenant, "prefix": "corr", "address_label": "Koresp. adresa",
        "street": tenant.corr_street, "district": tenant.corr_district,
        "city": tenant.corr_city, "zip": tenant.corr_zip,
        "country": tenant.corr_country, **extra,
    }


@router.get("/{tenant_id}/adresa/{prefix}/formular")
async def tenant_address_form(
    tenant_id: int,
    prefix: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Formulář editace adresy nájemce."""
    if prefix not in ("perm", "corr"):
        return RedirectResponse(f"/najemci/{tenant_id}", status_code=302)
    tenant = db.query(Tenant).options(joinedload(Tenant.owner)).get(tenant_id)
    if not tenant:
        return RedirectResponse("/najemci", status_code=302)
    return templates.TemplateResponse("tenants/partials/_tenant_address_form.html", {
        "request": request, **_address_ctx(tenant, prefix),
    })


@router.get("/{tenant_id}/adresa/{prefix}/info")
async def tenant_address_info(
    tenant_id: int,
    prefix: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Display adresy nájemce."""
    if prefix not in ("perm", "corr"):
        return RedirectResponse(f"/najemci/{tenant_id}", status_code=302)
    tenant = db.query(Tenant).options(joinedload(Tenant.owner)).get(tenant_id)
    if not tenant:
        return RedirectResponse("/najemci", status_code=302)
    return templates.TemplateResponse("tenants/partials/_tenant_address_info.html", {
        "request": request, **_address_ctx(tenant, prefix),
    })


@router.post("/{tenant_id}/adresa/{prefix}/upravit")
async def tenant_address_update(
    tenant_id: int,
    prefix: str,
    request: Request,
    street: str = Form(""),
    district: str = Form(""),
    city: str = Form(""),
    zip: str = Form(""),
    country: str = Form(""),
    db: Session = Depends(get_db),
):
    """Uložení adresy nájemce."""
    if prefix not in ("perm", "corr"):
        return RedirectResponse(f"/najemci/{tenant_id}", status_code=302)
    tenant = db.query(Tenant).options(joinedload(Tenant.owner)).get(tenant_id)
    if not tenant:
        return RedirectResponse("/najemci", status_code=302)
    if tenant.is_linked:
        return templates.TemplateResponse("tenants/partials/_tenant_address_info.html", {
            "request": request, **_address_ctx(tenant, prefix),
            "error": "Propojený nájemce — upravte údaje v detailu vlastníka.",
        })

    if prefix == "perm":
        tenant.perm_street = street.strip() or None
        tenant.perm_district = district.strip() or None
        tenant.perm_city = city.strip() or None
        tenant.perm_zip = zip.strip() or None
        tenant.perm_country = country.strip() or None
    else:
        tenant.corr_street = street.strip() or None
        tenant.corr_district = district.strip() or None
        tenant.corr_city = city.strip() or None
        tenant.corr_zip = zip.strip() or None
        tenant.corr_country = country.strip() or None

    tenant.updated_at = utcnow()
    db.commit()

    return templates.TemplateResponse("tenants/partials/_tenant_address_info.html", {
        "request": request, **_address_ctx(tenant, prefix, saved=True),
    })


# ── Link/Unlink owner ────────────────────────────────────────────────


@router.post("/{tenant_id}/propojit")
async def tenant_link_owner(
    tenant_id: int,
    request: Request,
    owner_id: int = Form(...),
    db: Session = Depends(get_db),
):
    """Propojit nájemce s existujícím vlastníkem."""
    tenant = db.query(Tenant).get(tenant_id)
    if not tenant:
        return RedirectResponse("/najemci", status_code=302)

    owner = db.query(Owner).get(owner_id)
    if not owner:
        return RedirectResponse(f"/najemci/{tenant_id}", status_code=302)

    tenant.owner_id = owner.id
    tenant.updated_at = utcnow()
    db.commit()

    return RedirectResponse(f"/najemci/{tenant_id}?flash=linked", status_code=302)


@router.post("/{tenant_id}/odpojit")
async def tenant_unlink_owner(
    tenant_id: int,
    db: Session = Depends(get_db),
):
    """Odpojit nájemce od vlastníka."""
    tenant = db.query(Tenant).options(joinedload(Tenant.owner)).get(tenant_id)
    if not tenant:
        return RedirectResponse("/najemci", status_code=302)

    # Copy data from owner before unlinking
    if tenant.owner:
        o = tenant.owner
        tenant.first_name = tenant.first_name or o.first_name
        tenant.last_name = tenant.last_name or o.last_name
        tenant.title = tenant.title or o.title
        tenant.name_with_titles = tenant.name_with_titles or o.name_with_titles
        tenant.name_normalized = tenant.name_normalized or o.name_normalized
        tenant.tenant_type = tenant.tenant_type or o.owner_type
        tenant.phone = tenant.phone or o.phone
        tenant.phone_secondary = tenant.phone_secondary or o.phone_secondary
        tenant.phone_landline = tenant.phone_landline or o.phone_landline
        tenant.email = tenant.email or o.email
        tenant.email_secondary = tenant.email_secondary or o.email_secondary
        tenant.birth_number = tenant.birth_number or o.birth_number
        tenant.company_id = tenant.company_id or o.company_id
        tenant.perm_street = tenant.perm_street or o.perm_street
        tenant.perm_district = tenant.perm_district or o.perm_district
        tenant.perm_city = tenant.perm_city or o.perm_city
        tenant.perm_zip = tenant.perm_zip or o.perm_zip
        tenant.perm_country = tenant.perm_country or o.perm_country
        tenant.corr_street = tenant.corr_street or o.corr_street
        tenant.corr_district = tenant.corr_district or o.corr_district
        tenant.corr_city = tenant.corr_city or o.corr_city
        tenant.corr_zip = tenant.corr_zip or o.corr_zip
        tenant.corr_country = tenant.corr_country or o.corr_country

    tenant.owner_id = None
    tenant.updated_at = utcnow()
    db.commit()

    return RedirectResponse(f"/najemci/{tenant_id}?flash=unlinked", status_code=302)


@router.post("/{tenant_id}/smazat")
async def tenant_delete(
    tenant_id: int,
    db: Session = Depends(get_db),
):
    """Smazání nájemce."""
    tenant = db.query(Tenant).get(tenant_id)
    if tenant:
        db.delete(tenant)
        db.commit()
    return RedirectResponse("/najemci?flash=deleted", status_code=302)


# ── List ──────────────────────────────────────────────────────────────


@router.get("/")
async def tenant_list(
    request: Request,
    q: str = Query("", alias="q"),
    typ: str = Query("", alias="typ"),
    stav: str = Query("", alias="stav"),
    sort: str = Query("name", alias="sort"),
    order: str = Query("asc", alias="order"),
    back: str = Query("", alias="back"),
    flash: str = Query("", alias="flash"),
    db: Session = Depends(get_db),
):
    """Seznam nájemců s filtry, hledáním a řazením."""
    tenants = _filter_tenants(db, q, typ, stav, sort, order)

    list_url = build_list_url(request)

    if is_htmx_partial(request):
        return templates.TemplateResponse("tenants/partials/_tbody.html", {
            "request": request,
            "tenants": tenants,
            "list_url": list_url,
        })

    stats = _tenant_stats(db)

    flash_message = None
    if flash == "deleted":
        flash_message = "Nájemce byl smazán."

    return templates.TemplateResponse("tenants/list.html", {
        "request": request,
        "active_nav": "tenants",
        "tenants": tenants,
        "list_url": list_url,
        "back_url": back,
        "q": q,
        "typ": typ,
        "stav": stav,
        "sort": sort,
        "order": order,
        "stats": stats,
        "flash_message": flash_message,
    })


# ── Export ─────────────────────────────────────────────────────────────


@router.get("/exportovat/{fmt}")
async def tenant_export(
    fmt: str,
    q: str = Query("", alias="q"),
    typ: str = Query("", alias="typ"),
    stav: str = Query("", alias="stav"),
    sort: str = Query("name", alias="sort"),
    order: str = Query("asc", alias="order"),
    db: Session = Depends(get_db),
):
    """Export filtered tenants to Excel or CSV."""
    if fmt not in ("xlsx", "csv"):
        return RedirectResponse("/najemci", status_code=302)

    tenants = _filter_tenants(db, q, typ, stav, sort, order)

    headers = ["Jméno", "Typ", "RČ/IČ", "Telefon", "Email", "Prostor", "Nájemné", "VS"]
    type_labels = {"physical": "FO", "legal": "PO"}

    def _row(t):
        asr = t.active_space_rel
        return [
            t.display_name,
            type_labels.get((t.resolved_type or OwnerType.PHYSICAL).value, ""),
            t.birth_number or t.company_id or "",
            t.resolved_phone or "",
            t.resolved_email or "",
            asr.space.designation if asr else "",
            asr.monthly_rent if asr else "",
            asr.variable_symbol if asr else "",
        ]

    timestamp = datetime.now().strftime("%Y%m%d")

    typ_labels = {"physical": "fyzicke", "legal": "pravnicke", "linked": "propojeni", "standalone": "vlastni"}
    if typ and typ in typ_labels:
        suffix = f"_{typ_labels[typ]}"
    elif stav:
        suffix = f"_{stav}"
    elif q:
        suffix = "_hledani"
    else:
        suffix = "_vsichni"
    filename = f"najemci{suffix}_{timestamp}"

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Nájemci"
        bold = Font(bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold

        for row_idx, t in enumerate(tenants, 2):
            for col_idx, val in enumerate(_row(t), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        excel_auto_width(ws)

        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    else:
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(headers)
        for t in tenants:
            writer.writerow(_row(t))
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )


# ── Detail (catch-all — must be last) ─────────────────────────────────


@router.get("/{tenant_id}")
async def tenant_detail(
    tenant_id: int,
    request: Request,
    back: str = Query("", alias="back"),
    flash: str = Query("", alias="flash"),
    db: Session = Depends(get_db),
):
    """Detail nájemce s prostorovými vazbami."""
    tenant = db.query(Tenant).options(
        joinedload(Tenant.owner),
        joinedload(Tenant.spaces).joinedload(SpaceTenant.space),
    ).get(tenant_id)
    if not tenant:
        return RedirectResponse("/najemci", status_code=302)

    # Back label
    if "/prostory/" in back:
        back_label = "Zpět na detail prostoru"
    elif back.startswith("/prostory"):
        back_label = "Zpět na seznam prostorů"
    elif "/platby" in back:
        back_label = "Zpět na platby"
    elif back:
        back_label = "Zpět"
    else:
        back_label = "Zpět na seznam nájemců"

    # Active + history spaces
    active_rel = tenant.active_space_rel
    history = [st for st in tenant.spaces if not st.is_active]
    history.sort(key=lambda st: st.contract_end or st.created_at, reverse=True)

    # Owners for link modal
    owners = db.query(Owner).filter(Owner.is_active == True).order_by(Owner.name_normalized).all()  # noqa: E712

    flash_message = None
    if flash == "linked":
        flash_message = "Nájemce propojen s vlastníkem."
    elif flash == "unlinked":
        flash_message = "Propojení s vlastníkem zrušeno."

    return templates.TemplateResponse("tenants/detail.html", {
        "request": request,
        "active_nav": "tenants",
        "tenant": tenant,
        "active_rel": active_rel,
        "history": history,
        "owners": owners,
        "back_url": back or "/najemci",
        "back_label": back_label,
        "flash_message": flash_message,
    })
