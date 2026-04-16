import logging
from contextlib import asynccontextmanager
from pathlib import Path

from fastapi import FastAPI
from fastapi.staticfiles import StaticFiles
from sqlalchemy import text

from app.config import settings
from app.database import Base, engine, SessionLocal

logger = logging.getLogger(__name__)


def _migrate_units_table():
    """Ensure units table has correct schema with INTEGER PRIMARY KEY.

    SQLite only auto-generates rowid for 'INTEGER PRIMARY KEY' (exact syntax).
    If the table was created with bare 'INT' or missing PK, ids will be NULL.
    Fixes this by recreating the table and re-linking owner_units via unit_number.
    """
    with engine.connect() as conn:
        result = conn.execute(text(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='units'"
        )).scalar()
        if not result:
            return
        if "PRIMARY KEY" in result.upper():
            return
        logger.info("Migrating units table: fixing INTEGER PRIMARY KEY")
        conn.execute(text("PRAGMA foreign_keys = OFF"))
        # Create new table with proper PK
        conn.execute(text("""
            CREATE TABLE units_new (
                id INTEGER NOT NULL PRIMARY KEY,
                unit_number INTEGER NOT NULL,
                building_number VARCHAR(20),
                podil_scd INTEGER,
                floor_area FLOAT,
                room_count VARCHAR(20),
                space_type VARCHAR(50),
                section VARCHAR(10),
                orientation_number INTEGER,
                address VARCHAR(200),
                lv_number INTEGER,
                created_at DATETIME
            )
        """))
        # Copy data (new ids will be auto-generated)
        conn.execute(text("""
            INSERT INTO units_new (
                unit_number, building_number, podil_scd, floor_area,
                room_count, space_type, section, orientation_number,
                address, lv_number, created_at
            )
            SELECT CAST(unit_number AS INTEGER), building_number, podil_scd,
                   floor_area, room_count, space_type, section,
                   orientation_number, address, lv_number, created_at
            FROM units
        """))
        # Re-link owner_units: old unit_id was rowid, map via unit_number
        conn.execute(text("""
            UPDATE owner_units SET unit_id = (
                SELECT un.id FROM units_new un
                WHERE un.unit_number = (
                    SELECT CAST(u.unit_number AS INTEGER) FROM units u
                    WHERE u.rowid = owner_units.unit_id
                )
            )
        """))
        conn.execute(text("DROP TABLE units"))
        conn.execute(text("ALTER TABLE units_new RENAME TO units"))
        conn.execute(text("CREATE UNIQUE INDEX ix_units_unit_number ON units (unit_number)"))
        conn.execute(text("CREATE INDEX ix_units_building_number ON units (building_number)"))
        conn.execute(text("CREATE INDEX ix_units_space_type ON units (space_type)"))
        conn.execute(text("CREATE INDEX ix_units_section ON units (section)"))
        conn.execute(text("CREATE INDEX ix_units_lv_number ON units (lv_number)"))
        conn.execute(text("PRAGMA foreign_keys = ON"))
        conn.commit()
        logger.info("Units table migration complete")


def _migrate_owner_units_history():
    """Add valid_from / valid_to columns to owner_units for ownership history."""
    with engine.connect() as conn:
        columns = [
            row[1] for row in
            conn.execute(text("PRAGMA table_info(owner_units)")).fetchall()
        ]
        if "valid_from" not in columns:
            conn.execute(text("ALTER TABLE owner_units ADD COLUMN valid_from DATE"))
            logger.info("Added valid_from column to owner_units")
        if "valid_to" not in columns:
            conn.execute(text("ALTER TABLE owner_units ADD COLUMN valid_to DATE"))
            logger.info("Added valid_to column to owner_units")
        conn.commit()


def _migrate_tax_tables():
    """Add new columns to tax_sessions and tax_distributions for send workflow."""
    with engine.connect() as conn:
        # tax_sessions
        columns = [
            row[1] for row in
            conn.execute(text("PRAGMA table_info(tax_sessions)")).fetchall()
        ]
        for col, ddl in [
            ("send_batch_size", "ALTER TABLE tax_sessions ADD COLUMN send_batch_size INTEGER DEFAULT 10"),
            ("send_batch_interval", "ALTER TABLE tax_sessions ADD COLUMN send_batch_interval INTEGER DEFAULT 5"),
            ("send_scheduled_at", "ALTER TABLE tax_sessions ADD COLUMN send_scheduled_at DATETIME"),
            ("send_status", "ALTER TABLE tax_sessions ADD COLUMN send_status VARCHAR(20) DEFAULT 'DRAFT'"),
            ("test_email_passed", "ALTER TABLE tax_sessions ADD COLUMN test_email_passed BOOLEAN DEFAULT 0"),
            ("test_email_address", "ALTER TABLE tax_sessions ADD COLUMN test_email_address VARCHAR"),
            ("send_confirm_each_batch", "ALTER TABLE tax_sessions ADD COLUMN send_confirm_each_batch BOOLEAN DEFAULT 0"),
        ]:
            if col not in columns:
                conn.execute(text(ddl))
                logger.info("Added %s column to tax_sessions", col)

        # Fix lowercase enum values from earlier migration
        conn.execute(text(
            "UPDATE tax_sessions SET send_status = 'DRAFT' WHERE send_status = 'draft'"
        ))

        # tax_distributions
        columns = [
            row[1] for row in
            conn.execute(text("PRAGMA table_info(tax_distributions)")).fetchall()
        ]
        for col, ddl in [
            ("email_status", "ALTER TABLE tax_distributions ADD COLUMN email_status VARCHAR(20) DEFAULT 'PENDING'"),
            ("email_address_used", "ALTER TABLE tax_distributions ADD COLUMN email_address_used VARCHAR(200)"),
            ("email_error", "ALTER TABLE tax_distributions ADD COLUMN email_error TEXT"),
            ("ad_hoc_name", "ALTER TABLE tax_distributions ADD COLUMN ad_hoc_name VARCHAR(300)"),
            ("ad_hoc_email", "ALTER TABLE tax_distributions ADD COLUMN ad_hoc_email VARCHAR(200)"),
        ]:
            if col not in columns:
                conn.execute(text(ddl))
                logger.info("Added %s column to tax_distributions", col)

        # Fix lowercase enum values from earlier migration
        conn.execute(text(
            "UPDATE tax_distributions SET email_status = 'PENDING' WHERE email_status = 'pending'"
        ))

        conn.commit()


def _migrate_owners_phone_secondary():
    """Add phone_secondary column to owners table."""
    with engine.connect() as conn:
        columns = [
            row[1] for row in
            conn.execute(text("PRAGMA table_info(owners)")).fetchall()
        ]
        if "phone_secondary" not in columns:
            conn.execute(text(
                "ALTER TABLE owners ADD COLUMN phone_secondary VARCHAR(50)"
            ))
            logger.info("Added phone_secondary column to owners")
        conn.commit()


def _migrate_ballots_shared_owners():
    """Add shared_owners_text column to ballots table."""
    with engine.connect() as conn:
        columns = [
            row[1] for row in
            conn.execute(text("PRAGMA table_info(ballots)")).fetchall()
        ]
        if "shared_owners_text" not in columns:
            conn.execute(text(
                "ALTER TABLE ballots ADD COLUMN shared_owners_text VARCHAR(500)"
            ))
            logger.info("Added shared_owners_text column to ballots")
        conn.commit()


def _migrate_svj_info_voting_mapping():
    """Add voting_import_mapping column to svj_info table."""
    with engine.connect() as conn:
        columns = [
            row[1] for row in
            conn.execute(text("PRAGMA table_info(svj_info)")).fetchall()
        ]
        if "voting_import_mapping" not in columns:
            conn.execute(text(
                "ALTER TABLE svj_info ADD COLUMN voting_import_mapping TEXT"
            ))
            logger.info("Added voting_import_mapping column to svj_info")
        conn.commit()


def _migrate_svj_import_mappings():
    """Add owner_import_mapping and contact_import_mapping columns to svj_info table."""
    with engine.connect() as conn:
        columns = [
            row[1] for row in
            conn.execute(text("PRAGMA table_info(svj_info)")).fetchall()
        ]
        for col_name in ("owner_import_mapping", "contact_import_mapping"):
            if col_name not in columns:
                # safe: col_name from hardcoded tuple above
                conn.execute(text(
                    f"ALTER TABLE svj_info ADD COLUMN {col_name} TEXT"
                ))
                logger.info("Added %s column to svj_info", col_name)
        conn.commit()


def _migrate_owners_email_invalid():
    """Add email_invalid + email_invalid_reason columns to owners table."""
    with engine.connect() as conn:
        columns = [
            row[1] for row in
            conn.execute(text("PRAGMA table_info(owners)")).fetchall()
        ]
        if "email_invalid" not in columns:
            conn.execute(text(
                "ALTER TABLE owners ADD COLUMN email_invalid BOOLEAN DEFAULT 0"
            ))
            logger.info("Added email_invalid column to owners")
        if "email_invalid_reason" not in columns:
            conn.execute(text(
                "ALTER TABLE owners ADD COLUMN email_invalid_reason VARCHAR(500)"
            ))
            logger.info("Added email_invalid_reason column to owners")
        conn.commit()


def _migrate_email_log_name_normalized():
    """Add name_normalized column to email_logs and backfill from recipient_name."""
    with engine.connect() as conn:
        columns = [
            row[1] for row in
            conn.execute(text("PRAGMA table_info(email_logs)")).fetchall()
        ]
        if "name_normalized" not in columns:
            conn.execute(text(
                "ALTER TABLE email_logs ADD COLUMN name_normalized VARCHAR(300)"
            ))
            logger.info("Added name_normalized column to email_logs")
            # Backfill — SQLite doesn't have strip_diacritics, do in Python
            from app.utils import strip_diacritics
            rows = conn.execute(text(
                "SELECT id, recipient_name FROM email_logs WHERE recipient_name IS NOT NULL"
            )).fetchall()
            for row_id, name in rows:
                conn.execute(text(
                    "UPDATE email_logs SET name_normalized = :norm WHERE id = :id"
                ), {"norm": strip_diacritics(name), "id": row_id})
            if rows:
                logger.info("Backfilled name_normalized for %d email_logs", len(rows))
            conn.commit()


def _ensure_indexes():
    """Create indexes defined in models that may be missing on existing tables."""
    _INDEXES = [
        # voting.py
        ("ix_votings_status", "votings", "status"),
        ("ix_voting_items_voting_id", "voting_items", "voting_id"),
        ("ix_ballots_voting_id", "ballots", "voting_id"),
        ("ix_ballots_owner_id", "ballots", "owner_id"),
        ("ix_ballots_status", "ballots", "status"),
        ("ix_ballot_votes_ballot_id", "ballot_votes", "ballot_id"),
        ("ix_ballot_votes_voting_item_id", "ballot_votes", "voting_item_id"),
        # administration.py
        ("ix_svj_addresses_svj_info_id", "svj_addresses", "svj_info_id"),
        ("ix_board_members_group", "board_members", "\"group\""),
        # tax.py
        ("ix_tax_documents_session_id", "tax_documents", "session_id"),
        ("ix_tax_distributions_document_id", "tax_distributions", "document_id"),
        ("ix_tax_distributions_owner_id", "tax_distributions", "owner_id"),
        ("ix_tax_distributions_match_status", "tax_distributions", "match_status"),
        ("ix_tax_sessions_send_status", "tax_sessions", "send_status"),
        ("ix_tax_distributions_email_status", "tax_distributions", "email_status"),
        # sync.py
        ("ix_sync_records_session_id", "sync_records", "session_id"),
        ("ix_sync_records_status", "sync_records", "status"),
        ("ix_sync_records_resolution", "sync_records", "resolution"),
        # common.py
        ("ix_email_logs_status", "email_logs", "status"),
        ("ix_email_logs_module", "email_logs", "module"),
        ("ix_email_logs_reference_id", "email_logs", "reference_id"),
        ("ix_email_logs_name_normalized", "email_logs", "name_normalized"),
        # email_bounces
        ("ix_email_bounces_recipient_email", "email_bounces", "recipient_email"),
        ("ix_email_bounces_owner_id", "email_bounces", "owner_id"),
        ("ix_email_bounces_email_log_id", "email_bounces", "email_log_id"),
        ("ix_email_bounces_bounce_type", "email_bounces", "bounce_type"),
        ("ix_email_bounces_module", "email_bounces", "module"),
        ("ix_email_bounces_reference_id", "email_bounces", "reference_id"),
        ("ix_email_bounces_bounced_at", "email_bounces", "bounced_at"),
        ("ix_email_bounces_imap_uid", "email_bounces", "imap_uid"),
        ("ix_email_bounces_imap_message_id", "email_bounces", "imap_message_id"),
        ("ix_email_bounces_created_at", "email_bounces", "created_at"),
        ("ix_owners_email_invalid", "owners", "email_invalid"),
        ("ix_import_logs_import_type", "import_logs", "import_type"),
        # owner_units history
        ("ix_owner_units_valid_from", "owner_units", "valid_from"),
        ("ix_owner_units_valid_to", "owner_units", "valid_to"),
        # share_check.py
        ("ix_share_check_records_session_id", "share_check_records", "session_id"),
        ("ix_share_check_records_status", "share_check_records", "status"),
        ("ix_share_check_records_resolution", "share_check_records", "resolution"),
        # administration.py — code lists
        ("ix_code_list_items_category", "code_list_items", "category"),
        # activity_logs
        ("ix_activity_logs_action", "activity_logs", "action"),
        ("ix_activity_logs_entity_type", "activity_logs", "entity_type"),
        ("ix_activity_logs_module", "activity_logs", "module"),
        ("ix_activity_logs_created_at", "activity_logs", "created_at"),
        # payment.py
        ("ix_variable_symbol_mappings_unit_id", "variable_symbol_mappings", "unit_id"),
        ("ix_variable_symbol_mappings_source", "variable_symbol_mappings", "source"),
        ("ix_unit_balances_unit_id", "unit_balances", "unit_id"),
        ("ix_unit_balances_year", "unit_balances", "year"),
        ("ix_unit_balances_owner_id", "unit_balances", "owner_id"),
        ("ix_prescription_years_year", "prescription_years", "year"),
        ("ix_prescriptions_prescription_year_id", "prescriptions", "prescription_year_id"),
        ("ix_prescriptions_unit_id", "prescriptions", "unit_id"),
        ("ix_prescriptions_variable_symbol", "prescriptions", "variable_symbol"),
        ("ix_prescription_items_prescription_id", "prescription_items", "prescription_id"),
        ("ix_bank_statements_import_status", "bank_statements", "import_status"),
        ("ix_payments_statement_id", "payments", "statement_id"),
        ("ix_payments_date", "payments", "date"),
        ("ix_payments_vs", "payments", "vs"),
        ("ix_payments_match_status", "payments", "match_status"),
        ("ix_payments_unit_id", "payments", "unit_id"),
        ("ix_payments_owner_id", "payments", "owner_id"),
        ("ix_payments_prescription_id", "payments", "prescription_id"),
        # payment_allocations
        ("ix_payment_allocations_payment_id", "payment_allocations", "payment_id"),
        ("ix_payment_allocations_unit_id", "payment_allocations", "unit_id"),
        ("ix_payment_allocations_owner_id", "payment_allocations", "owner_id"),
        ("ix_payment_allocations_prescription_id", "payment_allocations", "prescription_id"),
        ("ix_settlements_year", "settlements", "year"),
        ("ix_settlements_unit_id", "settlements", "unit_id"),
        ("ix_settlements_status", "settlements", "status"),
        ("ix_settlement_items_settlement_id", "settlement_items", "settlement_id"),
        # space.py
        ("ix_spaces_section", "spaces", "section"),
        ("ix_spaces_status", "spaces", "status"),
        ("ix_tenants_owner_id", "tenants", "owner_id"),
        ("ix_tenants_is_active", "tenants", "is_active"),
        ("ix_tenants_name_normalized", "tenants", "name_normalized"),
        ("ix_space_tenants_space_id", "space_tenants", "space_id"),
        ("ix_space_tenants_tenant_id", "space_tenants", "tenant_id"),
        ("ix_space_tenants_is_active", "space_tenants", "is_active"),
        ("ix_space_tenants_variable_symbol", "space_tenants", "variable_symbol"),
        # space_id on payment tables
        ("ix_variable_symbol_mappings_space_id", "variable_symbol_mappings", "space_id"),
        ("ix_prescriptions_space_id", "prescriptions", "space_id"),
        ("ix_payments_space_id", "payments", "space_id"),
        ("ix_payment_allocations_space_id", "payment_allocations", "space_id"),
        ("ix_unit_balances_space_id", "unit_balances", "space_id"),
        # smtp_profiles
        ("ix_smtp_profiles_is_default", "smtp_profiles", "is_default"),
        ("ix_tax_sessions_smtp_profile_id", "tax_sessions", "smtp_profile_id"),
        ("ix_bank_statements_smtp_profile_id", "bank_statements", "smtp_profile_id"),
    ]
    # Složené indexy: (název, tabulka, "sloupec1, sloupec2")
    _COMPOUND_INDEXES = [
        ("ix_email_logs_module_reference", "email_logs", "module, reference_id"),
    ]
    import re
    _SAFE_IDENT = re.compile(r'^"?[a-z_][a-z0-9_]*"?$')
    with engine.connect() as conn:
        for idx_name, table, column in _INDEXES:
            if not all(_SAFE_IDENT.match(p) for p in (idx_name, table, column)):
                logger.warning("Skipping index %s — invalid identifier", idx_name)
                continue
            try:
                conn.execute(text(
                    f"CREATE INDEX IF NOT EXISTS {idx_name} ON {table} ({column})"
                ))
            except Exception as e:
                logger.warning("Index %s creation failed: %s", idx_name, e)
        for idx_name, table, columns in _COMPOUND_INDEXES:
            parts = [c.strip() for c in columns.split(",")]
            if not _SAFE_IDENT.match(idx_name) or not _SAFE_IDENT.match(table) or not all(_SAFE_IDENT.match(p) for p in parts):
                logger.warning("Skipping compound index %s — invalid identifier", idx_name)
                continue
            try:
                conn.execute(text(
                    f"CREATE INDEX IF NOT EXISTS {idx_name} ON {table} ({columns})"
                ))
            except Exception as e:
                logger.warning("Compound index %s creation failed: %s", idx_name, e)
        conn.commit()
    logger.info("Database indexes ensured")


def _seed_code_lists():
    """Populate code_list_items from existing unique values if table is empty."""
    from sqlalchemy.orm import Session as _Session
    from app.models.administration import CodeListItem
    from app.models.owner import Unit, OwnerUnit

    with _Session(engine) as session:
        if session.query(CodeListItem).first() is not None:
            return  # already seeded

        _SEED_SOURCES = [
            ("space_type", Unit, "space_type"),
            ("section", Unit, "section"),
            ("room_count", Unit, "room_count"),
            ("ownership_type", OwnerUnit, "ownership_type"),
        ]
        for category, model, column in _SEED_SOURCES:
            col = getattr(model, column)
            values = (
                session.query(col)
                .filter(col.isnot(None), col != "")
                .distinct()
                .order_by(col)
                .all()
            )
            for idx, (val,) in enumerate(values):
                session.add(CodeListItem(
                    category=category,
                    value=val,
                    order=idx,
                ))
        session.commit()
        logger.info("Code lists seeded from existing data")


_TH = 'style="padding:4px 8px;text-align:{align};border:1px solid #ddd;"'
_TD = 'style="padding:4px 8px;{extra}border:1px solid #ddd;"'
_TABLE = 'style="border-collapse:collapse;width:100%;font-size:13px;margin:8px 0;"'

_WATER_EMAIL_BODY_V2 = (
    "Dobrý den, {{ jmeno }},\n\n"
    "přehled odečtů vodoměrů — jednotka {{ jednotka }}."
    "{% if odecty_sv %}\n\n"
    "<strong>Studená voda:</strong>\n"
    f'<table {_TABLE}>'
    f'<tr style="background:#f0f4f8;">'
    f'<th {_TH.format(align="left")}>Vodoměr</th>'
    f'<th {_TH.format(align="right")}>Předchozí</th>'
    f'<th {_TH.format(align="right")}>Aktuální</th>'
    f'<th {_TH.format(align="right")}>Spotřeba</th>'
    f'<th {_TH.format(align="right")}>Průměr SVJ</th>'
    f'<th {_TH.format(align="center")}>Srovnání</th>'
    '</tr>'
    '{% for m in odecty_sv %}'
    '<tr>'
    f'<td {_TD.format(extra="")}>{{{{ m.cislo }}}}</td>'
    f'<td {_TD.format(extra="text-align:right;")}>{{{{ m.predchozi }}}}</td>'
    f'<td {_TD.format(extra="text-align:right;")}>{{{{ m.aktualni }}}}</td>'
    f'<td {_TD.format(extra="text-align:right;font-weight:bold;")}>{{{{ m.spotreba }}}}</td>'
    f'<td {_TD.format(extra="text-align:right;")}>{{{{ m.prumer }}}}</td>'
    f'<td {_TD.format(extra="text-align:center;")}>{{{{ m.srovnani }}}}</td>'
    '</tr>'
    '{% endfor %}'
    '</table>'
    "{% endif %}"
    "{% if odecty_tv %}\n\n"
    "<strong>Teplá voda:</strong>\n"
    f'<table {_TABLE}>'
    f'<tr style="background:#fff4f0;">'
    f'<th {_TH.format(align="left")}>Vodoměr</th>'
    f'<th {_TH.format(align="right")}>Předchozí</th>'
    f'<th {_TH.format(align="right")}>Aktuální</th>'
    f'<th {_TH.format(align="right")}>Spotřeba</th>'
    f'<th {_TH.format(align="right")}>Průměr SVJ</th>'
    f'<th {_TH.format(align="center")}>Srovnání</th>'
    '</tr>'
    '{% for m in odecty_tv %}'
    '<tr>'
    f'<td {_TD.format(extra="")}>{{{{ m.cislo }}}}</td>'
    f'<td {_TD.format(extra="text-align:right;")}>{{{{ m.predchozi }}}}</td>'
    f'<td {_TD.format(extra="text-align:right;")}>{{{{ m.aktualni }}}}</td>'
    f'<td {_TD.format(extra="text-align:right;font-weight:bold;")}>{{{{ m.spotreba }}}}</td>'
    f'<td {_TD.format(extra="text-align:right;")}>{{{{ m.prumer }}}}</td>'
    f'<td {_TD.format(extra="text-align:center;")}>{{{{ m.srovnani }}}}</td>'
    '</tr>'
    '{% endfor %}'
    '</table>'
    "{% endif %}\n\n"
    "V případě dotazů nás kontaktujte.\n\n"
    "S pozdravem,\nSVJ"
)


def _seed_email_templates():
    """Seed default email templates if missing."""
    from sqlalchemy.orm import Session as _Session
    from app.models.administration import EmailTemplate

    with _Session(engine) as session:
        existing = {t.name for t in session.query(EmailTemplate).all()}

        if "Rozúčtování příjmů" not in existing:
            session.add(EmailTemplate(
                name="Rozúčtování příjmů",
                subject_template="Rozúčtování příjmů za rok {rok}",
                body_template="Dobrý den,\n\nv příloze zasíláme rozúčtování příjmů za rok {rok}.\n\nS pozdravem,\nSVJ",
                order=0,
            ))

        if "Upozornění na nesrovnalost v platbě" not in existing:
            session.add(EmailTemplate(
                name="Upozornění na nesrovnalost v platbě",
                subject_template="Upozornění na nesrovnalost v platbě za {{ mesic_nazev }} {{ rok }}",
                body_template=(
                    "Dobrý den, {{ jmeno }},\n\n"
                    "při zpracování plateb za {{ mesic_nazev }} {{ rok }} jsme zjistili nesrovnalost:\n\n"
                    "{% for chyba in chyby %}"
                    "- {{ chyba }}\n"
                    "{% endfor %}\n"
                    "Vaše platba:\n"
                    "- Datum: {{ datum_platby }}\n"
                    "- Částka: {{ castka_zaplaceno }} Kč\n"
                    "- Variabilní symbol: {{ vs_platby }}\n\n"
                    "Očekávaný předpis:\n"
                    "- {{ entita }}\n"
                    "- Měsíční předpis: {{ castka_predpis }} Kč\n"
                    "- Variabilní symbol: {{ vs_predpisu }}\n\n"
                    "Prosíme o úpravu platby tak, aby odpovídala předpisu. "
                    "V případě dotazů nás kontaktujte.\n\n"
                    "S pozdravem,\n"
                    "{{ svj_nazev }}"
                ),
                order=10,
            ))

        if "Odečty vodoměrů" not in existing:
            session.add(EmailTemplate(
                name="Odečty vodoměrů",
                subject_template="Odečty vodoměrů — {{ jednotka }}",
                body_template=_WATER_EMAIL_BODY_V2,
                order=20,
            ))

        session.commit()
        logger.info("Email templates seeded")


def _migrate_payment_allocations():
    """Vytvořit PaymentAllocation záznamy pro existující napárované platby."""
    from sqlalchemy.orm import Session as _Session
    from app.models.payment import Payment, PaymentAllocation, PaymentMatchStatus

    with _Session(engine) as session:
        # Zjistit zda tabulka existuje a má data
        existing_count = session.query(PaymentAllocation).count()
        if existing_count > 0:
            return  # Už migrováno

        # Pro každou napárovanou platbu s unit_id vytvořit alokaci
        payments = (
            session.query(Payment)
            .filter(Payment.unit_id.isnot(None))
            .filter(Payment.match_status != PaymentMatchStatus.UNMATCHED)
            .all()
        )
        if not payments:
            return

        for p in payments:
            session.add(PaymentAllocation(
                payment_id=p.id,
                unit_id=p.unit_id,
                owner_id=p.owner_id,
                prescription_id=p.prescription_id,
                amount=p.amount,
            ))
        session.commit()
        logger.info("Migrated %d payments → payment_allocations", len(payments))


def _migrate_bank_statement_locked():
    """Přidat sloupec locked_at do bank_statements."""
    with engine.connect() as conn:
        cols = [r[1] for r in conn.execute(text("PRAGMA table_info('bank_statements')")).fetchall()]
        if "locked_at" not in cols:
            conn.execute(text("ALTER TABLE bank_statements ADD COLUMN locked_at DATETIME"))
            conn.commit()
            logger.info("Added locked_at column to bank_statements")


def _migrate_unit_balances_owner():
    """Přidat sloupce owner_id a owner_name do unit_balances + balance_import_mapping do svj_info."""
    with engine.connect() as conn:
        cols = [r[1] for r in conn.execute(text("PRAGMA table_info('unit_balances')")).fetchall()]
        if "owner_id" not in cols:
            conn.execute(text("ALTER TABLE unit_balances ADD COLUMN owner_id INTEGER REFERENCES owners(id)"))
            conn.execute(text("ALTER TABLE unit_balances ADD COLUMN owner_name VARCHAR(300)"))
            conn.commit()
            logger.info("Added owner_id, owner_name columns to unit_balances")
        svj_cols = [r[1] for r in conn.execute(text("PRAGMA table_info('svj_info')")).fetchall()]
        if "balance_import_mapping" not in svj_cols:
            conn.execute(text("ALTER TABLE svj_info ADD COLUMN balance_import_mapping TEXT"))
            conn.commit()
            logger.info("Added balance_import_mapping column to svj_info")


def _migrate_spaces_tables():
    """Add space_id columns to payment tables + space_import_mapping to svj_info."""
    _SPACE_COLUMNS = [
        ("variable_symbol_mappings", "space_id", "INTEGER REFERENCES spaces(id)"),
        ("prescriptions", "space_id", "INTEGER REFERENCES spaces(id)"),
        ("payments", "space_id", "INTEGER REFERENCES spaces(id)"),
        ("payment_allocations", "space_id", "INTEGER REFERENCES spaces(id)"),
        ("unit_balances", "space_id", "INTEGER REFERENCES spaces(id)"),
    ]
    with engine.connect() as conn:
        for table, col, col_type in _SPACE_COLUMNS:
            cols = [r[1] for r in conn.execute(text(f"PRAGMA table_info('{table}')")).fetchall()]
            if col not in cols:
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN {col} {col_type}"))
                logger.info("Added %s column to %s", col, table)
        # space_import_mapping on svj_info
        svj_cols = [r[1] for r in conn.execute(text("PRAGMA table_info('svj_info')")).fetchall()]
        if "space_import_mapping" not in svj_cols:
            conn.execute(text("ALTER TABLE svj_info ADD COLUMN space_import_mapping TEXT"))
            logger.info("Added space_import_mapping column to svj_info")
        # Fix unit_id NOT NULL → nullable on variable_symbol_mappings (needed for space-only VS)
        vsm_cols = conn.execute(text("PRAGMA table_info('variable_symbol_mappings')")).fetchall()
        unit_id_col = next((c for c in vsm_cols if c[1] == "unit_id"), None)
        if unit_id_col and unit_id_col[3] == 1:  # notnull=1 → needs fix
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS _vsm_new (
                    id INTEGER PRIMARY KEY,
                    variable_symbol VARCHAR(20) NOT NULL UNIQUE,
                    unit_id INTEGER REFERENCES units(id),
                    space_id INTEGER REFERENCES spaces(id),
                    source VARCHAR(6),
                    description TEXT,
                    is_active BOOLEAN DEFAULT 1,
                    created_at DATETIME,
                    updated_at DATETIME
                )
            """))
            conn.execute(text("INSERT INTO _vsm_new SELECT id, variable_symbol, unit_id, space_id, source, description, is_active, created_at, updated_at FROM variable_symbol_mappings"))
            conn.execute(text("DROP TABLE variable_symbol_mappings"))
            conn.execute(text("ALTER TABLE _vsm_new RENAME TO variable_symbol_mappings"))
            conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_vsm_variable_symbol ON variable_symbol_mappings(variable_symbol)"))
            conn.execute(text("CREATE INDEX IF NOT EXISTS ix_vsm_unit_id ON variable_symbol_mappings(unit_id)"))
            conn.execute(text("CREATE INDEX IF NOT EXISTS ix_vsm_space_id ON variable_symbol_mappings(space_id)"))
            logger.info("Recreated variable_symbol_mappings with nullable unit_id")
        # Fix unit_id NOT NULL → nullable on payment_allocations (needed for space-only allocations)
        pa_cols = conn.execute(text("PRAGMA table_info('payment_allocations')")).fetchall()
        pa_unit_col = next((c for c in pa_cols if c[1] == "unit_id"), None)
        if pa_unit_col and pa_unit_col[3] == 1:  # notnull=1 → needs fix
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS _pa_new (
                    id INTEGER PRIMARY KEY,
                    payment_id INTEGER NOT NULL REFERENCES payments(id) ON DELETE CASCADE,
                    unit_id INTEGER REFERENCES units(id),
                    space_id INTEGER REFERENCES spaces(id),
                    owner_id INTEGER REFERENCES owners(id),
                    prescription_id INTEGER REFERENCES prescriptions(id),
                    amount FLOAT NOT NULL
                )
            """))
            conn.execute(text("INSERT INTO _pa_new SELECT id, payment_id, unit_id, space_id, owner_id, prescription_id, amount FROM payment_allocations"))
            conn.execute(text("DROP TABLE payment_allocations"))
            conn.execute(text("ALTER TABLE _pa_new RENAME TO payment_allocations"))
            conn.execute(text("CREATE INDEX IF NOT EXISTS ix_pa_payment_id ON payment_allocations(payment_id)"))
            conn.execute(text("CREATE INDEX IF NOT EXISTS ix_pa_unit_id ON payment_allocations(unit_id)"))
            conn.execute(text("CREATE INDEX IF NOT EXISTS ix_pa_space_id ON payment_allocations(space_id)"))
            conn.execute(text("CREATE INDEX IF NOT EXISTS ix_pa_owner_id ON payment_allocations(owner_id)"))
            conn.execute(text("CREATE INDEX IF NOT EXISTS ix_pa_prescription_id ON payment_allocations(prescription_id)"))
            logger.info("Recreated payment_allocations with nullable unit_id")
        conn.commit()


def _migrate_svj_send_settings():
    """Přidat sdílená nastavení odesílání do svj_info."""
    with engine.connect() as conn:
        cols = [r[1] for r in conn.execute(text("PRAGMA table_info('svj_info')")).fetchall()]
        for col_name, col_def in [
            ("send_batch_size", "INTEGER DEFAULT 10"),
            ("send_batch_interval", "INTEGER DEFAULT 5"),
            ("send_confirm_each_batch", "BOOLEAN DEFAULT 0"),
            ("send_test_email_address", "VARCHAR(200)"),
            ("vs_prefix", "VARCHAR(10) DEFAULT '1098'"),
        ]:
            if col_name not in cols:
                conn.execute(text(f"ALTER TABLE svj_info ADD COLUMN {col_name} {col_def}"))
                logger.info("Added %s column to svj_info", col_name)
        conn.commit()


def _migrate_payment_notified_at():
    """Přidat sloupec notified_at do payments a discrepancy_test_passed do bank_statements."""
    with engine.connect() as conn:
        cols = [r[1] for r in conn.execute(text("PRAGMA table_info('payments')")).fetchall()]
        if "notified_at" not in cols:
            conn.execute(text("ALTER TABLE payments ADD COLUMN notified_at DATETIME"))
            logger.info("Added notified_at column to payments")
        bs_cols = [r[1] for r in conn.execute(text("PRAGMA table_info('bank_statements')")).fetchall()]
        if "discrepancy_test_passed" not in bs_cols:
            conn.execute(text("ALTER TABLE bank_statements ADD COLUMN discrepancy_test_passed BOOLEAN DEFAULT 0"))
            logger.info("Added discrepancy_test_passed column to bank_statements")
        conn.commit()


def _migrate_dedupe_tenants():
    """Sloučit duplicitní Tenant záznamy.

    Klíč duplicity (priorita): owner_id → birth_number → company_id → (name_normalized).
    Vítěz = záznam s nejvíce vyplněnými poli. SpaceTenants se přesunou k vítězi,
    ostatní Tenanti se smažou. Idempotentní.
    """
    from sqlalchemy.orm import Session as _Session
    from app.models import Tenant, SpaceTenant
    from app.utils import strip_diacritics

    def _norm_name(s):
        if not s:
            return ""
        s = s.strip()
        if s == "*":
            return ""
        return strip_diacritics(s)

    def _score(t: Tenant) -> int:
        fields = [
            t.first_name, t.last_name, t.title,
            t.birth_number, t.company_id,
            t.phone, t.phone_landline, t.phone_secondary,
            t.email, t.email_secondary,
            t.perm_street, t.perm_city, t.perm_zip,
            t.corr_street, t.corr_city, t.corr_zip,
            t.note,
        ]
        return sum(1 for f in fields if f and str(f).strip() and str(f).strip() != "*")

    def _key(t: Tenant):
        if t.owner_id:
            return ("owner", t.owner_id)
        if t.birth_number and t.birth_number.strip():
            return ("rc", t.birth_number.strip())
        if t.company_id and t.company_id.strip():
            return ("ic", t.company_id.strip())
        ln = _norm_name(t.last_name)
        fn = _norm_name(t.first_name)
        ttype = t.tenant_type.value if t.tenant_type else "physical"
        if ln or fn:
            return ("name", ln, fn, ttype)
        return None  # nelze deduplikovat

    db = _Session(bind=engine)
    try:
        tenants = db.query(Tenant).all()
        groups: dict = {}
        skipped = 0
        for t in tenants:
            k = _key(t)
            if k is None:
                skipped += 1
                continue
            groups.setdefault(k, []).append(t)
        if skipped:
            logger.info(
                "_migrate_dedupe_tenants: přeskočeno %d Tenantů bez dedup klíče (chybí owner_id, RČ/IČ i jméno)",
                skipped,
            )

        merged_count = 0
        for key, group in groups.items():
            if len(group) < 2:
                continue
            group.sort(key=lambda t: (-_score(t), -(t.updated_at.timestamp() if t.updated_at else 0), t.id))
            winner = group[0]
            losers = group[1:]
            for loser in losers:
                # přenos space_tenants
                for st in db.query(SpaceTenant).filter(SpaceTenant.tenant_id == loser.id).all():
                    st.tenant_id = winner.id
                # doplnit prázdná pole vítěze z losera
                for attr in [
                    "first_name", "last_name", "title", "name_with_titles", "name_normalized",
                    "tenant_type", "birth_number", "company_id",
                    "phone", "phone_landline", "phone_secondary", "email", "email_secondary",
                    "perm_street", "perm_district", "perm_city", "perm_zip", "perm_country",
                    "corr_street", "corr_district", "corr_city", "corr_zip", "corr_country",
                    "note",
                ]:
                    wv = getattr(winner, attr, None)
                    lv = getattr(loser, attr, None)
                    if (not wv or (isinstance(wv, str) and (not wv.strip() or wv.strip() == "*"))) and lv:
                        setattr(winner, attr, lv)
                db.delete(loser)
                merged_count += 1
            # opravit „*" v jméně vítěze
            if winner.first_name and winner.first_name.strip() == "*":
                winner.first_name = None
            if winner.last_name and winner.last_name.strip() == "*":
                winner.last_name = None
        if merged_count:
            db.commit()
            logger.info("Sloučeno %d duplicitních Tenant záznamů", merged_count)
        else:
            db.rollback()
    finally:
        db.close()


def _migrate_bank_statement_send_settings():
    """Přidat per-statement nastavení odesílání do bank_statements."""
    with engine.connect() as conn:
        cols = [r[1] for r in conn.execute(text("PRAGMA table_info('bank_statements')")).fetchall()]
        for col_name, col_def in [
            ("send_batch_size", "INTEGER"),
            ("send_batch_interval", "INTEGER"),
            ("send_confirm_each_batch", "BOOLEAN"),
        ]:
            if col_name not in cols:
                conn.execute(text(f"ALTER TABLE bank_statements ADD COLUMN {col_name} {col_def}"))
                logger.info("Added %s column to bank_statements", col_name)
        conn.commit()


def _migrate_smtp_profiles():
    """Vytvořit tabulku smtp_profiles + přidat smtp_profile_id FK do tax_sessions a bank_statements + seed z .env."""
    with engine.connect() as conn:
        # Tabulka — create_all ji vytvoří pro nové DB, ale pro existující potřebujeme explicitně
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS smtp_profiles (
                id INTEGER NOT NULL PRIMARY KEY,
                name VARCHAR(100) NOT NULL,
                smtp_host VARCHAR(255) NOT NULL,
                smtp_port INTEGER DEFAULT 465,
                smtp_user VARCHAR(255) NOT NULL,
                smtp_password_b64 TEXT NOT NULL,
                smtp_from_name VARCHAR(255) DEFAULT '',
                smtp_from_email VARCHAR(255) NOT NULL,
                smtp_use_tls BOOLEAN DEFAULT 1,
                is_default BOOLEAN DEFAULT 0,
                created_at DATETIME,
                updated_at DATETIME
            )
        """))

        # imap_save_sent sloupec
        sp_cols = [r[1] for r in conn.execute(text("PRAGMA table_info('smtp_profiles')")).fetchall()]
        if "imap_save_sent" not in sp_cols:
            conn.execute(text("ALTER TABLE smtp_profiles ADD COLUMN imap_save_sent BOOLEAN DEFAULT 0"))
            logger.info("Added imap_save_sent to smtp_profiles")

        # FK sloupce na tax_sessions
        ts_cols = [r[1] for r in conn.execute(text("PRAGMA table_info('tax_sessions')")).fetchall()]
        if "smtp_profile_id" not in ts_cols:
            conn.execute(text("ALTER TABLE tax_sessions ADD COLUMN smtp_profile_id INTEGER"))
            logger.info("Added smtp_profile_id to tax_sessions")

        # FK sloupce na bank_statements
        bs_cols = [r[1] for r in conn.execute(text("PRAGMA table_info('bank_statements')")).fetchall()]
        if "smtp_profile_id" not in bs_cols:
            conn.execute(text("ALTER TABLE bank_statements ADD COLUMN smtp_profile_id INTEGER"))
            logger.info("Added smtp_profile_id to bank_statements")

        conn.commit()

    # Seed z .env — pouze pokud tabulka je prázdná a .env má SMTP host
    from app.utils import encode_smtp_password, utcnow as _utcnow
    db = SessionLocal()
    try:
        count = db.execute(text("SELECT COUNT(*) FROM smtp_profiles")).scalar()
        if count == 0 and settings.smtp_host not in ("smtp.example.com", ""):
            from app.models.smtp_profile import SmtpProfile
            profile = SmtpProfile(
                name="Hlavní SMTP",
                smtp_host=settings.smtp_host,
                smtp_port=settings.smtp_port,
                smtp_user=settings.smtp_user or "",
                smtp_password_b64=encode_smtp_password(settings.smtp_password) if settings.smtp_password else "",
                smtp_from_name=settings.smtp_from_name or "",
                smtp_from_email=settings.smtp_from_email or "",
                smtp_use_tls=settings.smtp_use_tls,
                is_default=True,
                created_at=_utcnow(),
                updated_at=_utcnow(),
            )
            db.add(profile)
            db.commit()
            logger.info("Seeded SMTP profile from .env: %s", settings.smtp_host)
    except Exception:
        db.rollback()
        logger.warning("SMTP profile seeding failed", exc_info=True)
    finally:
        db.close()


def _migrate_fix_activity_log_modules():
    """Fix activity_logs where profile ID was stored as module instead of 'nastaveni'."""
    db = SessionLocal()
    try:
        db.execute(
            text("""
                UPDATE activity_logs
                SET module = 'nastaveni'
                WHERE entity_type = 'smtp_profile'
                  AND module NOT IN ('nastaveni', 'sprava')
            """)
        )
        db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()


def _migrate_water_meter_import_mapping():
    """Add water_meter_import_mapping column to svj_info table."""
    with engine.connect() as conn:
        cols = [r[1] for r in conn.execute(text("PRAGMA table_info('svj_info')")).fetchall()]
        if "water_meter_import_mapping" not in cols:
            conn.execute(text("ALTER TABLE svj_info ADD COLUMN water_meter_import_mapping TEXT"))
            conn.commit()
            logger.info("Added water_meter_import_mapping column to svj_info")


def _migrate_fix_water_meter_unit_links():
    """Fix water meter unit_id links — re-match by building_number.

    Previous import matched by unit_number (katastrální číslo) instead of
    building_number (číslo jednotky). This corrects the links.
    Uses raw label from last import Excel file when available.
    """
    db = SessionLocal()
    try:
        from app.models import WaterMeter, WaterReading, Unit

        meters = db.query(WaterMeter).all()
        if not meters:
            return

        # Build building_number lookup (normalized)
        from app.routers.water_meters._helpers import normalize_unit_label
        all_units = db.query(Unit).all()
        bn_lookup = {}
        for u in all_units:
            if u.building_number:
                bn_lookup[normalize_unit_label(u.building_number)] = u

        # Try to read raw labels from last imported Excel file
        raw_labels = {}  # meter_serial → raw_label
        import glob as _glob
        upload_dir = str(settings.upload_dir / "water_meters")
        files = sorted(_glob.glob(f"{upload_dir}/*.xlsx"), reverse=True)
        if files:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(files[0], read_only=True, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > 9:
                        label = str(row[3]).strip() if row[3] else ""
                        serial = str(row[9]).strip() if row[9] else ""
                        if serial.endswith(".0"):
                            serial = serial[:-2]
                        if serial and label:
                            raw_labels[serial] = label
                wb.close()
            except Exception as e:
                logger.warning("Could not read Excel for migration: %s", e)

        fixed = 0
        for m in meters:
            # Prefer raw label from Excel, fallback to reconstructed
            raw = raw_labels.get(m.meter_serial, "")
            if not raw:
                if m.unit_letter and m.unit_number:
                    raw = f"{m.unit_letter} {m.unit_number}"
                elif m.unit_number:
                    raw = str(m.unit_number)
            norm = normalize_unit_label(raw)
            if not norm:
                continue

            expected = bn_lookup.get(norm)
            if expected and m.unit_id != expected.id:
                m.unit_id = expected.id
                fixed += 1
            elif not expected and m.unit_id is not None:
                # No match found — unlink wrong assignment
                m.unit_id = None
                fixed += 1

        if fixed:
            db.commit()
            logger.info("Fixed %d water meter unit links (building_number re-match)", fixed)
    finally:
        db.close()


def _migrate_water_meter_unit_suffix():
    """Add unit_suffix column and populate from Excel labels."""
    db = SessionLocal()
    try:
        cur = db.execute(text("PRAGMA table_info(water_meters)"))
        cols = {row[1] for row in cur}
        if "unit_suffix" not in cols:
            db.execute(text("ALTER TABLE water_meters ADD COLUMN unit_suffix VARCHAR(5) DEFAULT ''"))
            db.commit()

        # Populate suffix from last uploaded Excel
        from app.models import WaterMeter
        from app.routers.water_meters._helpers import parse_unit_label
        import glob as _glob

        meters = db.query(WaterMeter).all()
        if not meters:
            return

        # Read labels from latest Excel
        raw_labels = {}  # serial → raw_label
        upload_dir = str(settings.upload_dir / "water_meters")
        files = sorted(_glob.glob(f"{upload_dir}/*.xlsx"), reverse=True)
        if files:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(files[0], read_only=True, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > 9:
                        label = str(row[3]).strip() if row[3] else ""
                        serial = str(row[9]).strip() if row[9] else ""
                        if serial.endswith(".0"):
                            serial = serial[:-2]
                        if serial and label:
                            raw_labels[serial] = label
                wb.close()
            except Exception as e:
                logger.warning("Could not read Excel for suffix migration: %s", e)

        updated = 0
        for m in meters:
            raw = raw_labels.get(m.meter_serial, "")
            if raw:
                number, letter, suffix = parse_unit_label(raw)
                if number is not None and m.unit_number != number:
                    m.unit_number = number
                    updated += 1
                if letter and m.unit_letter != letter:
                    m.unit_letter = letter
                    updated += 1
                if m.unit_suffix != (suffix or ""):
                    m.unit_suffix = suffix
                    updated += 1

        if updated:
            db.commit()
            logger.info("Updated %d water meter suffix/letter values", updated)
    finally:
        db.close()


def _migrate_water_email_template_v2():
    """Aktualizovat email šablonu vodoměrů na verzi s HTML tabulkou."""
    from sqlalchemy.orm import Session as _Session
    from app.models.administration import EmailTemplate

    with _Session(engine) as session:
        tpl = session.query(EmailTemplate).filter_by(name="Odečty vodoměrů").first()
        if not tpl:
            return
        # Aktualizovat jen pokud je stará defaultní šablona (obsahuje statický TV řádek)
        if "<strong>Teplá voda (TV):</strong>" in (tpl.body_template or ""):
            tpl.body_template = _WATER_EMAIL_BODY_V2
            session.commit()
            logger.info("Updated water meter email template to v2 (HTML table)")


def _migrate_water_meter_notified_at():
    """Přidat notified_at do water_meters a water_notified_at do owners."""
    with engine.connect() as conn:
        cols = [r[1] for r in conn.execute(text("PRAGMA table_info('water_meters')")).fetchall()]
        if "notified_at" not in cols:
            conn.execute(text("ALTER TABLE water_meters ADD COLUMN notified_at DATETIME"))
            logger.info("Added notified_at column to water_meters")
        owner_cols = [r[1] for r in conn.execute(text("PRAGMA table_info('owners')")).fetchall()]
        if "water_notified_at" not in owner_cols:
            conn.execute(text("ALTER TABLE owners ADD COLUMN water_notified_at DATETIME"))
            logger.info("Added water_notified_at column to owners")
        # water_test_passed flag on svj_info for gating send
        svj_cols = [r[1] for r in conn.execute(text("PRAGMA table_info('svj_info')")).fetchall()]
        if "water_test_passed" not in svj_cols:
            conn.execute(text("ALTER TABLE svj_info ADD COLUMN water_test_passed BOOLEAN DEFAULT 0"))
            logger.info("Added water_test_passed column to svj_info")
        conn.commit()


_ALL_MIGRATIONS = [
    ("units table", _migrate_units_table),
    ("owner_units history", _migrate_owner_units_history),
    ("tax tables", _migrate_tax_tables),
    ("owners phone_secondary", _migrate_owners_phone_secondary),
    ("ballots shared_owners", _migrate_ballots_shared_owners),
    ("svj_info voting_import_mapping", _migrate_svj_info_voting_mapping),
    ("svj_info import_mappings", _migrate_svj_import_mappings),
    ("email_logs name_normalized", _migrate_email_log_name_normalized),
    ("owners email_invalid", _migrate_owners_email_invalid),
    ("payment_allocations migration", _migrate_payment_allocations),
    ("bank_statement locked_at", _migrate_bank_statement_locked),
    ("unit_balances owner columns", _migrate_unit_balances_owner),
    ("spaces tables migration", _migrate_spaces_tables),
    ("svj send settings", _migrate_svj_send_settings),
    ("payment notified_at", _migrate_payment_notified_at),
    ("dedupe tenants", _migrate_dedupe_tenants),
    ("bank_statement send settings", _migrate_bank_statement_send_settings),
    ("smtp profiles", _migrate_smtp_profiles),
    ("fix activity_log modules", _migrate_fix_activity_log_modules),
    ("water meter import mapping", _migrate_water_meter_import_mapping),
    ("fix water meter unit links", _migrate_fix_water_meter_unit_links),
    ("water meter unit_suffix", _migrate_water_meter_unit_suffix),
    ("water meter notified_at", _migrate_water_meter_notified_at),
    ("water email template v2", _migrate_water_email_template_v2),
    ("index creation", _ensure_indexes),
    ("code list seeding", _seed_code_lists),
    ("email template seeding", _seed_email_templates),
]


def run_post_restore_migrations() -> list[str]:
    """Re-connect to the (possibly replaced) database and run all migrations.

    Called after every backup restore so the server keeps running even when
    the restored DB is missing new columns or tables.

    Returns list of warning messages (empty if all migrations succeeded).
    """
    warnings = []
    engine.dispose()  # drop stale connections to the old file

    # Create any missing tables (e.g. share_check_* from newer code)
    import app.models  # noqa: F401
    Base.metadata.create_all(bind=engine)

    for name, func in _ALL_MIGRATIONS:
        try:
            func()
        except Exception:
            msg = f"post-restore: {name} migration skipped"
            logger.warning(msg)
            warnings.append(msg)

    try:
        from app.routers.tax import recover_stuck_sending_sessions
        recover_stuck_sending_sessions()
    except Exception:
        msg = "post-restore: sending session recovery skipped"
        logger.warning(msg)
        warnings.append(msg)

    return warnings


@asynccontextmanager
async def lifespan(app: FastAPI):
    if settings.debug:
        logger.warning("DEBUG mode je zapnutý — SQL echo aktivní. Pro produkci nastavte DEBUG=false v .env")

    # Import models so they register with Base
    import app.models  # noqa: F401

    Base.metadata.create_all(bind=engine)

    # Run all migrations (shared list with post-restore)
    for name, func in _ALL_MIGRATIONS:
        try:
            func()
        except Exception:
            logger.warning("%s migration skipped", name)

    # Recover stuck SENDING sessions (server restart recovery)
    try:
        from app.routers.tax import recover_stuck_sending_sessions
        recover_stuck_sending_sessions()
    except Exception:
        logger.warning("sending session recovery skipped")

    # Ensure data directories exist
    for d in [settings.upload_dir, settings.generated_dir, settings.temp_dir]:
        d.mkdir(parents=True, exist_ok=True)
    for sub in ["excel", "word_templates", "scanned_ballots", "tax_pdfs", "csv", "share_check", "contracts"]:
        (settings.upload_dir / sub).mkdir(exist_ok=True)
    for sub in ["ballots", "exports"]:
        (settings.generated_dir / sub).mkdir(exist_ok=True)

    yield


app = FastAPI(title=settings.app_name, lifespan=lifespan)


# Custom error pages
from fastapi.templating import Jinja2Templates

_error_templates = Jinja2Templates(directory="app/templates")


from sqlalchemy.exc import IntegrityError, OperationalError  # noqa: E402


@app.exception_handler(IntegrityError)
async def integrity_error_handler(request, exc):
    logger.warning("DB IntegrityError: %s", exc.orig)
    return _error_templates.TemplateResponse(request, "error.html", {
        "status_code": 409,
        "title": "Konflikt dat",
        "message": "Operace nemohla být dokončena — data kolidují s existujícím záznamem.",
    }, status_code=409)


@app.exception_handler(OperationalError)
async def operational_error_handler(request, exc):
    logger.error("DB OperationalError: %s", exc.orig)
    return _error_templates.TemplateResponse(request, "error.html", {
        "status_code": 500,
        "title": "Chyba databáze",
        "message": "Nastala chyba při práci s databází. Zkuste to prosím znovu.",
    }, status_code=500)


@app.exception_handler(404)
async def not_found_handler(request, exc):
    return _error_templates.TemplateResponse(request, "error.html", {
        "status_code": 404,
        "title": "Stránka nenalezena",
        "message": "Požadovaná stránka neexistuje nebo byla přesunuta.",
    }, status_code=404)


@app.exception_handler(500)
async def server_error_handler(request, exc):
    return _error_templates.TemplateResponse(request, "error.html", {
        "status_code": 500,
        "title": "Chyba serveru",
        "message": "Nastala neočekávaná chyba. Zkuste to prosím znovu.",
    }, status_code=500)

# Security headers middleware
@app.middleware("http")
async def add_security_headers(request, call_next):
    response = await call_next(request)
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return response


# Global debtor count for sidebar badge — runs on full-page HTML requests only
@app.middleware("http")
async def inject_debtor_count(request, call_next):
    path = request.url.path
    is_htmx_partial = (
        request.headers.get("hx-request") == "true"
        and request.headers.get("hx-boosted") != "true"
    )
    # Skip static files, HTMX partials, and non-page requests (exports, API)
    if path.startswith("/static") or is_htmx_partial:
        return await call_next(request)
    # Compute debtor count
    try:
        from app.routers.payments._helpers import _count_debtors_fast
        from app.models import PrescriptionYear
        db = SessionLocal()
        try:
            py = db.query(PrescriptionYear).order_by(PrescriptionYear.year.desc()).first()
            request.state.nav_debtor_count = _count_debtors_fast(db, py.year) if py else 0
        finally:
            db.close()
    except Exception:
        request.state.nav_debtor_count = 0
    return await call_next(request)

# Raise default Starlette multipart limits (default max_files=1000 is too low
# for large PDF directories uploaded via webkitdirectory)
try:
    from starlette.requests import Request as _StarletteRequest
    _StarletteRequest.form.__kwdefaults__["max_files"] = 5000
    _StarletteRequest.form.__kwdefaults__["max_fields"] = 5000
except (AttributeError, KeyError, TypeError):
    logging.getLogger(__name__).warning("Cannot override Starlette max_files limit")

# Static files
app.mount("/static", StaticFiles(directory=Path(__file__).parent / "static"), name="static")

# Register routers
from app.routers import dashboard, owners, units, voting, tax, sync, share_check, settings_page, administration, payments, spaces, tenants, bounces, water_meters  # noqa: E402

app.include_router(dashboard.router)
app.include_router(owners.router, prefix="/vlastnici", tags=["Vlastníci"])
app.include_router(units.router, prefix="/jednotky", tags=["Jednotky"])
app.include_router(voting.router, prefix="/hlasovani", tags=["Hlasování"])
app.include_router(tax.router, prefix="/dane", tags=["Daně"])
app.include_router(sync.router, prefix="/synchronizace", tags=["Synchronizace"])
app.include_router(share_check.router, prefix="/kontrola-podilu", tags=["Kontrola podílu"])
app.include_router(administration.router, prefix="/sprava", tags=["Administrace"])
app.include_router(settings_page.router, prefix="/nastaveni", tags=["Nastavení"])
app.include_router(payments.router, prefix="/platby", tags=["Platby"])
app.include_router(spaces.router, prefix="/prostory", tags=["Prostory"])
app.include_router(tenants.router, prefix="/najemci", tags=["Nájemci"])
app.include_router(water_meters.router, prefix="/vodometry", tags=["Vodoměry"])
app.include_router(bounces.router, tags=["Nedoručené emaily"])
